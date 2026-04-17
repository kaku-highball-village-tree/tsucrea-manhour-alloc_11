from __future__ import annotations

import argparse
import csv
import re
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List


INVALID_FILE_CHARS_PATTERN: re.Pattern[str] = re.compile(r'[\\/:*?"<>|]')
YEAR_MONTH_PATTERN: re.Pattern[str] = re.compile(r"(\d{2})\.(\d{1,2})月")
DURATION_TEXT_PATTERN: re.Pattern[str] = re.compile(r"^\s*(\d+)\s+day(?:s)?,\s*(\d+):(\d{2}):(\d{2})\s*$")
TIME_TEXT_PATTERN: re.Pattern[str] = re.compile(r"^\d+:\d{2}:\d{2}$")
HM_PATTERN: re.Pattern[str] = re.compile(r"^(\d+):(\d{2})$")
SALARY_PAYMENT_STEP0001_FILE_PATTERN: re.Pattern[str] = re.compile(r"^支給・控除等一覧表_給与_step0001_.+\.tsv$")
NEW_RAWDATA_STEP0001_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0001_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0002_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0002_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0003_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0003_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0004_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0004_\d{4}年\d{2}月\.tsv$")
SALARY_PAYMENT_DEDUCTION_REQUIRED_HEADERS: tuple[str, ...] = (
    "従業員名",
    "スタッフコード",
    "基本給",
    "課税通勤手当",
    "非課税通勤手当",
    "残業手当",
    "残業時間(60時間以上)",
    "深夜労働手当",
    "休日労働手当",
    "固定残業代",
    "赴任手当",
    "テレワーク手当",
    "プロジェクトリーダー手当",
    "その他支給",
    "欠勤控除",
    "遅刻早退控除",
    "立替経費",
    "その他手当",
    "その他控除",
    "健保事業主負担",
    "介護事業主負担",
    "厚年事業主負担",
    "雇保事業主負担",
    "労災保険料",
    "一般拠出金",
    "子育拠出金",
)
MANAGEMENT_ACCOUNTING_MANHOUR_REQUIRED_HEADERS: tuple[str, ...] = (
    "日時",
    "スタッフコード",
    "姓 名",
    "所属グループ名",
    "スタッフ種別",
    "総労働時間",
    "プロジェクトコード",
    "プロジェクト名",
    "タスクコード",
    "タスク名",
    "工数",
)


def build_candidate_paths(pszInputPath: str) -> List[Path]:
    objInputPath: Path = Path(pszInputPath)
    objScriptDirectoryPath: Path = Path(__file__).resolve().parent
    objInputDirectoryPath: Path = Path.cwd() / "input"
    return [
        objInputPath,
        objScriptDirectoryPath / pszInputPath,
        objInputDirectoryPath / pszInputPath,
    ]


def resolve_existing_input_path(pszInputPath: str) -> Path:
    for objCandidatePath in build_candidate_paths(pszInputPath):
        if objCandidatePath.exists():
            return objCandidatePath
    raise FileNotFoundError(f"Input file not found: {pszInputPath}")


def sanitize_sheet_name_for_file_name(pszSheetName: str) -> str:
    pszSanitized: str = INVALID_FILE_CHARS_PATTERN.sub("_", pszSheetName).strip()
    if pszSanitized == "":
        return "Sheet"
    return pszSanitized


def build_unique_output_path(
    objBaseDirectoryPath: Path,
    pszExcelStem: str,
    pszSanitizedSheetName: str,
    objUsedPaths: set[Path],
) -> Path:
    objOutputPath: Path = objBaseDirectoryPath / f"{pszExcelStem}_{pszSanitizedSheetName}.tsv"
    if objOutputPath not in objUsedPaths:
        objUsedPaths.add(objOutputPath)
        return objOutputPath

    iSuffix: int = 2
    while True:
        objCandidatePath: Path = (
            objBaseDirectoryPath / f"{pszExcelStem}_{pszSanitizedSheetName}_{iSuffix}.tsv"
        )
        if objCandidatePath not in objUsedPaths:
            objUsedPaths.add(objCandidatePath)
            return objCandidatePath
        iSuffix += 1


def format_timedelta_as_h_mm_ss(objDuration: timedelta) -> str:
    iTotalSeconds: int = int(objDuration.total_seconds())
    iSign: int = -1 if iTotalSeconds < 0 else 1
    iAbsTotalSeconds: int = abs(iTotalSeconds)
    iHours: int = iAbsTotalSeconds // 3600
    iMinutes: int = (iAbsTotalSeconds % 3600) // 60
    iSeconds: int = iAbsTotalSeconds % 60
    pszPrefix: str = "-" if iSign < 0 else ""
    return f"{pszPrefix}{iHours}:{iMinutes:02d}:{iSeconds:02d}"


def normalize_duration_text_if_needed(pszText: str) -> str:
    objMatch = DURATION_TEXT_PATTERN.match(pszText)
    if objMatch is None:
        return pszText
    iDays: int = int(objMatch.group(1))
    iHours: int = int(objMatch.group(2))
    iMinutes: int = int(objMatch.group(3))
    iSeconds: int = int(objMatch.group(4))
    iTotalHours: int = iDays * 24 + iHours
    return f"{iTotalHours}:{iMinutes:02d}:{iSeconds:02d}"


def normalize_cell_value(objValue: object) -> str:
    if objValue is None:
        return ""
    if isinstance(objValue, timedelta):
        return format_timedelta_as_h_mm_ss(objValue)
    pszText: str = str(objValue)
    pszText = normalize_duration_text_if_needed(pszText)
    return pszText.replace("\t", "_")


def write_sheet_to_tsv(objOutputPath: Path, objRows: List[List[object]]) -> None:
    with open(objOutputPath, mode="w", encoding="utf-8", newline="") as objFile:
        objWriter: csv.writer = csv.writer(objFile, delimiter="\t", lineterminator="\n")
        for objRow in objRows:
            objWriter.writerow([normalize_cell_value(objValue) for objValue in objRow])


def convert_csv_rows_to_tsv_file(objOutputPath: Path, objRows: List[List[str]]) -> None:
    write_sheet_to_tsv(objOutputPath, objRows)


def format_xlsx_cell_value_for_tsv(objValue: object) -> object:
    if isinstance(objValue, datetime):
        if (
            objValue.hour == 0
            and objValue.minute == 0
            and objValue.second == 0
            and objValue.microsecond == 0
        ):
            return objValue.strftime("%Y/%m/%d")
        return objValue.strftime("%Y/%m/%d %H:%M:%S")

    if isinstance(objValue, date):
        return objValue.strftime("%Y/%m/%d")

    if isinstance(objValue, time):
        if objValue.second == 0 and objValue.microsecond == 0:
            return f"{objValue.hour}:{objValue.minute:02d}"
        return f"{objValue.hour}:{objValue.minute:02d}:{objValue.second:02d}"

    if isinstance(objValue, timedelta):
        pszText: str = format_timedelta_as_h_mm_ss(objValue)
        return re.sub(r"^(\d+):(\d{2}):00$", r"\1:\2", pszText)

    return objValue


def convert_xlsx_rows_to_tsv_file(objOutputPath: Path, objRows: List[List[object]]) -> None:
    objNormalizedRows: List[List[object]] = []
    for objRow in objRows:
        objNormalizedRows.append([
            format_xlsx_cell_value_for_tsv(objValue) for objValue in objRow
        ])
    write_sheet_to_tsv(objOutputPath, objNormalizedRows)


def read_tsv_rows(objInputPath: Path) -> List[List[str]]:
    objRows: List[List[str]] = []
    with open(objInputPath, mode="r", encoding="utf-8-sig", newline="") as objFile:
        objReader = csv.reader(objFile, delimiter="\t")
        for objRow in objReader:
            objRows.append(list(objRow))
    return objRows


def is_blank_text(pszText: str) -> bool:
    return (pszText or "").strip().replace("\u3000", "") == ""


def get_effective_column_count(objRow: List[str]) -> int:
    for iIndex in range(len(objRow) - 1, -1, -1):
        if not is_blank_text(objRow[iIndex]):
            return iIndex + 1
    return 0


def is_jobcan_long_format_tsv(objRows: List[List[str]]) -> bool:
    objNonEmptyRows: List[List[str]] = [
        objRow for objRow in objRows if any(not is_blank_text(pszCell) for pszCell in objRow)
    ]
    if not objNonEmptyRows:
        return False

    iTotal: int = len(objNonEmptyRows)
    iFourColumnsLike: int = 0
    iTimeTextRows: int = 0
    iProjectCodeRows: int = 0
    for objRow in objNonEmptyRows:
        iEffectiveColumns: int = get_effective_column_count(objRow)
        if 3 <= iEffectiveColumns <= 5:
            iFourColumnsLike += 1

        if len(objRow) >= 4:
            pszTimeText: str = (objRow[3] or "").strip()
            if TIME_TEXT_PATTERN.match(pszTimeText) is not None or DURATION_TEXT_PATTERN.match(pszTimeText) is not None:
                iTimeTextRows += 1

        if len(objRow) >= 2:
            pszProjectText: str = (objRow[1] or "").strip()
            if re.match(r"^(P\d{5}|[A-OQ-Z]\d{3})", pszProjectText) is not None:
                iProjectCodeRows += 1

    return (
        iFourColumnsLike / iTotal >= 0.7
        and iTimeTextRows / iTotal >= 0.5
        and iProjectCodeRows / iTotal >= 0.5
    )


def is_salary_payment_deduction_list_tsv(objRows: List[List[str]]) -> bool:
    if len(objRows) < 2:
        return False

    objHeaderRow: List[str] = objRows[0]
    objHeaderSet: set[str] = {
        (pszCell or "").strip()
        for pszCell in objHeaderRow
        if (pszCell or "").strip() != ""
    }
    if not all(pszRequiredHeader in objHeaderSet for pszRequiredHeader in SALARY_PAYMENT_DEDUCTION_REQUIRED_HEADERS):
        return False

    iStaffCodeIndex: int = objHeaderRow.index("スタッフコード")
    bHasStaffCodeValue: bool = False
    for objRow in objRows[1:]:
        if iStaffCodeIndex >= len(objRow):
            continue
        pszStaffCode: str = (objRow[iStaffCodeIndex] or "").strip()
        if re.match(r"^\d+$", pszStaffCode) is not None:
            bHasStaffCodeValue = True
            break
    return bHasStaffCodeValue


def is_management_accounting_manhour_csv(objRows: List[List[str]]) -> bool:
    if len(objRows) < 2:
        return False

    objHeaderRow: List[str] = objRows[0]
    objHeaderSet: set[str] = {
        (pszCell or "").strip()
        for pszCell in objHeaderRow
        if (pszCell or "").strip() != ""
    }
    if not all(
        pszRequiredHeader in objHeaderSet
        for pszRequiredHeader in MANAGEMENT_ACCOUNTING_MANHOUR_REQUIRED_HEADERS
    ):
        return False

    iStaffCodeIndex: int = objHeaderRow.index("スタッフコード")
    iManhourIndex: int = objHeaderRow.index("工数")

    bHasStaffCode: bool = False
    bHasManhour: bool = False
    for objRow in objRows[1:]:
        if iStaffCodeIndex < len(objRow):
            pszStaffCode: str = (objRow[iStaffCodeIndex] or "").strip()
            if re.match(r"^\d+$", pszStaffCode) is not None:
                bHasStaffCode = True
        if iManhourIndex < len(objRow):
            pszManhour: str = (objRow[iManhourIndex] or "").strip()
            if re.match(r"^\d+:\d{2}(?::\d{2})?$", pszManhour) is not None:
                bHasManhour = True
        if bHasStaffCode and bHasManhour:
            return True

    return False


def is_management_accounting_manhour_tsv(objRows: List[List[str]]) -> bool:
    return is_management_accounting_manhour_csv(objRows)


def is_management_accounting_manhour_xlsx_sheet(objRows: List[List[object]]) -> bool:
    objStringRows: List[List[str]] = []
    for objRow in objRows:
        objStringRows.append([
            ("" if objValue is None else str(format_xlsx_cell_value_for_tsv(objValue))).strip()
            for objValue in objRow
        ])
    return is_management_accounting_manhour_csv(objStringRows)


def build_staff_code_by_name_from_management_accounting_rows(
    objRows: List[List[str]],
) -> dict[str, str]:
    if not objRows:
        return {}

    objHeaderRow: List[str] = [(pszCell or "").strip() for pszCell in objRows[0]]
    if "スタッフコード" not in objHeaderRow or "姓 名" not in objHeaderRow:
        return {}

    iStaffCodeIndex: int = objHeaderRow.index("スタッフコード")
    iStaffNameIndex: int = objHeaderRow.index("姓 名")

    objStaffCodeByName: dict[str, str] = {}
    for objRow in objRows[1:]:
        if iStaffCodeIndex >= len(objRow) or iStaffNameIndex >= len(objRow):
            continue
        pszStaffCode: str = (objRow[iStaffCodeIndex] or "").strip()
        pszStaffName: str = (objRow[iStaffNameIndex] or "").strip()
        if pszStaffName == "" or pszStaffCode == "":
            continue
        if re.match(r"^\d+$", pszStaffCode) is None:
            continue
        if pszStaffName not in objStaffCodeByName:
            objStaffCodeByName[pszStaffName] = pszStaffCode

    return objStaffCodeByName


def load_staff_code_by_name_from_management_accounting_file(
    objManagementAccountingPath: Path,
) -> dict[str, str]:
    pszSuffix: str = objManagementAccountingPath.suffix.lower()

    if pszSuffix == ".tsv":
        objRows: List[List[str]] = read_tsv_rows(objManagementAccountingPath)
        if not is_management_accounting_manhour_tsv(objRows):
            raise ValueError(f"Not management accounting manhour TSV: {objManagementAccountingPath}")
        return build_staff_code_by_name_from_management_accounting_rows(objRows)

    if pszSuffix == ".csv":
        objRows = []
        with open(objManagementAccountingPath, mode="r", encoding="utf-8-sig", newline="") as objFile:
            objReader = csv.reader(objFile)
            for objRow in objReader:
                objRows.append(list(objRow))
        if not is_management_accounting_manhour_csv(objRows):
            raise ValueError(f"Not management accounting manhour CSV: {objManagementAccountingPath}")
        return build_staff_code_by_name_from_management_accounting_rows(objRows)

    if pszSuffix == ".xlsx":
        try:
            import openpyxl
        except Exception as objException:
            raise RuntimeError(f"Failed to import openpyxl: {objException}") from objException

        objWorkbook = openpyxl.load_workbook(
            filename=objManagementAccountingPath,
            read_only=True,
            data_only=True,
        )
        try:
            for objWorksheet in objWorkbook.worksheets:
                objRowsXlsx: List[List[object]] = [
                    list(objRow)
                    for objRow in objWorksheet.iter_rows(values_only=True)
                ]
                if not is_management_accounting_manhour_xlsx_sheet(objRowsXlsx):
                    continue

                objRowsString: List[List[str]] = []
                for objRow in objRowsXlsx:
                    objRowsString.append([
                        "" if objValue is None else str(format_xlsx_cell_value_for_tsv(objValue)).strip()
                        for objValue in objRow
                    ])
                return build_staff_code_by_name_from_management_accounting_rows(objRowsString)
        finally:
            objWorkbook.close()

        raise ValueError(f"No management accounting manhour sheet found in XLSX: {objManagementAccountingPath}")

    raise ValueError(f"Unsupported management accounting extension: {objManagementAccountingPath}")


def build_new_rawdata_step0003_output_path_from_step0002(objStep0002Path: Path) -> Path:
    pszFileName: str = objStep0002Path.name
    if "_step0002_" not in pszFileName:
        raise ValueError(f"Input is not step0002 file: {objStep0002Path}")
    pszOutputFileName: str = pszFileName.replace("_step0002_", "_step0003_", 1)
    return objStep0002Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0004_output_path_from_step0003(objStep0003Path: Path) -> Path:
    pszFileName: str = objStep0003Path.name
    if "_step0003_" not in pszFileName:
        raise ValueError(f"Input is not step0003 file: {objStep0003Path}")
    pszOutputFileName: str = pszFileName.replace("_step0003_", "_step0004_", 1)
    return objStep0003Path.resolve().parent / pszOutputFileName


def process_new_rawdata_step0004_from_step0003(
    objNewRawdataStep0003Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0003Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0003Path}")

    objOutputRows: List[List[str]] = []
    objSeenStaffCodes: set[str] = set()
    for objRow in objInputRows:
        objOriginalRow: List[str] = list(objRow)

        pszDisplayStaffCode: str = ""
        if objOriginalRow:
            pszStaffCodeCell: str = (objOriginalRow[0] or "").strip()
            if pszStaffCodeCell != "":
                if pszStaffCodeCell not in objSeenStaffCodes:
                    pszDisplayStaffCode = pszStaffCodeCell
                    objSeenStaffCodes.add(pszStaffCodeCell)

        objOutputRows.append([pszDisplayStaffCode] + objOriginalRow)

    objOutputPath: Path = build_new_rawdata_step0004_output_path_from_step0003(objNewRawdataStep0003Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def build_new_rawdata_step0005_output_path_from_step0004(objStep0004Path: Path) -> Path:
    pszFileName: str = objStep0004Path.name
    if "_step0004_" not in pszFileName:
        raise ValueError(f"Input is not step0004 file: {objStep0004Path}")
    pszOutputFileName: str = pszFileName.replace("_step0004_", "_step0005_", 1)
    return objStep0004Path.resolve().parent / pszOutputFileName


def parse_numeric_text(pszText: str) -> float | None:
    pszValue: str = (pszText or "").strip()
    if pszValue == "":
        return None
    if re.match(r"^-?\d+(?:\.\d+)?$", pszValue) is None:
        return None
    try:
        return float(pszValue)
    except Exception:
        return None


def format_decimal_for_tsv_cell(objValue: Decimal) -> str:
    if objValue == objValue.to_integral_value():
        return str(int(objValue))
    return format(objValue.normalize(), "f")


def detect_prepayed_commute_period_label(
    objRows: List[List[str]],
) -> str:
    if not objRows:
        raise ValueError("Input TSV has no rows")
    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    if "前払支給分" not in objHeaderRow:
        raise ValueError("Missing header: 前払支給分")
    iAdvanceIndex: int = objHeaderRow.index("前払支給分")
    objMonthValues: List[int] = []
    for iOffset in range(1, 7):
        iIndex: int = iAdvanceIndex + iOffset
        if iIndex >= len(objHeaderRow):
            break
        objMatch = re.match(r"^(\d{1,2})月$", objHeaderRow[iIndex])
        if objMatch is not None:
            objMonthValues.append(int(objMatch.group(1)))
    if objMonthValues == [4, 5, 6, 7, 8, 9]:
        return "04-09月"
    if objMonthValues == [10, 11, 12, 1, 2, 3]:
        return "10-03月"
    raise ValueError(f"Could not detect prepaid commute period from month headers: {objMonthValues}")


def _build_new_rawdata_step0005_monthly_output_path_from_step0004(
    objStep0004Path: Path,
    iTargetMonth: int,
) -> Path:
    objBaseStep0005Path: Path = build_new_rawdata_step0005_output_path_from_step0004(objStep0004Path)
    pszBaseName: str = objBaseStep0005Path.name
    pszSuffix: str = "_前払通勤交通費按分表.tsv"
    if not pszBaseName.endswith(pszSuffix):
        raise ValueError(f"Input is not prepaid commute step0004 file: {objStep0004Path}")
    pszMonthLabel: str = f"{iTargetMonth:02d}月"
    objTenToThreeMatch = re.match(
        r"^新_ローデータ_シート_step0005_(\d{4})年(10-03月)_前払通勤交通費按分表\.tsv$",
        pszBaseName,
    )
    if objTenToThreeMatch is not None:
        iBaseYear: int = int(objTenToThreeMatch.group(1))
        pszPeriodLabel: str = objTenToThreeMatch.group(2)
        iOutputYear: int = iBaseYear - 1 if iTargetMonth in (10, 11, 12) else iBaseYear
        pszOutputName: str = (
            f"新_ローデータ_シート_step0005_{iOutputYear}年{pszPeriodLabel}_{pszMonthLabel}_前払通勤交通費按分表.tsv"
        )
    else:
        pszOutputName = pszBaseName[: -len(pszSuffix)] + f"_{pszMonthLabel}" + pszSuffix
    return objBaseStep0005Path.resolve().parent / pszOutputName


def _parse_decimal_text(pszText: str) -> Decimal | None:
    pszValue: str = (pszText or "").strip()
    if pszValue == "":
        return None
    try:
        return Decimal(pszValue)
    except (InvalidOperation, ValueError):
        return None


def _build_prepayed_commute_step0005_rows_for_target_month(
    objInputRows: List[List[str]],
    iTargetMonth: int,
) -> List[List[str]]:
    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objInputRows[0]]
    iAdvanceIndex: int = objHeaderRow.index("前払支給分")
    iApplicationIndex: int = objHeaderRow.index("申請の有無")
    iTotalIndex: int = objHeaderRow.index("合計") if "合計" in objHeaderRow else -1
    iRemainingIndex: int = objHeaderRow.index("残り") if "残り" in objHeaderRow else -1

    pszPeriodLabel: str = detect_prepayed_commute_period_label(objInputRows)
    objPeriodMonths: List[int] = [4, 5, 6, 7, 8, 9] if pszPeriodLabel == "04-09月" else [10, 11, 12, 1, 2, 3]
    objMonthIndicesByMonth: dict[int, int] = {iMonth: objHeaderRow.index(f"{iMonth}月") for iMonth in objPeriodMonths}
    objMonthOrderIndexByMonth: dict[int, int] = {
        iMonth: iOrderIndex for iOrderIndex, iMonth in enumerate(objPeriodMonths)
    }
    if iTargetMonth not in objMonthOrderIndexByMonth:
        raise ValueError(f"Target month {iTargetMonth} is out of period months {objPeriodMonths}")
    iTargetOrderIndex: int = objMonthOrderIndexByMonth[iTargetMonth]

    objOutputRows: List[List[str]] = [list(objInputRows[0])]
    iRequiredLength: int = max(
        iAdvanceIndex,
        iApplicationIndex,
        iTotalIndex if iTotalIndex >= 0 else 0,
        iRemainingIndex if iRemainingIndex >= 0 else 0,
        max(objMonthIndicesByMonth.values()),
    )
    for objRow in objInputRows[1:]:
        objNewRow: List[str] = list(objRow)
        while len(objNewRow) <= iRequiredLength:
            objNewRow.append("")

        pszAdvanceText: str = (objNewRow[iAdvanceIndex] or "").strip()
        if pszAdvanceText == "":
            objOutputRows.append(objNewRow)
            continue

        pszApplicationText: str = (objNewRow[iApplicationIndex] or "").strip()
        objMatch = re.match(r"^(\d{1,2})/\d{1,2}支給$", pszApplicationText)
        if objMatch is None:
            objNewRow[iApplicationIndex] = "?/??支給_不明"
            objOutputRows.append(objNewRow)
            continue

        iPaymentMonth: int = int(objMatch.group(1))
        iStartMonth: int = iPaymentMonth
        if pszPeriodLabel == "04-09月" and iPaymentMonth in (3, 4):
            iStartMonth = 4
        if pszPeriodLabel == "10-03月" and iPaymentMonth in (9, 10):
            iStartMonth = 10
        if iStartMonth not in objMonthOrderIndexByMonth:
            continue
        if iTargetOrderIndex < objMonthOrderIndexByMonth[iStartMonth]:
            continue

        objMonthSum: Decimal = Decimal("0")
        for iMonth in objPeriodMonths:
            iMonthIndex: int = objMonthIndicesByMonth[iMonth]
            pszMonthText: str = (objNewRow[iMonthIndex] or "").strip()
            if objMonthOrderIndexByMonth[iMonth] > iTargetOrderIndex:
                objNewRow[iMonthIndex] = ""
                continue
            objMonthValue: Decimal | None = _parse_decimal_text(pszMonthText)
            if objMonthValue is not None:
                objMonthSum += objMonthValue

        if iTotalIndex >= 0:
            objNewRow[iTotalIndex] = format_decimal_for_tsv_cell(objMonthSum)
        if iRemainingIndex >= 0:
            objAdvanceValue: Decimal | None = _parse_decimal_text(pszAdvanceText)
            if objAdvanceValue is not None:
                objNewRow[iRemainingIndex] = format_decimal_for_tsv_cell(objAdvanceValue - objMonthSum)

        objOutputRows.append(objNewRow)

    return objOutputRows


def _process_new_rawdata_step0005_monthly_prepayed_commute_from_step0004(
    objNewRawdataStep0004Path: Path,
    objInputRows: List[List[str]],
) -> int:
    pszPeriodLabel: str = detect_prepayed_commute_period_label(objInputRows)
    objPeriodMonths: List[int] = [4, 5, 6, 7, 8, 9] if pszPeriodLabel == "04-09月" else [10, 11, 12, 1, 2, 3]
    for iTargetMonth in objPeriodMonths:
        objOutputRows: List[List[str]] = _build_prepayed_commute_step0005_rows_for_target_month(
            objInputRows,
            iTargetMonth,
        )
        objOutputPath: Path = _build_new_rawdata_step0005_monthly_output_path_from_step0004(
            objNewRawdataStep0004Path,
            iTargetMonth,
        )
        write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_new_rawdata_step0005_from_step0004(
    objNewRawdataStep0004Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0004Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0004Path}")
    if "前払通勤交通費按分表" in objNewRawdataStep0004Path.name:
        return _process_new_rawdata_step0005_monthly_prepayed_commute_from_step0004(
            objNewRawdataStep0004Path,
            objInputRows,
        )

    objRankTargets: List[float] = []
    for objRow in objInputRows:
        if not objRow:
            continue
        fValue = parse_numeric_text(objRow[0])
        if fValue is not None:
            objRankTargets.append(fValue)

    objOutputRows: List[List[str]] = []
    for objRow in objInputRows:
        objNewRow: List[str] = list(objRow)
        pszRankText: str = ""
        if objNewRow:
            fValue = parse_numeric_text(objNewRow[0])
            if fValue is not None:
                iRank: int = 1 + sum(1 for fTarget in objRankTargets if fTarget < fValue)
                pszRankText = str(iRank)
        objOutputRows.append([pszRankText] + objNewRow)

    objOutputPath: Path = build_new_rawdata_step0005_output_path_from_step0004(objNewRawdataStep0004Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def fill_missing_staff_codes_in_new_rawdata_step0002_by_management_accounting(
    objNewRawdataStep0002Path: Path,
    objStaffCodeByName: dict[str, str],
) -> int:
    if not objStaffCodeByName:
        raise ValueError("No staff code mapping from management accounting file")

    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0002Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0002Path}")

    objOutputRows: List[List[str]] = []
    pszCurrentStaffName: str = ""
    for objRow in objInputRows:
        objNewRow: List[str] = list(objRow)
        if not objNewRow:
            objOutputRows.append(objNewRow)
            continue

        if len(objNewRow) >= 2:
            pszStaffNameCell: str = (objNewRow[1] or "").strip()
            if pszStaffNameCell != "":
                pszCurrentStaffName = pszStaffNameCell

        pszStaffCodeCell: str = (objNewRow[0] or "").strip()
        if pszStaffCodeCell == "" and pszCurrentStaffName != "":
            pszFilledCode: str = objStaffCodeByName.get(pszCurrentStaffName, "")
            if pszFilledCode != "":
                objNewRow[0] = pszFilledCode

        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0003_output_path_from_step0002(objNewRawdataStep0002Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_management_accounting_manhour_csv_input(
    objResolvedInputPath: Path,
    objRows: List[List[str]],
) -> int:
    objOutputPath: Path = objResolvedInputPath.resolve().with_suffix(".tsv")
    convert_csv_rows_to_tsv_file(objOutputPath, objRows)
    return 0


def extract_year_month_text_from_path(objInputPath: Path) -> str:
    objMatch = YEAR_MONTH_PATTERN.search(str(objInputPath))
    if objMatch is None:
        raise ValueError(f"Could not extract YY.MM月 from input path: {objInputPath}")
    iYear: int = 2000 + int(objMatch.group(1))
    iMonth: int = int(objMatch.group(2))
    return f"{iYear}年{iMonth:02d}月"


def normalize_project_name_for_jobcan_long_tsv(pszProjectName: str) -> str:
    pszNormalized: str = pszProjectName or ""
    pszNormalized = pszNormalized.replace("\t", "_")
    pszNormalized = re.sub(r"(P\d{5})(?![ _\t　【])", r"\1_", pszNormalized)
    pszNormalized = re.sub(r"([A-OQ-Z]\d{3})(?![ _\t　【])", r"\1_", pszNormalized)
    pszNormalized = re.sub(r"^([A-OQ-Z]\d{3}) +", r"\1_", pszNormalized)
    pszNormalized = re.sub(r"([A-OQ-Z]\d{3})[ 　]+", r"\1_", pszNormalized)
    pszNormalized = re.sub(r"(P\d{5})[ 　]+", r"\1_", pszNormalized)
    return pszNormalized


def process_jobcan_long_tsv_input_rawdata_sheet_step0001(
    objResolvedInputPath: Path,
    objRows: List[List[str]],
) -> int:
    pszYearMonthText: str = extract_year_month_text_from_path(objResolvedInputPath)

    objOutputRows: List[List[str]] = []
    pszCurrentStaffName: str = ""
    pszLastOutputStaffName: str = ""
    for objRow in objRows:
        if not any(not is_blank_text(pszCell) for pszCell in objRow):
            continue
        if len(objRow) < 4:
            continue

        pszStaffName: str = (objRow[0] or "").strip()
        if pszStaffName != "":
            pszCurrentStaffName = pszStaffName
        if pszCurrentStaffName == "":
            continue

        pszProjectName: str = normalize_project_name_for_jobcan_long_tsv((objRow[1] or "").strip())
        pszManhour: str = (objRow[3] or "").strip()
        if pszProjectName == "" and pszManhour == "":
            continue

        pszOutputStaffName: str = pszCurrentStaffName
        if pszCurrentStaffName == pszLastOutputStaffName:
            pszOutputStaffName = ""
        else:
            pszLastOutputStaffName = pszCurrentStaffName

        objOutputRows.append([pszOutputStaffName, pszProjectName, pszManhour])

    if not objOutputRows:
        raise ValueError("No output rows generated for Jobcan long-format TSV")

    objOutputPath: Path = (
        objResolvedInputPath.resolve().parent
        / f"ローデータ_シート_step0001_{pszYearMonthText}.tsv"
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_jobcan_long_tsv_input_new_rawdata_sheet_step0001(
    objResolvedInputPath: Path,
    objRows: List[List[str]],
) -> int:
    pszYearMonthText: str = extract_year_month_text_from_path(objResolvedInputPath)

    objOutputRows: List[List[str]] = []
    pszCurrentStaffName: str = ""
    pszLastOutputStaffName: str = ""
    for objRow in objRows:
        if not any(not is_blank_text(pszCell) for pszCell in objRow):
            continue
        if len(objRow) < 4:
            continue

        pszStaffName: str = (objRow[0] or "").strip()
        if pszStaffName != "":
            pszCurrentStaffName = pszStaffName
        if pszCurrentStaffName == "":
            continue

        pszProjectName: str = normalize_project_name_for_jobcan_long_tsv((objRow[1] or "").strip())
        pszManhour: str = (objRow[3] or "").strip()
        if pszProjectName == "" and pszManhour == "":
            continue

        pszOutputStaffName: str = pszCurrentStaffName
        if pszCurrentStaffName == pszLastOutputStaffName:
            pszOutputStaffName = ""
        else:
            pszLastOutputStaffName = pszCurrentStaffName

        objOutputRows.append([pszOutputStaffName, pszProjectName, pszManhour])

    if not objOutputRows:
        raise ValueError("No output rows generated for Jobcan long-format TSV")

    objOutputPath: Path = (
        objResolvedInputPath.resolve().parent
        / f"新_ローデータ_シート_step0001_{pszYearMonthText}.tsv"
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_jobcan_long_tsv_input(objResolvedInputPath: Path, objRows: List[List[str]]) -> int:
    process_jobcan_long_tsv_input_rawdata_sheet_step0001(objResolvedInputPath, objRows)
    process_jobcan_long_tsv_input_new_rawdata_sheet_step0001(objResolvedInputPath, objRows)
    return 0


def build_new_rawdata_step0002_output_path_from_step0001(objStep0001Path: Path) -> Path:
    pszFileName: str = objStep0001Path.name
    if "_step0001_" not in pszFileName:
        raise ValueError(f"Input is not step0001 file: {objStep0001Path}")
    pszOutputFileName: str = pszFileName.replace("_step0001_", "_step0002_", 1)
    return objStep0001Path.resolve().parent / pszOutputFileName


def build_staff_code_by_name_from_salary_step0001(objSalaryStep0001Path: Path) -> dict[str, str]:
    objRows: List[List[str]] = read_tsv_rows(objSalaryStep0001Path)
    objStaffCodeByName: dict[str, str] = {}
    for iRowIndex, objRow in enumerate(objRows):
        if len(objRow) < 2:
            continue
        pszStaffName: str = (objRow[0] or "").strip()
        pszStaffCode: str = (objRow[1] or "").strip()
        if pszStaffName == "" or pszStaffCode == "":
            continue
        if iRowIndex == 0 and pszStaffName == "従業員名" and pszStaffCode == "スタッフコード":
            continue
        objStaffCodeByName[pszStaffName] = pszStaffCode
    if not objStaffCodeByName:
        raise ValueError(f"No staff codes found in salary step0001 TSV: {objSalaryStep0001Path}")
    return objStaffCodeByName


def build_staff_name_by_code_from_salary_step0001(objSalaryStep0001Path: Path) -> dict[str, str]:
    objRows: List[List[str]] = read_tsv_rows(objSalaryStep0001Path)
    objStaffNameByCode: dict[str, str] = {}
    for iRowIndex, objRow in enumerate(objRows):
        if len(objRow) < 2:
            continue
        pszStaffName: str = (objRow[0] or "").strip()
        pszStaffCode: str = (objRow[1] or "").strip()
        if pszStaffName == "" or pszStaffCode == "":
            continue
        if iRowIndex == 0 and pszStaffName == "従業員名" and pszStaffCode == "スタッフコード":
            continue
        if pszStaffCode not in objStaffNameByCode:
            objStaffNameByCode[pszStaffCode] = pszStaffName
    if not objStaffNameByCode:
        raise ValueError(f"No staff names found in salary step0001 TSV: {objSalaryStep0001Path}")
    return objStaffNameByCode


def build_new_rawdata_step0003_name_mapping_output_path(objStep0003Path: Path) -> Path:
    pszFileName: str = objStep0003Path.name
    if not NEW_RAWDATA_STEP0003_FILE_PATTERN.match(pszFileName):
        raise ValueError(f"Input is not step0003 file: {objStep0003Path}")
    pszStem: str = objStep0003Path.stem
    return objStep0003Path.resolve().parent / f"{pszStem}_工数の姓_給与の姓_対応表.tsv"


def build_new_rawdata_step0003_name_mapping_sorted_output_path(objStep0003NameMappingPath: Path) -> Path:
    pszFileName: str = objStep0003NameMappingPath.name
    pszSuffix: str = "_工数の姓_給与の姓_対応表.tsv"
    if not pszFileName.endswith(pszSuffix):
        raise ValueError(f"Input is not step0003 name mapping file: {objStep0003NameMappingPath}")
    pszOutputFileName: str = pszFileName[:-4] + "_昇順.tsv"
    return objStep0003NameMappingPath.resolve().parent / pszOutputFileName


def process_new_rawdata_step0003_name_mapping_sorted_by_staff_code(
    objStep0003NameMappingPath: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0003NameMappingPath)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0003NameMappingPath}")

    objHeaderRow: List[str] = list(objInputRows[0])
    objDataRows: List[List[str]] = [list(objRow) for objRow in objInputRows[1:]]

    objDataRows.sort(key=lambda objRow: int((objRow[0] or "").strip()) if len(objRow) >= 1 and (objRow[0] or "").strip().isdigit() else 10 ** 18)

    objOutputRows: List[List[str]] = [objHeaderRow] + objDataRows
    objOutputPath: Path = build_new_rawdata_step0003_name_mapping_sorted_output_path(objStep0003NameMappingPath)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_salary_step0001_for_step0003_old_new_name_mapping(
    objNewRawdataStep0003Path: Path,
    objSalaryStep0001Path: Path,
) -> int:
    objStaffNameByCode: dict[str, str] = build_staff_name_by_code_from_salary_step0001(objSalaryStep0001Path)

    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0003Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0003Path}")

    objOutputRows: List[List[str]] = [["スタッフコード", "氏名", "氏名"]]
    objSeenStaffCodes: set[str] = set()
    for objRow in objInputRows:
        pszStaffCode: str = (objRow[0] or "").strip() if len(objRow) >= 1 else ""
        pszStep0003StaffName: str = (objRow[1] or "").strip() if len(objRow) >= 2 else ""
        if pszStep0003StaffName == "":
            continue
        if pszStaffCode == "":
            continue
        if pszStaffCode in objSeenStaffCodes:
            continue

        pszSalaryStaffName: str = objStaffNameByCode.get(pszStaffCode, "")
        objOutputRows.append([pszStaffCode, pszStep0003StaffName, pszSalaryStaffName])
        objSeenStaffCodes.add(pszStaffCode)

    objOutputPath: Path = build_new_rawdata_step0003_name_mapping_output_path(objNewRawdataStep0003Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0003_name_mapping_sorted_by_staff_code(objOutputPath)
    return 0


def process_new_rawdata_step0002_from_salary_and_new_rawdata_step0001(
    objSalaryStep0001Path: Path,
    objNewRawdataStep0001Path: Path,
) -> int:
    objStaffCodeByName: dict[str, str] = build_staff_code_by_name_from_salary_step0001(objSalaryStep0001Path)

    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0001Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0001Path}")

    objOutputRows: List[List[str]] = []
    pszCurrentStaffName: str = ""
    for objRow in objInputRows:
        objNewRow: List[str] = list(objRow)
        if objNewRow:
            pszStaffNameCell: str = (objNewRow[0] or "").strip()
            if pszStaffNameCell != "":
                pszCurrentStaffName = pszStaffNameCell
        pszStaffCode: str = objStaffCodeByName.get(pszCurrentStaffName, "") if pszCurrentStaffName != "" else ""
        objOutputRows.append([pszStaffCode] + objNewRow)

    objOutputPath: Path = build_new_rawdata_step0002_output_path_from_step0001(objNewRawdataStep0001Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0




def build_diagnose_log_path_from_input_tsv(objInputPath: Path) -> Path:
    return objInputPath.resolve().with_name(f"{objInputPath.stem}_log.txt")


def analyze_fourth_column_manhour_h_mm_tsv(objRows: List[List[str]]) -> tuple[bool, int, int, float]:
    objNonEmptyRows: List[List[str]] = [
        objRow for objRow in objRows if any(not is_blank_text(pszCell) for pszCell in objRow)
    ]
    if not objNonEmptyRows:
        return (False, 0, 0, 0.0)

    iTotal: int = len(objNonEmptyRows)
    iHmRows: int = 0
    for objRow in objNonEmptyRows:
        if len(objRow) < 4:
            continue
        pszTimeText: str = (objRow[3] or "").strip()
        if HM_PATTERN.match(pszTimeText) is not None:
            iHmRows += 1
    fHmRatio: float = iHmRows / iTotal
    return (fHmRatio >= 0.5, iHmRows, iTotal, fHmRatio)


def analyze_third_column_task_keyword_tsv(objRows: List[List[str]]) -> tuple[int, int, float]:
    objNonEmptyRows: List[List[str]] = [
        objRow for objRow in objRows if any(not is_blank_text(pszCell) for pszCell in objRow)
    ]
    if not objNonEmptyRows:
        return (0, 0, 0.0)

    iTotal: int = len(objNonEmptyRows)
    iTaskKeywordRows: int = 0
    for objRow in objNonEmptyRows:
        if len(objRow) < 3:
            continue
        pszTaskText: str = (objRow[2] or "").strip()
        if "タスク" in pszTaskText:
            iTaskKeywordRows += 1
    return (iTaskKeywordRows, iTotal, iTaskKeywordRows / iTotal)


def is_jobcan_four_column_candidate_tsv(
    objRows: List[List[str]],
    bIsJobcanLongFormat: bool,
    bIsFourthColumnManhourHm: bool,
    fTaskKeywordRatio: float,
) -> bool:
    if bIsJobcanLongFormat:
        return True
    return bIsFourthColumnManhourHm and fTaskKeywordRatio >= 0.5


def validate_h_mm_ss_tsv_for_step0001(objRows: List[List[str]]) -> tuple[bool, str]:
    if len(objRows) < 2:
        return (False, "rows_too_few")

    objNonEmptyRows: List[List[str]] = [
        objRow for objRow in objRows if any(not is_blank_text(pszCell) for pszCell in objRow)
    ]
    if not objNonEmptyRows:
        return (False, "non_empty_rows_zero")

    iFourColumnRows: int = 0
    iFourthNonEmptyRows: int = 0
    iFourthHmsRows: int = 0
    for objRow in objNonEmptyRows:
        if len(objRow) >= 4:
            iFourColumnRows += 1
            pszFourthText: str = (objRow[3] or "").strip()
            if pszFourthText != "":
                iFourthNonEmptyRows += 1
                if TIME_TEXT_PATTERN.match(pszFourthText) is not None:
                    iFourthHmsRows += 1
    if iFourColumnRows / len(objNonEmptyRows) < 0.8:
        return (False, "four_column_ratio_too_low")
    if iFourthNonEmptyRows == 0:
        return (False, "fourth_column_non_empty_zero")
    if iFourthHmsRows / iFourthNonEmptyRows < 0.5:
        return (False, "fourth_column_h_mm_ss_ratio_too_low")
    return (True, "ok")


def append_tsv_diagnose_log(
    objResolvedInputPath: Path,
    objRows: List[List[str]],
    bIsJobcanLongFormat: bool,
    bIsFourthColumnManhourHm: bool,
    iFourthColHmRows: int,
    iNonEmptyRows: int,
    fHmRatio: float,
    iTaskKeywordRows: int,
    fTaskKeywordRatio: float,
    pszClassification: str,
    pszSafetyCheckResult: str,
    pszBranch: str,
) -> None:
    objLogPath: Path = build_diagnose_log_path_from_input_tsv(objResolvedInputPath)
    iRowsTotal: int = len(objRows)
    pszTimestamp: str = datetime.now().astimezone().isoformat(timespec="seconds")
    pszLine: str = (
        f"{pszTimestamp}\n"
        f'resolved_input_path="{objResolvedInputPath.resolve()}"\n'
        f"rows_total={iRowsTotal}\n"
        f"non_empty_rows={iNonEmptyRows}\n"
        f"is_jobcan_long_format_tsv={str(bIsJobcanLongFormat).lower()}\n"
        f"is_fourth_column_manhour_h_mm_tsv={str(bIsFourthColumnManhourHm).lower()}\n"
        f"fourth_col_hm_rows={iFourthColHmRows}\n"
        f"hm_ratio={fHmRatio:.4f}\n"
        f"task_keyword_rows={iTaskKeywordRows}\n"
        f"task_keyword_ratio={fTaskKeywordRatio:.4f}\n"
        f"classification=\"{pszClassification}\"\n"
        f"safety_check_result=\"{pszSafetyCheckResult}\"\n"
        f'branch="{pszBranch}"\n'
    )
    with open(objLogPath, mode="a", encoding="utf-8", newline="") as objLogFile:
        objLogFile.write(pszLine)


def is_fourth_column_manhour_h_mm_tsv(objRows: List[List[str]]) -> bool:
    bIsHm, _iHmRows, _iTotal, _fHmRatio = analyze_fourth_column_manhour_h_mm_tsv(objRows)
    return bIsHm


def build_h_mm_ss_output_path_from_input_tsv(objInputPath: Path) -> Path:
    return objInputPath.resolve().with_name(f"{objInputPath.stem}_h_mm_ss.tsv")


def convert_manhour_h_mm_to_h_mm_ss_rows(objRows: List[List[str]]) -> List[List[str]]:
    objConvertedRows: List[List[str]] = []
    for objRow in objRows:
        objNewRow: List[str] = list(objRow)
        if len(objNewRow) >= 4:
            pszManhour: str = (objNewRow[3] or "").strip()
            objMatch = HM_PATTERN.match(pszManhour)
            if objMatch is not None:
                iHours: int = int(objMatch.group(1))
                iMinutes: int = int(objMatch.group(2))
                objNewRow[3] = f"{iHours}:{iMinutes:02d}:00"
        objConvertedRows.append(objNewRow)
    return objConvertedRows


def process_tsv_input(objResolvedInputPath: Path) -> int:
    objRows: List[List[str]] = read_tsv_rows(objResolvedInputPath)
    if len(objRows) < 2:
        raise ValueError(f"Input TSV has too few rows: {objResolvedInputPath}")

    bIsJobcanLongFormat: bool = is_jobcan_long_format_tsv(objRows)
    bIsFourthColumnManhourHm, iFourthColHmRows, iNonEmptyRows, fHmRatio = analyze_fourth_column_manhour_h_mm_tsv(objRows)
    iTaskKeywordRows, _iTaskTotalRows, fTaskKeywordRatio = analyze_third_column_task_keyword_tsv(objRows)
    bIsJobcanFourColCandidate: bool = is_jobcan_four_column_candidate_tsv(
        objRows,
        bIsJobcanLongFormat,
        bIsFourthColumnManhourHm,
        fTaskKeywordRatio,
    )
    pszClassification: str = "jobcan_4col" if bIsJobcanFourColCandidate else "unknown"
    bIsHmSsInput: bool = objResolvedInputPath.stem.endswith("_h_mm_ss")

    if bIsJobcanFourColCandidate and bIsHmSsInput:
        bSafetyCheckOk, pszSafetyCheckResult = validate_h_mm_ss_tsv_for_step0001(objRows)
        if not bSafetyCheckOk:
            append_tsv_diagnose_log(
                objResolvedInputPath,
                objRows,
                bIsJobcanLongFormat,
                bIsFourthColumnManhourHm,
                iFourthColHmRows,
                iNonEmptyRows,
                fHmRatio,
                iTaskKeywordRows,
                fTaskKeywordRatio,
                pszClassification,
                pszSafetyCheckResult,
                "safety_check_failed",
            )
            raise ValueError(f"Safety check failed for step0001 TSV processing: {objResolvedInputPath}")
        append_tsv_diagnose_log(
            objResolvedInputPath,
            objRows,
            bIsJobcanLongFormat,
            bIsFourthColumnManhourHm,
            iFourthColHmRows,
            iNonEmptyRows,
            fHmRatio,
            iTaskKeywordRows,
            fTaskKeywordRatio,
            pszClassification,
            pszSafetyCheckResult,
            "to_step0001",
        )
        return process_jobcan_long_tsv_input(objResolvedInputPath, objRows)

    if bIsJobcanFourColCandidate:
        append_tsv_diagnose_log(
            objResolvedInputPath,
            objRows,
            bIsJobcanLongFormat,
            bIsFourthColumnManhourHm,
            iFourthColHmRows,
            iNonEmptyRows,
            fHmRatio,
            iTaskKeywordRows,
            fTaskKeywordRatio,
            pszClassification,
            "not_applicable",
            "to_h_mm_ss",
        )
        objOutputPath: Path = build_h_mm_ss_output_path_from_input_tsv(objResolvedInputPath)
        objConvertedRows: List[List[str]] = convert_manhour_h_mm_to_h_mm_ss_rows(objRows)
        write_sheet_to_tsv(objOutputPath, objConvertedRows)
        return 0

    if is_salary_payment_deduction_list_tsv(objRows):
        append_tsv_diagnose_log(
            objResolvedInputPath,
            objRows,
            bIsJobcanLongFormat,
            bIsFourthColumnManhourHm,
            iFourthColHmRows,
            iNonEmptyRows,
            fHmRatio,
            iTaskKeywordRows,
            fTaskKeywordRatio,
            "salary_tsv",
            "not_applicable",
            "salary_tsv_not_supported",
        )
        raise ValueError(f"Salary payment/deduction list TSV is not supported yet: {objResolvedInputPath}")

    append_tsv_diagnose_log(
        objResolvedInputPath,
        objRows,
        bIsJobcanLongFormat,
        bIsFourthColumnManhourHm,
        iFourthColHmRows,
        iNonEmptyRows,
        fHmRatio,
        iTaskKeywordRows,
        fTaskKeywordRatio,
        pszClassification,
        "not_applicable",
        "unsupported_tsv",
    )
    raise ValueError(f"Unsupported TSV format: {objResolvedInputPath}")


def build_salary_payment_deduction_step0001_output_path_from_csv(
    objResolvedInputPath: Path,
) -> Path:
    pszStem: str = objResolvedInputPath.stem
    pszStem = re.sub(r"^作成用データ：", "", pszStem)

    pszBaseName: str
    pszDateLabel: str
    pszBaseName, pszSeparator, pszDateLabel = pszStem.rpartition("_")
    if pszSeparator == "" or pszBaseName == "" or pszDateLabel == "":
        raise ValueError(f"Could not build salary step0001 output name from csv: {objResolvedInputPath}")

    pszOutputFileName: str = f"{pszBaseName}_step0001_{pszDateLabel}.tsv"
    return objResolvedInputPath.resolve().with_name(pszOutputFileName)


def process_csv_input(objResolvedInputPath: Path) -> int:
    objRows: List[List[str]] = []
    with open(objResolvedInputPath, mode="r", encoding="utf-8-sig", newline="") as objFile:
        objReader = csv.reader(objFile)
        for objRow in objReader:
            objRows.append(list(objRow))

    if is_management_accounting_manhour_csv(objRows):
        return process_management_accounting_manhour_csv_input(
            objResolvedInputPath,
            objRows,
        )

    objOutputPath: Path = objResolvedInputPath.resolve().with_suffix(".tsv")
    convert_csv_rows_to_tsv_file(objOutputPath, objRows)

    if is_salary_payment_deduction_list_tsv(objRows):
        objSalaryStep0001OutputPath: Path = build_salary_payment_deduction_step0001_output_path_from_csv(
            objResolvedInputPath
        )
        convert_csv_rows_to_tsv_file(objSalaryStep0001OutputPath, objRows)

    return 0


def process_single_input(pszInputXlsxPath: str) -> int:
    objResolvedInputPath: Path = resolve_existing_input_path(pszInputXlsxPath)
    pszSuffix: str = objResolvedInputPath.suffix.lower()

    if pszSuffix == ".tsv":
        return process_tsv_input(objResolvedInputPath)

    if pszSuffix == ".csv":
        return process_csv_input(objResolvedInputPath)

    if pszSuffix != ".xlsx":
        raise ValueError(f"Unsupported extension (only .xlsx/.tsv/.csv): {objResolvedInputPath}")

    objBaseDirectoryPath: Path = objResolvedInputPath.resolve().parent
    pszExcelStem: str = objResolvedInputPath.stem

    try:
        import openpyxl
    except Exception as objException:
        raise RuntimeError(f"Failed to import openpyxl: {objException}") from objException

    try:
        objWorkbook = openpyxl.load_workbook(
            filename=objResolvedInputPath,
            read_only=True,
            data_only=True,
        )
    except Exception as objException:
        raise RuntimeError(f"Failed to read workbook: {objResolvedInputPath}. Detail = {objException}") from objException

    objUsedPaths: set[Path] = set()
    try:
        for objWorksheet in objWorkbook.worksheets:
            pszSanitizedSheetName: str = sanitize_sheet_name_for_file_name(objWorksheet.title)
            objOutputPath: Path = build_unique_output_path(
                objBaseDirectoryPath,
                pszExcelStem,
                pszSanitizedSheetName,
                objUsedPaths,
            )
            objRows: List[List[object]] = [list(objRow) for objRow in objWorksheet.iter_rows(values_only=True)]
            convert_xlsx_rows_to_tsv_file(objOutputPath, objRows)
    finally:
        objWorkbook.close()

    return 0


def main() -> int:
    objParser: argparse.ArgumentParser = argparse.ArgumentParser()
    objParser.add_argument(
        "pszInputXlsxPaths",
        nargs="+",
        help="Input file paths (.xlsx or .tsv or .csv)",
    )
    objArgs: argparse.Namespace = objParser.parse_args()

    iExitCode: int = 0
    objHandledInputPaths: set[Path] = set()

    objSalaryStep0001Paths: List[Path] = []
    objNewRawdataStep0001Paths: List[Path] = []
    objNewRawdataStep0002Paths: List[Path] = []
    objNewRawdataStep0003Paths: List[Path] = []
    objNewRawdataStep0004Paths: List[Path] = []
    objManagementAccountingCandidatePaths: List[Path] = []
    for pszInputXlsxPath in objArgs.pszInputXlsxPaths:
        try:
            objResolvedInputPath: Path = resolve_existing_input_path(pszInputXlsxPath)
        except Exception:
            continue

        if SALARY_PAYMENT_STEP0001_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objSalaryStep0001Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0001_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0001Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0002_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0002Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0003_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0003Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0004_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0004Paths.append(objResolvedInputPath)

        if objResolvedInputPath.suffix.lower() in (".tsv", ".csv", ".xlsx"):
            objManagementAccountingCandidatePaths.append(objResolvedInputPath)

    if objSalaryStep0001Paths:
        objSalaryStep0001Path: Path = objSalaryStep0001Paths[0]
        objKnownStep0002PathResolves: set[Path] = {objPath.resolve() for objPath in objNewRawdataStep0002Paths}
        for objNewRawdataStep0001Path in objNewRawdataStep0001Paths:
            try:
                process_new_rawdata_step0002_from_salary_and_new_rawdata_step0001(
                    objSalaryStep0001Path,
                    objNewRawdataStep0001Path,
                )
                objGeneratedStep0002Path: Path = build_new_rawdata_step0002_output_path_from_step0001(
                    objNewRawdataStep0001Path
                )
                if (
                    objGeneratedStep0002Path.exists()
                    and objGeneratedStep0002Path.resolve() not in objKnownStep0002PathResolves
                ):
                    objNewRawdataStep0002Paths.append(objGeneratedStep0002Path)
                    objKnownStep0002PathResolves.add(objGeneratedStep0002Path.resolve())
                objHandledInputPaths.add(objSalaryStep0001Path.resolve())
                objHandledInputPaths.add(objNewRawdataStep0001Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0002 pair: {0} / {1}. Detail = {2}".format(
                        objSalaryStep0001Path,
                        objNewRawdataStep0001Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0002Paths:
        if not objSalaryStep0001Paths:
            print("Error: salary step0001 TSV is required to process step0003 from step0002")
            iExitCode = 1
        else:
            for objNewRawdataStep0002Path in objNewRawdataStep0002Paths:
                for objManagementAccountingCandidatePath in objManagementAccountingCandidatePaths:
                    if objManagementAccountingCandidatePath.resolve() == objNewRawdataStep0002Path.resolve():
                        continue
                    try:
                        objStaffCodeByName: dict[str, str] = load_staff_code_by_name_from_management_accounting_file(
                            objManagementAccountingCandidatePath
                        )
                    except Exception:
                        continue

                    try:
                        fill_missing_staff_codes_in_new_rawdata_step0002_by_management_accounting(
                            objNewRawdataStep0002Path,
                            objStaffCodeByName,
                        )
                        objNewRawdataStep0003Path: Path = build_new_rawdata_step0003_output_path_from_step0002(
                            objNewRawdataStep0002Path
                        )
                        process_salary_step0001_for_step0003_old_new_name_mapping(
                            objNewRawdataStep0003Path,
                            objSalaryStep0001Paths[0],
                        )
                        process_new_rawdata_step0004_from_step0003(objNewRawdataStep0003Path)
                        objNewRawdataStep0004Path: Path = build_new_rawdata_step0004_output_path_from_step0003(
                            objNewRawdataStep0003Path
                        )
                        process_new_rawdata_step0005_from_step0004(objNewRawdataStep0004Path)
                        objHandledInputPaths.add(objNewRawdataStep0002Path.resolve())
                        objHandledInputPaths.add(objNewRawdataStep0003Path.resolve())
                        objHandledInputPaths.add(objNewRawdataStep0004Path.resolve())
                        objHandledInputPaths.add(objManagementAccountingCandidatePath.resolve())
                    except Exception as objException:
                        print(
                            "Error: failed to fill missing step0002 staff codes: {0} / {1}. Detail = {2}".format(
                                objNewRawdataStep0002Path,
                                objManagementAccountingCandidatePath,
                                objException,
                            )
                        )
                        iExitCode = 1
                    break
    if objNewRawdataStep0003Paths:
        for objNewRawdataStep0003Path in objNewRawdataStep0003Paths:
            if objNewRawdataStep0003Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0004_from_step0003(objNewRawdataStep0003Path)
                objNewRawdataStep0004Path: Path = build_new_rawdata_step0004_output_path_from_step0003(
                    objNewRawdataStep0003Path
                )
                process_new_rawdata_step0005_from_step0004(objNewRawdataStep0004Path)
                objHandledInputPaths.add(objNewRawdataStep0003Path.resolve())
                objHandledInputPaths.add(objNewRawdataStep0004Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0004 from step0003: {0}. Detail = {1}".format(
                        objNewRawdataStep0003Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0004Paths:
        for objNewRawdataStep0004Path in objNewRawdataStep0004Paths:
            if objNewRawdataStep0004Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0005_from_step0004(objNewRawdataStep0004Path)
                objHandledInputPaths.add(objNewRawdataStep0004Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0005 from step0004: {0}. Detail = {1}".format(
                        objNewRawdataStep0004Path,
                        objException,
                    )
                )
                iExitCode = 1

    for pszInputXlsxPath in objArgs.pszInputXlsxPaths:
        try:
            objResolvedInputPath: Path = resolve_existing_input_path(pszInputXlsxPath)
            if objResolvedInputPath.resolve() in objHandledInputPaths:
                continue
            if SALARY_PAYMENT_STEP0001_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
                continue
            process_single_input(pszInputXlsxPath)
        except Exception as objException:
            print(
                "Error: failed to process input file: {0}. Detail = {1}".format(
                    pszInputXlsxPath,
                    objException,
                )
            )
            iExitCode = 1
            continue

    return iExitCode


if __name__ == "__main__":
    raise SystemExit(main())
