# -*- coding: utf-8 -*-
"""
SellGeneralAdminCost_Allocation_Cmd.py

販管費配賦（工数付与）TSV を生成する。

入力:
  1) 工数_yyyy年mm月_step10_各プロジェクトの工数.tsv
  2) 損益計算書_yyyy年mm月_A∪B_プロジェクト名_C∪D_vertical.tsv

出力:
  損益計算書_yyyy年mm月_販管費配賦_A∪B_プロジェクト名_C∪D_vertical.tsv

処理:
  損益計算書TSVの各行に対し、
  プロジェクト行(A/C/J/Pで始まり、"_"までのキー)のみ
  工数TSVに同一キーがある場合は工数を、
  無い場合は 0:00:00 を末尾に追加する。
  非プロジェクト行はそのまま出力する。
"""

from __future__ import annotations

import os
import shutil
import re
import sys
import csv
from datetime import datetime
from copy import copy
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.styles import Border, Side


def print_usage() -> None:
    pszUsage: str = (
        "Usage: python SellGeneralAdminCost_Allocation_Cmd.py "
        "<manhour_tsv_path> <pl_tsv_path> [output_tsv_path]\n"
        "   or: python SellGeneralAdminCost_Allocation_Cmd.py "
        "<manhour_tsv_path> <pl_tsv_path> <manhour_tsv_path> <pl_tsv_path> ...\n"
        "   or: python SellGeneralAdminCost_Allocation_Cmd.py "
        "<manhour_tsv_path> ... <pl_tsv_path> ..."
    )
    print(pszUsage)


EXECUTION_ROOT_DIRECTORY: Optional[str] = None
PERIOD_BUTTON_ID_BASE: int = 1001
PERIOD_BUTTON_ID_OFFSET: int = 0
BN_DOUBLECLICKED: int = 5


def get_script_base_directory() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def create_execution_folders() -> str:
    global EXECUTION_ROOT_DIRECTORY
    pszScriptDirectory: str = get_script_base_directory()
    pszTimestamp: str = datetime.now().strftime("%Y年%m月%d日%H時%M分%S秒")
    pszRootDirectory: str = os.path.join(
        pszScriptDirectory,
        f"{pszTimestamp}_損益工数実行表",
    )
    objSubdirectories: List[str] = [
        "期間",
        "カンパニー実績",
        "カンパニー利益率順位",
        "プロジェクト損益",
        "グループ別損益",
        "カンパニー別損益",
    ]
    for pszSubdirectory in objSubdirectories:
        os.makedirs(os.path.join(pszRootDirectory, pszSubdirectory), exist_ok=True)
    EXECUTION_ROOT_DIRECTORY = pszRootDirectory
    return pszRootDirectory


def find_latest_execution_root_directory() -> Optional[str]:
    pszBaseDirectory = get_script_base_directory()
    objCandidates: List[str] = []
    for pszName in os.listdir(pszBaseDirectory):
        if not pszName.endswith("_損益工数実行表"):
            continue
        pszPath = os.path.join(pszBaseDirectory, pszName)
        if os.path.isdir(pszPath):
            objCandidates.append(pszPath)
    if not objCandidates:
        return None
    objCandidates.sort()
    return objCandidates[-1]


def handle_period_left_double_click() -> None:
    pszExecutionRoot = find_latest_execution_root_directory()
    if pszExecutionRoot is None:
        print("Error: 出力フォルダーがまだ作成されていません。")
        return

    pszPeriodDirectory = os.path.join(pszExecutionRoot, "期間")
    pszTargetPath = os.path.join(
        pszPeriodDirectory,
        "SellGeneralAdminCost_Allocation_Cmd_SelectedRange_And_AccountPeriodRange.txt",
    )
    if not os.path.isfile(pszTargetPath):
        print("Error: ファイルが見つかりません。\n" + pszTargetPath)
        return

    os.startfile(pszTargetPath)


def handle_period_button_left_double_click_event(
    iButtonId: int,
    iNotifyCode: int,
) -> bool:
    if iButtonId != PERIOD_BUTTON_ID_BASE + PERIOD_BUTTON_ID_OFFSET:
        return False
    if iNotifyCode != BN_DOUBLECLICKED:
        return False
    handle_period_left_double_click()
    return True


def build_default_output_path(pszInputPlPath: str) -> str:
    pszScriptDirectory: str = get_script_base_directory()
    pszFileName: str
    _, pszFileName = os.path.split(pszInputPlPath)

    pszStem: str
    pszExt: str
    pszStem, pszExt = os.path.splitext(pszFileName)
    if pszExt == "":
        pszExt = ".tsv"

    pszTargetMarker: str = "損益計算書_"
    pszSuffix: str = "販管費配賦_"
    pszStepMarker: str = "販管費配賦_step0010_"
    pszStepMarkerOld: str = "販管費配賦_step0001_"
    pszStepMarkerPrevious: str = "販管費配賦_step0002_"
    pszStepMarkerCurrent: str = "販管費配賦_step0007_"
    pszStepMarkerNext: str = "販管費配賦_step0008_"
    pszStepMarkerAfterNext: str = "販管費配賦_step0009_"

    if pszStepMarkerOld in pszStem:
        pszOutputStem = pszStem.replace(pszStepMarkerOld, pszStepMarker, 1)
    elif pszStepMarkerPrevious in pszStem:
        pszOutputStem = pszStem.replace(pszStepMarkerPrevious, pszStepMarker, 1)
    elif pszStepMarkerCurrent in pszStem:
        pszOutputStem = pszStem.replace(pszStepMarkerCurrent, pszStepMarker, 1)
    elif pszStepMarkerNext in pszStem:
        pszOutputStem = pszStem.replace(pszStepMarkerNext, pszStepMarker, 1)
    elif pszStepMarkerAfterNext in pszStem:
        pszOutputStem = pszStem.replace(pszStepMarkerAfterNext, pszStepMarker, 1)
    elif pszTargetMarker in pszStem and pszSuffix not in pszStem:
        pszOutputStem = pszStem.replace(pszTargetMarker, pszTargetMarker + pszStepMarker, 1)
    elif pszSuffix in pszStem and pszStepMarker not in pszStem:
        pszOutputStem = pszStem.replace(pszSuffix, pszStepMarker, 1)
    else:
        pszOutputStem = pszStem + "_販管費配賦"

    pszOutputFileName: str = pszOutputStem + pszExt
    pszOutputPath: str = os.path.join(pszScriptDirectory, pszOutputFileName)
    return pszOutputPath


def build_output_path_with_step(pszInputPlPath: str, pszStepMarker: str) -> str:
    pszScriptDirectory: str = get_script_base_directory()
    pszFileName: str
    _, pszFileName = os.path.split(pszInputPlPath)

    pszStem: str
    pszExt: str
    pszStem, pszExt = os.path.splitext(pszFileName)
    if pszExt == "":
        pszExt = ".tsv"

    pszTargetMarker: str = "損益計算書_"
    pszSuffix: str = "販管費配賦_"
    pszStepMarkerOld: str = "販管費配賦_step0001_"
    pszStepMarkerCurrent: str = "販管費配賦_step0002_"
    pszStepMarkerNext: str = "販管費配賦_step0007_"
    pszStepMarkerAfterNext: str = "販管費配賦_step0008_"
    pszStepMarkerAfterAfterNext: str = "販管費配賦_step0009_"

    if pszStepMarkerOld in pszStem:
        pszOutputStem: str = pszStem.replace(pszStepMarkerOld, pszStepMarker, 1)
    elif pszStepMarkerCurrent in pszStem:
        pszOutputStem = pszStem.replace(pszStepMarkerCurrent, pszStepMarker, 1)
    elif pszStepMarkerNext in pszStem:
        pszOutputStem = pszStem.replace(pszStepMarkerNext, pszStepMarker, 1)
    elif pszStepMarkerAfterNext in pszStem:
        pszOutputStem = pszStem.replace(pszStepMarkerAfterNext, pszStepMarker, 1)
    elif pszStepMarkerAfterAfterNext in pszStem:
        pszOutputStem = pszStem.replace(pszStepMarkerAfterAfterNext, pszStepMarker, 1)
    elif pszTargetMarker in pszStem and pszSuffix not in pszStem:
        pszOutputStem = pszStem.replace(pszTargetMarker, pszTargetMarker + pszStepMarker, 1)
    elif pszSuffix in pszStem and pszStepMarker not in pszStem:
        pszOutputStem = pszStem.replace(pszSuffix, pszStepMarker, 1)
    else:
        pszOutputStem = pszStem + "_販管費配賦"

    pszOutputFileName: str = pszOutputStem + pszExt
    pszOutputPath: str = os.path.join(pszScriptDirectory, pszOutputFileName)
    return pszOutputPath


def extract_project_key(pszProjectName: str) -> Optional[str]:
    pszText: str = (pszProjectName or "").strip()
    if pszText == "":
        return None
    if re.match(r"^C\d{3}_", pszText):
        return None

    iUnderscoreIndex: int = pszText.find("_")
    if iUnderscoreIndex <= 0:
        pszKey = pszText.split(" ", 1)[0]
    else:
        pszKey = pszText[:iUnderscoreIndex]
    if pszKey == "":
        return None

    cPrefix: str = pszKey[0]
    if cPrefix in ("A", "C", "J", "P"):
        return pszKey
    return None


def load_manhour_map(pszManhourPath: str) -> Dict[str, List[str]]:
    objManhourMap: Dict[str, List[str]] = {}
    with open(pszManhourPath, "r", encoding="utf-8", newline="") as objInputFile:
        for pszLine in objInputFile:
            pszLineText: str = pszLine.rstrip("\n").rstrip("\r")
            if pszLineText == "":
                continue

            objParts: List[str] = pszLineText.split("\t")
            pszFirstColumn: str = objParts[0] if objParts else ""

            pszKey: Optional[str] = extract_project_key(pszFirstColumn)
            if pszKey is None:
                continue

            objManhourValues: List[str] = objParts[-6:] if len(objParts) >= 7 else []
            if len(objManhourValues) < 6:
                objManhourValues.extend([""] * (6 - len(objManhourValues)))
            objManhourMap[pszKey] = objManhourValues

    return objManhourMap


def load_company_map(pszManhourPath: str) -> Dict[str, str]:
    objCompanyMap: Dict[str, str] = {}
    with open(pszManhourPath, "r", encoding="utf-8", newline="") as objInputFile:
        for pszLine in objInputFile:
            pszLineText: str = pszLine.rstrip("\n").rstrip("\r")
            if pszLineText == "":
                continue
            objParts: List[str] = pszLineText.split("\t")
            if not objParts:
                continue
            pszKey: Optional[str] = extract_project_key(objParts[0])
            if pszKey is None:
                continue
            pszCompany: str = objParts[1] if len(objParts) >= 2 else ""
            objCompanyMap[pszKey] = pszCompany
    return objCompanyMap


def parse_number(pszText: str) -> float:
    pszValue: str = (pszText or "").strip()
    if pszValue == "":
        return 0.0
    try:
        return float(pszValue)
    except ValueError:
        return 0.0


def parse_time_to_seconds(pszTimeText: str) -> float:
    pszValue: str = (pszTimeText or "").strip()
    if pszValue == "":
        return 0.0

    objParts: List[str] = pszValue.split(":")
    if len(objParts) != 3:
        return 0.0
    try:
        iHours: int = int(objParts[0])
        iMinutes: int = int(objParts[1])
        iSeconds: int = int(objParts[2])
    except ValueError:
        return 0.0

    return float(iHours * 3600 + iMinutes * 60 + iSeconds)


def is_time_text_or_blank(pszTimeText: str) -> bool:
    pszValue: str = (pszTimeText or "").strip()
    if pszValue == "":
        return True
    return re.match(r"^\d+:\d{2}:\d{2}$", pszValue) is not None


def format_seconds_as_time_text(fSeconds: float) -> str:
    iTotalSeconds: int = int(round(fSeconds))
    if iTotalSeconds < 0:
        iTotalSeconds = 0
    iHours: int = iTotalSeconds // 3600
    iRemain: int = iTotalSeconds % 3600
    iMinutes: int = iRemain // 60
    iSeconds: int = iRemain % 60
    return f"{iHours}:{iMinutes:02d}:{iSeconds:02d}"


def format_number(fValue: float) -> str:
    if abs(fValue - round(fValue)) < 0.0000001:
        return str(int(round(fValue)))
    pszText: str = f"{fValue:.6f}"
    pszText = pszText.rstrip("0").rstrip(".")
    return pszText


def find_total_row_index(objRows: List[List[str]]) -> int:
    for iRowIndex, objRow in enumerate(objRows):
        pszFirstColumn: str = objRow[0] if objRow else ""
        if pszFirstColumn == "合計":
            return iRowIndex
    return 1


def collect_allocation_target_row_indices(objRows: List[List[str]]) -> List[int]:
    objCompanyPattern = re.compile(r"^C\d{3}(?:_|$)")

    iLastCompanyRowIndex: int = -1
    for iRowIndex in range(1, len(objRows)):
        objRow = objRows[iRowIndex]
        pszFirstColumn: str = (objRow[0] if objRow else "").strip()
        if objCompanyPattern.match(pszFirstColumn):
            iLastCompanyRowIndex = iRowIndex

    if iLastCompanyRowIndex < 0:
        return []

    objTargetRowIndices: List[int] = []
    for iRowIndex in range(iLastCompanyRowIndex + 1, len(objRows)):
        objRow = objRows[iRowIndex]
        pszFirstColumn: str = (objRow[0] if objRow else "").strip()
        if pszFirstColumn == "":
            break
        objTargetRowIndices.append(iRowIndex)
    return objTargetRowIndices


def calculate_allocation(
    objRows: List[List[str]],
    iSellGeneralAdminCostColumnIndex: int,
    iAllocationColumnIndex: int,
    iManhourColumnIndex: int,
    objDeductionCodes: Optional[List[str]] = None,
    bUseHamiltonRounding: bool = True,
) -> None:
    iRowIndexTotal: int = find_total_row_index(objRows)

    fSellGeneralAdminCostTotal: float = 0.0
    if iRowIndexTotal < len(objRows) and iSellGeneralAdminCostColumnIndex >= 0:
        objRowTotal: List[str] = objRows[iRowIndexTotal]
        if iSellGeneralAdminCostColumnIndex < len(objRowTotal):
            fSellGeneralAdminCostTotal = parse_number(objRowTotal[iSellGeneralAdminCostColumnIndex])

    objDeductionSet = set(objDeductionCodes or [])
    fDeductionSum: float = 0.0
    if objDeductionSet:
        objCompanyPattern = re.compile(r"^C(\d{3})(?:_|$)")
        for iRowIndex in range(1, len(objRows)):
            objRow: List[str] = objRows[iRowIndex]
            pszFirstColumn: str = (objRow[0] if objRow else "").strip()
            objMatch = objCompanyPattern.match(pszFirstColumn)
            if objMatch is None:
                continue
            pszCode: str = f"C{objMatch.group(1)}"
            if pszCode not in objDeductionSet:
                continue
            if iSellGeneralAdminCostColumnIndex < len(objRow):
                fDeductionSum += parse_number(objRow[iSellGeneralAdminCostColumnIndex])

    fSellGeneralAdminCostAllocation: float = fSellGeneralAdminCostTotal - fDeductionSum

    objTargetRowIndices: List[int] = collect_allocation_target_row_indices(objRows)
    if not objTargetRowIndices:
        return

    objFilteredTargetRowIndices: List[int] = []
    objManhourSeconds: List[float] = []
    fTotalManhours: float = 0.0
    for iRowIndex in objTargetRowIndices:
        objRow: List[str] = objRows[iRowIndex]
        fManhourSeconds: float = 0.0
        if iManhourColumnIndex < len(objRow):
            fManhourSeconds = parse_time_to_seconds(objRow[iManhourColumnIndex])
        if fManhourSeconds <= 0.0:
            continue
        objFilteredTargetRowIndices.append(iRowIndex)
        objManhourSeconds.append(fManhourSeconds)
        fTotalManhours += fManhourSeconds

    if fTotalManhours <= 0.0:
        return

    objAllocations: List[int] = []
    if bUseHamiltonRounding:
        iTargetTotal: int = int(round(fSellGeneralAdminCostAllocation))
        objRawValues: List[float] = [
            fSellGeneralAdminCostAllocation * fManhourSeconds / fTotalManhours
            for fManhourSeconds in objManhourSeconds
        ]
        objBaseValues: List[int] = [int(fRawValue // 1) for fRawValue in objRawValues]
        iRemain: int = iTargetTotal - sum(objBaseValues)

        objRankIndices: List[int] = list(range(len(objFilteredTargetRowIndices)))
        objRankIndices.sort(
            key=lambda iIndex: (
                objRawValues[iIndex] - objBaseValues[iIndex],
                objManhourSeconds[iIndex],
                -objFilteredTargetRowIndices[iIndex],
            ),
            reverse=True,
        )

        if iRemain > 0:
            for iIndex in objRankIndices[:iRemain]:
                objBaseValues[iIndex] += 1
        elif iRemain < 0:
            objRankIndicesAsc: List[int] = list(reversed(objRankIndices))
            for iIndex in objRankIndicesAsc[:(-iRemain)]:
                objBaseValues[iIndex] -= 1
        objAllocations = objBaseValues
    else:
        objAllocations = [
            int(round(fSellGeneralAdminCostAllocation * fManhourSeconds / fTotalManhours))
            for fManhourSeconds in objManhourSeconds
        ]

    for iTargetIndex, iRowIndex in enumerate(objFilteredTargetRowIndices):
        objRow = objRows[iRowIndex]
        if iAllocationColumnIndex >= len(objRow):
            iAppendCount: int = iAllocationColumnIndex + 1 - len(objRow)
            objRow.extend([""] * iAppendCount)
        objRow[iAllocationColumnIndex] = format_number(float(objAllocations[iTargetIndex]))
        objRows[iRowIndex] = objRow


def load_tsv_rows(pszInputPath: str) -> List[List[str]]:
    objRows: List[List[str]] = []
    with open(pszInputPath, "r", encoding="utf-8", newline="") as objInputFile:
        for pszLine in objInputFile:
            pszLineText: str = pszLine.rstrip("\n").rstrip("\r")
            objRows.append(pszLineText.split("\t") if pszLineText != "" else [""])
    return objRows


def resolve_step0002_column_indices(objRows: List[List[str]]) -> Tuple[int, int, int]:
    iSellGeneralAdminCostColumnIndex: int = -1
    iAllocationColumnIndex: int = -1
    iManhourColumnIndex: int = -1
    if objRows:
        objHeaderRow: List[str] = objRows[0]
        for iColumnIndex, pszColumnName in enumerate(objHeaderRow):
            if pszColumnName == "販売費及び一般管理費計":
                iSellGeneralAdminCostColumnIndex = iColumnIndex
            elif pszColumnName == "配賦販管費":
                iAllocationColumnIndex = iColumnIndex
            elif pszColumnName == "工数":
                iManhourColumnIndex = iColumnIndex
    return (
        iSellGeneralAdminCostColumnIndex,
        iAllocationColumnIndex,
        iManhourColumnIndex,
    )


def write_tsv_rows(pszOutputPath: str, objRows: List[List[str]]) -> None:
    with open(pszOutputPath, "w", encoding="utf-8", newline="") as objOutputFile:
        for objRow in objRows:
            objOutputFile.write("\t".join(objRow) + "\n")


def zero_sell_general_admin_cost_for_step0002_targets(objRows: List[List[str]]) -> None:
    if not objRows:
        return

    iSellGeneralAdminCostColumnIndex: int = -1
    objHeaderRow: List[str] = objRows[0]
    for iColumnIndex, pszColumnName in enumerate(objHeaderRow):
        if pszColumnName == "販売費及び一般管理費計":
            iSellGeneralAdminCostColumnIndex = iColumnIndex
            break
    if iSellGeneralAdminCostColumnIndex < 0:
        return

    objTargetNames: set[str] = {
        "本部",
        "C006_社長室カンパニー販管費",
        "C007_本部カンパニー販管費",
    }
    for iRowIndex in range(1, len(objRows)):
        objRow: List[str] = objRows[iRowIndex]
        pszFirstColumn: str = (objRow[0] if objRow else "").strip()
        if pszFirstColumn not in objTargetNames:
            continue
        if iSellGeneralAdminCostColumnIndex >= len(objRow):
            iAppendCount: int = iSellGeneralAdminCostColumnIndex + 1 - len(objRow)
            objRow.extend([""] * iAppendCount)
        objRow[iSellGeneralAdminCostColumnIndex] = "0"


def zero_sell_general_admin_cost_for_step0006_targets(objRows: List[List[str]]) -> None:
    if not objRows:
        return

    iSellGeneralAdminCostColumnIndex: int = -1
    objHeaderRow: List[str] = objRows[0]
    for iColumnIndex, pszColumnName in enumerate(objHeaderRow):
        if pszColumnName == "販売費及び一般管理費計":
            iSellGeneralAdminCostColumnIndex = iColumnIndex
            break
    if iSellGeneralAdminCostColumnIndex < 0:
        return

    objTargetNames: set[str] = {
        "合計",
        "C001_1Cカンパニー販管費",
        "C002_2Cカンパニー販管費",
        "C003_3Cカンパニー販管費",
        "C004_4Cカンパニー販管費",
        "C005_事業開発カンパニー販管費",
    }
    for iRowIndex in range(1, len(objRows)):
        objRow: List[str] = objRows[iRowIndex]
        pszFirstColumn: str = (objRow[0] if objRow else "").strip()
        if pszFirstColumn not in objTargetNames:
            continue
        if iSellGeneralAdminCostColumnIndex >= len(objRow):
            iAppendCount: int = iSellGeneralAdminCostColumnIndex + 1 - len(objRow)
            objRow.extend([""] * iAppendCount)
        objRow[iSellGeneralAdminCostColumnIndex] = "0"


def build_step0002_variant_path(pszOutputStep0002Path: str, pszSuffix: str) -> str:
    if pszOutputStep0002Path.lower().endswith(".tsv"):
        return pszOutputStep0002Path[: -len(".tsv")] + pszSuffix + ".tsv"
    return pszOutputStep0002Path + pszSuffix + ".tsv"


def generate_step0002_variant_from_step0001(
    pszOutputStep0001Path: str,
    pszOutputStep0002Path: str,
    pszSuffix: str,
    objDeductionCodes: Optional[List[str]],
    bUseHamiltonRounding: bool,
) -> None:
    pszVariantPath: str = build_step0002_variant_path(pszOutputStep0002Path, pszSuffix)
    objRows: List[List[str]] = load_tsv_rows(pszOutputStep0001Path)
    zero_sell_general_admin_cost_for_step0002_targets(objRows)

    (
        iSellGeneralAdminCostColumnIndex,
        iAllocationColumnIndex,
        iManhourColumnIndex,
    ) = resolve_step0002_column_indices(objRows)

    if iSellGeneralAdminCostColumnIndex >= 0 and iAllocationColumnIndex >= 0 and iManhourColumnIndex >= 0:
        calculate_allocation(
            objRows,
            iSellGeneralAdminCostColumnIndex,
            iAllocationColumnIndex,
            iManhourColumnIndex,
            objDeductionCodes,
            bUseHamiltonRounding,
        )

    write_tsv_rows(pszVariantPath, objRows)


def generate_step0002_old_output(
    pszOutputStep0001Path: str,
    pszOutputStep0002Path: str,
) -> None:
    generate_step0002_variant_from_step0001(
        pszOutputStep0001Path,
        pszOutputStep0002Path,
        "_old",
        ["C001", "C002", "C003", "C004", "C005"],
        False,
    )


def generate_step0002_total_output(
    pszOutputStep0001Path: str,
    pszOutputStep0002Path: str,
) -> None:
    generate_step0002_variant_from_step0001(
        pszOutputStep0001Path,
        pszOutputStep0002Path,
        "_合計",
        [],
        True,
    )


def generate_step0002_msd3_09_output(
    pszOutputStep0001Path: str,
    pszOutputStep0002Path: str,
) -> None:
    generate_step0002_variant_from_step0001(
        pszOutputStep0001Path,
        pszOutputStep0002Path,
        "_MSD3_09",
        ["C001", "C002", "C003", "C004", "C005"],
        True,
    )


def generate_step0002_msd3_12_output(
    pszOutputStep0001Path: str,
    pszOutputStep0002Path: str,
) -> None:
    generate_step0002_variant_from_step0001(
        pszOutputStep0001Path,
        pszOutputStep0002Path,
        "_MSD3_12",
        ["C001", "C002", "C003", "C004", "C005", "C006", "C007"],
        True,
    )


def recalculate_operating_profit(
    objRows: List[List[str]],
    iGrossProfitColumnIndex: int,
    iOperatingProfitColumnIndex: int,
    objExcludeColumns: Optional[List[int]] = None,
) -> None:
    if iGrossProfitColumnIndex < 0 or iOperatingProfitColumnIndex < 0:
        return
    if iOperatingProfitColumnIndex <= iGrossProfitColumnIndex:
        return
    objExcludeSet = set(objExcludeColumns or [])

    for iRowIndex in range(1, len(objRows)):
        objRow: List[str] = objRows[iRowIndex]
        if iGrossProfitColumnIndex >= len(objRow):
            continue

        fGrossProfit: float = parse_number(objRow[iGrossProfitColumnIndex])
        fDeductionSum: float = 0.0
        for iColumnIndex in range(iGrossProfitColumnIndex + 1, iOperatingProfitColumnIndex):
            if iColumnIndex in objExcludeSet:
                continue
            if iColumnIndex >= len(objRow):
                continue
            fDeductionSum += parse_number(objRow[iColumnIndex])

        fOperatingProfit: float = fGrossProfit - fDeductionSum
        if iOperatingProfitColumnIndex >= len(objRow):
            iAppendCount: int = iOperatingProfitColumnIndex + 1 - len(objRow)
            objRow.extend([""] * iAppendCount)
        objRow[iOperatingProfitColumnIndex] = format_number(fOperatingProfit)
        objRows[iRowIndex] = objRow


def recalculate_ordinary_profit(
    objRows: List[List[str]],
    iOperatingProfitColumnIndex: int,
    iNonOperatingIncomeColumnIndex: int,
    iNonOperatingExpenseColumnIndex: int,
    iOrdinaryProfitColumnIndex: int,
) -> None:
    if (
        iOperatingProfitColumnIndex < 0
        or iNonOperatingIncomeColumnIndex < 0
        or iNonOperatingExpenseColumnIndex < 0
        or iOrdinaryProfitColumnIndex < 0
    ):
        return
    if iOperatingProfitColumnIndex >= iNonOperatingIncomeColumnIndex:
        return
    if iNonOperatingIncomeColumnIndex >= iNonOperatingExpenseColumnIndex:
        return
    if iNonOperatingExpenseColumnIndex >= iOrdinaryProfitColumnIndex:
        return

    for iRowIndex in range(1, len(objRows)):
        objRow: List[str] = objRows[iRowIndex]
        if iOperatingProfitColumnIndex >= len(objRow):
            continue

        fNonOperatingIncome: float = 0.0
        for iColumnIndex in range(iOperatingProfitColumnIndex + 1, iNonOperatingIncomeColumnIndex):
            if iColumnIndex >= len(objRow):
                continue
            fNonOperatingIncome += parse_number(objRow[iColumnIndex])

        fNonOperatingExpense: float = 0.0
        for iColumnIndex in range(iNonOperatingIncomeColumnIndex + 1, iNonOperatingExpenseColumnIndex):
            if iColumnIndex >= len(objRow):
                continue
            fNonOperatingExpense += parse_number(objRow[iColumnIndex])

        if iNonOperatingIncomeColumnIndex >= len(objRow):
            iAppendCount = iNonOperatingIncomeColumnIndex + 1 - len(objRow)
            objRow.extend([""] * iAppendCount)
        objRow[iNonOperatingIncomeColumnIndex] = format_number(fNonOperatingIncome)

        if iNonOperatingExpenseColumnIndex >= len(objRow):
            iAppendCount = iNonOperatingExpenseColumnIndex + 1 - len(objRow)
            objRow.extend([""] * iAppendCount)
        objRow[iNonOperatingExpenseColumnIndex] = format_number(fNonOperatingExpense)

        fOperatingProfit: float = parse_number(objRow[iOperatingProfitColumnIndex])
        fOrdinaryProfit: float = fOperatingProfit + fNonOperatingIncome - fNonOperatingExpense
        if iOrdinaryProfitColumnIndex >= len(objRow):
            iAppendCount = iOrdinaryProfitColumnIndex + 1 - len(objRow)
            objRow.extend([""] * iAppendCount)
        objRow[iOrdinaryProfitColumnIndex] = format_number(fOrdinaryProfit)
        objRows[iRowIndex] = objRow


def recalculate_pre_tax_profit(
    objRows: List[List[str]],
    iOrdinaryProfitColumnIndex: int,
    iExtraordinaryIncomeColumnIndex: int,
    iExtraordinaryLossColumnIndex: int,
    iPreTaxProfitColumnIndex: int,
) -> None:
    if (
        iOrdinaryProfitColumnIndex < 0
        or iExtraordinaryIncomeColumnIndex < 0
        or iExtraordinaryLossColumnIndex < 0
        or iPreTaxProfitColumnIndex < 0
    ):
        return

    for iRowIndex in range(1, len(objRows)):
        objRow: List[str] = objRows[iRowIndex]
        if iOrdinaryProfitColumnIndex >= len(objRow):
            continue

        fOrdinaryProfit: float = parse_number(objRow[iOrdinaryProfitColumnIndex])
        fExtraordinaryIncome: float = 0.0
        if iExtraordinaryIncomeColumnIndex < len(objRow):
            fExtraordinaryIncome = parse_number(objRow[iExtraordinaryIncomeColumnIndex])
        fExtraordinaryLoss: float = 0.0
        if iExtraordinaryLossColumnIndex < len(objRow):
            fExtraordinaryLoss = parse_number(objRow[iExtraordinaryLossColumnIndex])

        fPreTaxProfit: float = fOrdinaryProfit + fExtraordinaryIncome - fExtraordinaryLoss
        if iPreTaxProfitColumnIndex >= len(objRow):
            iAppendCount: int = iPreTaxProfitColumnIndex + 1 - len(objRow)
            objRow.extend([""] * iAppendCount)
        objRow[iPreTaxProfitColumnIndex] = format_number(fPreTaxProfit)
        objRows[iRowIndex] = objRow


def recalculate_net_profit(
    objRows: List[List[str]],
    iCorporateTaxColumnIndex: int,
    iCorporateTaxTotalColumnIndex: int,
    iPreTaxProfitColumnIndex: int,
    iNetProfitColumnIndex: int,
) -> None:
    if (
        iCorporateTaxColumnIndex < 0
        or iCorporateTaxTotalColumnIndex < 0
        or iPreTaxProfitColumnIndex < 0
        or iNetProfitColumnIndex < 0
    ):
        return

    for iRowIndex in range(1, len(objRows)):
        objRow: List[str] = objRows[iRowIndex]
        if iCorporateTaxColumnIndex >= len(objRow):
            continue

        fCorporateTax: float = parse_number(objRow[iCorporateTaxColumnIndex])
        if iCorporateTaxTotalColumnIndex >= len(objRow):
            iAppendCount: int = iCorporateTaxTotalColumnIndex + 1 - len(objRow)
            objRow.extend([""] * iAppendCount)
        objRow[iCorporateTaxTotalColumnIndex] = format_number(fCorporateTax)

        fPreTaxProfit: float = 0.0
        if iPreTaxProfitColumnIndex < len(objRow):
            fPreTaxProfit = parse_number(objRow[iPreTaxProfitColumnIndex])

        fNetProfit: float = fPreTaxProfit - fCorporateTax
        if iNetProfitColumnIndex >= len(objRow):
            iAppendCount = iNetProfitColumnIndex + 1 - len(objRow)
            objRow.extend([""] * iAppendCount)
        objRow[iNetProfitColumnIndex] = format_number(fNetProfit)
        objRows[iRowIndex] = objRow


def apply_step0006_second_row_totals(objRows: List[List[str]]) -> List[List[str]]:
    if len(objRows) < 2:
        return objRows

    objTargetColumns: List[str] = [
        "配賦販管費",
        "1Cカンパニー販管費",
        "2Cカンパニー販管費",
        "3Cカンパニー販管費",
        "4Cカンパニー販管費",
        "事業開発カンパニー販管費",
    ]
    objHeader: List[str] = objRows[0]
    objOutputRows: List[List[str]] = [list(objRow) for objRow in objRows]
    objTotalRow: List[str] = objOutputRows[1]

    for pszColumnName in objTargetColumns:
        iColumnIndex: int = find_column_index(objHeader, pszColumnName)
        if iColumnIndex < 0:
            continue

        if iColumnIndex >= len(objTotalRow):
            objTotalRow.extend([""] * (iColumnIndex + 1 - len(objTotalRow)))

        fTotalValue: float = 0.0
        for iRowIndex in range(2, len(objOutputRows)):
            objRow: List[str] = objOutputRows[iRowIndex]
            if iColumnIndex < len(objRow):
                fTotalValue += parse_number(objRow[iColumnIndex])

        objTotalRow[iColumnIndex] = format_number(fTotalValue)

    objOutputRows[1] = objTotalRow
    return objOutputRows


def allocate_company_sg_admin_cost(objRows: List[List[str]]) -> List[List[str]]:
    if not objRows:
        return objRows

    objHeader: List[str] = objRows[0]
    objCompanyColumns: List[str] = [
        "1Cカンパニー販管費",
        "2Cカンパニー販管費",
        "3Cカンパニー販管費",
        "4Cカンパニー販管費",
        "事業開発カンパニー販管費",
    ]
    objCompanyManhourColumns: List[str] = [
        "1Cカンパニー販管費の工数",
        "2Cカンパニー販管費の工数",
        "3Cカンパニー販管費の工数",
        "4Cカンパニー販管費の工数",
        "事業開発カンパニー販管費の工数",
    ]
    objCompanyRows: List[str] = [
        "C001_1Cカンパニー販管費",
        "C002_2Cカンパニー販管費",
        "C003_3Cカンパニー販管費",
        "C004_4Cカンパニー販管費",
        "C005_事業開発カンパニー販管費",
    ]

    objCompanyIndices: List[int] = [find_column_index(objHeader, pszName) for pszName in objCompanyColumns]
    objManhourIndices: List[int] = [find_column_index(objHeader, pszName) for pszName in objCompanyManhourColumns]

    objCompanyTotals: List[float] = [0.0] * len(objCompanyColumns)
    for iRowIndex, objRow in enumerate(objRows):
        if iRowIndex == 0:
            continue
        pszRowName: str = objRow[0] if objRow else ""
        if pszRowName in objCompanyRows:
            iCompany: int = objCompanyRows.index(pszRowName)
            iCompanyColumn: int = objCompanyIndices[iCompany]
            if 0 <= iCompanyColumn < len(objRow):
                objCompanyTotals[iCompany] = parse_number(objRow[iCompanyColumn])

    # zero initialize all company cost columns
    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex > 0:
            for iCompanyColumn in objCompanyIndices:
                if iCompanyColumn >= 0:
                    if len(objNewRow) <= iCompanyColumn:
                        objNewRow.extend([""] * (iCompanyColumn + 1 - len(objNewRow)))
                    objNewRow[iCompanyColumn] = "0"
        objOutputRows.append(objNewRow)

    # allocate per company
    for iCompany, pszCompanyColumn in enumerate(objCompanyColumns):
        iCompanyColumn: int = objCompanyIndices[iCompany]
        iManhourColumn: int = objManhourIndices[iCompany]
        if iCompanyColumn < 0 or iManhourColumn < 0:
            continue

        fCompanyTotal: float = objCompanyTotals[iCompany]
        fTotalSeconds: float = 0.0
        for iRowIndex, objRow in enumerate(objOutputRows):
            if iRowIndex == 0 or iManhourColumn >= len(objRow):
                continue
            fSeconds: float = parse_time_to_seconds(objRow[iManhourColumn])
            if fSeconds > 0.0:
                fTotalSeconds += fSeconds

        if fTotalSeconds <= 0.0:
            continue

        objTargetRows: List[int] = []
        objTargetSeconds: List[float] = []
        for iRowIndex, objRow in enumerate(objOutputRows):
            if iRowIndex == 0 or iManhourColumn >= len(objRow) or iCompanyColumn >= len(objRow):
                continue
            fSeconds = parse_time_to_seconds(objRow[iManhourColumn])
            if fSeconds <= 0.0:
                continue
            objTargetRows.append(iRowIndex)
            objTargetSeconds.append(fSeconds)

        if not objTargetRows:
            continue

        iTargetTotal: int = int(round(fCompanyTotal))
        objRawValues: List[float] = [
            fCompanyTotal * fSeconds / fTotalSeconds for fSeconds in objTargetSeconds
        ]
        objBaseValues: List[int] = [int(fRawValue // 1) for fRawValue in objRawValues]
        iRemain: int = iTargetTotal - sum(objBaseValues)

        objRankIndices: List[int] = list(range(len(objTargetRows)))
        objRankIndices.sort(
            key=lambda iIndex: (
                objRawValues[iIndex] - objBaseValues[iIndex],
                objTargetSeconds[iIndex],
                -objTargetRows[iIndex],
            ),
            reverse=True,
        )

        if iRemain > 0:
            for iIndex in objRankIndices[:iRemain]:
                objBaseValues[iIndex] += 1
        elif iRemain < 0:
            objRankIndicesAsc: List[int] = list(reversed(objRankIndices))
            for iIndex in objRankIndicesAsc[:(-iRemain)]:
                objBaseValues[iIndex] -= 1

        for iTargetIndex, iRowIndex in enumerate(objTargetRows):
            objRow = objOutputRows[iRowIndex]
            objRow[iCompanyColumn] = format_number(float(objBaseValues[iTargetIndex]))
            objOutputRows[iRowIndex] = objRow

    return objOutputRows


def _build_pj_summary_group_total_paths() -> Tuple[str, str]:
    pszScriptDirectory: str = os.path.dirname(os.path.abspath(__file__))
    pszTemplatePath: str = os.path.join(
        pszScriptDirectory,
        "TEMPLATE_PJサマリ_グループ別合計.xlsx",
    )
    pszOutputPath: str = os.path.join(
        pszScriptDirectory,
        "PJサマリ",
        "PJサマリ_グループ別合計.xlsx",
    )
    return pszTemplatePath, pszOutputPath


def _build_pj_summary_company_total_paths(pszOrgMode: str) -> Tuple[str, str]:
    pszScriptDirectory: str = os.path.dirname(os.path.abspath(__file__))
    if pszOrgMode == "new":
        pszTemplatePath: str = os.path.join(
            pszScriptDirectory,
            "TEMPLATE_PJサマリ_Div別合計.xlsx",
        )
        pszOutputPath: str = os.path.join(
            pszScriptDirectory,
            "PJサマリ",
            "PJサマリ_Div別合計.xlsx",
        )
    else:
        pszTemplatePath = os.path.join(
            pszScriptDirectory,
            "TEMPLATE_PJサマリ_カンパニー別合計.xlsx",
        )
        pszOutputPath = os.path.join(
            pszScriptDirectory,
            "PJサマリ",
            "PJサマリ_カンパニー別合計.xlsx",
        )
    return pszTemplatePath, pszOutputPath


def _build_pj_summary_group_sheet_name(
    objStart: Tuple[int, int],
    objEnd: Tuple[int, int],
) -> str:
    pszSummaryStartMonth: str = f"{objStart[1]:02d}"
    pszSummaryEndMonth: str = f"{objEnd[1]:02d}"
    return f"{objStart[0]}年{pszSummaryStartMonth}月-{objEnd[0]}年{pszSummaryEndMonth}月"


def insert_step0006_rows_into_group_summary_excel(
    objRows: List[List[str]],
    objStart: Tuple[int, int],
    objEnd: Tuple[int, int],
) -> None:
    pszTemplatePath, pszOutputPath = _build_pj_summary_group_total_paths()
    pszBaseSheetName: str = _build_pj_summary_group_sheet_name(objStart, objEnd)
    pszSheetName: str = f"グループ別損益_{pszBaseSheetName}"
    if not os.path.isfile(pszTemplatePath):
        return
    if os.path.isfile(pszOutputPath):
        objWorkbook = load_workbook(pszOutputPath)
    else:
        objWorkbook = load_workbook(pszTemplatePath)
    if pszSheetName not in objWorkbook.sheetnames:
        pszSourceSheetName: str = "Sheet1" if objStart[1] == 4 else "Sheet2"
        if pszSourceSheetName in objWorkbook.sheetnames:
            objWorkbook[pszSourceSheetName].title = pszSheetName
    if pszSheetName not in objWorkbook.sheetnames:
        return
    objSheet = objWorkbook[pszSheetName]

    for iRow, objRow in enumerate(objRows, start=1):
        for iCol, pszValue in enumerate(objRow, start=1):
            objCellValue = parse_tsv_value_for_excel(pszValue)
            objSheet.cell(row=iRow, column=iCol, value=objCellValue)

    os.makedirs(os.path.dirname(pszOutputPath), exist_ok=True)
    objWorkbook.save(pszOutputPath)
    if EXECUTION_ROOT_DIRECTORY:
        pszGroupProfitDirectory = os.path.join(
            EXECUTION_ROOT_DIRECTORY,
            "グループ別損益",
        )
        os.makedirs(pszGroupProfitDirectory, exist_ok=True)
        shutil.copy2(
            pszOutputPath,
            os.path.join(pszGroupProfitDirectory, os.path.basename(pszOutputPath)),
        )


def insert_step0006_rows_into_company_summary_excel(
    objRows: List[List[str]],
    objStart: Tuple[int, int],
    objEnd: Tuple[int, int],
    pszOrgMode: str,
) -> None:
    pszTemplatePath, pszOutputPath = _build_pj_summary_company_total_paths(pszOrgMode)
    pszBaseSheetName: str = _build_pj_summary_group_sheet_name(objStart, objEnd)
    pszSheetName: str = (
        f"Div別損益_{pszBaseSheetName}" if pszOrgMode == "new" else pszBaseSheetName
    )
    if not os.path.isfile(pszTemplatePath):
        return
    if os.path.isfile(pszOutputPath):
        objWorkbook = load_workbook(pszOutputPath)
    else:
        objWorkbook = load_workbook(pszTemplatePath)
    if pszSheetName not in objWorkbook.sheetnames:
        pszSourceSheetName: str = "Sheet1" if objStart[1] == 4 else "Sheet2"
        if pszSourceSheetName in objWorkbook.sheetnames:
            objWorkbook[pszSourceSheetName].title = pszSheetName
    if pszSheetName not in objWorkbook.sheetnames:
        return
    objSheet = objWorkbook[pszSheetName]

    for iRow, objRow in enumerate(objRows, start=1):
        for iCol, pszValue in enumerate(objRow, start=1):
            objCellValue = parse_tsv_value_for_excel(pszValue)
            objSheet.cell(row=iRow, column=iCol, value=objCellValue)

    os.makedirs(os.path.dirname(pszOutputPath), exist_ok=True)
    objWorkbook.save(pszOutputPath)
    if EXECUTION_ROOT_DIRECTORY:
        pszOutputSubDirectoryName: str = "Div別損益" if pszOrgMode == "new" else "カンパニー別損益"
        pszCompanyProfitDirectory = os.path.join(
            EXECUTION_ROOT_DIRECTORY,
            pszOutputSubDirectoryName,
        )
        os.makedirs(pszCompanyProfitDirectory, exist_ok=True)
        shutil.copy2(
            pszOutputPath,
            os.path.join(pszCompanyProfitDirectory, os.path.basename(pszOutputPath)),
        )

def insert_company_sg_admin_cost_columns(objRows: List[List[str]]) -> List[List[str]]:
    if not objRows:
        return objRows

    objHeader: List[str] = objRows[0]
    iAllocationIndex: int = find_column_index(objHeader, "配賦販管費")
    iOperatingProfitIndex: int = find_column_index(objHeader, "営業利益")
    if iAllocationIndex < 0 or iOperatingProfitIndex < 0:
        return objRows

    iInsertIndex: int = iAllocationIndex + 1
    objNewColumns: List[str] = [
        "1Cカンパニー販管費",
        "2Cカンパニー販管費",
        "3Cカンパニー販管費",
        "4Cカンパニー販管費",
        "事業開発カンパニー販管費",
    ]
    objNewHeader: List[str] = (
        objHeader[:iInsertIndex] + objNewColumns + objHeader[iInsertIndex:]
    )

    iGrossProfitIndex: int = find_column_index(objNewHeader, "売上総利益")
    iSellGeneralAdminTotalIndex: int = find_column_index(objNewHeader, "販売費及び一般管理費計")

    objTargetMap: Dict[str, str] = {
        "C001_1Cカンパニー販管費": "1Cカンパニー販管費",
        "C002_2Cカンパニー販管費": "2Cカンパニー販管費",
        "C003_3Cカンパニー販管費": "3Cカンパニー販管費",
        "C004_4Cカンパニー販管費": "4Cカンパニー販管費",
        "C005_事業開発カンパニー販管費": "事業開発カンパニー販管費",
    }
    objTargetColumnIndices: Dict[str, int] = {
        pszColumnName: find_column_index(objNewHeader, pszColumnName)
        for pszColumnName in objNewColumns
    }

    objOutputRows: List[List[str]] = [objNewHeader]
    for objRow in objRows[1:]:
        objNewRow: List[str] = objRow[:iInsertIndex] + [""] * len(objNewColumns) + objRow[iInsertIndex:]

        pszRowName: str = objNewRow[0] if objNewRow else ""
        pszTargetColumn: Optional[str] = objTargetMap.get(pszRowName)
        if (
            pszTargetColumn is not None
            and iGrossProfitIndex >= 0
            and iSellGeneralAdminTotalIndex > iGrossProfitIndex + 1
        ):
            fSum: float = 0.0
            iEndIndex: int = min(iSellGeneralAdminTotalIndex, len(objNewRow))
            for iColumnIndex in range(iGrossProfitIndex + 1, iEndIndex):
                fSum += parse_number(objNewRow[iColumnIndex])
            iTargetIndex: int = objTargetColumnIndices.get(pszTargetColumn, -1)
            if iTargetIndex >= 0:
                objNewRow[iTargetIndex] = format_number(fSum)

        if len(objNewRow) < len(objNewHeader):
            objNewRow.extend([""] * (len(objNewHeader) - len(objNewRow)))
        objOutputRows.append(objNewRow)

    return objOutputRows


def process_pl_tsv(
    pszPlPath: str,
    pszOutputPath: str,
    pszOutputStep0001Path: str,
    pszOutputStep0002Path: str,
    pszOutputStep0003ZeroPath: str,
    pszOutputStep0007Path: str,
    pszOutputStep0008Path: str,
    pszOutputStep0009Path: str,
    pszOutputStep0005Path: str,
    pszOutputStep0006Path: str,
    pszOutputStep0010Path: str,
    pszOutputFinalPath: str,
    objManhourMap: Dict[str, List[str]],
    objCompanyMap: Dict[str, str],
) -> None:
    objRows: List[List[str]] = []
    with open(pszPlPath, "r", encoding="utf-8", newline="") as objInputFile:
        for pszLine in objInputFile:
            pszLineText: str = pszLine.rstrip("\n").rstrip("\r")
            objRows.append(pszLineText.split("\t") if pszLineText != "" else [""])

    for iRowIndex, objRow in enumerate(objRows):
        pszFirstColumn: str = objRow[0] if objRow else ""
        if iRowIndex == 0:
            if len(objRow) == 0:
                objRow = [""]
            objRow.extend(
                [
                    "工数",
                    "1Cカンパニー販管費の工数",
                    "2Cカンパニー販管費の工数",
                    "3Cカンパニー販管費の工数",
                    "4Cカンパニー販管費の工数",
                    "事業開発カンパニー販管費の工数",
                ]
            )
            objRows[iRowIndex] = objRow
            continue

        pszKey: Optional[str] = extract_project_key(pszFirstColumn)
        if pszKey is None:
            continue

        objManhours: List[str] = objManhourMap.get(pszKey, [])
        if len(objManhours) < 6:
            objManhours = objManhours + ["0:00:00"] * (6 - len(objManhours))

        objRow.extend(objManhours[:6])
        objRows[iRowIndex] = objRow

    with open(pszOutputStep0001Path, "w", encoding="utf-8", newline="") as objOutputFile:
        for objRow in objRows:
            objOutputFile.write("\t".join(objRow) + "\n")

    zero_sell_general_admin_cost_for_step0002_targets(objRows)

    (
        iSellGeneralAdminCostColumnIndex,
        iAllocationColumnIndex,
        iManhourColumnIndex,
    ) = resolve_step0002_column_indices(objRows)

    if iSellGeneralAdminCostColumnIndex >= 0 and iAllocationColumnIndex >= 0 and iManhourColumnIndex >= 0:
        calculate_allocation(
            objRows,
            iSellGeneralAdminCostColumnIndex,
            iAllocationColumnIndex,
            iManhourColumnIndex,
            ["C001", "C002", "C003", "C004", "C005"],
            True,
        )

    write_tsv_rows(pszOutputStep0002Path, objRows)

    generate_step0002_old_output(pszOutputStep0001Path, pszOutputStep0002Path)
    generate_step0002_total_output(pszOutputStep0001Path, pszOutputStep0002Path)
    generate_step0002_msd3_09_output(pszOutputStep0001Path, pszOutputStep0002Path)
    generate_step0002_msd3_12_output(pszOutputStep0001Path, pszOutputStep0002Path)

    # step0004の処理
    # ここから
    objZeroRows: List[List[str]] = [list(objRow) for objRow in objRows]
    if objZeroRows:
        objHeaderZero: List[str] = objZeroRows[0]
        objTargetColumns: List[str] = [
            "1Cカンパニー販管費の工数",
            "2Cカンパニー販管費の工数",
            "3Cカンパニー販管費の工数",
            "4Cカンパニー販管費の工数",
            "事業開発カンパニー販管費の工数",
        ]
        objTargetIndices: List[int] = [
            find_column_index(objHeaderZero, pszColumn) for pszColumn in objTargetColumns
        ]
        for iRowIndex, objRow in enumerate(objZeroRows):
            if iRowIndex < 3:
                continue
            for iColumnIndex in objTargetIndices:
                if 0 <= iColumnIndex < len(objRow):
                    objRow[iColumnIndex] = "0:00:00"
            objZeroRows[iRowIndex] = objRow

    with open(pszOutputStep0003ZeroPath, "w", encoding="utf-8", newline="") as objOutputFile:
        for objRow in objZeroRows:
            objOutputFile.write("\t".join(objRow) + "\n")

    pszOutputStep0004Path: str = pszOutputStep0003ZeroPath.replace("step0003_", "step0004_", 1)
    iManhourColumnIndexZero: int = find_column_index(objZeroRows[0], "工数") if objZeroRows else -1
    objTargetColumnsZero: List[str] = [
        "1Cカンパニー販管費の工数",
        "2Cカンパニー販管費の工数",
        "3Cカンパニー販管費の工数",
        "4Cカンパニー販管費の工数",
        "事業開発カンパニー販管費の工数",
    ]
    objTargetIndicesZero: List[int] = [
        find_column_index(objZeroRows[0], pszColumn) if objZeroRows else -1 for pszColumn in objTargetColumnsZero
    ]
    bSeenHeadquarter: bool = False
    for iRowIndex, objRow in enumerate(objZeroRows):
        if not objRow:
            continue
        pszName: str = objRow[0]
        if pszName == "本部":
            bSeenHeadquarter = True
            continue
        if not bSeenHeadquarter:
            continue
        if pszName.startswith("C"):
            for iColumnIndex in objTargetIndicesZero:
                if 0 <= iColumnIndex < len(objRow):
                    objRow[iColumnIndex] = "0:00:00"
            objZeroRows[iRowIndex] = objRow
            continue
        pszKey: Optional[str] = extract_project_key(pszName)
        if pszKey is None:
            continue
        pszCompany: str = objCompanyMap.get(pszKey, "")
        iTargetColumn: int = -1
        if pszCompany == "第一インキュ":
            iTargetColumn = objTargetIndicesZero[0] if len(objTargetIndicesZero) > 0 else -1
        elif pszCompany == "第二インキュ":
            iTargetColumn = objTargetIndicesZero[1] if len(objTargetIndicesZero) > 1 else -1
        elif pszCompany == "第三インキュ":
            iTargetColumn = objTargetIndicesZero[2] if len(objTargetIndicesZero) > 2 else -1
        elif pszCompany == "第四インキュ":
            iTargetColumn = objTargetIndicesZero[3] if len(objTargetIndicesZero) > 3 else -1
        elif pszCompany == "事業開発":
            iTargetColumn = objTargetIndicesZero[4] if len(objTargetIndicesZero) > 4 else -1
        if iTargetColumn >= 0:
            if len(objRow) <= iTargetColumn:
                objRow.extend([""] * (iTargetColumn + 1 - len(objRow)))
            pszManhourValue: str = "0:00:00"
            if iManhourColumnIndexZero >= 0 and iManhourColumnIndexZero < len(objRow):
                pszManhourValue = objRow[iManhourColumnIndexZero] or "0:00:00"
            for iColumnIndex in objTargetIndicesZero:
                if 0 <= iColumnIndex < len(objRow):
                    objRow[iColumnIndex] = "0:00:00"
            objRow[iTargetColumn] = pszManhourValue
        else:
            for iColumnIndex in objTargetIndicesZero:
                if 0 <= iColumnIndex < len(objRow):
                    objRow[iColumnIndex] = "0:00:00"
        objZeroRows[iRowIndex] = objRow

    with open(pszOutputStep0004Path, "w", encoding="utf-8", newline="") as objOutputFile:
        for objRow in objZeroRows:
            objOutputFile.write("\t".join(objRow) + "\n")
    # step0004の処理
    # ここまで

    iGrossProfitColumnIndex: int = -1
    iOperatingProfitColumnIndex: int = -1
    if objRows:
        objHeaderRow = objRows[0]
        for iColumnIndex, pszColumnName in enumerate(objHeaderRow):
            if pszColumnName == "売上総利益":
                iGrossProfitColumnIndex = iColumnIndex
            elif pszColumnName == "営業利益":
                iOperatingProfitColumnIndex = iColumnIndex

    if iGrossProfitColumnIndex >= 0 and iOperatingProfitColumnIndex >= 0:
        recalculate_operating_profit(
            objRows,
            iGrossProfitColumnIndex,
            iOperatingProfitColumnIndex,
            [],
        )

    iNonOperatingIncomeColumnIndex: int = -1
    iNonOperatingExpenseColumnIndex: int = -1
    iOrdinaryProfitColumnIndex: int = -1
    if objRows:
        objHeaderRow = objRows[0]
        for iColumnIndex, pszColumnName in enumerate(objHeaderRow):
            if pszColumnName == "営業外収益":
                iNonOperatingIncomeColumnIndex = iColumnIndex
            elif pszColumnName == "営業外費用":
                iNonOperatingExpenseColumnIndex = iColumnIndex
            elif pszColumnName == "経常利益":
                iOrdinaryProfitColumnIndex = iColumnIndex

    if (
        iOperatingProfitColumnIndex >= 0
        and iNonOperatingIncomeColumnIndex >= 0
        and iNonOperatingExpenseColumnIndex >= 0
        and iOrdinaryProfitColumnIndex >= 0
    ):
        recalculate_ordinary_profit(
            objRows,
            iOperatingProfitColumnIndex,
            iNonOperatingIncomeColumnIndex,
            iNonOperatingExpenseColumnIndex,
            iOrdinaryProfitColumnIndex,
        )

    iExtraordinaryIncomeColumnIndex: int = -1
    iExtraordinaryLossColumnIndex: int = -1
    iPreTaxProfitColumnIndex: int = -1
    if objRows:
        objHeaderRow = objRows[0]
        for iColumnIndex, pszColumnName in enumerate(objHeaderRow):
            if pszColumnName == "特別利益":
                iExtraordinaryIncomeColumnIndex = iColumnIndex
            elif pszColumnName == "特別損失":
                iExtraordinaryLossColumnIndex = iColumnIndex
            elif pszColumnName == "税引前当期純利益":
                iPreTaxProfitColumnIndex = iColumnIndex

    if (
        iOrdinaryProfitColumnIndex >= 0
        and iExtraordinaryIncomeColumnIndex >= 0
        and iExtraordinaryLossColumnIndex >= 0
        and iPreTaxProfitColumnIndex >= 0
    ):
        recalculate_pre_tax_profit(
            objRows,
            iOrdinaryProfitColumnIndex,
            iExtraordinaryIncomeColumnIndex,
            iExtraordinaryLossColumnIndex,
            iPreTaxProfitColumnIndex,
        )

    objRows = insert_company_sg_admin_cost_columns(objRows)

    with open(pszOutputStep0005Path, "w", encoding="utf-8", newline="") as objOutputFile:
        for objRow in objRows:
            objOutputFile.write("\t".join(objRow) + "\n")

    objRows = allocate_company_sg_admin_cost(objRows)
    zero_sell_general_admin_cost_for_step0006_targets(objRows)

    with open(pszOutputStep0006Path, "w", encoding="utf-8", newline="") as objOutputFile:
        for objRow in objRows:
            objOutputFile.write("\t".join(objRow) + "\n")

    # step0007: 営業利益の再計算（入力は step0006）
    objStep0007Rows: List[List[str]] = [list(objRow) for objRow in objRows]
    iGrossProfitColumnIndex: int = -1
    iOperatingProfitColumnIndex: int = -1
    iSellGeneralAdminTotalIndex: int = -1
    if objStep0007Rows:
        objHeaderRow = objStep0007Rows[0]
        for iColumnIndex, pszColumnName in enumerate(objHeaderRow):
            if pszColumnName == "売上総利益":
                iGrossProfitColumnIndex = iColumnIndex
            elif pszColumnName == "営業利益":
                iOperatingProfitColumnIndex = iColumnIndex
            elif pszColumnName == "販売費及び一般管理費計":
                iSellGeneralAdminTotalIndex = iColumnIndex

    if iGrossProfitColumnIndex >= 0 and iOperatingProfitColumnIndex >= 0:
        recalculate_operating_profit(
            objStep0007Rows,
            iGrossProfitColumnIndex,
            iOperatingProfitColumnIndex,
            [iSellGeneralAdminTotalIndex] if iSellGeneralAdminTotalIndex >= 0 else [],
        )

    with open(pszOutputStep0007Path, "w", encoding="utf-8", newline="") as objOutputFile:
        for objRow in objStep0007Rows:
            objOutputFile.write("\t".join(objRow) + "\n")

    # step0008: 営業外収益・費用、経常利益の再計算（入力は step0007）
    objStep0008Rows: List[List[str]] = [list(objRow) for objRow in objStep0007Rows]
    iNonOperatingIncomeColumnIndex: int = -1
    iNonOperatingExpenseColumnIndex: int = -1
    iOrdinaryProfitColumnIndex: int = -1
    if objStep0008Rows:
        objHeaderRow = objStep0008Rows[0]
        for iColumnIndex, pszColumnName in enumerate(objHeaderRow):
            if pszColumnName == "営業外収益":
                iNonOperatingIncomeColumnIndex = iColumnIndex
            elif pszColumnName == "営業外費用":
                iNonOperatingExpenseColumnIndex = iColumnIndex
            elif pszColumnName == "経常利益":
                iOrdinaryProfitColumnIndex = iColumnIndex

    if (
        iOperatingProfitColumnIndex >= 0
        and iNonOperatingIncomeColumnIndex >= 0
        and iNonOperatingExpenseColumnIndex >= 0
        and iOrdinaryProfitColumnIndex >= 0
    ):
        recalculate_ordinary_profit(
            objStep0008Rows,
            iOperatingProfitColumnIndex,
            iNonOperatingIncomeColumnIndex,
            iNonOperatingExpenseColumnIndex,
            iOrdinaryProfitColumnIndex,
        )

    with open(pszOutputStep0008Path, "w", encoding="utf-8", newline="") as objOutputFile:
        for objRow in objStep0008Rows:
            objOutputFile.write("\t".join(objRow) + "\n")

    # step0009: 税引前当期純利益の再計算（入力は step0008）
    objStep0009Rows: List[List[str]] = [list(objRow) for objRow in objStep0008Rows]
    iExtraordinaryIncomeColumnIndex: int = -1
    iExtraordinaryLossColumnIndex: int = -1
    iPreTaxProfitColumnIndex: int = -1
    if objStep0009Rows:
        objHeaderRow = objStep0009Rows[0]
        for iColumnIndex, pszColumnName in enumerate(objHeaderRow):
            if pszColumnName == "特別利益":
                iExtraordinaryIncomeColumnIndex = iColumnIndex
            elif pszColumnName == "特別損失":
                iExtraordinaryLossColumnIndex = iColumnIndex
            elif pszColumnName == "税引前当期純利益":
                iPreTaxProfitColumnIndex = iColumnIndex

    if (
        iOrdinaryProfitColumnIndex >= 0
        and iExtraordinaryIncomeColumnIndex >= 0
        and iExtraordinaryLossColumnIndex >= 0
        and iPreTaxProfitColumnIndex >= 0
    ):
        recalculate_pre_tax_profit(
            objStep0009Rows,
            iOrdinaryProfitColumnIndex,
            iExtraordinaryIncomeColumnIndex,
            iExtraordinaryLossColumnIndex,
            iPreTaxProfitColumnIndex,
        )

    with open(pszOutputStep0009Path, "w", encoding="utf-8", newline="") as objOutputFile:
        for objRow in objStep0009Rows:
            objOutputFile.write("\t".join(objRow) + "\n")

    objStep0010Rows: List[List[str]] = [list(objRow) for objRow in objStep0009Rows]
    iCorporateTaxColumnIndexStep0010: int = -1
    iCorporateTaxTotalColumnIndexStep0010: int = -1
    iNetProfitColumnIndexStep0010: int = -1
    iPreTaxProfitColumnIndexStep0010: int = -1
    if objStep0010Rows:
        objHeaderRow = objStep0010Rows[0]
        for iColumnIndex, pszColumnName in enumerate(objHeaderRow):
            if pszColumnName == "法人税、住民税及び事業税":
                iCorporateTaxColumnIndexStep0010 = iColumnIndex
            elif pszColumnName == "法人税等":
                iCorporateTaxTotalColumnIndexStep0010 = iColumnIndex
            elif pszColumnName == "税引前当期純利益":
                iPreTaxProfitColumnIndexStep0010 = iColumnIndex
            elif pszColumnName == "当期純利益":
                iNetProfitColumnIndexStep0010 = iColumnIndex

    if (
        iCorporateTaxColumnIndexStep0010 >= 0
        and iCorporateTaxTotalColumnIndexStep0010 >= 0
        and iPreTaxProfitColumnIndexStep0010 >= 0
        and iNetProfitColumnIndexStep0010 >= 0
    ):
        recalculate_net_profit(
            objStep0010Rows,
            iCorporateTaxColumnIndexStep0010,
            iCorporateTaxTotalColumnIndexStep0010,
            iPreTaxProfitColumnIndexStep0010,
            iNetProfitColumnIndexStep0010,
        )

    with open(pszOutputStep0010Path, "w", encoding="utf-8", newline="") as objOutputFile:
        for objRow in objStep0010Rows:
            objOutputFile.write("\t".join(objRow) + "\n")

    write_transposed_tsv(pszOutputStep0010Path)
    pszOutputStep0010HorizontalPath: str = pszOutputStep0010Path.replace("_vertical", "")
    move_files_to_temp_and_copy_back(
        [pszOutputStep0010Path, pszOutputStep0010HorizontalPath],
        get_script_base_directory(),
    )

    with open(pszOutputFinalPath, "w", encoding="utf-8", newline="") as objOutputFile:
        for objRow in objStep0010Rows:
            objOutputFile.write("\t".join(objRow) + "\n")
    write_transposed_tsv(pszOutputFinalPath)


def transpose_rows(objRows: List[List[str]]) -> List[List[str]]:
    if not objRows:
        return []
    iMaxColumns: int = max(len(objRow) for objRow in objRows)
    objNormalized: List[List[str]] = []
    for objRow in objRows:
        objNormalized.append(objRow + [""] * (iMaxColumns - len(objRow)))

    objTransposed: List[List[str]] = []
    for iColumnIndex in range(iMaxColumns):
        objTransposed.append([objRow[iColumnIndex] for objRow in objNormalized])
    return objTransposed


def write_transposed_tsv(pszInputPath: str) -> None:
    pszDirectory: str
    pszFileName: str
    pszDirectory, pszFileName = os.path.split(pszInputPath)
    pszOutputFileName: str = pszFileName.replace("_vertical", "")
    pszOutputPath: str = os.path.join(pszDirectory, pszOutputFileName)

    objRows: List[List[str]] = []
    with open(pszInputPath, "r", encoding="utf-8", newline="") as objInputFile:
        for pszLine in objInputFile:
            pszLineText: str = pszLine.rstrip("\n").rstrip("\r")
            objRows.append(pszLineText.split("\t"))

    objTransposed = transpose_rows(objRows)
    with open(pszOutputPath, "w", encoding="utf-8", newline="") as objOutputFile:
        for objRow in objTransposed:
            objOutputFile.write("\t".join(objRow) + "\n")


def move_files_to_temp_and_copy_back(objFilePaths: List[str], pszBaseDirectory: str) -> None:
    if not objFilePaths:
        return

    pszTempDirectory: str = os.path.join(pszBaseDirectory, "temp")
    os.makedirs(pszTempDirectory, exist_ok=True)

    for pszFilePath in objFilePaths:
        if not os.path.isfile(pszFilePath):
            continue
        pszFileName: str = os.path.basename(pszFilePath)
        pszTempPath: str = os.path.join(pszTempDirectory, pszFileName)
        shutil.move(pszFilePath, pszTempPath)
        shutil.copy2(pszTempPath, os.path.join(pszBaseDirectory, pszFileName))


def move_files_to_temp(objFilePaths: List[str], pszBaseDirectory: str) -> None:
    if not objFilePaths:
        return

    pszTempDirectory: str = os.path.join(pszBaseDirectory, "temp")
    os.makedirs(pszTempDirectory, exist_ok=True)

    for pszFilePath in objFilePaths:
        if not os.path.isfile(pszFilePath):
            continue
        pszFileName: str = os.path.basename(pszFilePath)
        pszTempPath: str = os.path.join(pszTempDirectory, pszFileName)
        shutil.move(pszFilePath, pszTempPath)


def move_cp_step_tsv_files_to_temp_subfolders(pszBaseDirectory: str) -> None:
    pszTempDirectory: str = os.path.join(pszBaseDirectory, "temp")
    os.makedirs(pszTempDirectory, exist_ok=True)

    objSelectedRangePath: Optional[str] = find_selected_range_path(pszBaseDirectory)
    objSelectedRange: Optional[Tuple[Tuple[int, int], Tuple[int, int]]] = None
    if objSelectedRangePath is not None:
        objSelectedRange = parse_selected_range(objSelectedRangePath)

    objMonths: List[str] = []
    objCumulativeRanges: List[str] = []

    def format_month_label(objYearMonth: Tuple[int, int]) -> str:
        iYear, iMonth = objYearMonth
        return f"{iYear}年{iMonth:02d}月"

    def format_range_label(objRange: Tuple[Tuple[int, int], Tuple[int, int]]) -> str:
        return f"{format_month_label(objRange[0])}-{format_month_label(objRange[1])}"

    if objSelectedRange is not None:
        objMonths = [format_month_label(objMonth) for objMonth in build_month_sequence(*objSelectedRange)]

        objRangeItems: List[Tuple[Tuple[int, int], Tuple[int, int]]] = []

        def append_unique_range(objRangeItem: Tuple[Tuple[int, int], Tuple[int, int]]) -> None:
            if objRangeItem not in objRangeItems:
                objRangeItems.append(objRangeItem)

        append_unique_range(objSelectedRange)
        for objRangeItem in split_by_fiscal_boundary(objSelectedRange[0], objSelectedRange[1], 3):
            append_unique_range(objRangeItem)
        for objRangeItem in split_by_fiscal_boundary(objSelectedRange[0], objSelectedRange[1], 8):
            append_unique_range(objRangeItem)

        objCumulativeRanges = [format_range_label(objRangeItem) for objRangeItem in objRangeItems]

    if not objMonths:
        objMonths = [
            "2025年04月",
            "2025年05月",
            "2025年06月",
            "2025年07月",
            "2025年08月",
            "2025年09月",
            "2025年10月",
            "2025年11月",
            "2025年12月",
        ]
    if not objCumulativeRanges:
        objCumulativeRanges = [
            "2025年04月-2025年08月",
            "2025年04月-2025年12月",
            "2025年09月-2025年12月",
        ]

    def build_step0001_to_step0004_names(pszPrefix: str, iStartStep: int) -> List[str]:
        objNames: List[str] = []
        for iStep in range(iStartStep, 5):
            pszStep: str = f"step{iStep:04d}"
            for pszMonth in objMonths:
                objNames.append(
                    f"{pszPrefix}{pszStep}_単月_損益計算書_{pszMonth}.tsv"
                )
            for pszRange in objCumulativeRanges:
                objNames.append(
                    f"{pszPrefix}{pszStep}_累計_損益計算書_{pszRange}.tsv"
                )
        return objNames

    def build_step0005_names(pszPrefix: str) -> List[str]:
        objNames: List[str] = []
        for pszMonth in objMonths:
            objNames.append(
                f"{pszPrefix}step0005_単月_損益計算書_{pszMonth}_vertical.tsv"
            )
        for pszRange in objCumulativeRanges:
            objNames.append(
                f"{pszPrefix}step0005_累計_損益計算書_{pszRange}_vertical.tsv"
            )
        return objNames

    objConfig: List[Tuple[str, List[str]]] = [
        (
            "0001_CP別_step0001-0005",
            build_step0001_to_step0004_names("0001_CP別_", 1)
            + build_step0005_names("0001_CP別_"),
        ),
        (
            "0002_CP別_step0001-0005",
            build_step0001_to_step0004_names("0002_CP別_", 2)
            + build_step0005_names("0002_CP別_"),
        ),
    ]

    for pszFolderName, objFileNames in objConfig:
        pszTargetDirectory: str = os.path.join(pszTempDirectory, pszFolderName)
        os.makedirs(pszTargetDirectory, exist_ok=True)
        for pszFileName in objFileNames:
            pszSourcePath: str = os.path.join(pszBaseDirectory, pszFileName)
            if not os.path.isfile(pszSourcePath):
                continue
            pszDestinationPath: str = os.path.join(pszTargetDirectory, pszFileName)
            shutil.move(pszSourcePath, pszDestinationPath)

    objStep0004VerticalPatterns: List[Tuple[re.Pattern[str], str]] = [
        (
            re.compile(r"^0001_CP別_step0004_(?:単月|累計)_損益計算書_.*_vertical\.tsv$"),
            "0001_CP別_step0001-0005",
        ),
        (
            re.compile(r"^0002_CP別_step0004_(?:単月|累計)_損益計算書_.*_vertical\.tsv$"),
            "0002_CP別_step0001-0005",
        ),
    ]
    for pszFileName in sorted(os.listdir(pszTempDirectory)):
        pszSourcePath: str = os.path.join(pszTempDirectory, pszFileName)
        if not os.path.isfile(pszSourcePath):
            continue
        pszTargetFolderName: Optional[str] = None
        for objPattern, pszFolderName in objStep0004VerticalPatterns:
            if objPattern.match(pszFileName):
                pszTargetFolderName = pszFolderName
                break
        if pszTargetFolderName is None:
            continue
        pszTargetDirectory = os.path.join(pszTempDirectory, pszTargetFolderName)
        os.makedirs(pszTargetDirectory, exist_ok=True)
        pszDestinationPath: str = os.path.join(pszTargetDirectory, pszFileName)
        if os.path.exists(pszDestinationPath):
            os.remove(pszDestinationPath)
        shutil.move(pszSourcePath, pszDestinationPath)


def move_pl_tsv_files_into_income_statement_temp_subfolder(pszBaseDirectory: str) -> None:
    pszTempDirectory: str = os.path.join(pszBaseDirectory, "temp")
    pszTargetDirectory: str = os.path.join(pszTempDirectory, "損益計算書_販管費配賦後")
    os.makedirs(pszTargetDirectory, exist_ok=True)

    objPattern = re.compile(r"^(損益計算書_販管費配賦_|累計_損益計算書_販管費配賦_).*.tsv$")
    objSourceDirectories: List[str] = [pszBaseDirectory, pszTempDirectory]
    for pszSourceDirectory in objSourceDirectories:
        if not os.path.isdir(pszSourceDirectory):
            continue
        for pszFileName in sorted(os.listdir(pszSourceDirectory)):
            if not objPattern.match(pszFileName):
                continue
            pszSourcePath: str = os.path.join(pszSourceDirectory, pszFileName)
            if not os.path.isfile(pszSourcePath):
                continue
            pszDestinationPath: str = os.path.join(pszTargetDirectory, pszFileName)
            if os.path.abspath(pszSourcePath) == os.path.abspath(pszDestinationPath):
                continue
            if os.path.exists(pszDestinationPath):
                os.remove(pszDestinationPath)
            shutil.move(pszSourcePath, pszDestinationPath)


def move_monthly_income_statement_tsv_files_into_temp_subfolder(pszBaseDirectory: str) -> None:
    pszTempDirectory: str = os.path.join(pszBaseDirectory, "temp")
    pszTargetDirectory: str = os.path.join(pszTempDirectory, "損益計算書系")
    os.makedirs(pszTargetDirectory, exist_ok=True)

    objPatterns: List[re.Pattern[str]] = [
        re.compile(r"^損益計算書_\d{4}年\d{2}月.*\.tsv$"),
        re.compile(r"^累計_損益計算書_.*\.tsv$"),
    ]
    for pszFileName in sorted(os.listdir(pszTempDirectory)):
        if not any(objPattern.match(pszFileName) for objPattern in objPatterns):
            continue
        if pszFileName.startswith("累計_損益計算書_販管費配賦_"):
            continue
        pszSourcePath: str = os.path.join(pszTempDirectory, pszFileName)
        if not os.path.isfile(pszSourcePath):
            continue
        pszDestinationPath: str = os.path.join(pszTargetDirectory, pszFileName)
        if os.path.exists(pszDestinationPath):
            os.remove(pszDestinationPath)
        shutil.move(pszSourcePath, pszDestinationPath)


def move_cost_report_tsv_files_into_temp_subfolder(pszBaseDirectory: str) -> None:
    pszTempDirectory: str = os.path.join(pszBaseDirectory, "temp")
    pszTargetDirectory: str = os.path.join(pszTempDirectory, "製造原価報告書系")
    os.makedirs(pszTargetDirectory, exist_ok=True)

    objPatterns: List[re.Pattern[str]] = [
        re.compile(r"^累計_製造原価報告書_.*\.tsv$"),
        re.compile(r"^製造原価報告書_.*\.tsv$"),
    ]
    for pszFileName in sorted(os.listdir(pszTempDirectory)):
        if not any(objPattern.match(pszFileName) for objPattern in objPatterns):
            continue
        pszSourcePath: str = os.path.join(pszTempDirectory, pszFileName)
        if not os.path.isfile(pszSourcePath):
            continue
        pszDestinationPath: str = os.path.join(pszTargetDirectory, pszFileName)
        if os.path.exists(pszDestinationPath):
            os.remove(pszDestinationPath)
        shutil.move(pszSourcePath, pszDestinationPath)


def move_step0007_split_files_into_0003_pj_summary_temp_subfolder(pszBaseDirectory: str) -> None:
    pszTempDirectory: str = os.path.join(pszBaseDirectory, "temp")
    pszTargetDirectory: str = os.path.join(pszTempDirectory, "0003_PJサマリ")
    os.makedirs(pszTargetDirectory, exist_ok=True)

    objFileNames: List[str] = [
        "0003_PJサマリ_step0007_単月_0001.tsv",
        "0003_PJサマリ_step0007_単月_0002.tsv",
        "0003_PJサマリ_step0007_単月_0003.tsv",
        "0003_PJサマリ_step0007_単月_PL_CR.tsv",
        "0003_PJサマリ_step0007_累計_0001.tsv",
        "0003_PJサマリ_step0007_累計_0002.tsv",
        "0003_PJサマリ_step0007_累計_0003.tsv",
        "0003_PJサマリ_step0007_累計_PL_CR.tsv",
    ]

    for pszFileName in objFileNames:
        pszSourcePath: str = os.path.join(pszTempDirectory, pszFileName)
        if not os.path.isfile(pszSourcePath):
            continue
        pszDestinationPath: str = os.path.join(pszTargetDirectory, pszFileName)
        if os.path.exists(pszDestinationPath):
            os.remove(pszDestinationPath)
        shutil.move(pszSourcePath, pszDestinationPath)


def move_pj_summary_tsv_files_to_temp_subfolders(pszBaseDirectory: str) -> None:
    pszTempDirectory: str = os.path.join(pszBaseDirectory, "temp")
    os.makedirs(pszTempDirectory, exist_ok=True)

    objFolderPatterns: List[Tuple[str, List[re.Pattern[str]]]] = [
        (
            "0001_PJサマリ",
            [
                re.compile(r"^0001_PJサマリ_step0001_.*_(単月|累計)_(損益計算書|製造原価報告書)\.tsv$"),
                re.compile(r"^0001_PJサマリ_step000[2-5]_.*_(単月|累計)_損益計算書\.tsv$"),
                re.compile(r"^0001_PJサマリ_step000[6-9]_.*_単月・累計_損益計算書\.tsv$"),
            ],
        ),
        (
            "0002_PJサマリ",
            [
                re.compile(r"^0002_PJサマリ_step000[12]_(単月|累計)_粗利金額ランキング\.tsv$"),
                re.compile(r"^0002_PJサマリ_step0007_(単月|累計|単月・累計)_粗利金額ランキング\.tsv$"),
                re.compile(r"^0002_PJサマリ_step00(08|09|10)_単月・累計_粗利金額ランキング\.tsv$"),
                re.compile(r"^0002_PJサマリ_単月・累計_粗利金額ランキング\.tsv$"),
            ],
        ),
        (
            "0003_PJサマリ",
            [
                re.compile(r"^0003_PJサマリ_step000[1-3]_(単月|累計)_(損益計算書|製造原価報告書)\.tsv$"),
                re.compile(r"^0003_PJサマリ_step0004_(単月|累計)_(損益計算書|製造原価報告書)(_vertical)?\.tsv$"),
                re.compile(r"^0003_PJサマリ_step0005_(単月|累計)_(損益計算書|製造原価報告書)_E∪F(_vertical)?\.tsv$"),
                re.compile(r"^0003_PJサマリ_step0006_(単月|累計)_(損益計算書|製造原価報告書)_E∪F\.tsv$"),
                re.compile(r"^0003_PJサマリ_step0007_(単月|累計)_PL_CR\.tsv$"),
            ],
        ),
        (
            "0004_PJサマリ",
            [
                re.compile(r"^0004_PJサマリ_step000[1-5]_単月_損益計算書_\d{4}年\d{2}月\.tsv$"),
                re.compile(r"^0004_PJサマリ_step000[1-5]_累計_損益計算書_\d{4}年\d{2}月-\d{4}年\d{2}月\.tsv$"),
                re.compile(r"^0004_PJサマリ_step000[67]_単・累_損益計算書_\d{4}年\d{2}月-\d{4}年\d{2}月\.tsv$"),
            ],
        ),
        (
            "0005_PJサマリ",
            [
                re.compile(r"^0005_PJサマリ_step000[1-4]_単月_損益計算書_\d{4}年\d{2}月\.tsv$"),
                re.compile(r"^0005_PJサマリ_step000[1-4]_累計_損益計算書_\d{4}年\d{2}月-\d{4}年\d{2}月\.tsv$"),
                re.compile(r"^0005_PJサマリ_step000[5-7]_単・累_損益計算書_\d{4}年\d{2}月-\d{4}年\d{2}月\.tsv$"),
            ],
        ),
    ]

    objFileNames: List[str] = sorted(os.listdir(pszBaseDirectory))
    for pszFolderName, objPatterns in objFolderPatterns:
        pszTargetDirectory: str = os.path.join(pszTempDirectory, pszFolderName)
        os.makedirs(pszTargetDirectory, exist_ok=True)
        for pszFileName in objFileNames:
            if not pszFileName.lower().endswith(".tsv"):
                continue
            if not any(objPattern.match(pszFileName) for objPattern in objPatterns):
                continue
            pszSourcePath: str = os.path.join(pszBaseDirectory, pszFileName)
            if not os.path.isfile(pszSourcePath):
                continue
            pszDestinationPath: str = os.path.join(pszTargetDirectory, pszFileName)
            shutil.move(pszSourcePath, pszDestinationPath)


def move_cp_step_folders_to_temp(pszBaseDirectory: str) -> None:
    pszTempDirectory: str = os.path.join(pszBaseDirectory, "temp")
    os.makedirs(pszTempDirectory, exist_ok=True)

    objTargetFolderNames: List[str] = [
        "0001_CP別_step0006",
        "0001_CP別_step0007",
        "0001_CP別_step0008",
        "0001_CP別_step0009",
        "0002_CP別_step0006",
        "0002_CP別_step0007",
        "0002_CP別_step0008",
        "0002_CP別_step0009",
        "PJ_Summary_step0008_Project",
        "PJ_Summary_step0009_Project",
        "PJ_Summary_step0010_Project",
        "PJ_Summary_step0011_Project",
        "PJサマリ",
        "DragAndDropManhourAndPl",
        #// -----------------------------------------------------------------------------
        #// 設計方針:
        #// - log ディレクトリは運用上「移動対象外」とする（恒久方針）。
        #// - 理由: 実行中に log 配下ファイルが他プロセスで使用中（ロック中）となり得るため、
        #//         move 対象にすると PermissionError による処理失敗の原因になる。
        #// - 期待動作: log はベースディレクトリ直下に残置し、他の対象フォルダのみ temp へ移動する。
        #// - 注意: 本方針は一時回避ではなく仕様固定。将来も log を移動対象へ戻さない。
        #// -----------------------------------------------------------------------------
        #        "log",
    ]

    for pszFolderName in objTargetFolderNames:
        pszSourcePath: str = os.path.join(pszBaseDirectory, pszFolderName)
        if not os.path.isdir(pszSourcePath):
            continue
        pszDestinationPath: str = os.path.join(pszTempDirectory, pszFolderName)
        if os.path.isdir(pszDestinationPath):
            shutil.rmtree(pszDestinationPath)
        shutil.move(pszSourcePath, pszDestinationPath)


def remove_bycompany_managementcontrol_step0005_directory() -> None:
    pszDirectoryPath: str = os.path.join(
        get_script_base_directory(),
        "ByCompany_ManagementControl_step0005",
    )
    if not os.path.isdir(pszDirectoryPath):
        return
    shutil.rmtree(pszDirectoryPath)


def find_selected_range_path(pszBaseDirectory: str) -> Optional[str]:
    objFileNames: List[str] = [
        "SellGeneralAdminCost_Allocation_Cmd_SelectedRange.txt",
        "SellGeneralAdminCost_Allocation_DnD_SelectedRange.txt",
    ]
    objCandidates: List[str] = []
    for pszFileName in objFileNames:
        objCandidates.append(os.path.join(pszBaseDirectory, pszFileName))
        objCandidates.append(os.path.join(os.path.dirname(__file__), pszFileName))
    for pszCandidate in objCandidates:
        if os.path.isfile(pszCandidate):
            return pszCandidate
    return None


def parse_selected_range(pszRangePath: str) -> Optional[Tuple[Tuple[int, int], Tuple[int, int]]]:
    try:
        with open(pszRangePath, "r", encoding="utf-8", newline="") as objFile:
            pszContent: str = objFile.read()
    except OSError:
        return None

    objMatch = re.search(r"開始:\s*(\d{4})/(\d{2}).*?終了:\s*(\d{4})/(\d{2})", pszContent, re.DOTALL)
    if objMatch is None:
        objMatch = re.search(r"採用範囲:\s*(\d{4})年(\d{1,2})月〜(\d{4})年(\d{1,2})月", pszContent)
        if objMatch is None:
            return None
        iStartYear: int = int(objMatch.group(1))
        iStartMonth: int = int(objMatch.group(2))
        iEndYear: int = int(objMatch.group(3))
        iEndMonth: int = int(objMatch.group(4))
    else:
        iStartYear = int(objMatch.group(1))
        iStartMonth = int(objMatch.group(2))
        iEndYear = int(objMatch.group(3))
        iEndMonth = int(objMatch.group(4))

    if not (1 <= iStartMonth <= 12 and 1 <= iEndMonth <= 12):
        return None
    return (iStartYear, iStartMonth), (iEndYear, iEndMonth)


def extract_year_month_from_path(pszPath: str) -> Optional[Tuple[int, int]]:
    pszBaseName: str = os.path.basename(pszPath)
    objMatch = re.search(r"_(\d{4})年(\d{1,2})月", pszBaseName)
    if objMatch is None:
        return None
    try:
        iYear: int = int(objMatch.group(1))
        iMonth: int = int(objMatch.group(2))
    except ValueError:
        return None
    if iMonth < 1 or iMonth > 12:
        return None
    return iYear, iMonth


def find_best_continuous_range(objYearMonths: List[Tuple[int, int]]) -> Optional[Tuple[Tuple[int, int], Tuple[int, int]]]:
    if not objYearMonths:
        return None

    objSorted: List[Tuple[int, int]] = sorted(set(objYearMonths))
    objBestRange: Optional[Tuple[Tuple[int, int], Tuple[int, int]]] = None
    iBestLength: int = 0
    objCurrentStart: Tuple[int, int] = objSorted[0]
    objCurrentEnd: Tuple[int, int] = objSorted[0]

    for objMonth in objSorted[1:]:
        if next_year_month(*objCurrentEnd) == objMonth:
            objCurrentEnd = objMonth
            continue

        iCurrentLength: int = _range_length(objCurrentStart, objCurrentEnd)
        objBestRange, iBestLength = _update_best_range(objBestRange, iBestLength, objCurrentStart, objCurrentEnd)
        objCurrentStart = objMonth
        objCurrentEnd = objMonth

    objBestRange, _ = _update_best_range(objBestRange, iBestLength, objCurrentStart, objCurrentEnd)
    return objBestRange


def _range_length(objStart: Tuple[int, int], objEnd: Tuple[int, int]) -> int:
    iYearStart, iMonthStart = objStart
    iYearEnd, iMonthEnd = objEnd
    return (iYearEnd * 12 + iMonthEnd) - (iYearStart * 12 + iMonthStart) + 1


def _update_best_range(
    objCurrentBest: Optional[Tuple[Tuple[int, int], Tuple[int, int]]],
    iCurrentBestLength: int,
    objCandidateStart: Tuple[int, int],
    objCandidateEnd: Tuple[int, int],
) -> Tuple[Optional[Tuple[Tuple[int, int], Tuple[int, int]]], int]:
    iCandidateLength: int = _range_length(objCandidateStart, objCandidateEnd)
    if iCandidateLength > iCurrentBestLength:
        return (objCandidateStart, objCandidateEnd), iCandidateLength
    if iCandidateLength == iCurrentBestLength and objCurrentBest is not None:
        if objCandidateEnd > objCurrentBest[1]:
            return (objCandidateStart, objCandidateEnd), iCandidateLength
    if objCurrentBest is None:
        return (objCandidateStart, objCandidateEnd), iCandidateLength
    return objCurrentBest, iCurrentBestLength


def ensure_selected_range_file(pszDirectory: str, objRange: Tuple[Tuple[int, int], Tuple[int, int]]) -> str:
    iStartYear, iStartMonth = objRange[0]
    iEndYear, iEndMonth = objRange[1]
    pszOutputPath: str = os.path.join(pszDirectory, "SellGeneralAdminCost_Allocation_Cmd_SelectedRange.txt")
    pszStartText: str = f"{iStartYear:04d}/{iStartMonth:02d}"
    pszEndText: str = f"{iEndYear:04d}/{iEndMonth:02d}"
    objLines: List[str] = [
        "採用範囲:",
        f"開始: {pszStartText}",
        f"終了: {pszEndText}",
    ]
    with open(pszOutputPath, "w", encoding="utf-8", newline="") as objFile:
        objFile.write("\n".join(objLines) + "\n")

    def format_period_block(
        pszLabel: str,
        objRanges: List[Tuple[Tuple[int, int], Tuple[int, int]]],
    ) -> List[str]:
        objResultLines: List[str] = [f"{pszLabel}:", ""]
        if len(objRanges) >= 3:
            (iTwoPeriodsAgoStartYear, iTwoPeriodsAgoStartMonth), (iTwoPeriodsAgoEndYear, iTwoPeriodsAgoEndMonth) = objRanges[-3]
            objResultLines.extend(
                [
                    "2期前(前期の前期)",
                    f"開始: {iTwoPeriodsAgoStartYear:04d}/{iTwoPeriodsAgoStartMonth:02d}",
                    f"終了: {iTwoPeriodsAgoEndYear:04d}/{iTwoPeriodsAgoEndMonth:02d}",
                    "",
                ]
            )
        else:
            objResultLines.extend(["2期前(前期の前期)", "なし。", ""])
        if len(objRanges) >= 2:
            (iPriorStartYear, iPriorStartMonth), (iPriorEndYear, iPriorEndMonth) = objRanges[-2]
            objResultLines.extend(
                [
                    "前期",
                    f"開始: {iPriorStartYear:04d}/{iPriorStartMonth:02d}",
                    f"終了: {iPriorEndYear:04d}/{iPriorEndMonth:02d}",
                    "",
                ]
            )
        else:
            objResultLines.extend(["前期", "なし。", ""])
        if objRanges:
            (iCurrentStartYear, iCurrentStartMonth), (iCurrentEndYear, iCurrentEndMonth) = objRanges[-1]
            objResultLines.extend(
                [
                    "当期",
                    f"開始: {iCurrentStartYear:04d}/{iCurrentStartMonth:02d}",
                    f"終了: {iCurrentEndYear:04d}/{iCurrentEndMonth:02d}",
                ]
            )
        else:
            objResultLines.extend(["当期", "なし。"])
        return objResultLines

    objStart, objEnd = objRange
    objFiscalARanges = split_by_fiscal_boundary(objStart, objEnd, 3)
    objFiscalBRanges = split_by_fiscal_boundary(objStart, objEnd, 8)
    objAccountPeriodLines: List[str] = []
    objAccountPeriodLines.extend(format_period_block("3月決算の会計期間", objFiscalARanges))
    objAccountPeriodLines.append("")
    objAccountPeriodLines.extend(format_period_block("8月決算の会計期間", objFiscalBRanges))

    pszAccountPeriodPath: str = os.path.join(
        pszDirectory,
        "SellGeneralAdminCost_Allocation_Cmd_AccountPeriodRange.txt",
    )
    with open(pszAccountPeriodPath, "w", encoding="utf-8", newline="") as objFile:
        objFile.write("\n".join(objAccountPeriodLines) + "\n")

    write_cp_previous_period_range_file(pszDirectory, objRange)
    if EXECUTION_ROOT_DIRECTORY:
        pszPeriodDirectory = os.path.join(EXECUTION_ROOT_DIRECTORY, "期間")
        os.makedirs(pszPeriodDirectory, exist_ok=True)
        pszSelectedRangeCopyPath = os.path.join(pszPeriodDirectory, os.path.basename(pszOutputPath))
        pszAccountPeriodCopyPath = os.path.join(pszPeriodDirectory, os.path.basename(pszAccountPeriodPath))
        shutil.copy2(
            pszOutputPath,
            pszSelectedRangeCopyPath,
        )
        shutil.copy2(
            pszAccountPeriodPath,
            pszAccountPeriodCopyPath,
        )
        if os.path.isfile(pszSelectedRangeCopyPath) and os.path.isfile(pszAccountPeriodCopyPath):
            pszMergedPath = os.path.join(
                pszPeriodDirectory,
                "SellGeneralAdminCost_Allocation_Cmd_SelectedRange_And_AccountPeriodRange.txt",
            )
            with open(pszSelectedRangeCopyPath, "r", encoding="utf-8") as objFile:
                pszSelectedRangeText = objFile.read()
            with open(pszAccountPeriodCopyPath, "r", encoding="utf-8") as objFile:
                pszAccountPeriodText = objFile.read()
            with open(pszMergedPath, "w", encoding="utf-8", newline="") as objFile:
                objFile.write(pszSelectedRangeText)
                objFile.write("\n")
                objFile.write(pszAccountPeriodText)
    return pszOutputPath


def record_created_file(_pszPath: str) -> None:
    return


def month_to_ordinal(objMonth: Tuple[int, int]) -> int:
    iYear, iMonth = objMonth
    return iYear * 12 + iMonth


def is_month_in_range(objMonth: Tuple[int, int], objRange: Tuple[Tuple[int, int], Tuple[int, int]]) -> bool:
    iValue: int = month_to_ordinal(objMonth)
    iStart: int = month_to_ordinal(objRange[0])
    iEnd: int = month_to_ordinal(objRange[1])
    return iStart <= iValue <= iEnd


def next_year_month(iYear: int, iMonth: int) -> Tuple[int, int]:
    iMonth += 1
    if iMonth > 12:
        iMonth = 1
        iYear += 1
    return iYear, iMonth


def build_month_sequence(
    objStart: Tuple[int, int],
    objEnd: Tuple[int, int],
) -> List[Tuple[int, int]]:
    objMonths: List[Tuple[int, int]] = []
    iYear, iMonth = objStart
    while True:
        objMonths.append((iYear, iMonth))
        if (iYear, iMonth) == objEnd:
            break
        iYear, iMonth = next_year_month(iYear, iMonth)
    return objMonths


def split_by_fiscal_boundary(
    objStart: Tuple[int, int],
    objEnd: Tuple[int, int],
    iBoundaryEndMonth: int,
) -> List[Tuple[Tuple[int, int], Tuple[int, int]]]:
    objMonths = build_month_sequence(objStart, objEnd)
    if not objMonths:
        return []

    objRanges: List[Tuple[Tuple[int, int], Tuple[int, int]]] = []
    objRangeStart: Tuple[int, int] = objMonths[0]
    for iIndex, objMonth in enumerate(objMonths):
        iYear, iMonth = objMonth
        is_last: bool = iIndex == len(objMonths) - 1
        if iMonth == iBoundaryEndMonth and not is_last:
            objRanges.append((objRangeStart, objMonth))
            objRangeStart = objMonths[iIndex + 1]
    objRanges.append((objRangeStart, objMonths[-1]))
    return objRanges


def shift_year_of_period_range(
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
    iYearOffset: int,
) -> Tuple[Tuple[int, int], Tuple[int, int]]:
    (iStartYear, iStartMonth), (iEndYear, iEndMonth) = objRange
    return (iStartYear + iYearOffset, iStartMonth), (iEndYear + iYearOffset, iEndMonth)


def build_cp_previous_period_range_from_selected_range(
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
    iBoundaryEndMonth: int,
) -> Optional[Tuple[Tuple[int, int], Tuple[int, int]]]:
    objStart, objEnd = objRange
    objFiscalRanges = split_by_fiscal_boundary(objStart, objEnd, iBoundaryEndMonth)
    if not objFiscalRanges:
        return None
    objCurrentRange = objFiscalRanges[-1]
    return shift_year_of_period_range(objCurrentRange, -1)




def build_cp_current_period_range_from_selected_range(
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
    iBoundaryEndMonth: int,
) -> Optional[Tuple[Tuple[int, int], Tuple[int, int]]]:
    objStart, objEnd = objRange
    objFiscalRanges = split_by_fiscal_boundary(objStart, objEnd, iBoundaryEndMonth)
    if not objFiscalRanges:
        return None
    return objFiscalRanges[-1]

def write_cp_previous_period_range_file(
    pszDirectory: str,
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
) -> str:
    def build_prior_block(
        pszLabel: str,
        iBoundaryEndMonth: int,
    ) -> List[str]:
        objPriorRange = build_cp_previous_period_range_from_selected_range(objRange, iBoundaryEndMonth)
        objCurrentRange = build_cp_current_period_range_from_selected_range(objRange, iBoundaryEndMonth)
        objResultLines: List[str] = [f"{pszLabel}:", "", "前期"]
        if objPriorRange is not None and is_month_in_range(objPriorRange[0], objRange) and is_month_in_range(objPriorRange[1], objRange):
            (iStartYear, iStartMonth), (iEndYear, iEndMonth) = objPriorRange
            objResultLines.extend(
                [
                    f"開始: {iStartYear:04d}/{iStartMonth:02d}",
                    f"終了: {iEndYear:04d}/{iEndMonth:02d}",
                    "",
                ]
            )
        else:
            objResultLines.extend(["なし。", ""])
        objResultLines.append("当期")
        if objCurrentRange is not None and is_month_in_range(objCurrentRange[0], objRange) and is_month_in_range(objCurrentRange[1], objRange):
            (iStartYear, iStartMonth), (iEndYear, iEndMonth) = objCurrentRange
            objResultLines.extend(
                [
                    f"開始: {iStartYear:04d}/{iStartMonth:02d}",
                    f"終了: {iEndYear:04d}/{iEndMonth:02d}",
                ]
            )
        else:
            objResultLines.append("なし。")
        return objResultLines

    pszOutputPath: str = os.path.join(
        pszDirectory,
        "SellGeneralAdminCost_Allocation_Cmd_CP別用PreviousPeriodRange.txt",
    )
    objLines: List[str] = []
    objLines.extend(build_prior_block("3月決算の会計期間", 3))
    objLines.append("")
    objLines.extend(build_prior_block("8月決算の会計期間", 8))

    with open(pszOutputPath, "w", encoding="utf-8", newline="") as objFile:
        objFile.write("\n".join(objLines) + "\n")

    if EXECUTION_ROOT_DIRECTORY:
        pszPeriodDirectory = os.path.join(EXECUTION_ROOT_DIRECTORY, "期間")
        os.makedirs(pszPeriodDirectory, exist_ok=True)
        pszCopyPath = os.path.join(pszPeriodDirectory, os.path.basename(pszOutputPath))
        shutil.copy2(pszOutputPath, pszCopyPath)

    return pszOutputPath


def build_cp_period_ranges_from_selected_range(
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
) -> List[Tuple[Tuple[int, int], Tuple[int, int]]]:
    objStart, objEnd = objRange
    objResult: List[Tuple[Tuple[int, int], Tuple[int, int]]] = []

    def add_range(objRangeItem: Tuple[Tuple[int, int], Tuple[int, int]]) -> None:
        if objRangeItem not in objResult:
            objResult.append(objRangeItem)

    objFiscalARanges = split_by_fiscal_boundary(objStart, objEnd, 3)
    if objFiscalARanges:
        add_range(objFiscalARanges[-1])

    objFiscalBRanges = split_by_fiscal_boundary(objStart, objEnd, 8)
    if objFiscalBRanges:
        add_range(objFiscalBRanges[-1])

    return objResult


def build_cp_period_ranges_from_previous_period_range_file(
    pszDirectory: str,
) -> Tuple[List[Tuple[Tuple[int, int], Tuple[int, int]]], List[Tuple[Tuple[int, int], Tuple[int, int]]]]:
    pszPath: str = os.path.join(
        pszDirectory,
        "SellGeneralAdminCost_Allocation_Cmd_CP別用PreviousPeriodRange.txt",
    )
    if not os.path.isfile(pszPath):
        return [], []

    try:
        with open(pszPath, "r", encoding="utf-8", newline="") as objFile:
            objLines: List[str] = [pszLine.strip() for pszLine in objFile.readlines()]
    except OSError:
        return [], []

    objPriorRanges: List[Tuple[Tuple[int, int], Tuple[int, int]]] = []
    objCurrentRanges: List[Tuple[Tuple[int, int], Tuple[int, int]]] = []

    def parse_range_lines(iIndexStart: int) -> Optional[Tuple[Tuple[int, int], Tuple[int, int]]]:
        iStartLineIndex: int = iIndexStart + 1
        while iStartLineIndex < len(objLines) and objLines[iStartLineIndex] == "":
            iStartLineIndex += 1

        iEndLineIndex: int = iStartLineIndex + 1
        while iEndLineIndex < len(objLines) and objLines[iEndLineIndex] == "":
            iEndLineIndex += 1

        if iStartLineIndex >= len(objLines) or iEndLineIndex >= len(objLines):
            return None

        objStartMatch = re.match(r"^開始:\s*(\d{4})/(\d{2})$", objLines[iStartLineIndex])
        objEndMatch = re.match(r"^終了:\s*(\d{4})/(\d{2})$", objLines[iEndLineIndex])
        if objStartMatch is None or objEndMatch is None:
            return None
        iStartYear = int(objStartMatch.group(1))
        iStartMonth = int(objStartMatch.group(2))
        iEndYear = int(objEndMatch.group(1))
        iEndMonth = int(objEndMatch.group(2))
        if not (1 <= iStartMonth <= 12 and 1 <= iEndMonth <= 12):
            return None
        if (iStartYear, iStartMonth) > (iEndYear, iEndMonth):
            return None
        return (iStartYear, iStartMonth), (iEndYear, iEndMonth)

    def append_unique(
        objTarget: List[Tuple[Tuple[int, int], Tuple[int, int]]],
        objRangeItem: Tuple[Tuple[int, int], Tuple[int, int]],
    ) -> None:
        if objRangeItem not in objTarget:
            objTarget.append(objRangeItem)

    iIndex: int = 0
    while iIndex < len(objLines):
        pszLine: str = objLines[iIndex]
        if pszLine == "前期":
            objPriorRange = parse_range_lines(iIndex)
            if objPriorRange is not None:
                append_unique(objPriorRanges, objPriorRange)
            iIndex += 1
            continue
        if pszLine == "当期":
            objCurrentRange = parse_range_lines(iIndex)
            if objCurrentRange is not None:
                append_unique(objCurrentRanges, objCurrentRange)
            iIndex += 1
            continue
        iIndex += 1

    objAllRanges: List[Tuple[Tuple[int, int], Tuple[int, int]]] = []
    for objRangeItem in objPriorRanges:
        append_unique(objAllRanges, objRangeItem)
    for objRangeItem in objCurrentRanges:
        append_unique(objAllRanges, objRangeItem)

    return objAllRanges, objCurrentRanges


def build_current_period_ranges_for_pj_summary_totals(
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
) -> List[Tuple[Tuple[int, int], Tuple[int, int]]]:
    objRanges = build_cp_period_ranges_from_selected_range(objRange)
    _, objSelectedEnd = objRange
    objCurrentRanges: List[Tuple[Tuple[int, int], Tuple[int, int]]] = []
    for objRangeItem in objRanges:
        if objRangeItem[1] != objSelectedEnd:
            continue
        if objRangeItem in objCurrentRanges:
            continue
        objCurrentRanges.append(objRangeItem)
    return objCurrentRanges


def try_parse_float(pszText: str) -> Optional[float]:
    pszValue: str = (pszText or "").strip()
    if pszValue == "":
        return None
    try:
        return float(pszValue)
    except ValueError:
        return None


def read_tsv_rows(pszPath: str) -> List[List[str]]:
    objRows: List[List[str]] = []
    with open(pszPath, "r", encoding="utf-8", newline="") as objFile:
        for pszLine in objFile:
            pszLineText: str = pszLine.rstrip("\n").rstrip("\r")
            objRows.append(pszLineText.split("\t") if pszLineText != "" else [""])
    return objRows


def sum_tsv_rows(objBaseRows: List[List[str]], objAddRows: List[List[str]]) -> List[List[str]]:
    if not objBaseRows:
        return [list(objRow) for objRow in objAddRows]
    if not objAddRows:
        return objBaseRows

    objBaseKeyIndices: Dict[str, List[int]] = {}
    for iRowIndex, objRow in enumerate(objBaseRows):
        pszKey: str = objRow[0] if objRow else ""
        objBaseKeyIndices.setdefault(pszKey, []).append(iRowIndex)

    objBaseKeyCursor: Dict[str, int] = {pszKey: 0 for pszKey in objBaseKeyIndices}

    for iRowIndex, objAddRow in enumerate(objAddRows):
        if iRowIndex == 0:
            objBaseHeader: List[str] = objBaseRows[0]
            iColumnCount: int = max(len(objBaseHeader), len(objAddRow))
            if len(objBaseHeader) < iColumnCount:
                objBaseHeader.extend([""] * (iColumnCount - len(objBaseHeader)))
            if len(objAddRow) < iColumnCount:
                objAddRow = objAddRow + [""] * (iColumnCount - len(objAddRow))
            for iColumnIndex in range(iColumnCount):
                if objBaseHeader[iColumnIndex].strip() == "" and objAddRow[iColumnIndex].strip() != "":
                    objBaseHeader[iColumnIndex] = objAddRow[iColumnIndex]
            objBaseRows[0] = objBaseHeader
            continue

        pszKey = objAddRow[0] if objAddRow else ""
        objIndices: List[int] = objBaseKeyIndices.get(pszKey, [])
        iCursor: int = objBaseKeyCursor.get(pszKey, 0)
        if iCursor < len(objIndices):
            iTargetIndex = objIndices[iCursor]
            objBaseKeyCursor[pszKey] = iCursor + 1
        else:
            iTargetIndex = len(objBaseRows)
            objBaseRows.append(list(objAddRow))
            objBaseKeyIndices.setdefault(pszKey, []).append(iTargetIndex)
            objBaseKeyCursor[pszKey] = objBaseKeyCursor.get(pszKey, 0) + 1
            continue

        objBaseRow = objBaseRows[iTargetIndex]
        iColumnCount = max(len(objBaseRow), len(objAddRow))
        if len(objBaseRow) < iColumnCount:
            objBaseRow.extend([""] * (iColumnCount - len(objBaseRow)))
        if len(objAddRow) < iColumnCount:
            objAddRow = objAddRow + [""] * (iColumnCount - len(objAddRow))

        for iColumnIndex in range(1, iColumnCount):
            pszBaseValue = objBaseRow[iColumnIndex]
            pszAddValue = objAddRow[iColumnIndex]
            fBase = try_parse_float(pszBaseValue)
            fAdd = try_parse_float(pszAddValue)

            if fBase is not None and fAdd is not None:
                objBaseRow[iColumnIndex] = format_number(fBase + fAdd)
            elif fBase is None and pszBaseValue.strip() == "" and fAdd is not None:
                objBaseRow[iColumnIndex] = format_number(fAdd)
            elif fBase is None and pszBaseValue.strip() == "" and pszAddValue.strip() != "":
                objBaseRow[iColumnIndex] = pszAddValue

        objBaseRows[iTargetIndex] = objBaseRow

    return objBaseRows

def can_use_simple_position_sum(
    objBaseRows: List[List[str]],
    objAddRows: List[List[str]],
) -> bool:
    if not objBaseRows or not objAddRows:
        return True
    if len(objBaseRows) != len(objAddRows):
        return False

    objManhourColumns: set[str] = {
        "工数",
        "1Cカンパニー販管費の工数",
        "2Cカンパニー販管費の工数",
        "3Cカンパニー販管費の工数",
        "4Cカンパニー販管費の工数",
        "事業開発カンパニー販管費の工数",
    }

    iRowCount: int = len(objBaseRows)
    objHeader: List[str] = objBaseRows[0] if objBaseRows else []
    for iRowIndex in range(iRowCount):
        objBaseRow: List[str] = objBaseRows[iRowIndex]
        objAddRow: List[str] = objAddRows[iRowIndex]
        if len(objBaseRow) != len(objAddRow):
            return False
        if iRowIndex == 0:
            if objBaseRow != objAddRow:
                return False
            continue
        if (objBaseRow[0] if objBaseRow else "") != (objAddRow[0] if objAddRow else ""):
            return False
        for iColumnIndex in range(1, len(objBaseRow)):
            pszBaseValue: str = objBaseRow[iColumnIndex].strip()
            pszAddValue: str = objAddRow[iColumnIndex].strip()
            pszColumnName: str = objHeader[iColumnIndex] if iColumnIndex < len(objHeader) else ""
            if pszColumnName in objManhourColumns:
                if not is_time_text_or_blank(pszBaseValue):
                    return False
                if not is_time_text_or_blank(pszAddValue):
                    return False
                continue
            if pszBaseValue != "" and try_parse_float(pszBaseValue) is None:
                return False
            if pszAddValue != "" and try_parse_float(pszAddValue) is None:
                return False
    return True


def sum_tsv_rows_by_position(
    objBaseRows: List[List[str]],
    objAddRows: List[List[str]],
) -> List[List[str]]:
    if not objBaseRows:
        return [list(objRow) for objRow in objAddRows]
    if not objAddRows:
        return objBaseRows

    objManhourColumns: set[str] = {
        "工数",
        "1Cカンパニー販管費の工数",
        "2Cカンパニー販管費の工数",
        "3Cカンパニー販管費の工数",
        "4Cカンパニー販管費の工数",
        "事業開発カンパニー販管費の工数",
    }

    objHeader: List[str] = objBaseRows[0] if objBaseRows else []
    iRowCount: int = len(objBaseRows)
    for iRowIndex in range(1, iRowCount):
        objBaseRow: List[str] = objBaseRows[iRowIndex]
        objAddRow: List[str] = objAddRows[iRowIndex]
        for iColumnIndex in range(1, len(objBaseRow)):
            pszColumnName: str = objHeader[iColumnIndex] if iColumnIndex < len(objHeader) else ""
            if pszColumnName in objManhourColumns:
                fBaseSeconds: float = parse_time_to_seconds(objBaseRow[iColumnIndex])
                fAddSeconds: float = parse_time_to_seconds(objAddRow[iColumnIndex])
                objBaseRow[iColumnIndex] = format_seconds_as_time_text(fBaseSeconds + fAddSeconds)
                continue
            fBase: float = try_parse_float(objBaseRow[iColumnIndex]) or 0.0
            fAdd: float = try_parse_float(objAddRow[iColumnIndex]) or 0.0
            objBaseRow[iColumnIndex] = format_number(fBase + fAdd)
        objBaseRows[iRowIndex] = objBaseRow
    return objBaseRows



def write_tsv_rows(pszPath: str, objRows: List[List[str]]) -> None:
    with open(pszPath, "w", encoding="utf-8", newline="") as objFile:
        for objRow in objRows:
            objFile.write("\t".join(objRow) + "\n")


def format_sales_ratio(fValue: float) -> str:
    objDecimal = Decimal(str(fValue))
    objRounded = objDecimal.quantize(Decimal("0.000"), rounding=ROUND_HALF_UP)
    return f"{objRounded:.3f}"


def add_sales_ratio_column(objRows: List[List[str]]) -> List[List[str]]:
    if not objRows:
        return []

    iSalesRowIndex: int = find_row_index_by_name(objRows, "純売上高")
    fSales: float = 0.0
    if iSalesRowIndex >= 0 and len(objRows[iSalesRowIndex]) >= 2:
        fSales = parse_number(objRows[iSalesRowIndex][1])

    iLastRatioRowIndex: int = find_row_index_by_name(objRows, "当期製品製造原価")
    if iLastRatioRowIndex < 0:
        iLastRatioRowIndex = len(objRows) - 1

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objRows):
        pszName: str = objRow[0] if objRow else ""
        pszValue: str = objRow[1] if len(objRow) >= 2 else ""
        if iRowIndex == 0:
            objOutputRows.append([pszName, pszValue, "売上比率"])
            continue

        if iRowIndex > iLastRatioRowIndex:
            objOutputRows.append([pszName, pszValue, ""])
            continue

        fValue: float = parse_number(pszValue)
        if abs(fSales) > 0.0000001:
            fRatio: float = fValue / fSales
            objOutputRows.append([pszName, pszValue, format_sales_ratio(fRatio)])
            continue

        if fValue > 0.0:
            pszRatio = "'＋∞"
        elif fValue < 0.0:
            pszRatio = "'－∞"
        else:
            pszRatio = format_sales_ratio(0.0)
        objOutputRows.append([pszName, pszValue, pszRatio])

    return objOutputRows


def build_step0010_rows(
    objSingleRows: List[List[str]],
    objCumulativeRows: List[List[str]],
) -> List[List[str]]:
    if not objSingleRows and not objCumulativeRows:
        return []

    iSingleColumnCount: int = max((len(objRow) for objRow in objSingleRows), default=0)
    iCumulativeColumnCount: int = max((len(objRow) for objRow in objCumulativeRows), default=0)
    iMaxRows: int = max(len(objSingleRows), len(objCumulativeRows))
    objOutputRows: List[List[str]] = []

    for iRowIndex in range(iMaxRows):
        objSingleRow: List[str] = (
            list(objSingleRows[iRowIndex]) if iRowIndex < len(objSingleRows) else []
        )
        objCumulativeRow: List[str] = (
            list(objCumulativeRows[iRowIndex]) if iRowIndex < len(objCumulativeRows) else []
        )
        if len(objSingleRow) < iSingleColumnCount:
            objSingleRow.extend([""] * (iSingleColumnCount - len(objSingleRow)))
        if len(objCumulativeRow) < iCumulativeColumnCount:
            objCumulativeRow.extend([""] * (iCumulativeColumnCount - len(objCumulativeRow)))
        objOutputRows.append(objSingleRow + [""] + objCumulativeRow)

    return objOutputRows


def build_step0011_rows(objRows: List[List[str]]) -> List[List[str]]:
    if not objRows:
        return []

    objOutputRows: List[List[str]] = [list(objRow) for objRow in objRows]
    iBlankColumnIndex: int = -1
    for objRow in objOutputRows:
        if "" in objRow:
            iBlankColumnIndex = objRow.index("")
            break
    if iBlankColumnIndex < 0:
        return objOutputRows

    iCumulativeNameIndex: int = iBlankColumnIndex + 1
    iMaterialsIndex: int = find_row_index_by_name(objOutputRows, "材料費")
    iLaborIndex: int = find_row_index_by_name(objOutputRows, "労務費")
    iOutsourceIndex: int = find_row_index_by_name(objOutputRows, "外注加工費")
    iManufacturingIndex: int = find_row_index_by_name(objOutputRows, "製造経費")

    for iRowIndex, objRow in enumerate(objOutputRows):
        pszName: str = objRow[0] if objRow else ""
        if iRowIndex == 0:
            objRow[0] = "単月"
            if iCumulativeNameIndex < len(objRow):
                objRow[iCumulativeNameIndex] = "累計"
        if pszName == "純売上高" and iBlankColumnIndex < len(objRow):
            objRow[iBlankColumnIndex] = "損益計算書"
        if pszName == "材料費" and iBlankColumnIndex < len(objRow):
            objRow[iBlankColumnIndex] = "製造原価報告書"
        if pszName == "工数行(時間)" and iBlankColumnIndex < len(objRow):
            objRow[iBlankColumnIndex] = "工数表"
        if pszName in (
            "工数行(時間)",
            "工数1時間当たり純売上高",
            "工数1時間当たり営業利益",
            "工数行(h:mm:ss)",
        ):
            objUnitsMap: Dict[str, str] = {
                "工数行(時間)": "時間",
                "工数1時間当たり純売上高": "円",
                "工数1時間当たり営業利益": "円",
                "工数行(h:mm:ss)": "h:mm:ss",
            }
            pszUnit: str = objUnitsMap.get(pszName, "")
            iSingleUnitIndex: int = iBlankColumnIndex - 1
            iCumulativeUnitIndex: int = iBlankColumnIndex + 3
            if 0 <= iSingleUnitIndex < len(objRow):
                objRow[iSingleUnitIndex] = pszUnit
            if 0 <= iCumulativeUnitIndex < len(objRow):
                objRow[iCumulativeUnitIndex] = pszUnit

        if (
            iMaterialsIndex >= 0
            and iLaborIndex >= 0
            and iMaterialsIndex < iRowIndex < iLaborIndex
        ):
            if pszName != "":
                objRow[0] = f"　　{pszName}"
            if iCumulativeNameIndex < len(objRow) and objRow[iCumulativeNameIndex] != "":
                objRow[iCumulativeNameIndex] = f"　　{objRow[iCumulativeNameIndex]}"

        if (
            iOutsourceIndex >= 0
            and iManufacturingIndex >= 0
            and iOutsourceIndex < iRowIndex < iManufacturingIndex
        ):
            if pszName != "":
                objRow[0] = f"　　{pszName}"
            if iCumulativeNameIndex < len(objRow) and objRow[iCumulativeNameIndex] != "":
                objRow[iCumulativeNameIndex] = f"　　{objRow[iCumulativeNameIndex]}"

        objOutputRows[iRowIndex] = objRow

    return objOutputRows


def append_gross_margin_column(objRows: List[List[str]]) -> List[List[str]]:
    if not objRows:
        return []
    objHeader: List[str] = objRows[0]
    iSalesIndex: int = find_column_index(objHeader, "純売上高")
    iGrossProfitIndex: int = find_column_index(objHeader, "売上総利益")
    objOutputRows: List[List[str]] = []

    for iRowIndex, objRow in enumerate(objRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex == 0:
            objNewRow.append("粗利益率")
            objOutputRows.append(objNewRow)
            continue

        fSales: float = 0.0
        fGrossProfit: float = 0.0
        if 0 <= iSalesIndex < len(objRow):
            fSales = parse_number(objRow[iSalesIndex])
        if 0 <= iGrossProfitIndex < len(objRow):
            fGrossProfit = parse_number(objRow[iGrossProfitIndex])

        if abs(fSales) < 0.0000001:
            if fGrossProfit > 0:
                objNewRow.append("'＋∞")
            elif fGrossProfit < 0:
                objNewRow.append("'－∞")
            else:
                objNewRow.append("0")
        else:
            objNewRow.append(format_number(fGrossProfit / fSales))

        objOutputRows.append(objNewRow)
    return objOutputRows


def write_step0006_pj_summary(
    pszOutputPath: str,
    objSingleRows: List[List[str]],
    objCumulativeRows: List[List[str]],
) -> None:
    objSingleHeader: List[str] = objSingleRows[0] if objSingleRows else []
    objCumulativeHeader: List[str] = objCumulativeRows[0] if objCumulativeRows else []
    iMaxColumns: int = max(len(objSingleHeader), len(objCumulativeHeader))
    objSingleOnlyColumnNames = {"計上div", "計上カンパニー", "計上グループ", "科目名"}
    objSingleOnlyIndices = {
        iColumnIndex
        for iColumnIndex in range(1, iMaxColumns)
        if (
            (objSingleHeader[iColumnIndex] if iColumnIndex < len(objSingleHeader) else "")
            or (
                objCumulativeHeader[iColumnIndex]
                if iColumnIndex < len(objCumulativeHeader)
                else ""
            )
        )
        in objSingleOnlyColumnNames
    }
    objStep0004Rows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objSingleRows):
        objCumulativeRow = objCumulativeRows[iRowIndex] if iRowIndex < len(objCumulativeRows) else []
        if iRowIndex == 0:
            objHeader: List[str] = [objRow[0] if objRow else ""]
            iMaxColumns: int = max(len(objRow), len(objCumulativeRow))
            for iColumnIndex in range(1, iMaxColumns):
                pszSingleHeader: str = objRow[iColumnIndex] if iColumnIndex < len(objRow) else ""
                pszCumulativeHeader: str = (
                    objCumulativeRow[iColumnIndex] if iColumnIndex < len(objCumulativeRow) else ""
                )
                objHeader.append(pszSingleHeader)
                if iColumnIndex in objSingleOnlyIndices:
                    continue
                objHeader.append(pszCumulativeHeader)
            objStep0004Rows.append(objHeader)
            continue

        objOutputRow: List[str] = [objRow[0] if objRow else ""]
        iMaxColumns = max(len(objRow), len(objCumulativeRow))
        for iColumnIndex in range(1, iMaxColumns):
            objOutputRow.append(objRow[iColumnIndex] if iColumnIndex < len(objRow) else "")
            if iColumnIndex in objSingleOnlyIndices:
                continue
            objOutputRow.append(
                objCumulativeRow[iColumnIndex] if iColumnIndex < len(objCumulativeRow) else ""
            )
        objStep0004Rows.append(objOutputRow)

    write_tsv_rows(pszOutputPath, objStep0004Rows)
    if not objStep0004Rows:
        return


def insert_ratio_rows_for_vertical(
    objRows: List[List[str]],
) -> List[List[str]]:
    if not objRows:
        return []

    objBaseRows: List[List[str]] = [list(objRow) for objRow in objRows]
    iSalesRowIndex: int = find_row_index_by_name(objBaseRows, "純売上高")
    if iSalesRowIndex < 0:
        return objBaseRows

    objSalesRow: List[str] = objBaseRows[iSalesRowIndex]

    def build_ratio_row(
        pszTargetName: str,
        pszRatioName: str,
    ) -> Optional[Tuple[int, List[str]]]:
        iTargetRowIndex: int = find_row_index_by_name(objBaseRows, pszTargetName)
        if iTargetRowIndex < 0:
            return None
        objTargetRow: List[str] = objBaseRows[iTargetRowIndex]
        iColumnCount: int = max(len(objSalesRow), len(objTargetRow))
        objRatioRow: List[str] = [pszRatioName] + [""] * max(iColumnCount - 1, 0)
        for iColumnIndex in range(1, iColumnCount):
            fSales: float = 0.0
            fTarget: float = 0.0
            if iColumnIndex < len(objSalesRow):
                fSales = parse_number(objSalesRow[iColumnIndex])
            if iColumnIndex < len(objTargetRow):
                fTarget = parse_number(objTargetRow[iColumnIndex])

            if abs(fSales) < 0.0000001:
                if fTarget > 0.0:
                    objRatioRow[iColumnIndex] = "'＋∞"
                elif fTarget < 0.0:
                    objRatioRow[iColumnIndex] = "'－∞"
                else:
                    objRatioRow[iColumnIndex] = "0.0000"
                continue
            objRatioRow[iColumnIndex] = f"{fTarget / fSales:.4f}"
        return iTargetRowIndex, objRatioRow

    objInsertions: List[Tuple[int, List[str]]] = []
    for pszTargetName, pszRatioName in [
        ("売上総利益", "売上総利益率"),
        ("営業利益", "営業利益率"),
    ]:
        objResult = build_ratio_row(pszTargetName, pszRatioName)
        if objResult is not None:
            objInsertions.append(objResult)

    objInsertions.sort(key=lambda objItem: objItem[0])
    objOutputRows: List[List[str]] = [list(objRow) for objRow in objBaseRows]
    iOffset: int = 0
    for iTargetRowIndex, objRatioRow in objInsertions:
        iInsertIndex: int = iTargetRowIndex + 1 + iOffset
        if iInsertIndex < 0:
            iInsertIndex = 0
        if iInsertIndex > len(objOutputRows):
            iInsertIndex = len(objOutputRows)
        objOutputRows.insert(iInsertIndex, objRatioRow)
        iOffset += 1

    return objOutputRows


def build_report_file_path(
    pszDirectory: str,
    pszPrefix: str,
    objYearMonth: Tuple[int, int],
) -> str:
    iYear, iMonth = objYearMonth
    pszMonth: str = f"{iMonth:02d}"
    pszFileName: str = f"{pszPrefix}_{iYear}年{pszMonth}月_A∪B_プロジェクト名_C∪D.tsv"
    return os.path.join(pszDirectory, pszFileName)


def build_report_vertical_file_path(
    pszDirectory: str,
    pszPrefix: str,
    objYearMonth: Tuple[int, int],
) -> str:
    iYear, iMonth = objYearMonth
    pszMonth: str = f"{iMonth:02d}"
    pszFileName: str = f"{pszPrefix}_{iYear}年{pszMonth}月_A∪B_プロジェクト名_C∪D_vertical.tsv"
    return os.path.join(pszDirectory, pszFileName)


def build_cumulative_file_path(
    pszDirectory: str,
    pszPrefix: str,
    objStart: Tuple[int, int],
    objEnd: Tuple[int, int],
) -> str:
    iStartYear, iStartMonth = objStart
    iEndYear, iEndMonth = objEnd
    pszStartMonth: str = f"{iStartMonth:02d}"
    pszEndMonth: str = f"{iEndMonth:02d}"
    pszFileName: str = (
        f"累計_{pszPrefix}_{iStartYear}年{pszStartMonth}月_{iEndYear}年{pszEndMonth}月.tsv"
    )
    return os.path.join(pszDirectory, pszFileName)


def read_report_rows(
    pszDirectory: str,
    pszPrefix: str,
    objYearMonth: Tuple[int, int],
) -> Optional[List[List[str]]]:
    pszHorizontalPath: str = build_report_file_path(pszDirectory, pszPrefix, objYearMonth)
    if os.path.isfile(pszHorizontalPath):
        return read_tsv_rows(pszHorizontalPath)

    pszVerticalPath: str = build_report_vertical_file_path(pszDirectory, pszPrefix, objYearMonth)
    if os.path.isfile(pszVerticalPath):
        objVerticalRows: List[List[str]] = read_tsv_rows(pszVerticalPath)
        return transpose_rows(objVerticalRows)

    print(f"Input file not found: {pszHorizontalPath}")
    print(f"Input file not found: {pszVerticalPath}")
    return None


def find_column_index(objHeader: List[str], pszName: str) -> int:
    for iIndex, pszValue in enumerate(objHeader):
        if pszValue == pszName:
            return iIndex
    return -1


def is_company_project(pszProjectName: str) -> bool:
    return re.match(r"^C\d{3}_", pszProjectName) is not None


def is_summary_project(pszProjectName: str) -> bool:
    return pszProjectName.startswith("合計")


def is_project_code(pszProjectName: str, pszPrefix: str, iDigits: int) -> bool:
    return re.match(rf"^{pszPrefix}\d{{{iDigits}}}_", pszProjectName) is not None


def collect_project_rows(
    objRows: List[List[str]],
    iProjectNameColumnIndex: int,
) -> List[List[str]]:
    if not objRows:
        return []
    iStartIndex: int = 1
    for iIndex, objRow in enumerate(objRows[1:], start=1):
        pszName: str = ""
        if 0 <= iProjectNameColumnIndex < len(objRow):
            pszName = objRow[iProjectNameColumnIndex]
        if pszName.startswith("本部"):
            iStartIndex = iIndex
            break
    return objRows[iStartIndex:]


def build_project_rows_for_summary(
    objRows: List[List[str]],
    iProjectNameColumnIndex: int,
) -> List[List[str]]:
    objCandidateRows: List[List[str]] = collect_project_rows(objRows, iProjectNameColumnIndex)
    objOrderedRows: List[List[str]] = []
    objRules: List[Tuple[str, int]] = [("J", 3), ("P", 5)]
    for pszPrefix, iDigits in objRules:
        for objRow in objCandidateRows:
            pszName: str = ""
            if 0 <= iProjectNameColumnIndex < len(objRow):
                pszName = objRow[iProjectNameColumnIndex]
            if pszName == "" or is_company_project(pszName) or is_summary_project(pszName):
                continue
            if is_project_code(pszName, pszPrefix, iDigits):
                objOrderedRows.append(objRow)
    return objOrderedRows


def extract_project_values(
    objRows: List[List[str]],
    iProjectNameColumnIndex: int,
    iValueColumnIndex: int,
) -> List[str]:
    objValues: List[str] = []
    for objRow in build_project_rows_for_summary(objRows, iProjectNameColumnIndex):
        if iValueColumnIndex < 0 or iValueColumnIndex >= len(objRow):
            objValues.append("")
        else:
            objValues.append(objRow[iValueColumnIndex])
    return objValues


def extract_project_names(
    objRows: List[List[str]],
    iProjectNameColumnIndex: int,
) -> List[str]:
    objNames: List[str] = []
    for objRow in build_project_rows_for_summary(objRows, iProjectNameColumnIndex):
        pszName: str = ""
        if 0 <= iProjectNameColumnIndex < len(objRow):
            pszName = objRow[iProjectNameColumnIndex]
        objNames.append(pszName)
    return objNames


def build_gross_margin_values(
    objRows: List[List[str]],
    iProjectNameColumnIndex: int,
    iGrossProfitColumnIndex: int,
    iSalesColumnIndex: int,
) -> List[str]:
    objValues: List[str] = []
    for objRow in build_project_rows_for_summary(objRows, iProjectNameColumnIndex):
        fGrossProfit: float = 0.0
        fSales: float = 0.0
        if 0 <= iGrossProfitColumnIndex < len(objRow):
            fGrossProfit = parse_number(objRow[iGrossProfitColumnIndex])
        if 0 <= iSalesColumnIndex < len(objRow):
            fSales = parse_number(objRow[iSalesColumnIndex])
        if abs(fSales) < 0.0000001:
            if fGrossProfit > 0:
                objValues.append("'＋∞")
            elif fGrossProfit < 0:
                objValues.append("'－∞")
            else:
                objValues.append("0")
        else:
            objValues.append(format_number(fGrossProfit / fSales))
    return objValues


def write_pj_summary(
    pszOutputPath: str,
    objSingleRows: List[List[str]],
    objCumulativeRows: List[List[str]],
) -> None:
    if not objSingleRows or not objCumulativeRows:
        return
    objSingleHeader: List[str] = objSingleRows[0]
    objCumulativeHeader: List[str] = objCumulativeRows[0]

    iProjectNameColumnIndex: int = 0
    iSingleSalesIndex: int = 2
    iCumulativeSalesIndex: int = 2
    iSingleCostIndex: int = find_column_index(objSingleHeader, "売上原価")
    iCumulativeCostIndex: int = find_column_index(objCumulativeHeader, "売上原価")
    iSingleGrossIndex: int = find_column_index(objSingleHeader, "売上総利益")
    iCumulativeGrossIndex: int = find_column_index(objCumulativeHeader, "売上総利益")
    iSingleAllocationIndex: int = find_column_index(objSingleHeader, "配賦販管費")
    iCumulativeAllocationIndex: int = find_column_index(objCumulativeHeader, "配賦販管費")

    objProjectNames: List[str] = extract_project_names(objSingleRows, iProjectNameColumnIndex)
    objSingleSales: List[str] = extract_project_values(
        objSingleRows,
        iProjectNameColumnIndex,
        iSingleSalesIndex,
    )
    objCumulativeSales: List[str] = extract_project_values(
        objCumulativeRows,
        iProjectNameColumnIndex,
        iCumulativeSalesIndex,
    )
    objSingleCost: List[str] = extract_project_values(
        objSingleRows,
        iProjectNameColumnIndex,
        iSingleCostIndex,
    )
    objCumulativeCost: List[str] = extract_project_values(
        objCumulativeRows,
        iProjectNameColumnIndex,
        iCumulativeCostIndex,
    )
    objSingleGross: List[str] = extract_project_values(
        objSingleRows,
        iProjectNameColumnIndex,
        iSingleGrossIndex,
    )
    objCumulativeGross: List[str] = extract_project_values(
        objCumulativeRows,
        iProjectNameColumnIndex,
        iCumulativeGrossIndex,
    )
    objSingleAllocation: List[str] = extract_project_values(
        objSingleRows,
        iProjectNameColumnIndex,
        iSingleAllocationIndex,
    )
    objCumulativeAllocation: List[str] = extract_project_values(
        objCumulativeRows,
        iProjectNameColumnIndex,
        iCumulativeAllocationIndex,
    )
    objSingleMargin: List[str] = build_gross_margin_values(
        objSingleRows,
        iProjectNameColumnIndex,
        iSingleGrossIndex,
        iSingleSalesIndex,
    )
    objCumulativeMargin: List[str] = build_gross_margin_values(
        objCumulativeRows,
        iProjectNameColumnIndex,
        iCumulativeGrossIndex,
        iCumulativeSalesIndex,
    )

    objRows: List[List[str]] = []
    objRows.append(
        [
            "1",
            "PJ名称",
            "単月_純売上高",
            "累計_純売上高",
            "単月_売上原価",
            "累計_売上原価",
            "単月_売上総利益",
            "累計_売上総利益",
            "単月_カンパニー販管費",
            "累計_カンパニー販管費",
            "単月_配賦販管費",
            "累計_配賦販管費",
            "粗利益率",
            "粗利益率",
        ]
    )

    for iIndex, pszProjectName in enumerate(objProjectNames):
        objRows.append(
            [
                str(iIndex + 2),
                pszProjectName,
                objSingleSales[iIndex] if iIndex < len(objSingleSales) else "",
                objCumulativeSales[iIndex] if iIndex < len(objCumulativeSales) else "",
                objSingleCost[iIndex] if iIndex < len(objSingleCost) else "",
                objCumulativeCost[iIndex] if iIndex < len(objCumulativeCost) else "",
                objSingleGross[iIndex] if iIndex < len(objSingleGross) else "",
                objCumulativeGross[iIndex] if iIndex < len(objCumulativeGross) else "",
                "0",
                "0",
                objSingleAllocation[iIndex] if iIndex < len(objSingleAllocation) else "",
                objCumulativeAllocation[iIndex] if iIndex < len(objCumulativeAllocation) else "",
                objSingleMargin[iIndex] if iIndex < len(objSingleMargin) else "",
                objCumulativeMargin[iIndex] if iIndex < len(objCumulativeMargin) else "",
            ]
        )

    write_tsv_rows(pszOutputPath, objRows)


def filter_rows_by_columns(
    objRows: List[List[str]],
    objTargetColumns: List[str],
) -> List[List[str]]:
    if not objRows:
        return []
    objHeader: List[str] = objRows[0]
    objColumnIndices: List[int] = [
        find_column_index(objHeader, pszColumn)
        for pszColumn in objTargetColumns
    ]
    objFilteredRows: List[List[str]] = []
    for objRow in objRows:
        objFilteredRow: List[str] = []
        for iColumnIndex in objColumnIndices:
            if 0 <= iColumnIndex < len(objRow):
                objFilteredRow.append(objRow[iColumnIndex])
            else:
                objFilteredRow.append("")
        objFilteredRows.append(objFilteredRow)
    return objFilteredRows


def move_column_before(
    objRows: List[List[str]],
    pszMoveName: str,
    pszBeforeName: str,
) -> List[List[str]]:
    if not objRows:
        return objRows

    objHeader = objRows[0]
    iMoveIndex = find_column_index(objHeader, pszMoveName)
    iBeforeIndex = find_column_index(objHeader, pszBeforeName)
    if iMoveIndex < 0 or iBeforeIndex < 0 or iMoveIndex == iBeforeIndex:
        return objRows

    objOutputRows: List[List[str]] = []
    for objRow in objRows:
        objRowValues = list(objRow)
        pszValue = objRowValues.pop(iMoveIndex) if iMoveIndex < len(objRowValues) else ""
        if iMoveIndex < iBeforeIndex:
            iBeforeIndexAdjusted = iBeforeIndex - 1
        else:
            iBeforeIndexAdjusted = iBeforeIndex
        if iBeforeIndexAdjusted < 0:
            iBeforeIndexAdjusted = 0
        if iBeforeIndexAdjusted > len(objRowValues):
            iBeforeIndexAdjusted = len(objRowValues)
        objRowValues.insert(iBeforeIndexAdjusted, pszValue)
        objOutputRows.append(objRowValues)

    return objOutputRows


def combine_company_sg_admin_columns(
    objRows: List[List[str]],
) -> List[List[str]]:
    if not objRows:
        return []
    objHeader: List[str] = objRows[0]
    objCompanyColumns: List[str] = [
        "1Cカンパニー販管費",
        "2Cカンパニー販管費",
        "3Cカンパニー販管費",
        "4Cカンパニー販管費",
        "事業開発カンパニー販管費",
    ]
    objCompanyIndices: List[int] = [
        find_column_index(objHeader, pszColumn) for pszColumn in objCompanyColumns
    ]
    objCompanyIndexSet = {iIndex for iIndex in objCompanyIndices if iIndex >= 0}
    iAllocationIndex: int = find_column_index(objHeader, "配賦販管費")
    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objRows):
        if iRowIndex == 0:
            objOutputRow: List[str] = []
            bInserted: bool = False
            for iColumnIndex, pszValue in enumerate(objHeader):
                if iColumnIndex == iAllocationIndex:
                    objOutputRow.append("カンパニー販管費")
                    bInserted = True
                if iColumnIndex in objCompanyIndexSet:
                    continue
                objOutputRow.append(pszValue)
            if not bInserted:
                objOutputRow.append("カンパニー販管費")
            objOutputRows.append(objOutputRow)
            continue

        fCompanyTotal: float = 0.0
        for iColumnIndex in objCompanyIndices:
            if 0 <= iColumnIndex < len(objRow):
                fCompanyTotal += parse_number(objRow[iColumnIndex])
        pszCompanyTotal: str = format_number(fCompanyTotal)

        objOutputRow = []
        bInserted = False
        for iColumnIndex, pszValue in enumerate(objRow):
            if iColumnIndex == iAllocationIndex:
                objOutputRow.append(pszCompanyTotal)
                bInserted = True
            if iColumnIndex in objCompanyIndexSet:
                continue
            objOutputRow.append(pszValue)
        if not bInserted:
            objOutputRow.append(pszCompanyTotal)
        objOutputRows.append(objOutputRow)

    return objOutputRows


def load_org_table_group_map(pszOrgTablePath: str) -> Dict[str, str]:
    objGroupMap: Dict[str, str] = {}
    if not os.path.isfile(pszOrgTablePath):
        return objGroupMap

    objRows = read_tsv_rows(pszOrgTablePath)
    if not objRows:
        return objGroupMap

    objHeader = objRows[0]
    iCodeIndex = find_column_index(objHeader, "PJコード")
    objGroupColumnCandidates = ["計上グループ名", "計上グループ"]
    iGroupIndex = -1
    for pszColumn in objGroupColumnCandidates:
        iGroupIndex = find_column_index(objHeader, pszColumn)
        if iGroupIndex >= 0:
            break

    iStartIndex = 0
    if iCodeIndex >= 0:
        if iGroupIndex < 0:
            iGroupIndex = iCodeIndex + 2
        iStartIndex = 1
    else:
        iCodeIndex = 2
        iGroupIndex = 4

    for objRow in objRows[iStartIndex:]:
        if iCodeIndex >= len(objRow) or iGroupIndex >= len(objRow):
            continue
        pszProjectCode: str = normalize_org_table_project_code_prefix(objRow[iCodeIndex])
        pszGroupName: str = objRow[iGroupIndex].strip()
        if not pszProjectCode:
            continue

        objMatch = re.match(r"^(P\d{5}_|[A-OQ-Z]\d{3}_)", pszProjectCode)
        if objMatch is None:
            objMatch = re.match(r"^(P\d{5}|[A-OQ-Z]\d{3})", pszProjectCode)
        if objMatch is None:
            continue
        pszPrefix: str = objMatch.group(1)
        if not pszPrefix.endswith("_"):
            pszPrefix += "_"
        if pszPrefix not in objGroupMap:
            objGroupMap[pszPrefix] = pszGroupName

    return objGroupMap


def normalize_org_table_project_code_prefix(pszProjectCode: str) -> str:
    pszNormalized: str = (pszProjectCode or "").strip()
    return re.sub(
        r"^((?:P\d{5}|[A-OQ-Z]\d{3}))[\u0020\u3000]+",
        r"\1_",
        pszNormalized,
    )


def load_org_table_company_map(pszOrgTablePath: str) -> Dict[str, str]:
    objCompanyMap: Dict[str, str] = {}
    if not os.path.isfile(pszOrgTablePath):
        return objCompanyMap

    objRows = read_tsv_rows(pszOrgTablePath)
    if not objRows:
        return objCompanyMap

    objHeader = objRows[0]
    iCodeIndex = find_column_index(objHeader, "PJコード")
    objCompanyColumnCandidates = ["計上div名", "計上div", "計上カンパニー名", "計上カンパニー"]
    iCompanyIndex = -1
    for pszColumn in objCompanyColumnCandidates:
        iCompanyIndex = find_column_index(objHeader, pszColumn)
        if iCompanyIndex >= 0:
            break

    iStartIndex = 0
    if iCodeIndex >= 0:
        if iCompanyIndex < 0:
            iCompanyIndex = iCodeIndex + 2
        iStartIndex = 1
    else:
        iCodeIndex = 2
        iCompanyIndex = 4

    for objRow in objRows[iStartIndex:]:
        if iCodeIndex >= len(objRow) or iCompanyIndex >= len(objRow):
            continue
        pszProjectCode: str = normalize_org_table_project_code_prefix(objRow[iCodeIndex])
        pszCompanyName: str = objRow[iCompanyIndex].strip()
        if not pszProjectCode:
            continue

        objMatch = re.match(r"^(P\d{5}_|[A-OQ-Z]\d{3}_)", pszProjectCode)
        if objMatch is None:
            objMatch = re.match(r"^(P\d{5}|[A-OQ-Z]\d{3})", pszProjectCode)
        if objMatch is None:
            continue
        pszPrefix: str = objMatch.group(1)
        if not pszPrefix.endswith("_"):
            pszPrefix += "_"
        if pszPrefix not in objCompanyMap:
            objCompanyMap[pszPrefix] = pszCompanyName

    return objCompanyMap


def insert_accounting_group_column(
    objRows: List[List[str]],
    objGroupMap: Dict[str, str],
) -> List[List[str]]:
    objOutputRows: List[List[str]] = []
    for objRow in objRows:
        pszProjectName: str = objRow[0].strip() if objRow else ""
        if pszProjectName == "科目名":
            objOutputRows.append(
                ["計上グループ", pszProjectName] + (objRow[1:] if len(objRow) > 1 else [])
            )
            continue

        pszGroupName: str = ""
        objMatch = re.match(r"^(P\d{5}_|[A-OQ-Z]\d{3}_)", pszProjectName)
        if objMatch is not None:
            pszPrefix = objMatch.group(1)
            pszGroupName = objGroupMap.get(pszPrefix, "")
        elif pszProjectName == "本部":
            pszGroupName = objGroupMap.get("本部", "")

        objOutputRows.append(
            [pszGroupName, pszProjectName] + (objRow[1:] if len(objRow) > 1 else [])
        )

    return objOutputRows


def get_headquarters_group_from_org_table(pszOrgTablePath: str) -> str:
    if not os.path.isfile(pszOrgTablePath):
        print(f"Warning: org table not found: {pszOrgTablePath}")
        return ""

    objRows = read_tsv_rows(pszOrgTablePath)
    if not objRows:
        print(f"Warning: org table empty: {pszOrgTablePath}")
        return ""

    objHeader = objRows[0]
    iCodeIndex = find_column_index(objHeader, "PJコード")
    objGroupColumnCandidates = ["計上グループ名", "計上グループ"]
    iGroupIndex = -1
    for pszColumn in objGroupColumnCandidates:
        iGroupIndex = find_column_index(objHeader, pszColumn)
        if iGroupIndex >= 0:
            break

    iStartIndex = 0
    if iCodeIndex >= 0:
        if iGroupIndex < 0:
            iGroupIndex = iCodeIndex + 2
        iStartIndex = 1
    else:
        iCodeIndex = 2
        iGroupIndex = 4

    for objRow in objRows[iStartIndex:]:
        if iCodeIndex >= len(objRow) or iGroupIndex >= len(objRow):
            continue
        if objRow[iCodeIndex].strip() != "本部":
            continue
        return objRow[iGroupIndex].strip()

    print("Warning: 本部 row not found in org table.")
    return ""


def insert_accounting_company_column(
    objRows: List[List[str]],
    objCompanyMap: Dict[str, str],
) -> List[List[str]]:
    objOutputRows: List[List[str]] = []
    for objRow in objRows:
        pszProjectName: str = objRow[1].strip() if len(objRow) > 1 else ""
        if pszProjectName == "科目名":
            objOutputRows.append(
                ["計上div"] + (objRow if objRow else [])
            )
            continue

        pszCompanyName: str = ""
        objMatch = re.match(r"^(P\d{5}_|[A-OQ-Z]\d{3}_)", pszProjectName)
        if objMatch is not None:
            pszPrefix = objMatch.group(1)
            pszCompanyName = objCompanyMap.get(pszPrefix, "")

        objOutputRows.append([pszCompanyName] + (objRow if objRow else []))

    return objOutputRows


def get_headquarters_company_from_org_table(pszOrgTablePath: str) -> str:
    if not os.path.isfile(pszOrgTablePath):
        return ""

    objRows = read_tsv_rows(pszOrgTablePath)
    if not objRows:
        return ""

    objHeader = objRows[0]
    iCodeIndex = find_column_index(objHeader, "PJコード")
    objCompanyColumnCandidates = ["計上div名", "計上div", "計上カンパニー名", "計上カンパニー"]
    iCompanyIndex = -1
    for pszColumn in objCompanyColumnCandidates:
        iCompanyIndex = find_column_index(objHeader, pszColumn)
        if iCompanyIndex >= 0:
            break

    iStartIndex = 0
    if iCodeIndex >= 0:
        if iCompanyIndex < 0:
            iCompanyIndex = iCodeIndex + 2
        iStartIndex = 1
    else:
        iCodeIndex = 2
        iCompanyIndex = 3

    for objRow in objRows[iStartIndex:]:
        if iCodeIndex >= len(objRow) or iCompanyIndex >= len(objRow):
            continue
        if objRow[iCodeIndex].strip() != "本部":
            continue
        return objRow[iCompanyIndex].strip()

    return ""


def fill_headquarters_company_in_rows(
    objRows: List[List[str]],
    pszOrgTablePath: str,
) -> List[List[str]]:
    if not objRows:
        return objRows

    objHeader = objRows[0]
    iCompanyIndex = find_column_index(objHeader, "計上div")
    if iCompanyIndex < 0:
        iCompanyIndex = find_column_index(objHeader, "計上カンパニー")
    iGroupIndex = find_column_index(objHeader, "計上グループ")
    if iCompanyIndex < 0 or iGroupIndex < 0:
        return objRows

    pszHeadquartersCompany = get_headquarters_company_from_org_table(pszOrgTablePath)
    if pszHeadquartersCompany == "":
        pszHeadquartersCompany = "本部"

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objRows):
        if iRowIndex == 0:
            objOutputRows.append(list(objRow))
            continue
        objNewRow = list(objRow)
        if iGroupIndex < len(objNewRow) and iCompanyIndex < len(objNewRow):
            if objNewRow[iGroupIndex].strip() == "本部" and objNewRow[iCompanyIndex].strip() == "":
                objNewRow[iCompanyIndex] = pszHeadquartersCompany
        objOutputRows.append(objNewRow)

    return objOutputRows


def update_step0003_headquarters_group(
    pszStep0003Path: str,
    pszOrgTablePath: str,
) -> None:
    if not os.path.isfile(pszStep0003Path):
        return

    pszGroupName = get_headquarters_group_from_org_table(pszOrgTablePath)
    if pszGroupName == "":
        return

    objRows = read_tsv_rows(pszStep0003Path)
    if not objRows:
        return

    objOutputRows: List[List[str]] = []
    for objRow in objRows:
        if len(objRow) >= 2 and objRow[1].strip() == "本部":
            objRow = list(objRow)
            objRow[0] = pszGroupName
        objOutputRows.append(objRow)

    write_tsv_rows(pszStep0003Path, objOutputRows)


def update_step0005_headquarters_company(
    pszStep0005Path: str,
    pszOrgTablePath: str,
) -> None:
    if not os.path.isfile(pszStep0005Path):
        return

    pszCompanyName = get_headquarters_company_from_org_table(pszOrgTablePath)
    if pszCompanyName == "":
        return

    objRows = read_tsv_rows(pszStep0005Path)
    if not objRows:
        return

    objOutputRows: List[List[str]] = []
    for objRow in objRows:
        if len(objRow) >= 3 and objRow[2].strip() == "本部":
            objRow = list(objRow)
            objRow[0] = pszCompanyName
        objOutputRows.append(objRow)

    write_tsv_rows(pszStep0005Path, objOutputRows)


def build_step0003_rows_with_debug(
    objRows: List[List[str]],
    objGroupMap: Dict[str, str],
) -> Tuple[List[List[str]], List[List[str]]]:
    if not objRows:
        return [], []
    objRemovalTargets = {
        "C001_1Cカンパニー販管費",
        "C002_2Cカンパニー販管費",
        "C003_3Cカンパニー販管費",
        "C004_4Cカンパニー販管費",
        "C005_事業開発カンパニー販管費",
        "C006_社長室カンパニー販管費",
        "C007_本部カンパニー販管費",
    }
    iStartIndex = -1
    for iRowIndex, objRow in enumerate(objRows):
        pszName = objRow[0].strip() if objRow else ""
        if pszName == "本部":
            iStartIndex = iRowIndex
            break

    objOutputRows: List[List[str]] = []
    objDebugRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objRows):
        if objRow and objRow[0].strip() in objRemovalTargets:
            continue

        pszCompanyName = objRow[0].strip() if iRowIndex < 2 and objRow else ""
        pszReason = ""
        pszProjectName = objRow[0].strip() if objRow else ""
        bInAssignmentRange = iRowIndex >= 2 and iStartIndex >= 0 and iRowIndex >= iStartIndex

        if bInAssignmentRange:
            if pszProjectName == "本部":
                pszCompanyName = "本部"
            elif pszProjectName == "":
                pszReason = "その他の原因"
            else:
                objMatch = re.match(r"^(P\d{5}_|[A-OQ-Z]\d{3}_)", pszProjectName)
                if objMatch is not None:
                    pszPrefix = objMatch.group(1)
                    pszCompanyName = objGroupMap.get(pszPrefix, "")
                    if pszCompanyName == "":
                        pszReason = "管轄PJ表エラー"
                else:
                    pszReason = "PJ名エラー"

        if pszCompanyName == "" and pszReason == "":
            pszReason = "その他の原因"

        objOutputRow = [pszCompanyName] + (objRow[1:] if len(objRow) > 1 else [])
        objOutputRows.append(objOutputRow)

        pszDebugFirstColumn = pszCompanyName if pszCompanyName != "" else pszReason
        objDebugRows.append([pszDebugFirstColumn] + (objRow[1:] if len(objRow) > 1 else []))

    return objOutputRows, objDebugRows


def build_step0003_rows(
    objRows: List[List[str]],
    objGroupMap: Dict[str, str],
) -> List[List[str]]:
    objOutputRows, _ = build_step0003_rows_with_debug(objRows, objGroupMap)
    return objOutputRows


def detect_step0004_org_mode(objRows: List[List[str]]) -> str:
    objLegacyNameSet = {"第一インキュ", "第二インキュ", "第三インキュ", "第四インキュ"}
    objNewNameSet = {
        "テクノロジーインキュベーション",
        "コンテンツビジネス",
        "スタートアップサイド",
        "スタートアップコミュニティ",
        "スタートアップグロース",
        "経営管理",
    }
    objFirstColumnNames: set[str] = set()
    for objRow in objRows:
        if not objRow:
            continue
        pszName = objRow[0].strip()
        if pszName != "":
            objFirstColumnNames.add(pszName)

    bHasLegacyExclusive: bool = bool(objFirstColumnNames & objLegacyNameSet)
    bHasNewExclusive: bool = bool(objFirstColumnNames & objNewNameSet)
    if bHasLegacyExclusive and bHasNewExclusive:
        raise ValueError(
            "step0004 集計エラー: 旧組織名と新組織名が同時に存在します。"
        )
    if bHasLegacyExclusive:
        return "legacy"
    if bHasNewExclusive:
        return "new"
    raise ValueError(
        "step0004 集計エラー: 組織判定キー(旧4分類/新6分類)が見つかりません。"
    )


def build_step0004_rows_for_summary(objRows: List[List[str]]) -> List[List[str]]:
    if not objRows:
        return []
    objLegacyExclusiveNames: List[str] = [
        "第一インキュ",
        "第二インキュ",
        "第三インキュ",
        "第四インキュ",
    ]
    objNewExclusiveNames: List[str] = [
        "テクノロジーインキュベーション",
        "コンテンツビジネス",
        "スタートアップサイド",
        "スタートアップコミュニティ",
        "スタートアップグロース",
        "経営管理",
    ]
    objCommonNames: List[str] = [
        "事業開発",
        "子会社",
        "投資先",
        "本部",
    ]
    objHeaderRow: List[str] = objRows[0]
    objTotalRow: Optional[List[str]] = None
    objFirstColumnNames: set[str] = set()
    for objRow in objRows:
        if not objRow:
            continue
        pszName = objRow[0].strip()
        if pszName != "":
            objFirstColumnNames.add(pszName)
        if pszName == "科目名":
            objHeaderRow = objRow
        elif pszName == "合計" and objTotalRow is None:
            objTotalRow = objRow

    pszOrgMode: str = detect_step0004_org_mode(objRows)
    if pszOrgMode == "legacy":
        objTargetNames: List[str] = objLegacyExclusiveNames + objCommonNames
    else:
        objTargetNames = [
            "コンテンツビジネス",
            "スタートアップコミュニティ",
            "スタートアップグロース",
            "経営管理",
            "テクノロジーインキュベーション",
            "事業開発",
            "スタートアップサイド",
            "子会社",
            "投資先",
            "本部",
        ]

    objTargetSet = set(objTargetNames)
    iMaxColumns: int = max(len(objRow) for objRow in objRows) if objRows else 0
    objTotalsByName: Dict[str, List[float]] = {
        pszName: [0.0] * iMaxColumns for pszName in objTargetNames
    }
    for objRow in objRows[2:]:
        if not objRow:
            continue
        pszName = objRow[0].strip()
        if pszName not in objTargetSet:
            continue
        for iColumnIndex in range(1, iMaxColumns):
            if iColumnIndex < len(objRow):
                objTotalsByName[pszName][iColumnIndex] += parse_number(objRow[iColumnIndex])

    objOutputRows: List[List[str]] = []
    objOutputRows.append(list(objHeaderRow))
    for pszName in objTargetNames:
        objNewRow: List[str] = [""] * iMaxColumns
        objNewRow[0] = pszName
        for iColumnIndex in range(1, iMaxColumns):
            objNewRow[iColumnIndex] = format_number(objTotalsByName[pszName][iColumnIndex])
        objOutputRows.append(objNewRow)
    if objTotalRow is not None:
        objTotalOutputRow: List[str] = list(objTotalRow)
        iAllocationIndex: int = find_column_index(objHeaderRow, "配賦販管費")
        if iAllocationIndex >= 0:
            if len(objTotalOutputRow) <= iAllocationIndex:
                objTotalOutputRow.extend([""] * (iAllocationIndex + 1 - len(objTotalOutputRow)))
            fAllocationTotal: float = 0.0
            for pszName in objTargetNames:
                fAllocationTotal += objTotalsByName[pszName][iAllocationIndex]
            objTotalOutputRow[iAllocationIndex] = format_number(fAllocationTotal)

        iCompanySgAdminIndex: int = find_column_index(objHeaderRow, "カンパニー販管費")
        if iCompanySgAdminIndex >= 0:
            if len(objTotalOutputRow) <= iCompanySgAdminIndex:
                objTotalOutputRow.extend([""] * (iCompanySgAdminIndex + 1 - len(objTotalOutputRow)))
            fCompanySgAdminTotal: float = 0.0
            for pszName in objTargetNames:
                fCompanySgAdminTotal += objTotalsByName[pszName][iCompanySgAdminIndex]
            objTotalOutputRow[iCompanySgAdminIndex] = format_number(fCompanySgAdminTotal)

        objOutputRows.append(objTotalOutputRow)
    return objOutputRows


def build_step0004_rows_for_group_summary(objRows: List[List[str]]) -> List[List[str]]:
    if not objRows:
        return []
    objTargetNames: List[str] = [
        "自社-その他",
        "自社-施設運営",
        "受託事業-その他",
        "受託事業-施設運営",
    ]
    objTargetSet = set(objTargetNames)
    objHeaderRow: List[str] = objRows[0]
    objTotalRow: Optional[List[str]] = None
    for objRow in objRows:
        if not objRow:
            continue
        pszName = objRow[0].strip()
        if pszName == "科目名":
            objHeaderRow = objRow
        elif pszName == "合計" and objTotalRow is None:
            objTotalRow = objRow

    iMaxColumns: int = max(len(objRow) for objRow in objRows) if objRows else 0
    objTotalsByName: Dict[str, List[float]] = {
        pszName: [0.0] * iMaxColumns for pszName in objTargetNames
    }
    for objRow in objRows[2:]:
        if not objRow:
            continue
        pszName = objRow[0].strip()
        if pszName not in objTargetSet:
            continue
        for iColumnIndex in range(1, iMaxColumns):
            if iColumnIndex < len(objRow):
                objTotalsByName[pszName][iColumnIndex] += parse_number(objRow[iColumnIndex])

    objOutputRows: List[List[str]] = []
    objOutputRows.append(list(objHeaderRow))
    for pszName in objTargetNames:
        objNewRow: List[str] = [""] * iMaxColumns
        objNewRow[0] = pszName
        for iColumnIndex in range(1, iMaxColumns):
            objNewRow[iColumnIndex] = format_number(objTotalsByName[pszName][iColumnIndex])
        objOutputRows.append(objNewRow)
    if objTotalRow is not None:
        objTotalOutputRow: List[str] = list(objTotalRow)
        iAllocationIndex: int = find_column_index(objHeaderRow, "配賦販管費")
        if iAllocationIndex >= 0:
            if len(objTotalOutputRow) <= iAllocationIndex:
                objTotalOutputRow.extend([""] * (iAllocationIndex + 1 - len(objTotalOutputRow)))
            fAllocationTotal: float = 0.0
            for pszName in objTargetNames:
                fAllocationTotal += objTotalsByName[pszName][iAllocationIndex]
            objTotalOutputRow[iAllocationIndex] = format_number(fAllocationTotal)

        iCompanySgAdminIndex: int = find_column_index(objHeaderRow, "カンパニー販管費")
        if iCompanySgAdminIndex >= 0:
            if len(objTotalOutputRow) <= iCompanySgAdminIndex:
                objTotalOutputRow.extend([""] * (iCompanySgAdminIndex + 1 - len(objTotalOutputRow)))
            fCompanySgAdminTotal: float = 0.0
            for pszName in objTargetNames:
                fCompanySgAdminTotal += objTotalsByName[pszName][iCompanySgAdminIndex]
            objTotalOutputRow[iCompanySgAdminIndex] = format_number(fCompanySgAdminTotal)

        objOutputRows.append(objTotalOutputRow)
    return objOutputRows


def build_step0005_rows_for_summary(
    objSingleRows: List[List[str]],
    objCumulativeRows: List[List[str]],
) -> List[List[str]]:
    iMaxRows: int = max(len(objSingleRows), len(objCumulativeRows))
    objOutputRows: List[List[str]] = []
    for iRowIndex in range(iMaxRows):
        objSingleRow: List[str] = objSingleRows[iRowIndex] if iRowIndex < len(objSingleRows) else []
        objCumulativeRow: List[str] = (
            objCumulativeRows[iRowIndex] if iRowIndex < len(objCumulativeRows) else []
        )
        objOutputRows.append(list(objSingleRow) + [""] + list(objCumulativeRow))
    return objOutputRows


def add_profit_ratio_columns(objRows: List[List[str]]) -> List[List[str]]:
    if not objRows:
        return []
    objHeader: List[str] = objRows[0]
    iSalesIndex: int = find_column_index(objHeader, "純売上高")
    iGrossProfitIndex: int = find_column_index(objHeader, "売上総利益")
    iOperatingProfitIndex: int = find_column_index(objHeader, "営業利益")
    if iSalesIndex < 0 or iGrossProfitIndex < 0 or iOperatingProfitIndex < 0:
        return [list(objRow) for objRow in objRows]

    objInsertSpecs = [
        (iGrossProfitIndex + 1, "売上総利益率", iGrossProfitIndex),
        (iOperatingProfitIndex + 1, "営業利益率", iOperatingProfitIndex),
    ]
    objInsertSpecs.sort(key=lambda objSpec: objSpec[0], reverse=True)

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objRows):
        objNewRow: List[str] = list(objRow)
        for iInsertIndex, pszLabel, iProfitIndex in objInsertSpecs:
            if iRowIndex == 0:
                pszValue = pszLabel
            else:
                fSales: float = 0.0
                fProfit: float = 0.0
                if 0 <= iSalesIndex < len(objRow):
                    fSales = parse_number(objRow[iSalesIndex])
                if 0 <= iProfitIndex < len(objRow):
                    fProfit = parse_number(objRow[iProfitIndex])
                if abs(fSales) < 0.0000001:
                    if fProfit > 0:
                        pszValue = "＋∞"
                    elif fProfit < 0:
                        pszValue = "－∞"
                    else:
                        pszValue = "0"
                else:
                    pszValue = format_number(fProfit / fSales)
            if iInsertIndex <= len(objNewRow):
                objNewRow.insert(iInsertIndex, pszValue)
            else:
                objNewRow.extend([""] * (iInsertIndex - len(objNewRow)))
                objNewRow.append(pszValue)
        objOutputRows.append(objNewRow)
    return objOutputRows


def build_step0006_rows_for_summary(objRows: List[List[str]]) -> List[List[str]]:
    if not objRows:
        return []
    objHeaderRow: List[str] = objRows[0]
    iHeaderLength: int = len(objHeaderRow)
    objLabelRow: List[str] = [""] * iHeaderLength
    objSubjectIndices: List[int] = [
        iIndex for iIndex, pszValue in enumerate(objHeaderRow) if pszValue == "科目名"
    ]
    if len(objSubjectIndices) >= 1:
        objLabelRow[objSubjectIndices[0]] = "単月"
    if len(objSubjectIndices) >= 2:
        objLabelRow[objSubjectIndices[1]] = "累計"
    iAllocationIndex: int = next(
        (iIndex for iIndex, pszValue in enumerate(objHeaderRow) if pszValue == "営業利益"),
        -1,
    )
    if iAllocationIndex >= 0:
        objLabelRow[iAllocationIndex] = "div別合計"
    return [objLabelRow] + [list(objRow) for objRow in objRows]


def build_step0006_rows_for_summary_0005(pszStep0005Path: str) -> List[List[str]]:
    if not os.path.isfile(pszStep0005Path):
        return []
    objRows = read_tsv_rows(pszStep0005Path)
    return [list(objRow) for objRow in objRows]


def build_step0007_rows_for_summary_0005(pszStep0006Path: str) -> List[List[str]]:
    if not os.path.isfile(pszStep0006Path):
        return []
    objRows = read_tsv_rows(pszStep0006Path)
    if not objRows:
        return []
    objHeaderRow: List[str] = objRows[0]
    iHeaderLength: int = len(objHeaderRow)
    objLabelRow: List[str] = [""] * iHeaderLength
    objSubjectIndices: List[int] = [
        iIndex for iIndex, pszValue in enumerate(objHeaderRow) if pszValue == "科目名"
    ]
    if len(objSubjectIndices) >= 1:
        objLabelRow[objSubjectIndices[0]] = "単月"
    if len(objSubjectIndices) >= 2:
        objLabelRow[objSubjectIndices[1]] = "累計"
    iAllocationIndex: int = next(
        (iIndex for iIndex, pszValue in enumerate(objHeaderRow) if pszValue == "営業利益"),
        -1,
    )
    if iAllocationIndex >= 0:
        objLabelRow[iAllocationIndex] = "グループ別合計"
    return [objLabelRow] + [list(objRow) for objRow in objRows]


def build_step0007_rows_from_step0006_path(pszStep0006Path: str) -> None:
    if not os.path.isfile(pszStep0006Path):
        return

    objRows = read_tsv_rows(pszStep0006Path)
    if not objRows:
        return

    iColumnCount: int = max(len(objRow) for objRow in objRows)
    objHeaderRow: List[str] = ["単／累"]
    for iColumnIndex in range(1, iColumnCount):
        if iColumnIndex % 2 == 1:
            objHeaderRow.append("単月")
        else:
            objHeaderRow.append("累計")

    objOutputRows: List[List[str]] = [objHeaderRow]
    objOutputRows.extend([list(objRow) for objRow in objRows])

    pszStep0007Path: str = pszStep0006Path.replace("step0006_", "step0007_", 1)
    write_tsv_rows(pszStep0007Path, objOutputRows)


def build_step0008_rows_from_step0007_path(pszStep0007Path: str) -> None:
    if not os.path.isfile(pszStep0007Path):
        return

    objRows = read_tsv_rows(pszStep0007Path)
    if len(objRows) < 2:
        return

    objHeaderRow: List[str] = objRows[1]
    objTargetIndices: List[int] = [
        iIndex for iIndex, pszValue in enumerate(objHeaderRow) if pszValue == "粗利益率"
    ]
    if not objTargetIndices:
        return

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objRows):
        if iRowIndex < 2:
            objOutputRows.append(list(objRow))
            continue
        objNewRow = list(objRow)
        for iColumnIndex in objTargetIndices:
            if iColumnIndex >= len(objNewRow):
                continue
            if objNewRow[iColumnIndex] == "'+∞":
                objNewRow[iColumnIndex] = "＋∞"
            elif objNewRow[iColumnIndex] == "'－∞":
                objNewRow[iColumnIndex] = "－∞"
        objOutputRows.append(objNewRow)

    pszStep0008Path: str = pszStep0007Path.replace("step0007_", "step0008_", 1)
    write_tsv_rows(pszStep0008Path, objOutputRows)


def build_step0009_rows_from_step0008_path(pszStep0008Path: str) -> None:
    if not os.path.isfile(pszStep0008Path):
        return

    objRows = read_tsv_rows(pszStep0008Path)
    if not objRows:
        return

    objHeaderRow: List[str] = [
        "－",
        "－",
        "PJコード",
        "単月",
        "累計",
        "単月",
        "累計",
        "単月",
        "累計",
        "単月",
        "累計",
        "単月",
        "累計",
        "単月",
        "累計",
    ]
    objOutputRows: List[List[str]] = [objHeaderRow]
    objOutputRows.extend([list(objRow) for objRow in objRows[1:]])

    pszStep0009Path: str = pszStep0008Path.replace("step0008_", "step0009_", 1)
    write_tsv_rows(pszStep0009Path, objOutputRows)


def filter_rows_by_names(
    objRows: List[List[str]],
    objTargetNames: List[str],
) -> List[List[str]]:
    if not objRows:
        return []
    objTargetSet = set(objTargetNames)
    objFilteredRows: List[List[str]] = []
    for objRow in objRows:
        if not objRow:
            continue
        pszName: str = objRow[0].strip()
        if pszName in objTargetSet:
            objFilteredRows.append(objRow)
    return objFilteredRows


def add_company_sg_admin_cost_total_row(objRows: List[List[str]]) -> List[List[str]]:
    # step0003向けにカンパニー販管費の合計行を追加する
    if not objRows:
        return objRows

    objTargetNames: List[str] = [
        "1Cカンパニー販管費",
        "2Cカンパニー販管費",
        "3Cカンパニー販管費",
        "4Cカンパニー販管費",
        "事業開発カンパニー販管費",
    ]
    objTargetSet = set(objTargetNames)

    for objRow in objRows:
        if objRow and objRow[0].strip() == "カンパニー販管費":
            return objRows

    iMaxColumns: int = max(len(objRow) for objRow in objRows)
    objTotals: List[float] = [0.0] * iMaxColumns

    for objRow in objRows:
        if not objRow:
            continue
        if objRow[0].strip() not in objTargetSet:
            continue
        for iColumnIndex in range(1, iMaxColumns):
            if iColumnIndex < len(objRow):
                objTotals[iColumnIndex] += parse_number(objRow[iColumnIndex])

    objNewRow: List[str] = [""] * iMaxColumns
    objNewRow[0] = "カンパニー販管費"
    for iColumnIndex in range(1, iMaxColumns):
        objNewRow[iColumnIndex] = format_number(objTotals[iColumnIndex])

    objOutputRows: List[List[str]] = []
    for objRow in objRows:
        if objRow and objRow[0].strip() in objTargetSet:
            continue
        objOutputRows.append(list(objRow))
    objOutputRows.append(objNewRow)
    return objOutputRows


def move_row_between(
    objRows: List[List[str]],
    pszMoveName: str,
    pszBeforeName: str,
    pszAfterName: str,
) -> List[List[str]]:
    if not objRows:
        return objRows

    iMoveIndex: int = -1
    iBeforeIndex: int = -1
    iAfterIndex: int = -1
    for iRowIndex, objRow in enumerate(objRows):
        if not objRow:
            continue
        pszName: str = objRow[0].strip()
        if pszName == pszMoveName:
            iMoveIndex = iRowIndex
        elif pszName == pszBeforeName:
            iBeforeIndex = iRowIndex
        elif pszName == pszAfterName:
            iAfterIndex = iRowIndex

    if iMoveIndex < 0 or iBeforeIndex < 0 or iAfterIndex < 0:
        return objRows

    objOutputRows: List[List[str]] = [list(objRow) for objRow in objRows]
    objMoveRow: List[str] = objOutputRows.pop(iMoveIndex)
    if iMoveIndex < iBeforeIndex:
        iBeforeIndex -= 1
    if iMoveIndex < iAfterIndex:
        iAfterIndex -= 1

    iInsertIndex: int = min(iAfterIndex, iBeforeIndex) + 1
    iInsertIndex = max(iInsertIndex, 0)
    if iInsertIndex > len(objOutputRows):
        iInsertIndex = len(objOutputRows)
    objOutputRows.insert(iInsertIndex, objMoveRow)
    return objOutputRows


def align_vertical_rows_for_union(
    objLeftRows: List[List[str]],
    objRightRows: List[List[str]],
) -> Tuple[List[List[str]], List[List[str]]]:
    objExcludedNames: set[str] = {"合計", "本部"}
    objLeftOrder: List[str] = [
        objRow[0] if objRow else ""
        for objRow in objLeftRows
        if objRow and objRow[0] not in objExcludedNames
    ]
    objRightOrder: List[str] = [
        objRow[0] if objRow else ""
        for objRow in objRightRows
        if objRow and objRow[0] not in objExcludedNames
    ]
    objUnionOrder: List[str] = []
    objSeen: set[str] = set()
    for pszName in objLeftOrder:
        if pszName in objSeen:
            continue
        objSeen.add(pszName)
        objUnionOrder.append(pszName)

    objPositions: Dict[str, int] = {pszName: iIndex for iIndex, pszName in enumerate(objUnionOrder)}
    iLastInsertIndex: int = -1
    for pszName in objRightOrder:
        if pszName in objPositions:
            iLastInsertIndex = objPositions[pszName]
            continue
        iInsertIndex = iLastInsertIndex + 1
        if iInsertIndex < 0:
            iInsertIndex = 0
        if iInsertIndex > len(objUnionOrder):
            iInsertIndex = len(objUnionOrder)
        objUnionOrder.insert(iInsertIndex, pszName)
        objPositions = {pszName: iIndex for iIndex, pszName in enumerate(objUnionOrder)}
        iLastInsertIndex = objPositions[pszName]

    objLeftMap: Dict[str, List[str]] = {}
    for objRow in objLeftRows:
        if not objRow:
            continue
        pszName = objRow[0]
        if pszName in objExcludedNames:
            continue
        if pszName in objLeftMap:
            continue
        objLeftMap[pszName] = objRow

    objRightMap: Dict[str, List[str]] = {}
    for objRow in objRightRows:
        if not objRow:
            continue
        pszName = objRow[0]
        if pszName in objExcludedNames:
            continue
        if pszName in objRightMap:
            continue
        objRightMap[pszName] = objRow

    iLeftColumnCount: int = max((len(objRow) for objRow in objLeftRows), default=1)
    iRightColumnCount: int = max((len(objRow) for objRow in objRightRows), default=1)

    objAlignedLeft: List[List[str]] = []
    objAlignedRight: List[List[str]] = []
    for pszName in objUnionOrder:
        if pszName in objLeftMap:
            objAlignedLeft.append(list(objLeftMap[pszName]))
        else:
            objAlignedLeft.append([pszName] + ["0"] * max(iLeftColumnCount - 1, 0))

        if pszName in objRightMap:
            objAlignedRight.append(list(objRightMap[pszName]))
        else:
            objAlignedRight.append([pszName] + ["0"] * max(iRightColumnCount - 1, 0))

    return objAlignedLeft, objAlignedRight


def insert_per_hour_rows(
    objRows: List[List[str]],
) -> List[List[str]]:
    if not objRows:
        return objRows
    iManhourRowIndex: int = -1
    for iRowIndex, objRow in enumerate(objRows):
        if objRow and objRow[0] == "工数":
            iManhourRowIndex = iRowIndex
            break
    if iManhourRowIndex < 0:
        return objRows

    iSalesRowIndex: int = -1
    iOperatingProfitRowIndex: int = -1
    for iRowIndex, objRow in enumerate(objRows):
        if not objRow:
            continue
        if objRow[0] == "純売上高":
            iSalesRowIndex = iRowIndex
        elif objRow[0] == "営業利益":
            iOperatingProfitRowIndex = iRowIndex

    if iSalesRowIndex < 0 and iOperatingProfitRowIndex < 0:
        return objRows

    objOutputRows: List[List[str]] = [list(objRow) for objRow in objRows]
    objManhourRow = objOutputRows[iManhourRowIndex]
    objSalesRow = objOutputRows[iSalesRowIndex] if iSalesRowIndex >= 0 else []
    objOperatingProfitRow = (
        objOutputRows[iOperatingProfitRowIndex] if iOperatingProfitRowIndex >= 0 else []
    )
    iColumnCount = max(len(objManhourRow), len(objSalesRow), len(objOperatingProfitRow), 1)
    objManhourHoursRow: List[str] = ["工数行(時間)"] + [""] * (iColumnCount - 1)
    objManhourHmsRow: List[str] = ["工数行(h:mm:ss)"] + [""] * (iColumnCount - 1)
    objSalesPerHourRow: List[str] = ["工数1時間当たり純売上高"] + [""] * (iColumnCount - 1)
    objOperatingPerHourRow: List[str] = ["工数1時間当たり営業利益"] + [""] * (iColumnCount - 1)

    for iColumnIndex in range(1, iColumnCount):
        pszManhour = objManhourRow[iColumnIndex] if iColumnIndex < len(objManhourRow) else ""
        fSeconds = parse_time_to_seconds(pszManhour)
        fHours = fSeconds / 3600.0 if fSeconds > 0.0 else 0.0
        objManhourHoursRow[iColumnIndex] = f"{fHours:.1f}"
        objManhourHmsRow[iColumnIndex] = pszManhour

        if iSalesRowIndex >= 0:
            pszSales = objSalesRow[iColumnIndex] if iColumnIndex < len(objSalesRow) else ""
            fSales = parse_number(pszSales)
            fSalesPerHour = fSales / fHours if fHours > 0.0 else 0.0
            objSalesPerHourRow[iColumnIndex] = str(int(fSalesPerHour))

        if iOperatingProfitRowIndex >= 0:
            pszOperating = (
                objOperatingProfitRow[iColumnIndex]
                if iColumnIndex < len(objOperatingProfitRow)
                else ""
            )
            fOperating = parse_number(pszOperating)
            fOperatingPerHour = fOperating / fHours if fHours > 0.0 else 0.0
            objOperatingPerHourRow[iColumnIndex] = str(int(fOperatingPerHour))

    iInsertIndex: int = iManhourRowIndex + 1
    objOutputRows[iManhourRowIndex] = objManhourHoursRow
    objOutputRows.insert(iInsertIndex, objSalesPerHourRow)
    objOutputRows.insert(iInsertIndex + 1, objOperatingPerHourRow)
    objOutputRows.insert(iInsertIndex + 2, objManhourHmsRow)
    return objOutputRows


def select_columns(
    objRows: List[List[str]],
    objColumnIndices: List[int],
) -> List[List[str]]:
    objOutputRows: List[List[str]] = []
    for objRow in objRows:
        objSelectedRow: List[str] = []
        for iColumnIndex in objColumnIndices:
            if 0 <= iColumnIndex < len(objRow):
                objSelectedRow.append(objRow[iColumnIndex])
            else:
                objSelectedRow.append("")
        objOutputRows.append(objSelectedRow)
    return objOutputRows


def select_columns(
    objRows: List[List[str]],
    objColumnIndices: List[int],
) -> List[List[str]]:
    objOutputRows: List[List[str]] = []
    for objRow in objRows:
        objSelectedRow: List[str] = []
        for iColumnIndex in objColumnIndices:
            if 0 <= iColumnIndex < len(objRow):
                objSelectedRow.append(objRow[iColumnIndex])
            else:
                objSelectedRow.append("")
        objOutputRows.append(objSelectedRow)
    return objOutputRows


def find_row_index_by_name(objRows: List[List[str]], pszName: str) -> int:
    for iRowIndex, objRow in enumerate(objRows):
        if not objRow:
            continue
        if objRow[0].strip() == pszName:
            return iRowIndex
    return -1


def create_step0007_pl_cr(
    pszDirectory: str,
) -> None:
    pszSinglePlStep0006Path: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0006_単月_損益計算書_E∪F.tsv",
    )
    pszSingleCostStep0006Path: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0006_単月_製造原価報告書_E∪F.tsv",
    )
    pszCumulativePlStep0006Path: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0006_累計_損益計算書_E∪F.tsv",
    )
    pszCumulativeCostStep0006Path: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0006_累計_製造原価報告書_E∪F.tsv",
    )

    if not (
        os.path.isfile(pszSinglePlStep0006Path)
        and os.path.isfile(pszSingleCostStep0006Path)
        and os.path.isfile(pszCumulativePlStep0006Path)
        and os.path.isfile(pszCumulativeCostStep0006Path)
    ):
        return

    objSinglePlRows = read_tsv_rows(pszSinglePlStep0006Path)
    objSingleCostRows = read_tsv_rows(pszSingleCostStep0006Path)
    objCumulativePlRows = read_tsv_rows(pszCumulativePlStep0006Path)
    objCumulativeCostRows = read_tsv_rows(pszCumulativeCostStep0006Path)

    if not objSinglePlRows or not objSingleCostRows or not objCumulativePlRows or not objCumulativeCostRows:
        return

    iSingleOperatingRowIndex: int = find_row_index_by_name(objSinglePlRows, "営業利益")
    iSingleManhourRowIndex: int = find_row_index_by_name(objSinglePlRows, "工数行(時間)")
    iCumulativeOperatingRowIndex: int = find_row_index_by_name(objCumulativePlRows, "営業利益")
    iCumulativeManhourRowIndex: int = find_row_index_by_name(objCumulativePlRows, "工数行(時間)")

    if iSingleOperatingRowIndex < 0 or iSingleManhourRowIndex < 0:
        return
    if iCumulativeOperatingRowIndex < 0 or iCumulativeManhourRowIndex < 0:
        return

    objSingle0001Rows = [list(objRow) for objRow in objSinglePlRows[: iSingleOperatingRowIndex + 1]]
    objSingle0002Rows = [list(objRow) for objRow in objSingleCostRows[1:]]
    objSingle0003Rows = [list(objRow) for objRow in objSinglePlRows[iSingleManhourRowIndex:]]
    objCumulative0001Rows = [
        list(objRow) for objRow in objCumulativePlRows[: iCumulativeOperatingRowIndex + 1]
    ]
    objCumulative0002Rows = [list(objRow) for objRow in objCumulativeCostRows[1:]]
    objCumulative0003Rows = [
        list(objRow) for objRow in objCumulativePlRows[iCumulativeManhourRowIndex:]
    ]

    pszSingle0001Path: str = os.path.join(pszDirectory, "0003_PJサマリ_step0007_単月_0001.tsv")
    pszSingle0002Path: str = os.path.join(pszDirectory, "0003_PJサマリ_step0007_単月_0002.tsv")
    pszSingle0003Path: str = os.path.join(pszDirectory, "0003_PJサマリ_step0007_単月_0003.tsv")
    pszCumulative0001Path: str = os.path.join(pszDirectory, "0003_PJサマリ_step0007_累計_0001.tsv")
    pszCumulative0002Path: str = os.path.join(pszDirectory, "0003_PJサマリ_step0007_累計_0002.tsv")
    pszCumulative0003Path: str = os.path.join(pszDirectory, "0003_PJサマリ_step0007_累計_0003.tsv")

    write_tsv_rows(pszSingle0001Path, objSingle0001Rows)
    write_tsv_rows(pszSingle0002Path, objSingle0002Rows)
    write_tsv_rows(pszSingle0003Path, objSingle0003Rows)
    write_tsv_rows(pszCumulative0001Path, objCumulative0001Rows)
    write_tsv_rows(pszCumulative0002Path, objCumulative0002Rows)
    write_tsv_rows(pszCumulative0003Path, objCumulative0003Rows)

    # step0007: 0001/0002/0003 を縦に連結して出力する
    objSingleFinalRows: List[List[str]] = []
    objSingleFinalRows.extend(objSingle0001Rows)
    objSingleFinalRows.extend(objSingle0002Rows)
    objSingleFinalRows.extend(objSingle0003Rows)
    objCumulativeFinalRows: List[List[str]] = []
    objCumulativeFinalRows.extend(objCumulative0001Rows)
    objCumulativeFinalRows.extend(objCumulative0002Rows)
    objCumulativeFinalRows.extend(objCumulative0003Rows)

    pszSingleOutputPath: str = os.path.join(pszDirectory, "0003_PJサマリ_step0007_単月_PL_CR.tsv")
    pszCumulativeOutputPath: str = os.path.join(pszDirectory, "0003_PJサマリ_step0007_累計_PL_CR.tsv")
    write_tsv_rows(pszSingleOutputPath, objSingleFinalRows)
    write_tsv_rows(pszCumulativeOutputPath, objCumulativeFinalRows)

    if objSingleFinalRows and objCumulativeFinalRows:
        pszScriptDirectory: str = get_script_base_directory()
        pszStep0008Directory: str = os.path.join(pszScriptDirectory, "PJ_Summary_step0008_Project")
        pszStep0009Directory: str = os.path.join(pszScriptDirectory, "PJ_Summary_step0009_Project")
        pszStep0010Directory: str = os.path.join(pszScriptDirectory, "PJ_Summary_step0010_Project")
        pszStep0011Directory: str = os.path.join(pszScriptDirectory, "PJ_Summary_step0011_Project")
        os.makedirs(pszStep0008Directory, exist_ok=True)
        os.makedirs(pszStep0009Directory, exist_ok=True)
        os.makedirs(pszStep0010Directory, exist_ok=True)
        os.makedirs(pszStep0011Directory, exist_ok=True)
        objStep0011ProjectInputs: List[Tuple[str, str]] = []
        objSingleHeaderRow: List[str] = objSingleFinalRows[0]
        objCumulativeHeaderRow: List[str] = objCumulativeFinalRows[0]
        iMaxColumns: int = max(len(objSingleHeaderRow), len(objCumulativeHeaderRow))
        for iColumnIndex in range(1, iMaxColumns):
            objSingleRatioRows: List[List[str]] = []
            objCumulativeRatioRows: List[List[str]] = []
            if iColumnIndex < len(objSingleHeaderRow):
                pszColumnName = objSingleHeaderRow[iColumnIndex]
                pszOutputName = f"0003_PJサマリ_step0008_単月_{pszColumnName}.tsv"
                pszOutputPath = os.path.join(pszStep0008Directory, pszOutputName)
                objSingleColumnRows = [
                    [
                        objRow[0] if len(objRow) > 0 else "",
                        objRow[iColumnIndex] if iColumnIndex < len(objRow) else "",
                    ]
                    for objRow in objSingleFinalRows
                ]
                write_tsv_rows(pszOutputPath, objSingleColumnRows)
                pszStep0009Name = f"0003_PJサマリ_step0009_単月_{pszColumnName}.tsv"
                pszStep0009Path = os.path.join(pszStep0009Directory, pszStep0009Name)
                objSingleRatioRows = add_sales_ratio_column(objSingleColumnRows)
                write_tsv_rows(pszStep0009Path, objSingleRatioRows)
            if iColumnIndex < len(objCumulativeHeaderRow):
                pszColumnName = objCumulativeHeaderRow[iColumnIndex]
                pszOutputName = f"0003_PJサマリ_step0008_累計_{pszColumnName}.tsv"
                pszOutputPath = os.path.join(pszStep0008Directory, pszOutputName)
                objCumulativeColumnRows = [
                    [
                        objRow[0] if len(objRow) > 0 else "",
                        objRow[iColumnIndex] if iColumnIndex < len(objRow) else "",
                    ]
                    for objRow in objCumulativeFinalRows
                ]
                write_tsv_rows(pszOutputPath, objCumulativeColumnRows)
                pszStep0009Name = f"0003_PJサマリ_step0009_累計_{pszColumnName}.tsv"
                pszStep0009Path = os.path.join(pszStep0009Directory, pszStep0009Name)
                objCumulativeRatioRows = add_sales_ratio_column(objCumulativeColumnRows)
                write_tsv_rows(pszStep0009Path, objCumulativeRatioRows)
                if objSingleRatioRows and objCumulativeRatioRows:
                    pszStep0010Name = f"0003_PJサマリ_step0010_単・累計_{pszColumnName}.tsv"
                    pszStep0010Path = os.path.join(pszStep0010Directory, pszStep0010Name)
                    objStep0010Rows = build_step0010_rows(objSingleRatioRows, objCumulativeRatioRows)
                    write_tsv_rows(pszStep0010Path, objStep0010Rows)
                    pszStep0011Name = f"0003_PJサマリ_step0011_単・累計_{pszColumnName}.tsv"
                    pszStep0011Path = os.path.join(pszStep0011Directory, pszStep0011Name)
                    objStep0011Rows = build_step0011_rows(objStep0010Rows)
                    write_tsv_rows(pszStep0011Path, objStep0011Rows)
                    objStep0011ProjectInputs.append((pszColumnName, pszStep0011Path))
                    create_pj_summary_pl_cr_manhour_excel(
                        pszDirectory,
                        pszColumnName,
                        pszStep0011Path,
                    )
        create_pj_summary_pl_cr_manhour_all_project_excel(
            pszDirectory,
            objStep0011ProjectInputs,
        )

    move_files_to_temp(
        [
            pszSingle0001Path,
            pszSingle0002Path,
            pszSingle0003Path,
            pszCumulative0001Path,
            pszCumulative0002Path,
            pszCumulative0003Path,
        ],
        pszDirectory,
    )
    move_files_to_temp_and_copy_back(
        [pszSingleOutputPath, pszCumulativeOutputPath],
        pszDirectory,
    )


def create_pj_summary(
    pszPlPath: str,
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
    create_step0007: bool = True,
    bWriteTotalsExcel: bool = False,
) -> None:
    def write_company_or_division_file(pszOrgMode: str) -> None:
        if EXECUTION_ROOT_DIRECTORY is None:
            return
        pszModePath: str = os.path.join(
            EXECUTION_ROOT_DIRECTORY,
            "company_or_division.txt",
        )
        pszModeLabel: str = "division" if pszOrgMode == "new" else "company"
        try:
            with open(pszModePath, "w", encoding="utf-8", newline="") as objModeFile:
                objModeFile.write(pszModeLabel + "\n")
        except OSError:
            return

    def write_step0004_error_file(pszStep0004Path: str, exc: Exception) -> None:
        pszErrorPath: str = pszStep0004Path.replace(".tsv", "_error.txt")
        try:
            with open(pszErrorPath, "w", encoding="utf-8", newline="") as objErrorFile:
                objErrorFile.write(f"Error: {exc}\n")
                objErrorFile.write(
                    "旧組織判定キー: 第一インキュ, 第二インキュ, 第三インキュ, 第四インキュ\n"
                )
                objErrorFile.write(
                    "新組織判定キー: テクノロジーインキュベーション, コンテンツビジネス, "
                    "スタートアップサイド, スタートアップコミュニティ, スタートアップグロース, 経営管理\n"
                )
                objErrorFile.write(
                    "共通カテゴリ: 事業開発, 子会社, 投資先, 本部\n"
                )
        except OSError:
            pass

    pszSummaryOrgMode: Optional[str] = None
    objSingleSummaryStep0003RowsCpCache: Optional[List[List[str]]] = None
    objCumulativeSummaryStep0003RowsCpCache: Optional[List[List[str]]] = None
    objStart, objEnd = objRange
    pszDirectory: str = get_script_base_directory()
    iEndYear, iEndMonth = objEnd
    pszEndMonth: str = f"{iEndMonth:02d}"
    pszSinglePlPath: str = os.path.join(
        pszDirectory,
        f"損益計算書_販管費配賦_{iEndYear}年{pszEndMonth}月_A∪B_プロジェクト名_C∪D_vertical.tsv",
    )
    pszSingleSummaryStep0002PathCp: str = os.path.join(
        pszDirectory,
        (
            "0001_CP別_step0002_単月_損益計算書_"
            f"{iEndYear}年{pszEndMonth}月.tsv"
        ),
    )
    pszSingleSummaryStep0003PathCp: str = os.path.join(
        pszDirectory,
        (
            "0001_CP別_step0003_単月_損益計算書_"
            f"{iEndYear}年{pszEndMonth}月.tsv"
        ),
    )
    pszCumulativePlPath: str = build_cumulative_file_path(
        pszDirectory,
        "損益計算書_販管費配賦",
        objStart,
        objEnd,
    ).replace(".tsv", "_vertical.tsv")

    objSingleRows: Optional[List[List[str]]] = None
    pszSinglePlStep0010Path: str = os.path.join(
        pszDirectory,
        f"損益計算書_販管費配賦_step0010_{iEndYear}年{pszEndMonth}月_A∪B_プロジェクト名_C∪D.tsv",
    )
    pszSinglePlStep0010VerticalPath: str = pszSinglePlStep0010Path.replace(
        ".tsv",
        "_vertical.tsv",
    )
    if os.path.isfile(pszSinglePlStep0010VerticalPath):
        objSingleRows = read_tsv_rows(pszSinglePlStep0010VerticalPath)
    elif os.path.isfile(pszSinglePlStep0010Path):
        objSingleRows = transpose_rows(read_tsv_rows(pszSinglePlStep0010Path))
    elif os.path.isfile(pszSinglePlPath):
        objSingleRows = read_tsv_rows(pszSinglePlPath)

    objCumulativeRows: Optional[List[List[str]]] = None
    if os.path.isfile(pszCumulativePlPath):
        objCumulativeRows = read_tsv_rows(pszCumulativePlPath)
    else:
        pszCumulativePlPathHorizontal: str = pszCumulativePlPath.replace("_vertical.tsv", ".tsv")
        if os.path.isfile(pszCumulativePlPathHorizontal):
            objCumulativeRows = transpose_rows(read_tsv_rows(pszCumulativePlPathHorizontal))

    if objSingleRows is None:
        if objStart == objEnd and os.path.isfile(pszSingleSummaryStep0002PathCp):
            objCompanyMapCpSingle = load_org_table_company_map(os.path.join(pszDirectory, "管轄PJ表.tsv"))
            objSingleSummaryStep0003RowsCp, objSingleSummaryStep0003DebugRowsCp = build_step0003_rows_with_debug(
                read_tsv_rows(pszSingleSummaryStep0002PathCp),
                objCompanyMapCpSingle,
            )
            objSingleSummaryStep0003RowsCpCache = objSingleSummaryStep0003RowsCp
            write_tsv_rows(pszSingleSummaryStep0003PathCp, objSingleSummaryStep0003RowsCp)
            pszSingleSummaryStep0003DebugPathCp: str = pszSingleSummaryStep0003PathCp.replace(
                ".tsv",
                "_debug.tsv",
            )
            write_tsv_rows(pszSingleSummaryStep0003DebugPathCp, objSingleSummaryStep0003DebugRowsCp)
        return
    if objCumulativeRows is None and objStart != objEnd:
        return

    objSummaryTargetColumns: List[str] = [
        "科目名",
        "純売上高",
        "売上原価",
        "売上総利益",
        "配賦販管費",
        "1Cカンパニー販管費",
        "2Cカンパニー販管費",
        "3Cカンパニー販管費",
        "4Cカンパニー販管費",
        "事業開発カンパニー販管費",
        "営業利益",
    ]
    objSingleSummaryRows: List[List[str]] = filter_rows_by_columns(
        objSingleRows,
        objSummaryTargetColumns,
    )
    pszSummaryStartMonth: str = f"{objStart[1]:02d}"
    pszSummaryEndMonth: str = f"{objEnd[1]:02d}"
    pszSingleSummaryStep0005VerticalPathCp: Optional[str] = None
    pszSingleSummaryStep0005VerticalPathCp0002: Optional[str] = None
    if objStart == objEnd:
        pszSingleSummaryPathCp: str = os.path.join(
            pszDirectory,
            (
                "0001_CP別_step0001_単月_損益計算書_"
                f"{iEndYear}年{pszEndMonth}月.tsv"
            ),
        )
        write_tsv_rows(pszSingleSummaryPathCp, objSingleSummaryRows)
        pszSingleSummaryPathCp0002: str = os.path.join(
            pszDirectory,
            (
                "0002_CP別_step0001_単月_損益計算書_"
                f"{iEndYear}年{pszEndMonth}月.tsv"
            ),
        )
        write_tsv_rows(pszSingleSummaryPathCp0002, objSingleSummaryRows)
        pszSingleSummaryStep0002PathCp0002: str = os.path.join(
            pszDirectory,
            (
                "0002_CP別_step0002_単月_損益計算書_"
                f"{iEndYear}年{pszEndMonth}月.tsv"
            ),
        )
        objSingleSummaryStep0002RowsCp0002 = combine_company_sg_admin_columns(
            read_tsv_rows(pszSingleSummaryPathCp0002)
        )
        write_tsv_rows(pszSingleSummaryStep0002PathCp0002, objSingleSummaryStep0002RowsCp0002)
        objSingleSummaryStep0002RowsCp = combine_company_sg_admin_columns(
            read_tsv_rows(pszSingleSummaryPathCp)
        )
        write_tsv_rows(pszSingleSummaryStep0002PathCp, objSingleSummaryStep0002RowsCp)
        pszSingleSummaryStep0003PathCp0002: str = os.path.join(
            pszDirectory,
            (
                "0002_CP別_step0003_単月_損益計算書_"
                f"{iEndYear}年{pszEndMonth}月.tsv"
            ),
        )
        objGroupMapCpSingle = load_org_table_group_map(os.path.join(pszDirectory, "管轄PJ表.tsv"))
        objCompanyMapCpSingle = load_org_table_company_map(os.path.join(pszDirectory, "管轄PJ表.tsv"))
        objSingleSummaryStep0003RowsCp0002 = build_step0003_rows(
            read_tsv_rows(pszSingleSummaryStep0002PathCp0002),
            objGroupMapCpSingle,
        )
        write_tsv_rows(pszSingleSummaryStep0003PathCp0002, objSingleSummaryStep0003RowsCp0002)
        objSingleSummaryStep0003RowsCp, objSingleSummaryStep0003DebugRowsCp = build_step0003_rows_with_debug(
            read_tsv_rows(pszSingleSummaryStep0002PathCp),
            objCompanyMapCpSingle,
        )
        write_tsv_rows(pszSingleSummaryStep0003PathCp, objSingleSummaryStep0003RowsCp)
        pszSingleSummaryStep0003DebugPathCp: str = pszSingleSummaryStep0003PathCp.replace(
            ".tsv",
            "_debug.tsv",
        )
        write_tsv_rows(pszSingleSummaryStep0003DebugPathCp, objSingleSummaryStep0003DebugRowsCp)
        pszSingleSummaryStep0004PathCp0002: str = os.path.join(
            pszDirectory,
            (
                "0002_CP別_step0004_単月_損益計算書_"
                f"{iEndYear}年{pszEndMonth}月.tsv"
            ),
        )
        objSingleSummaryStep0004RowsCp0002 = build_step0004_rows_for_group_summary(
            objSingleSummaryStep0003RowsCp0002
        )
        write_tsv_rows(pszSingleSummaryStep0004PathCp0002, objSingleSummaryStep0004RowsCp0002)
        pszSingleSummaryStep0004VerticalPathCp0002: str = pszSingleSummaryStep0004PathCp0002.replace(
            ".tsv",
            "_vertical.tsv",
        )
        objSingleSummaryStep0004VerticalRowsCp0002 = transpose_rows(objSingleSummaryStep0004RowsCp0002)
        write_tsv_rows(
            pszSingleSummaryStep0004VerticalPathCp0002,
            objSingleSummaryStep0004VerticalRowsCp0002,
        )
        objSingleSummaryStep0005VerticalRowsCp0002 = insert_ratio_rows_for_vertical(
            objSingleSummaryStep0004VerticalRowsCp0002
        )
        pszSingleSummaryStep0005VerticalPathCp0002 = os.path.join(
            pszDirectory,
            (
                "0002_CP別_step0005_単月_損益計算書_"
                f"{iEndYear}年{pszEndMonth}月_vertical.tsv"
            ),
        )
        write_tsv_rows(
            pszSingleSummaryStep0005VerticalPathCp0002,
            objSingleSummaryStep0005VerticalRowsCp0002,
        )
        pszSingleSummaryStep0004PathCp: str = os.path.join(
            pszDirectory,
            (
                "0001_CP別_step0004_単月_損益計算書_"
                f"{iEndYear}年{pszEndMonth}月.tsv"
            ),
        )
        objSingleSummaryStep0004RowsCp = build_step0004_rows_for_summary(
            objSingleSummaryStep0003RowsCp
        )
        write_tsv_rows(pszSingleSummaryStep0004PathCp, objSingleSummaryStep0004RowsCp)
        pszSingleSummaryStep0004VerticalPathCp: str = pszSingleSummaryStep0004PathCp.replace(
            ".tsv",
            "_vertical.tsv",
        )
        objSingleSummaryStep0004VerticalRowsCp = transpose_rows(objSingleSummaryStep0004RowsCp)
        write_tsv_rows(
            pszSingleSummaryStep0004VerticalPathCp,
            objSingleSummaryStep0004VerticalRowsCp,
        )
        objSingleSummaryStep0005VerticalRowsCp = insert_ratio_rows_for_vertical(
            objSingleSummaryStep0004VerticalRowsCp
        )
        pszSingleSummaryStep0005VerticalPathCp: str = os.path.join(
            pszDirectory,
            (
                "0001_CP別_step0005_単月_損益計算書_"
                f"{iEndYear}年{pszEndMonth}月_vertical.tsv"
            ),
        )
        write_tsv_rows(
            pszSingleSummaryStep0005VerticalPathCp,
            objSingleSummaryStep0005VerticalRowsCp,
        )
        move_cp_step0001_to_step0004_vertical_files(
            pszDirectory,
            objStart,
            objEnd,
        )
    pszCumulativeSummaryStep0005VerticalPathCp: Optional[str] = None
    pszCumulativeSummaryStep0005VerticalPathCp0002: Optional[str] = None
    if objCumulativeRows is not None:
        objCumulativeSummaryRows: List[List[str]] = filter_rows_by_columns(
            objCumulativeRows,
            objSummaryTargetColumns,
        )
        pszCumulativeSummaryPathCp: str = os.path.join(
            pszDirectory,
            (
                "0001_CP別_step0001_累計_損益計算書_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
            ),
        )
        write_tsv_rows(pszCumulativeSummaryPathCp, objCumulativeSummaryRows)
        pszCumulativeSummaryPathCp0002: str = os.path.join(
            pszDirectory,
            (
                "0002_CP別_step0001_累計_損益計算書_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
            ),
        )
        write_tsv_rows(pszCumulativeSummaryPathCp0002, objCumulativeSummaryRows)
        pszCumulativeSummaryStep0002PathCp0002: str = os.path.join(
            pszDirectory,
            (
                "0002_CP別_step0002_累計_損益計算書_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
            ),
        )
        objCumulativeSummaryStep0002RowsCp0002 = combine_company_sg_admin_columns(
            read_tsv_rows(pszCumulativeSummaryPathCp0002)
        )
        write_tsv_rows(pszCumulativeSummaryStep0002PathCp0002, objCumulativeSummaryStep0002RowsCp0002)
        pszCumulativeSummaryStep0003PathCp0002: str = os.path.join(
            pszDirectory,
            (
                "0002_CP別_step0003_累計_損益計算書_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
            ),
        )
        pszCumulativeSummaryStep0002PathCp: str = os.path.join(
            pszDirectory,
            (
                "0001_CP別_step0002_累計_損益計算書_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
            ),
        )
        objCumulativeSummaryStep0002RowsCp = combine_company_sg_admin_columns(
            read_tsv_rows(pszCumulativeSummaryPathCp)
        )
        write_tsv_rows(pszCumulativeSummaryStep0002PathCp, objCumulativeSummaryStep0002RowsCp)
        pszCumulativeSummaryStep0003PathCp: str = os.path.join(
            pszDirectory,
            (
                "0001_CP別_step0003_累計_損益計算書_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
            ),
        )
        objGroupMapCp = load_org_table_group_map(os.path.join(pszDirectory, "管轄PJ表.tsv"))
        objCompanyMapCp = load_org_table_company_map(os.path.join(pszDirectory, "管轄PJ表.tsv"))
        objCumulativeSummaryStep0003RowsCp0002 = build_step0003_rows(
            read_tsv_rows(pszCumulativeSummaryStep0002PathCp0002),
            objGroupMapCp,
        )
        write_tsv_rows(pszCumulativeSummaryStep0003PathCp0002, objCumulativeSummaryStep0003RowsCp0002)
        objCumulativeSummaryStep0003RowsCp = build_step0003_rows(
            read_tsv_rows(pszCumulativeSummaryStep0002PathCp),
            objCompanyMapCp,
        )
        objCumulativeSummaryStep0003RowsCpCache = objCumulativeSummaryStep0003RowsCp
        write_tsv_rows(pszCumulativeSummaryStep0003PathCp, objCumulativeSummaryStep0003RowsCp)
        pszCumulativeSummaryStep0004PathCp0002: str = os.path.join(
            pszDirectory,
            (
                "0002_CP別_step0004_累計_損益計算書_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
            ),
        )
        objCumulativeSummaryStep0004RowsCp0002 = build_step0004_rows_for_group_summary(
            objCumulativeSummaryStep0003RowsCp0002
        )
        write_tsv_rows(pszCumulativeSummaryStep0004PathCp0002, objCumulativeSummaryStep0004RowsCp0002)
        pszCumulativeSummaryStep0004VerticalPathCp0002: str = pszCumulativeSummaryStep0004PathCp0002.replace(
            ".tsv",
            "_vertical.tsv",
        )
        objCumulativeSummaryStep0004VerticalRowsCp0002 = transpose_rows(
            objCumulativeSummaryStep0004RowsCp0002
        )
        write_tsv_rows(
            pszCumulativeSummaryStep0004VerticalPathCp0002,
            objCumulativeSummaryStep0004VerticalRowsCp0002,
        )
        objCumulativeSummaryStep0005VerticalRowsCp0002 = insert_ratio_rows_for_vertical(
            objCumulativeSummaryStep0004VerticalRowsCp0002
        )
        pszCumulativeSummaryStep0005VerticalPathCp0002 = os.path.join(
            pszDirectory,
            (
                "0002_CP別_step0005_累計_損益計算書_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月_vertical.tsv"
            ),
        )
        write_tsv_rows(
            pszCumulativeSummaryStep0005VerticalPathCp0002,
            objCumulativeSummaryStep0005VerticalRowsCp0002,
        )
        pszCumulativeSummaryStep0004PathCp: str = os.path.join(
            pszDirectory,
            (
                "0001_CP別_step0004_累計_損益計算書_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
            ),
        )
        objCumulativeSummaryStep0004RowsCp = build_step0004_rows_for_summary(
            objCumulativeSummaryStep0003RowsCp
        )
        write_tsv_rows(pszCumulativeSummaryStep0004PathCp, objCumulativeSummaryStep0004RowsCp)
        pszCumulativeSummaryStep0004VerticalPathCp: str = pszCumulativeSummaryStep0004PathCp.replace(
            ".tsv",
            "_vertical.tsv",
        )
        objCumulativeSummaryStep0004VerticalRowsCp = transpose_rows(objCumulativeSummaryStep0004RowsCp)
        write_tsv_rows(
            pszCumulativeSummaryStep0004VerticalPathCp,
            objCumulativeSummaryStep0004VerticalRowsCp,
        )
        objCumulativeSummaryStep0005VerticalRowsCp = insert_ratio_rows_for_vertical(
            objCumulativeSummaryStep0004VerticalRowsCp
        )
        pszCumulativeSummaryStep0005VerticalPathCp: str = os.path.join(
            pszDirectory,
            (
                "0001_CP別_step0005_累計_損益計算書_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月_vertical.tsv"
            ),
        )
        write_tsv_rows(
            pszCumulativeSummaryStep0005VerticalPathCp,
            objCumulativeSummaryStep0005VerticalRowsCp,
        )
        if objStart != objEnd and objStart[1] == 4 and objEnd[1] != 3:
            pszPriorCp0001Path = create_empty_previous_fiscal_cp_step0005_vertical(
                pszDirectory,
                objStart,
                objEnd,
                "0001_CP別",
            )
            if pszPriorCp0001Path:
                copy_company_step0006_files(
                    pszDirectory,
                    [pszPriorCp0001Path],
                    "0001_CP別_step0006",
                    create_step0007=False,
                )
            pszPriorCp0002Path = create_empty_previous_fiscal_cp_step0005_vertical(
                pszDirectory,
                objStart,
                objEnd,
                "0002_CP別",
            )
            if pszPriorCp0002Path:
                copy_company_step0006_files(
                    pszDirectory,
                    [pszPriorCp0002Path],
                    "0002_CP別_step0006",
                    create_step0007=False,
                )
    move_cp_step0001_to_step0004_vertical_files(
        pszDirectory,
        objStart,
        objEnd,
    )
    copy_cp_step0005_vertical_files(
        pszDirectory,
        [
            pszSingleSummaryStep0005VerticalPathCp,
            pszCumulativeSummaryStep0005VerticalPathCp,
        ],
    )
    if pszSummaryOrgMode is None:
        try:
            if objCumulativeSummaryStep0003RowsCpCache is not None:
                pszSummaryOrgMode = detect_step0004_org_mode(objCumulativeSummaryStep0003RowsCpCache)
            elif objSingleSummaryStep0003RowsCpCache is not None:
                pszSummaryOrgMode = detect_step0004_org_mode(objSingleSummaryStep0003RowsCpCache)
        except ValueError:
            pszSummaryOrgMode = None
    copy_company_step0006_files(
        pszDirectory,
        [pszSingleSummaryStep0005VerticalPathCp, pszCumulativeSummaryStep0005VerticalPathCp],
        "0001_CP別_step0006",
        create_step0007=create_step0007,
        pszOrgMode=pszSummaryOrgMode,
    )
    copy_group_step0006_files(
        pszDirectory,
        [
            pszSingleSummaryStep0005VerticalPathCp0002,
            pszCumulativeSummaryStep0005VerticalPathCp0002,
        ],
        create_step0007=create_step0007,
    )
    pszSingleSummaryPath: str = os.path.join(
        pszDirectory,
        f"0004_PJサマリ_step0001_単月_損益計算書_{iEndYear}年{pszEndMonth}月.tsv",
    )
    write_tsv_rows(pszSingleSummaryPath, objSingleSummaryRows)
    pszSingleSummaryPath0005: str = os.path.join(
        pszDirectory,
        f"0005_PJサマリ_step0001_単月_損益計算書_{iEndYear}年{pszEndMonth}月.tsv",
    )
    # 0005 は copy_pj_summary_0005_files によるコピーを行わず、
    # step0001 以降は必要なステップを本関数内で作成する。
    write_tsv_rows(pszSingleSummaryPath0005, objSingleSummaryRows)

    objSingleStep0002Rows0005 = combine_company_sg_admin_columns(
        read_tsv_rows(pszSingleSummaryPath0005)
    )
    pszSingleStep0002Path0005: str = os.path.join(
        pszDirectory,
        f"0005_PJサマリ_step0002_単月_損益計算書_{iEndYear}年{pszEndMonth}月.tsv",
    )
    write_tsv_rows(pszSingleStep0002Path0005, objSingleStep0002Rows0005)

    objSingleStep0002Rows = combine_company_sg_admin_columns(
        read_tsv_rows(pszSingleSummaryPath)
    )
    pszSingleStep0002Path: str = os.path.join(
        pszDirectory,
        f"0004_PJサマリ_step0002_単月_損益計算書_{iEndYear}年{pszEndMonth}月.tsv",
    )
    write_tsv_rows(pszSingleStep0002Path, objSingleStep0002Rows)

    pszOrgTablePath: str = os.path.join(pszDirectory, "管轄PJ表.tsv")
    objGroupMap = load_org_table_group_map(pszOrgTablePath)
    objCompanyMap = load_org_table_company_map(pszOrgTablePath)
    objSingleStep0003Rows = build_step0003_rows(
        read_tsv_rows(pszSingleStep0002Path),
        objCompanyMap,
    )
    pszSingleStep0003Path: str = os.path.join(
        pszDirectory,
        f"0004_PJサマリ_step0003_単月_損益計算書_{iEndYear}年{pszEndMonth}月.tsv",
    )
    write_tsv_rows(pszSingleStep0003Path, objSingleStep0003Rows)
    objSingleStep0003Rows0005 = build_step0003_rows(
        read_tsv_rows(pszSingleStep0002Path0005),
        objGroupMap,
    )
    pszSingleStep0003Path0005: str = os.path.join(
        pszDirectory,
        f"0005_PJサマリ_step0003_単月_損益計算書_{iEndYear}年{pszEndMonth}月.tsv",
    )
    write_tsv_rows(pszSingleStep0003Path0005, objSingleStep0003Rows0005)
    pszSingleStep0004Path0005: str = os.path.join(
        pszDirectory,
        f"0005_PJサマリ_step0004_単月_損益計算書_{iEndYear}年{pszEndMonth}月.tsv",
    )
    objSingleStep0003Rows0005 = read_tsv_rows(pszSingleStep0003Path0005)
    objSingleStep0004Rows0005 = build_step0004_rows_for_group_summary(objSingleStep0003Rows0005)
    write_tsv_rows(pszSingleStep0004Path0005, objSingleStep0004Rows0005)
    pszSingleStep0004Path: str = os.path.join(
        pszDirectory,
        f"0004_PJサマリ_step0004_単月_損益計算書_{iEndYear}年{pszEndMonth}月.tsv",
    )
    objSingleStep0003Rows = read_tsv_rows(pszSingleStep0003Path)
    pszSummaryOrgMode = "legacy"
    try:
        pszSummaryOrgMode = detect_step0004_org_mode(objSingleStep0003Rows)
        objSingleStep0004Rows = build_step0004_rows_for_summary(objSingleStep0003Rows)
    except ValueError as exc:
        write_step0004_error_file(pszSingleStep0004Path, exc)
        return
    write_tsv_rows(pszSingleStep0004Path, objSingleStep0004Rows)
    pszSingleStep0005Path: str = os.path.join(
        pszDirectory,
        f"0004_PJサマリ_step0005_単月_損益計算書_{iEndYear}年{pszEndMonth}月.tsv",
    )
    objSingleStep0005Rows = add_profit_ratio_columns(read_tsv_rows(pszSingleStep0004Path))
    write_tsv_rows(pszSingleStep0005Path, objSingleStep0005Rows)

    if objCumulativeRows is None:
        return

    pszCumulativeSummaryPath: str = os.path.join(
        pszDirectory,
        (
            "0004_PJサマリ_step0001_累計_損益計算書_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
        ),
    )
    write_tsv_rows(pszCumulativeSummaryPath, objCumulativeSummaryRows)
    pszCumulativeSummaryPath0005: str = os.path.join(
        pszDirectory,
        (
            "0005_PJサマリ_step0001_累計_損益計算書_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
        ),
    )
    write_tsv_rows(pszCumulativeSummaryPath0005, objCumulativeSummaryRows)

    objCumulativeStep0002Rows0005 = combine_company_sg_admin_columns(
        read_tsv_rows(pszCumulativeSummaryPath0005)
    )
    pszCumulativeStep0002Path0005: str = os.path.join(
        pszDirectory,
        (
            "0005_PJサマリ_step0002_累計_損益計算書_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
        ),
    )
    write_tsv_rows(pszCumulativeStep0002Path0005, objCumulativeStep0002Rows0005)

    objCumulativeStep0002Rows = combine_company_sg_admin_columns(
        read_tsv_rows(pszCumulativeSummaryPath)
    )
    pszCumulativeStep0002Path: str = os.path.join(
        pszDirectory,
        (
            "0004_PJサマリ_step0002_累計_損益計算書_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
        ),
    )
    write_tsv_rows(pszCumulativeStep0002Path, objCumulativeStep0002Rows)

    objCumulativeStep0003Rows = build_step0003_rows(
        read_tsv_rows(pszCumulativeStep0002Path),
        objCompanyMap,
    )
    pszCumulativeStep0003Path: str = os.path.join(
        pszDirectory,
        (
            "0004_PJサマリ_step0003_累計_損益計算書_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
        ),
    )
    write_tsv_rows(pszCumulativeStep0003Path, objCumulativeStep0003Rows)
    objCumulativeStep0003Rows0005 = build_step0003_rows(
        read_tsv_rows(pszCumulativeStep0002Path0005),
        objGroupMap,
    )
    pszCumulativeStep0003Path0005: str = os.path.join(
        pszDirectory,
        (
            "0005_PJサマリ_step0003_累計_損益計算書_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
        ),
    )
    write_tsv_rows(pszCumulativeStep0003Path0005, objCumulativeStep0003Rows0005)
    pszCumulativeStep0004Path0005: str = os.path.join(
        pszDirectory,
        (
            "0005_PJサマリ_step0004_累計_損益計算書_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
        ),
    )
    objCumulativeStep0003Rows0005 = read_tsv_rows(pszCumulativeStep0003Path0005)
    objCumulativeStep0004Rows0005 = build_step0004_rows_for_group_summary(
        objCumulativeStep0003Rows0005
    )
    write_tsv_rows(pszCumulativeStep0004Path0005, objCumulativeStep0004Rows0005)
    pszStep0005Path0005: str = os.path.join(
        pszDirectory,
        (
            "0005_PJサマリ_step0005_単・累_損益計算書_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
        ),
    )
    objSingleStep0005Rows0005 = add_profit_ratio_columns(read_tsv_rows(pszSingleStep0004Path0005))
    objCumulativeStep0005Rows0005 = add_profit_ratio_columns(read_tsv_rows(pszCumulativeStep0004Path0005))
    objStep0005Rows0005 = build_step0005_rows_for_summary(
        objSingleStep0005Rows0005,
        objCumulativeStep0005Rows0005,
    )
    write_tsv_rows(pszStep0005Path0005, objStep0005Rows0005)

    pszStep0006Path0005: str = os.path.join(
        pszDirectory,
        (
            "0005_PJサマリ_step0006_単・累_損益計算書_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
        ),
    )
    objStep0006Rows0005 = build_step0006_rows_for_summary_0005(pszStep0005Path0005)
    write_tsv_rows(pszStep0006Path0005, objStep0006Rows0005)
    pszStep0007Path0005: str = os.path.join(
        pszDirectory,
        (
            "0005_PJサマリ_step0007_単・累_損益計算書_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
        ),
    )
    objStep0007Rows0005 = build_step0007_rows_for_summary_0005(pszStep0006Path0005)
    write_tsv_rows(pszStep0007Path0005, objStep0007Rows0005)
    if objStart != objEnd and bWriteTotalsExcel:
        insert_step0006_rows_into_group_summary_excel(
            objStep0007Rows0005,
            objStart,
            objEnd,
        )
    pszCumulativeStep0004Path: str = os.path.join(
        pszDirectory,
        (
            "0004_PJサマリ_step0004_累計_損益計算書_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
        ),
    )
    objCumulativeStep0003Rows = read_tsv_rows(pszCumulativeStep0003Path)
    try:
        pszCumulativeOrgMode = detect_step0004_org_mode(objCumulativeStep0003Rows)
        if pszSummaryOrgMode != pszCumulativeOrgMode:
            raise ValueError("step0004 集計エラー: 単月と累計で組織判定結果が一致しません。")
        objCumulativeStep0004Rows = build_step0004_rows_for_summary(objCumulativeStep0003Rows)
    except ValueError as exc:
        write_step0004_error_file(pszCumulativeStep0004Path, exc)
        return
    write_tsv_rows(pszCumulativeStep0004Path, objCumulativeStep0004Rows)
    pszCumulativeStep0005Path: str = os.path.join(
        pszDirectory,
        (
            "0004_PJサマリ_step0005_累計_損益計算書_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
        ),
    )
    objCumulativeStep0005Rows = add_profit_ratio_columns(read_tsv_rows(pszCumulativeStep0004Path))
    write_tsv_rows(pszCumulativeStep0005Path, objCumulativeStep0005Rows)
    pszStep0006Path: str = os.path.join(
        pszDirectory,
        (
            "0004_PJサマリ_step0006_単・累_損益計算書_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月.tsv"
        ),
    )
    objStep0006Rows = build_step0005_rows_for_summary(
        read_tsv_rows(pszSingleStep0005Path),
        read_tsv_rows(pszCumulativeStep0005Path),
    )
    write_tsv_rows(pszStep0006Path, objStep0006Rows)
    objStep0007Rows = build_step0006_rows_for_summary(read_tsv_rows(pszStep0006Path))
    pszStep0007Path: str = pszStep0006Path.replace("step0006_", "step0007_", 1)
    write_tsv_rows(pszStep0007Path, objStep0007Rows)
    if objStart != objEnd and bWriteTotalsExcel:
        write_company_or_division_file(pszSummaryOrgMode)
        insert_step0006_rows_into_company_summary_excel(
            objStep0007Rows,
            objStart,
            objEnd,
            pszSummaryOrgMode,
        )

    objSingleOutputRows: List[List[str]] = []
    for objRow in objSingleRows:
        pszName: str = objRow[0] if objRow else ""
        if pszName == "合計" or pszName.startswith("C"):
            continue
        objSingleOutputRows.append(objRow)

    objCumulativeOutputRows: List[List[str]] = []
    for objRow in objCumulativeRows:
        pszName: str = objRow[0] if objRow else ""
        if pszName == "合計" or pszName.startswith("C"):
            continue
        objCumulativeOutputRows.append(objRow)

    pszSingleOutputPath: str = os.path.join(
        pszDirectory,
        f"0001_PJサマリ_step0001_{iEndYear}年{pszEndMonth}月_単月_損益計算書.tsv",
    )
    pszCumulativeOutputPath: str = os.path.join(
        pszDirectory,
        (
            "0001_PJサマリ_step0001_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月_累計_損益計算書.tsv"
        ),
    )
    write_tsv_rows(pszSingleOutputPath, objSingleOutputRows)
    write_tsv_rows(pszCumulativeOutputPath, objCumulativeOutputRows)

    pszSingleCostReportPath: str = os.path.join(
        pszDirectory,
        f"製造原価報告書_{iEndYear}年{pszEndMonth}月_A∪B_プロジェクト名_C∪D.tsv",
    )
    pszCumulativeCostReportPath: str = build_cumulative_file_path(
        pszDirectory,
        "製造原価報告書",
        objStart,
        objEnd,
    )
    if os.path.isfile(pszSingleCostReportPath):
        objCostReportSingleRows: List[List[str]] = read_tsv_rows(pszSingleCostReportPath)
        pszCostReportSingleOutputPath: str = os.path.join(
            pszDirectory,
            f"0001_PJサマリ_step0001_{iEndYear}年{pszEndMonth}月_単月_製造原価報告書.tsv",
        )
        write_tsv_rows(pszCostReportSingleOutputPath, objCostReportSingleRows)
    if os.path.isfile(pszCumulativeCostReportPath):
        objCostReportCumulativeRows: List[List[str]] = read_tsv_rows(pszCumulativeCostReportPath)
        pszCostReportCumulativeOutputPath: str = os.path.join(
            pszDirectory,
            (
                "0001_PJサマリ_step0001_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月_累計_製造原価報告書.tsv"
            ),
        )
        write_tsv_rows(pszCostReportCumulativeOutputPath, objCostReportCumulativeRows)

    objSingleOutputVerticalRows = transpose_rows(objSingleOutputRows)
    pszSingleOutputVerticalPath: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0001_単月_損益計算書.tsv",
    )
    write_tsv_rows(pszSingleOutputVerticalPath, objSingleOutputVerticalRows)

    objCumulativeOutputVerticalRows = transpose_rows(objCumulativeOutputRows)
    pszCumulativeOutputVerticalPath: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0001_累計_損益計算書.tsv",
    )
    write_tsv_rows(pszCumulativeOutputVerticalPath, objCumulativeOutputVerticalRows)

    if os.path.isfile(pszSingleCostReportPath):
        pszCostReportSingleOutputPath: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0001_単月_製造原価報告書.tsv",
        )
        shutil.copy2(pszSingleCostReportPath, pszCostReportSingleOutputPath)
    if os.path.isfile(pszCumulativeCostReportPath):
        pszCostReportCumulativeOutputPath: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0001_累計_製造原価報告書.tsv",
        )
        shutil.copy2(pszCumulativeCostReportPath, pszCostReportCumulativeOutputPath)

    #//
    #// PJサマリの損益計算書部分の作成
    #//
    #// 科目名
    #// 純売上高
    #// 売上総利益
    #// 配賦販管費
    #// 1Cカンパニー販管費
    #// 2Cカンパニー販管費
    #// 3Cカンパニー販管費
    #// 4Cカンパニー販管費
    #// 事業開発カンパニー販管費
    #// 営業利益
    #//
    objTargetNames: List[str] = [
        "科目名",
        "純売上高",
        "売上総利益",
        "配賦販管費",
        "1Cカンパニー販管費",
        "2Cカンパニー販管費",
        "3Cカンパニー販管費",
        "4Cカンパニー販管費",
        "事業開発カンパニー販管費",
        "工数",
        "営業利益",
    ]
    objSingleStep0002Rows = filter_rows_by_names(
        objSingleOutputVerticalRows,
        objTargetNames,
    )
    objCumulativeStep0002Rows = filter_rows_by_names(
        objCumulativeOutputVerticalRows,
        objTargetNames,
    )
    pszSingleStep0002Path: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0002_単月_損益計算書.tsv",
    )
    pszCumulativeStep0002Path: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0002_累計_損益計算書.tsv",
    )
    write_tsv_rows(pszSingleStep0002Path, objSingleStep0002Rows)
    write_tsv_rows(pszCumulativeStep0002Path, objCumulativeStep0002Rows)

    if os.path.isfile(pszSingleCostReportPath):
        pszCostReportSingleStep0002Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0002_単月_製造原価報告書.tsv",
        )
        shutil.copy2(pszSingleCostReportPath, pszCostReportSingleStep0002Path)
        pszCostReportSingleStep0003Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0003_単月_製造原価報告書.tsv",
        )
        shutil.copy2(pszSingleCostReportPath, pszCostReportSingleStep0003Path)
        pszCostReportSingleStep0004Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0004_単月_製造原価報告書.tsv",
        )
        shutil.copy2(pszCostReportSingleStep0003Path, pszCostReportSingleStep0004Path)
        pszCostReportSingleStep0004VerticalPath: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0004_単月_製造原価報告書_vertical.tsv",
        )
        write_tsv_rows(
            pszCostReportSingleStep0004VerticalPath,
            transpose_rows(read_tsv_rows(pszCostReportSingleStep0004Path)),
        )
    if os.path.isfile(pszCumulativeCostReportPath):
        pszCostReportCumulativeStep0002Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0002_累計_製造原価報告書.tsv",
        )
        shutil.copy2(pszCumulativeCostReportPath, pszCostReportCumulativeStep0002Path)
        pszCostReportCumulativeStep0003Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0003_累計_製造原価報告書.tsv",
        )
        shutil.copy2(pszCumulativeCostReportPath, pszCostReportCumulativeStep0003Path)
        pszCostReportCumulativeStep0004Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0004_累計_製造原価報告書.tsv",
        )
        shutil.copy2(pszCostReportCumulativeStep0003Path, pszCostReportCumulativeStep0004Path)
        pszCostReportCumulativeStep0004VerticalPath: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0004_累計_製造原価報告書_vertical.tsv",
        )
        write_tsv_rows(
            pszCostReportCumulativeStep0004VerticalPath,
            transpose_rows(read_tsv_rows(pszCostReportCumulativeStep0004Path)),
        )

    pszSingleStep0003Path: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0003_単月_損益計算書.tsv",
    )
    if os.path.isfile(pszSingleStep0002Path):
        objSingleStep0002Rows = read_tsv_rows(pszSingleStep0002Path)
        objSingleStep0003Rows = add_company_sg_admin_cost_total_row(objSingleStep0002Rows)
        write_tsv_rows(pszSingleStep0003Path, objSingleStep0003Rows)
        pszSingleStep0004Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0004_単月_損益計算書.tsv",
        )
        objSingleStep0004Rows = move_row_between(
            objSingleStep0003Rows,
            "カンパニー販管費",
            "配賦販管費",
            "営業利益",
        )
        # step0004の損益計算書として保存する
        write_tsv_rows(pszSingleStep0004Path, objSingleStep0004Rows)
        pszSingleStep0004VerticalPath: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0004_単月_損益計算書_vertical.tsv",
        )
        write_tsv_rows(pszSingleStep0004VerticalPath, transpose_rows(objSingleStep0004Rows))

    pszCumulativeStep0003Path: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0003_累計_損益計算書.tsv",
    )
    if os.path.isfile(pszCumulativeStep0002Path):
        objCumulativeStep0002Rows = read_tsv_rows(pszCumulativeStep0002Path)
        objCumulativeStep0003Rows = add_company_sg_admin_cost_total_row(objCumulativeStep0002Rows)
        write_tsv_rows(pszCumulativeStep0003Path, objCumulativeStep0003Rows)
        pszCumulativeStep0004Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0004_累計_損益計算書.tsv",
        )
        objCumulativeStep0004Rows = move_row_between(
            objCumulativeStep0003Rows,
            "カンパニー販管費",
            "配賦販管費",
            "営業利益",
        )
        # step0004の損益計算書として保存する
        write_tsv_rows(pszCumulativeStep0004Path, objCumulativeStep0004Rows)
        pszCumulativeStep0004VerticalPath: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0004_累計_損益計算書_vertical.tsv",
        )
        write_tsv_rows(
            pszCumulativeStep0004VerticalPath,
            transpose_rows(objCumulativeStep0004Rows),
        )

    pszSingleCostStep0004VerticalPath: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0004_単月_製造原価報告書_vertical.tsv",
    )
    pszSinglePlStep0004VerticalPath: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0004_単月_損益計算書_vertical.tsv",
    )
    if os.path.isfile(pszSingleCostStep0004VerticalPath) and os.path.isfile(
        pszSinglePlStep0004VerticalPath
    ):
        objSingleCostStep0004Rows = read_tsv_rows(pszSingleCostStep0004VerticalPath)
        objSinglePlStep0004Rows = read_tsv_rows(pszSinglePlStep0004VerticalPath)
        objAlignedCostRows, objAlignedPlRows = align_vertical_rows_for_union(
            objSingleCostStep0004Rows,
            objSinglePlStep0004Rows,
        )
        pszSingleCostStep0005VerticalPath: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0005_単月_製造原価報告書_E∪F_vertical.tsv",
        )
        pszSinglePlStep0005VerticalPath: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0005_単月_損益計算書_E∪F_vertical.tsv",
        )
        write_tsv_rows(pszSingleCostStep0005VerticalPath, objAlignedCostRows)
        write_tsv_rows(pszSinglePlStep0005VerticalPath, objAlignedPlRows)
        pszSingleCostStep0005Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0005_単月_製造原価報告書_E∪F.tsv",
        )
        pszSinglePlStep0005Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0005_単月_損益計算書_E∪F.tsv",
        )
        write_tsv_rows(pszSingleCostStep0005Path, transpose_rows(objAlignedCostRows))
        write_tsv_rows(pszSinglePlStep0005Path, transpose_rows(objAlignedPlRows))
        pszSingleCostStep0006Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0006_単月_製造原価報告書_E∪F.tsv",
        )
        pszSinglePlStep0006Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0006_単月_損益計算書_E∪F.tsv",
        )
        shutil.copy2(pszSingleCostStep0005Path, pszSingleCostStep0006Path)
        objSingleStep0006Rows = insert_per_hour_rows(read_tsv_rows(pszSinglePlStep0005Path))
        write_tsv_rows(pszSinglePlStep0006Path, objSingleStep0006Rows)

    pszCumulativeCostStep0004VerticalPath: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0004_累計_製造原価報告書_vertical.tsv",
    )
    pszCumulativePlStep0004VerticalPath: str = os.path.join(
        pszDirectory,
        "0003_PJサマリ_step0004_累計_損益計算書_vertical.tsv",
    )
    if os.path.isfile(pszCumulativeCostStep0004VerticalPath) and os.path.isfile(
        pszCumulativePlStep0004VerticalPath
    ):
        objCumulativeCostStep0004Rows = read_tsv_rows(pszCumulativeCostStep0004VerticalPath)
        objCumulativePlStep0004Rows = read_tsv_rows(pszCumulativePlStep0004VerticalPath)
        objAlignedCostRows, objAlignedPlRows = align_vertical_rows_for_union(
            objCumulativeCostStep0004Rows,
            objCumulativePlStep0004Rows,
        )
        pszCumulativeCostStep0005VerticalPath: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0005_累計_製造原価報告書_E∪F_vertical.tsv",
        )
        pszCumulativePlStep0005VerticalPath: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0005_累計_損益計算書_E∪F_vertical.tsv",
        )
        write_tsv_rows(pszCumulativeCostStep0005VerticalPath, objAlignedCostRows)
        write_tsv_rows(pszCumulativePlStep0005VerticalPath, objAlignedPlRows)
        pszCumulativeCostStep0005Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0005_累計_製造原価報告書_E∪F.tsv",
        )
        pszCumulativePlStep0005Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0005_累計_損益計算書_E∪F.tsv",
        )
        write_tsv_rows(pszCumulativeCostStep0005Path, transpose_rows(objAlignedCostRows))
        write_tsv_rows(pszCumulativePlStep0005Path, transpose_rows(objAlignedPlRows))
        pszCumulativeCostStep0006Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0006_累計_製造原価報告書_E∪F.tsv",
        )
        pszCumulativePlStep0006Path: str = os.path.join(
            pszDirectory,
            "0003_PJサマリ_step0006_累計_損益計算書_E∪F.tsv",
        )
        shutil.copy2(pszCumulativeCostStep0005Path, pszCumulativeCostStep0006Path)
        objCumulativeStep0006Rows = insert_per_hour_rows(
            read_tsv_rows(pszCumulativePlStep0005Path)
        )
        write_tsv_rows(pszCumulativePlStep0006Path, objCumulativeStep0006Rows)

    create_step0007_pl_cr(pszDirectory)

    objTargetColumns: List[str] = [
        "科目名",
        "純売上高",
        "売上原価",
        "売上総利益",
        "配賦販管費",
        "1Cカンパニー販管費",
        "2Cカンパニー販管費",
        "3Cカンパニー販管費",
        "4Cカンパニー販管費",
        "事業開発カンパニー販管費",
    ]
    objSingleStep0002Rows: List[List[str]] = filter_rows_by_columns(
        objSingleOutputRows,
        objTargetColumns,
    )
    objSingleStep0002Rows = combine_company_sg_admin_columns(objSingleStep0002Rows)
    objSingleStep0002Rows = move_column_before(
        objSingleStep0002Rows,
        "カンパニー販管費",
        "配賦販管費",
    )
    objCumulativeStep0002Rows: List[List[str]] = filter_rows_by_columns(
        objCumulativeOutputRows,
        objTargetColumns,
    )
    objCumulativeStep0002Rows = combine_company_sg_admin_columns(objCumulativeStep0002Rows)
    objCumulativeStep0002Rows = move_column_before(
        objCumulativeStep0002Rows,
        "カンパニー販管費",
        "配賦販管費",
    )
    pszSingleStep0002Path: str = os.path.join(
        pszDirectory,
        f"0001_PJサマリ_step0002_{iEndYear}年{pszEndMonth}月_単月_損益計算書.tsv",
    )
    pszCumulativeStep0002Path: str = os.path.join(
        pszDirectory,
        (
            "0001_PJサマリ_step0002_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月_累計_損益計算書.tsv"
        ),
    )
    write_tsv_rows(pszSingleStep0002Path, objSingleStep0002Rows)
    write_tsv_rows(pszCumulativeStep0002Path, objCumulativeStep0002Rows)

    pszOrgTablePath: str = os.path.join(pszDirectory, "管轄PJ表.tsv")
    objGroupMap = load_org_table_group_map(pszOrgTablePath)
    objSingleStep0003GroupRows = insert_accounting_group_column(
        objSingleStep0002Rows,
        objGroupMap,
    )
    objCumulativeStep0003GroupRows = insert_accounting_group_column(
        objCumulativeStep0002Rows,
        objGroupMap,
    )
    pszSingleStep0003Path: str = os.path.join(
        pszDirectory,
        f"0001_PJサマリ_step0003_{iEndYear}年{pszEndMonth}月_単月_損益計算書.tsv",
    )
    pszCumulativeStep0003Path: str = os.path.join(
        pszDirectory,
        (
            "0001_PJサマリ_step0003_"
            f"{objStart[0]}年{pszSummaryStartMonth}月-"
            f"{objEnd[0]}年{pszSummaryEndMonth}月_累計_損益計算書.tsv"
        ),
    )
    write_tsv_rows(pszSingleStep0003Path, objSingleStep0003GroupRows)
    write_tsv_rows(pszCumulativeStep0003Path, objCumulativeStep0003GroupRows)
    update_step0003_headquarters_group(pszSingleStep0003Path, pszOrgTablePath)
    update_step0003_headquarters_group(pszCumulativeStep0003Path, pszOrgTablePath)

    objCompanyMap = load_org_table_company_map(pszOrgTablePath)
    objSingleStep0004Rows: Optional[List[List[str]]] = None
    objCumulativeStep0004Rows: Optional[List[List[str]]] = None
    if os.path.isfile(pszSingleStep0003Path):
        objSingleStep0004Rows = insert_accounting_company_column(
            read_tsv_rows(pszSingleStep0003Path),
            objCompanyMap,
        )
        pszSingleStep0004Path: str = os.path.join(
            pszDirectory,
            f"0001_PJサマリ_step0004_{iEndYear}年{pszEndMonth}月_単月_損益計算書.tsv",
        )
        write_tsv_rows(pszSingleStep0004Path, objSingleStep0004Rows)
        update_step0005_headquarters_company(pszSingleStep0004Path, pszOrgTablePath)
        objSingleStep0004Rows = read_tsv_rows(pszSingleStep0004Path)
    if os.path.isfile(pszCumulativeStep0003Path):
        objCumulativeStep0004Rows = insert_accounting_company_column(
            read_tsv_rows(pszCumulativeStep0003Path),
            objCompanyMap,
        )
        pszCumulativeStep0004Path: str = os.path.join(
            pszDirectory,
            (
                "0001_PJサマリ_step0004_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月_累計_損益計算書.tsv"
            ),
        )
        write_tsv_rows(pszCumulativeStep0004Path, objCumulativeStep0004Rows)
        update_step0005_headquarters_company(pszCumulativeStep0004Path, pszOrgTablePath)
        objCumulativeStep0004Rows = read_tsv_rows(pszCumulativeStep0004Path)

    objSingleStep0005Rows: Optional[List[List[str]]] = None
    objCumulativeStep0005Rows: Optional[List[List[str]]] = None
    if objSingleStep0004Rows is not None:
        objSingleStep0005Rows = append_gross_margin_column(objSingleStep0004Rows)
        objSingleStep0005Rows = fill_headquarters_company_in_rows(
            objSingleStep0005Rows,
            pszOrgTablePath,
        )
        pszSingleStep0005Path: str = os.path.join(
            pszDirectory,
            f"0001_PJサマリ_step0005_{iEndYear}年{pszEndMonth}月_単月_損益計算書.tsv",
        )
        write_tsv_rows(pszSingleStep0005Path, objSingleStep0005Rows)

    if objCumulativeStep0004Rows is not None:
        objCumulativeStep0005Rows = append_gross_margin_column(objCumulativeStep0004Rows)
        objCumulativeStep0005Rows = fill_headquarters_company_in_rows(
            objCumulativeStep0005Rows,
            pszOrgTablePath,
        )
        pszCumulativeStep0005Path: str = os.path.join(
            pszDirectory,
            (
                "0001_PJサマリ_step0005_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月_累計_損益計算書.tsv"
            ),
        )
        write_tsv_rows(pszCumulativeStep0005Path, objCumulativeStep0005Rows)

    if (
        objStart != objEnd
        and objSingleStep0005Rows is not None
        and objCumulativeStep0005Rows is not None
    ):
        pszStep0006Path: str = os.path.join(
            pszDirectory,
            (
                "0001_PJサマリ_step0006_"
                f"{objStart[0]}年{pszSummaryStartMonth}月-"
                f"{objEnd[0]}年{pszSummaryEndMonth}月_単月・累計_損益計算書.tsv"
            ),
        )
        write_step0006_pj_summary(
            pszStep0006Path,
            objSingleStep0005Rows,
            objCumulativeStep0005Rows,
        )
        build_step0007_rows_from_step0006_path(pszStep0006Path)
        pszStep0007Path: str = pszStep0006Path.replace("step0006_", "step0007_", 1)
        build_step0008_rows_from_step0007_path(pszStep0007Path)
        pszStep0008Path: str = pszStep0007Path.replace("step0007_", "step0008_", 1)
        build_step0009_rows_from_step0008_path(pszStep0008Path)

    objGrossProfitColumns: List[str] = ["科目名", "売上総利益", "純売上高"]
    objGrossProfitSingleRows: List[List[str]] = filter_rows_by_columns(
        objSingleOutputRows,
        objGrossProfitColumns,
    )
    objGrossProfitCumulativeRows: List[List[str]] = filter_rows_by_columns(
        objCumulativeOutputRows,
        objGrossProfitColumns,
    )
    pszGrossProfitSinglePath: str = os.path.join(
        pszDirectory,
        "0002_PJサマリ_step0001_単月_粗利金額ランキング.tsv",
    )
    pszGrossProfitCumulativePath: str = os.path.join(
        pszDirectory,
        "0002_PJサマリ_step0001_累計_粗利金額ランキング.tsv",
    )
    write_tsv_rows(pszGrossProfitSinglePath, objGrossProfitSingleRows)
    write_tsv_rows(pszGrossProfitCumulativePath, objGrossProfitCumulativeRows)

    # 単月_粗利金額ランキング
    objGrossProfitSingleSortedRows: List[List[str]] = []
    objGrossProfitCumulativeSortedRows: List[List[str]] = []

    if objGrossProfitSingleRows:
        objSingleHeader: List[str] = objGrossProfitSingleRows[0]
        objSingleBody: List[List[str]] = objGrossProfitSingleRows[1:]
        objSingleBody.sort(
            key=lambda objRow: try_parse_float(objRow[1] if len(objRow) > 1 else "") or 0.0,
            reverse=True,
        )
        objGrossProfitSingleSortedRows = [objSingleHeader] + objSingleBody
        pszGrossProfitSingleSortedPath: str = os.path.join(
            pszDirectory,
            "0002_PJサマリ_step0002_単月_粗利金額ランキング.tsv",
        )
        write_tsv_rows(pszGrossProfitSingleSortedPath, objGrossProfitSingleSortedRows)

    # 累計_粗利金額ランキング
    if objGrossProfitCumulativeRows:
        objCumulativeHeader: List[str] = objGrossProfitCumulativeRows[0]
        objCumulativeBody: List[List[str]] = objGrossProfitCumulativeRows[1:]
        objCumulativeBody.sort(
            key=lambda objRow: try_parse_float(objRow[1] if len(objRow) > 1 else "") or 0.0,
            reverse=True,
        )
        objGrossProfitCumulativeSortedRows = [objCumulativeHeader] + objCumulativeBody
        pszGrossProfitCumulativeSortedPath: str = os.path.join(
            pszDirectory,
            "0002_PJサマリ_step0002_累計_粗利金額ランキング.tsv",
        )
        write_tsv_rows(
            pszGrossProfitCumulativeSortedPath,
            objGrossProfitCumulativeSortedRows,
        )

    if objGrossProfitSingleSortedRows and objGrossProfitCumulativeSortedRows:
        if len(objGrossProfitSingleSortedRows) != len(objGrossProfitCumulativeSortedRows):
            print("Error: gross profit ranking row count mismatch.")
            return

        objGrossProfitCombinedRows = [list(objRow) for objRow in objGrossProfitSingleSortedRows]
        for objRow in objGrossProfitCombinedRows:
            objRow.append("")

        for iRowIndex, objCumulativeRow in enumerate(objGrossProfitCumulativeSortedRows):
            if len(objGrossProfitCombinedRows[iRowIndex]) < 3:
                objGrossProfitCombinedRows[iRowIndex].extend(
                    [""] * (3 - len(objGrossProfitCombinedRows[iRowIndex]))
                )
            pszCumulativeProject: str = objCumulativeRow[0] if objCumulativeRow else ""
            pszCumulativeValue: str = objCumulativeRow[1] if len(objCumulativeRow) > 1 else ""
            objGrossProfitCombinedRows[iRowIndex].append(pszCumulativeProject)
            objGrossProfitCombinedRows[iRowIndex].append(pszCumulativeValue)

        pszGrossProfitCombinedPath: str = os.path.join(
            pszDirectory,
            "0002_PJサマリ_step0007_単月・累計_粗利金額ランキング.tsv",
        )
        write_tsv_rows(pszGrossProfitCombinedPath, objGrossProfitCombinedRows)

    if objGrossProfitSingleSortedRows:
        objGrossProfitSingleRankRows: List[List[str]] = []
        for iRowIndex, objRow in enumerate(objGrossProfitSingleSortedRows):
            if iRowIndex == 0:
                objGrossProfitSingleRankRows.append(
                    ["0", "プロジェクト名", "売上総利益", "純売上高", "利益率"]
                )
                continue
            pszGrossProfit: str = objRow[1] if len(objRow) > 1 else ""
            pszSales: str = objRow[2] if len(objRow) > 2 else ""
            fGrossProfit: float = parse_number(pszGrossProfit)
            fSales: float = parse_number(pszSales)
            if abs(fSales) < 0.0000001:
                if fGrossProfit > 0:
                    pszMargin = "'＋∞"
                elif fGrossProfit < 0:
                    pszMargin = "'－∞"
                else:
                    pszMargin = "0"
            else:
                pszMargin = format_number(fGrossProfit / fSales)
            objGrossProfitSingleRankRows.append(
                [
                    str(iRowIndex),
                    objRow[0] if objRow else "",
                    pszGrossProfit,
                    pszSales,
                    pszMargin,
                ]
            )
        pszGrossProfitSingleRankPath: str = os.path.join(
            pszDirectory,
            "0002_PJサマリ_step0007_単月_粗利金額ランキング.tsv",
        )
        write_tsv_rows(pszGrossProfitSingleRankPath, objGrossProfitSingleRankRows)

    if objGrossProfitCumulativeSortedRows:
        objGrossProfitCumulativeRankRows: List[List[str]] = []
        for iRowIndex, objRow in enumerate(objGrossProfitCumulativeSortedRows):
            if iRowIndex == 0:
                objGrossProfitCumulativeRankRows.append(
                    ["0", "プロジェクト名", "売上総利益", "純売上高", "利益率"]
                )
                continue
            pszGrossProfit = objRow[1] if len(objRow) > 1 else ""
            pszSales = objRow[2] if len(objRow) > 2 else ""
            fGrossProfit = parse_number(pszGrossProfit)
            fSales = parse_number(pszSales)
            if abs(fSales) < 0.0000001:
                if fGrossProfit > 0:
                    pszMargin = "'＋∞"
                elif fGrossProfit < 0:
                    pszMargin = "'－∞"
                else:
                    pszMargin = "0"
            else:
                pszMargin = format_number(fGrossProfit / fSales)
            objGrossProfitCumulativeRankRows.append(
                [
                    str(iRowIndex),
                    objRow[0] if objRow else "",
                    pszGrossProfit,
                    pszSales,
                    pszMargin,
                ]
            )
        pszGrossProfitCumulativeRankPath: str = os.path.join(
            pszDirectory,
            "0002_PJサマリ_step0007_累計_粗利金額ランキング.tsv",
        )
        write_tsv_rows(pszGrossProfitCumulativeRankPath, objGrossProfitCumulativeRankRows)

    if objGrossProfitSingleRankRows and objGrossProfitCumulativeRankRows:
        if len(objGrossProfitSingleRankRows) != len(objGrossProfitCumulativeRankRows):
            print("Error: gross profit ranking step0008 row count mismatch.")
            return

        objGrossProfitStep0004Rows: List[List[str]] = []
        for iRowIndex, objSingleRow in enumerate(objGrossProfitSingleRankRows):
            objCumulativeRow = objGrossProfitCumulativeRankRows[iRowIndex]
            objOutputRow: List[str] = list(objSingleRow)
            objOutputRow.append("")
            objOutputRow.extend(objCumulativeRow)
            objGrossProfitStep0004Rows.append(objOutputRow)

        pszGrossProfitStep0004Path: str = os.path.join(
            pszDirectory,
            "0002_PJサマリ_step0008_単月・累計_粗利金額ランキング.tsv",
        )
        write_tsv_rows(pszGrossProfitStep0004Path, objGrossProfitStep0004Rows)

        objGrossProfitStep0005Rows: List[List[str]] = [
            [
                "",
                "粗利金額ランキング",
                "粗利金額",
                "",
                "単月",
                "",
                "",
                "粗利金額ランキング",
                "粗利金額",
                "",
                "累計",
            ]
        ]
        objGrossProfitStep0005Rows.extend(objGrossProfitStep0004Rows)
        pszGrossProfitStep0005Path: str = os.path.join(
            pszDirectory,
            "0002_PJサマリ_step0009_単月・累計_粗利金額ランキング.tsv",
        )
        write_tsv_rows(pszGrossProfitStep0005Path, objGrossProfitStep0005Rows)

        if objGrossProfitStep0005Rows:
            objGrossProfitStep0006Rows: List[List[str]] = []
            objSalesIndices: List[int] = []
            objNumberColumnIndices: List[int] = []
            objHeaderRow: List[str] = objGrossProfitStep0005Rows[0]
            for iColumnIndex, pszColumnName in enumerate(objHeaderRow):
                if pszColumnName == "純売上高":
                    objSalesIndices.append(iColumnIndex)

            if objGrossProfitSingleRankRows:
                iSingleColumnCount = len(objGrossProfitSingleRankRows[0])
                objNumberColumnIndices = [0, iSingleColumnCount + 1]

            objSalesIndexSet = set(objSalesIndices)
            objNumberIndexSet = set(objNumberColumnIndices)
            for objRow in objGrossProfitStep0005Rows:
                objFilteredRow: List[str] = []
                for iColumnIndex, pszValue in enumerate(objRow):
                    if iColumnIndex in objSalesIndexSet:
                        continue
                    if iColumnIndex in objNumberIndexSet and pszValue == "0":
                        objFilteredRow.append("")
                    else:
                        objFilteredRow.append(pszValue)
                objGrossProfitStep0006Rows.append(objFilteredRow)

            pszGrossProfitStep0006Path: str = os.path.join(
                pszDirectory,
                "0002_PJサマリ_step0010_単月・累計_粗利金額ランキング.tsv",
            )
            write_tsv_rows(pszGrossProfitStep0006Path, objGrossProfitStep0006Rows)

            pszGrossProfitFinalPath: str = os.path.join(
                pszDirectory,
                "0002_PJサマリ_単月・累計_粗利金額ランキング.tsv",
            )
            write_tsv_rows(pszGrossProfitFinalPath, objGrossProfitStep0006Rows)


def create_cumulative_report(
    pszDirectory: str,
    pszPrefix: str,
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
    pszInputPrefix: Optional[str] = None,
) -> None:
    objStart, objEnd = objRange
    objMonths = build_month_sequence(objStart, objEnd)
    if not objMonths:
        return

    if pszInputPrefix is None:
        pszInputPrefix = pszPrefix

    objTotalRows: Optional[List[List[str]]] = None
    for objMonth in objMonths:
        objRows: Optional[List[List[str]]] = read_report_rows(
            pszDirectory,
            pszInputPrefix,
            objMonth,
        )
        if objRows is None:
            return
        if objTotalRows is None:
            objTotalRows = objRows
        else:
            if can_use_simple_position_sum(objTotalRows, objRows):
                objTotalRows = sum_tsv_rows_by_position(objTotalRows, objRows)
            else:
                objTotalRows = sum_tsv_rows(objTotalRows, objRows)

    if objTotalRows is None:
        return
    pszOutputPath: str = build_cumulative_file_path(pszDirectory, pszPrefix, objStart, objEnd)
    write_tsv_rows(pszOutputPath, objTotalRows)
    print(f"Output: {pszOutputPath}")
    pszVerticalOutputPath: str = pszOutputPath.replace(".tsv", "_vertical.tsv")
    objVerticalRows: List[List[str]] = transpose_rows(objTotalRows)
    write_tsv_rows(pszVerticalOutputPath, objVerticalRows)
    print(f"Output: {pszVerticalOutputPath}")


def create_cumulative_report_without_company_columns(
    pszDirectory: str,
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
) -> None:
    pszSourcePath: str = build_cumulative_file_path(
        pszDirectory,
        "損益計算書_販管費配賦",
        objRange[0],
        objRange[1],
    )
    if not os.path.isfile(pszSourcePath):
        return

    objRows: List[List[str]] = read_tsv_rows(pszSourcePath)
    if not objRows:
        return

    objRemovalTargets = {
        "1Cカンパニー販管費",
        "2Cカンパニー販管費",
        "3Cカンパニー販管費",
        "4Cカンパニー販管費",
        "事業開発カンパニー販管費",
    }
    objHeader: List[str] = objRows[0]
    objKeepIndices: List[int] = [
        iColumnIndex
        for iColumnIndex, pszColumnName in enumerate(objHeader)
        if pszColumnName not in objRemovalTargets
    ]

    objOutputRows: List[List[str]] = []
    for objRow in objRows:
        objOutputRows.append([
            objRow[iColumnIndex] if iColumnIndex < len(objRow) else ""
            for iColumnIndex in objKeepIndices
        ])

    pszOutputPath: str = build_cumulative_file_path(
        pszDirectory,
        "損益計算書_販管費配賦_カンパニー列なし",
        objRange[0],
        objRange[1],
    )
    write_tsv_rows(pszOutputPath, objOutputRows)
    print(f"Output: {pszOutputPath}")
    pszVerticalOutputPath: str = pszOutputPath.replace(".tsv", "_vertical.tsv")
    write_tsv_rows(pszVerticalOutputPath, transpose_rows(objOutputRows))
    print(f"Output: {pszVerticalOutputPath}")


def build_pj_summary_range(
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
) -> Tuple[Tuple[int, int], Tuple[int, int]]:
    _, objEnd = objRange
    iEndYear, iEndMonth = objEnd
    if iEndMonth >= 4:
        iStartYear: int = iEndYear
    else:
        iStartYear = iEndYear - 1
    return (iStartYear, 4), (iEndYear, iEndMonth)


def create_cumulative_reports(pszPlPath: str) -> None:
    pszInputDirectory: str = os.path.dirname(pszPlPath)
    pszDirectory: str = get_script_base_directory()
    pszRangePath: Optional[str] = find_selected_range_path(pszInputDirectory)
    if pszRangePath is None:
        return

    objRange = parse_selected_range(pszRangePath)
    if objRange is None:
        return
    ensure_selected_range_file(pszDirectory, objRange)

    objStart, objEnd = objRange
    objAllRanges, objCurrentRanges = build_cp_period_ranges_from_previous_period_range_file(pszDirectory)
    if not objAllRanges:
        objFiscalARanges = split_by_fiscal_boundary(objStart, objEnd, 3)
        objFiscalBRanges = split_by_fiscal_boundary(objStart, objEnd, 8)

        def append_unique_range(
            objTargetRange: Tuple[Tuple[int, int], Tuple[int, int]]
        ) -> None:
            if objTargetRange not in objAllRanges:
                objAllRanges.append(objTargetRange)

        if objFiscalARanges:
            if len(objFiscalARanges) >= 2:
                append_unique_range(objFiscalARanges[-2])
            append_unique_range(objFiscalARanges[-1])
        if objFiscalBRanges:
            if len(objFiscalBRanges) >= 2:
                append_unique_range(objFiscalBRanges[-2])
            append_unique_range(objFiscalBRanges[-1])
        objCurrentRanges = build_cp_period_ranges_from_selected_range(objRange)

    objTotalsExcelRanges = build_current_period_ranges_for_pj_summary_totals(objRange)

    for objRangeItem in objAllRanges:
        create_cumulative_report(
            pszDirectory,
            "損益計算書_販管費配賦",
            objRangeItem,
            pszInputPrefix="損益計算書_販管費配賦",
        )
        create_cumulative_report(pszDirectory, "製造原価報告書", objRangeItem)
        create_pj_summary(
            pszPlPath,
            objRangeItem,
            create_step0007=objRangeItem in objCurrentRanges,
            bWriteTotalsExcel=objRangeItem in objTotalsExcelRanges,
        )
    objMonths = build_month_sequence(objStart, objEnd)
    for objMonth in objMonths:
        create_pj_summary(pszPlPath, (objMonth, objMonth))
    pszCompanyManagementPath = try_create_cp_step0009_vertical(pszDirectory)
    pszGroupManagementPath = try_create_cp_group_step0009_vertical(pszDirectory)
    copy_cp_management_excels(pszCompanyManagementPath, pszGroupManagementPath)
    create_pj_summary_gross_profit_ranking_excel(pszDirectory)
    create_pj_summary_sales_cost_sg_admin_margin_excel(pszDirectory)
    move_monthly_income_statement_tsv_files_into_temp_subfolder(pszDirectory)
    move_pl_tsv_files_into_income_statement_temp_subfolder(pszDirectory)
    move_cost_report_tsv_files_into_temp_subfolder(pszDirectory)
    move_step0007_split_files_into_0003_pj_summary_temp_subfolder(pszDirectory)
    cleanup_cp_step_intermediate_tsv_files(pszDirectory)
    move_cp_step_tsv_files_to_temp_subfolders(pszDirectory)
    move_pj_summary_tsv_files_to_temp_subfolders(pszDirectory)
    move_cp_step_folders_to_temp(pszDirectory)
    remove_bycompany_managementcontrol_step0005_directory()


def cleanup_cp_step_intermediate_tsv_files(pszDirectory: str) -> None:
    objMonthPattern = r"\d{4}年\d{2}月"
    objRangePattern = rf"{objMonthPattern}-{objMonthPattern}"
    objPatterns = [
        re.compile(
            rf"^0001_CP別_step0006_単月_損益計算書_{objMonthPattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0006_累計_損益計算書_{objRangePattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0007_単月_損益計算書_{objMonthPattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0007_累計_損益計算書_{objRangePattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0008_単月_損益計算書_{objMonthPattern}_計上カンパニー_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0008_累計_損益計算書_{objRangePattern}_計上カンパニー_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0009_単月_損益計算書_{objMonthPattern}_計上カンパニー_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0009_累計_損益計算書_{objRangePattern}_計上カンパニー_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0001_単月_損益計算書_{objMonthPattern}\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0001_累計_損益計算書_{objRangePattern}\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0006_単月_損益計算書_{objMonthPattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0006_累計_損益計算書_{objRangePattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0007_単月_損益計算書_{objMonthPattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0007_累計_損益計算書_{objRangePattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0008_単月_損益計算書_{objMonthPattern}_計上グループ_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0008_累計_損益計算書_{objRangePattern}_計上グループ_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0009_単月_損益計算書_{objMonthPattern}_計上グループ_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0009_累計_損益計算書_{objRangePattern}_計上グループ_vertical\.tsv$"
        ),
    ]
    for pszName in os.listdir(pszDirectory):
        if not any(objPattern.match(pszName) for objPattern in objPatterns):
            continue
        pszPath = os.path.join(pszDirectory, pszName)
        if not os.path.isfile(pszPath):
            continue
        os.remove(pszPath)


def cleanup_cp_step_intermediate_tsv_files(pszDirectory: str) -> None:
    objMonthPattern = r"\d{4}年\d{2}月"
    objRangePattern = rf"{objMonthPattern}-{objMonthPattern}"
    objPatterns = [
        re.compile(
            rf"^0001_CP別_step0006_単月_損益計算書_{objMonthPattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0006_累計_損益計算書_{objRangePattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0007_単月_損益計算書_{objMonthPattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0007_累計_損益計算書_{objRangePattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0008_単月_損益計算書_{objMonthPattern}_計上カンパニー_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0008_累計_損益計算書_{objRangePattern}_計上カンパニー_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0009_単月_損益計算書_{objMonthPattern}_計上カンパニー_vertical\.tsv$"
        ),
        re.compile(
            rf"^0001_CP別_step0009_累計_損益計算書_{objRangePattern}_計上カンパニー_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0001_単月_損益計算書_{objMonthPattern}\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0001_累計_損益計算書_{objRangePattern}\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0006_単月_損益計算書_{objMonthPattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0006_累計_損益計算書_{objRangePattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0007_単月_損益計算書_{objMonthPattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0007_累計_損益計算書_{objRangePattern}_.+_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0008_単月_損益計算書_{objMonthPattern}_計上グループ_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0008_累計_損益計算書_{objRangePattern}_計上グループ_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0009_単月_損益計算書_{objMonthPattern}_計上グループ_vertical\.tsv$"
        ),
        re.compile(
            rf"^0002_CP別_step0009_累計_損益計算書_{objRangePattern}_計上グループ_vertical\.tsv$"
        ),
    ]
    for pszName in os.listdir(pszDirectory):
        if not any(objPattern.match(pszName) for objPattern in objPatterns):
            continue
        pszPath = os.path.join(pszDirectory, pszName)
        if not os.path.isfile(pszPath):
            continue
        os.remove(pszPath)


def copy_cp_step0005_vertical_files(pszDirectory: str, objPaths: List[Optional[str]]) -> None:
    pszTargetDirectory: str = os.path.join(get_script_base_directory(), "ByCompany_ManagementControl_step0005")
    os.makedirs(pszTargetDirectory, exist_ok=True)
    for pszPath in objPaths:
        if not pszPath:
            continue
        if not os.path.isfile(pszPath):
            continue
        pszFileName: str = os.path.basename(pszPath)
        pszTargetPath: str = os.path.join(pszTargetDirectory, pszFileName)
        shutil.copy2(pszPath, pszTargetPath)


def move_cp_step0001_to_step0004_vertical_files(
    pszDirectory: str,
    objStart: Tuple[int, int],
    objEnd: Tuple[int, int],
) -> None:
    pszTargetDirectory: str = os.path.join(pszDirectory, "temp")
    os.makedirs(pszTargetDirectory, exist_ok=True)
    iStartYear, iStartMonth = objStart
    iEndYear, iEndMonth = objEnd
    pszStartMonth: str = f"{iStartMonth:02d}"
    pszEndMonth: str = f"{iEndMonth:02d}"
    objTargets: List[str] = []
    if objStart == objEnd:
        for pszStep in ("step0001", "step0002", "step0003", "step0004"):
            objTargets.append(
                os.path.join(
                    pszDirectory,
                    f"0001_CP別_{pszStep}_単月_損益計算書_{iEndYear}年{pszEndMonth}月_vertical.tsv",
                )
            )
        objTargets.append(
            os.path.join(
                pszDirectory,
                f"0002_CP別_step0004_単月_損益計算書_{iEndYear}年{pszEndMonth}月_vertical.tsv",
            )
        )
    for pszStep in ("step0001", "step0002", "step0003", "step0004"):
        objTargets.append(
            os.path.join(
                pszDirectory,
                (
                    f"0001_CP別_{pszStep}_累計_損益計算書_"
                    f"{iStartYear}年{pszStartMonth}月-{iEndYear}年{pszEndMonth}月_vertical.tsv"
                ),
            )
        )
    objTargets.append(
        os.path.join(
            pszDirectory,
            (
                "0002_CP別_step0004_累計_損益計算書_"
                f"{iStartYear}年{pszStartMonth}月-{iEndYear}年{pszEndMonth}月_vertical.tsv"
            ),
        )
    )
    for pszPath in objTargets:
        if not os.path.isfile(pszPath):
            continue
        pszTargetPath: str = os.path.join(pszTargetDirectory, os.path.basename(pszPath))
        shutil.move(pszPath, pszTargetPath)


def copy_company_step0006_files(
    pszDirectory: str,
    objPaths: List[Optional[str]],
    pszTargetFolder: str,
    create_step0007: bool = True,
    pszOrgMode: Optional[str] = None,
) -> None:
    pszTargetDirectory: str = os.path.join(get_script_base_directory(), pszTargetFolder)
    os.makedirs(pszTargetDirectory, exist_ok=True)
    for pszPath in objPaths:
        if not pszPath or not os.path.isfile(pszPath):
            continue
        for pszOutputPath in build_company_step0006_files(
            pszPath,
            create_step0007=create_step0007,
            pszOrgMode=pszOrgMode,
        ):
            pszTargetPath: str = os.path.join(pszTargetDirectory, os.path.basename(pszOutputPath))
            shutil.copy2(pszOutputPath, pszTargetPath)


def build_company_step0006_files(
    pszStep0005Path: str,
    create_step0007: bool = True,
    pszOrgMode: Optional[str] = None,
) -> List[str]:
    objRows = read_tsv_rows(pszStep0005Path)
    if not objRows:
        return []
    objHeader = objRows[0]
    if len(objHeader) < 2:
        return []
    objCompanyIndices: List[Tuple[int, str]] = []
    for iColumnIndex, pszName in enumerate(objHeader[1:], start=1):
        if pszName == "":
            continue
        objCompanyIndices.append((iColumnIndex, pszName))

    objOutputPaths: List[str] = []
    for iColumnIndex, pszCompany in objCompanyIndices:
        objOutputRows: List[List[str]] = []
        for iRowIndex, objRow in enumerate(objRows):
            pszLabel: str = objRow[0] if objRow else ""
            if iRowIndex == 0:
                objOutputRows.append([pszLabel, pszCompany])
                continue
            pszValue: str = objRow[iColumnIndex] if iColumnIndex < len(objRow) else ""
            objOutputRows.append([pszLabel, pszValue])
        pszOutputPath: str = pszStep0005Path.replace("_step0005_", "_step0006_").replace(
            "_vertical.tsv",
            f"_{pszCompany}_vertical.tsv",
        )
        write_tsv_rows(pszOutputPath, objOutputRows)
        objOutputPaths.append(pszOutputPath)
        if create_step0007 and os.path.basename(pszOutputPath).startswith("0001_CP別_step0006_"):
            create_cp_step0007_file_0001(
                pszOutputPath,
                pszOrgMode=pszOrgMode,
            )
        if create_step0007 and os.path.basename(pszOutputPath).startswith("0002_CP別_step0006_"):
            create_cp_step0007_file_0002(pszOutputPath)
    return objOutputPaths


def copy_group_step0006_files(
    pszDirectory: str,
    objPaths: List[Optional[str]],
    create_step0007: bool = True,
) -> None:
    copy_company_step0006_files(
        pszDirectory,
        objPaths,
        "0002_CP別_step0006",
        create_step0007=create_step0007,
    )


def _apply_gross_profit_ranking_borders(
    objSheet,
    iLastRow: int,
) -> None:
    if iLastRow <= 0:
        return

    objThickSide = Side(style="medium", color="000000")
    objSolidSide = Side(style="thin", color="000000")
    objDottedSide = Side(style="dotted")

    def set_border(
        iRowIndex: int,
        iColumnIndex: int,
        objLeft: Optional[Side] = None,
        objRight: Optional[Side] = None,
        objTop: Optional[Side] = None,
        objBottom: Optional[Side] = None,
    ) -> None:
        objCell = objSheet.cell(row=iRowIndex, column=iColumnIndex)
        objBorder = copy(objCell.border)
        if objLeft is not None:
            objBorder.left = objLeft
        if objRight is not None:
            objBorder.right = objRight
        if objTop is not None:
            objBorder.top = objTop
        if objBottom is not None:
            objBorder.bottom = objBottom
        objCell.border = Border(
            left=objBorder.left,
            right=objBorder.right,
            top=objBorder.top,
            bottom=objBorder.bottom,
            diagonal=objBorder.diagonal,
            diagonal_direction=objBorder.diagonal_direction,
            outline=objBorder.outline,
            vertical=objBorder.vertical,
            horizontal=objBorder.horizontal,
        )

    for iRowIndex in range(1, iLastRow + 1):
        for iColumnIndex in range(1, 12):
            if iColumnIndex == 6:
                objSheet.cell(row=iRowIndex, column=iColumnIndex).border = Border()
                continue

            objLeft: Optional[Side] = None
            objRight: Optional[Side] = None
            objTop: Optional[Side] = None
            objBottom: Optional[Side] = None

            if iRowIndex == 1:
                objTop = objThickSide
            if iRowIndex == 2:
                objBottom = objThickSide
            elif 3 <= iRowIndex < iLastRow:
                objBottom = objDottedSide
            if iRowIndex == iLastRow:
                objBottom = objThickSide

            if iColumnIndex in (1, 7):
                objLeft = objThickSide
                objRight = objSolidSide
            elif iColumnIndex in (5, 11):
                objRight = objThickSide
            else:
                objRight = objSolidSide

            set_border(
                iRowIndex,
                iColumnIndex,
                objLeft=objLeft,
                objRight=objRight,
                objTop=objTop,
                objBottom=objBottom,
            )


def _clear_gross_profit_ranking_borders_below_last_row(
    objSheet,
    iStartRow: int,
) -> None:
    if iStartRow <= 0:
        return

    iSheetMaxRow: int = objSheet.max_row
    if iStartRow > iSheetMaxRow:
        return

    for iRowIndex in range(iStartRow, iSheetMaxRow + 1):
        for iColumnIndex in range(1, 12):
            objSheet.cell(row=iRowIndex, column=iColumnIndex).border = Border()


def create_pj_summary_gross_profit_ranking_excel(pszDirectory: str) -> Optional[str]:
    pszInputPath: str = os.path.join(
        pszDirectory,
        "0002_PJサマリ_step0010_単月・累計_粗利金額ランキング.tsv",
    )
    if not os.path.isfile(pszInputPath):
        return None
    pszTemplatePath: str = os.path.join(
        os.path.dirname(__file__),
        "TEMPLATE_PJサマリ_単月・累計_粗利金額ランキング.xlsx",
    )
    if not os.path.isfile(pszTemplatePath):
        return None
    objWorkbook = load_workbook(pszTemplatePath)
    objSheet = objWorkbook.worksheets[0]
    objSheet.title = "PJ別粗利金額ランキング"
    objRows = read_tsv_rows(pszInputPath)
    iFormatRowIndex: int = 2 if objSheet.max_row >= 2 else 1
    for iRowIndex, objRow in enumerate(objRows, start=1):
        for iColumnIndex, pszValue in enumerate(objRow, start=1):
            objCellValue = parse_tsv_value_for_excel(pszValue)
            objCell = objSheet.cell(
                row=iRowIndex,
                column=iColumnIndex,
                value=objCellValue,
            )
            if iRowIndex >= 2:
                objFormatCell = objSheet.cell(row=iFormatRowIndex, column=iColumnIndex)
                if objFormatCell.number_format:
                    objCell.number_format = objFormatCell.number_format

    _apply_gross_profit_ranking_borders(objSheet, len(objRows))
    _clear_gross_profit_ranking_borders_below_last_row(objSheet, len(objRows) + 1)

    pszTargetDirectory: str = os.path.join(pszDirectory, "PJサマリ")
    os.makedirs(pszTargetDirectory, exist_ok=True)
    pszOutputPath: str = os.path.join(
        pszTargetDirectory,
        "PJサマリ_単月・累計_粗利金額ランキング.xlsx",
    )
    objWorkbook.save(pszOutputPath)
    if EXECUTION_ROOT_DIRECTORY:
        pszRankingDirectory: str = os.path.join(EXECUTION_ROOT_DIRECTORY, "カンパニー利益率順位")
        os.makedirs(pszRankingDirectory, exist_ok=True)
        pszCopyPath: str = os.path.join(
            pszRankingDirectory,
            "PJサマリ_単月・累計_粗利金額ランキング.xlsx",
        )
        if os.path.abspath(pszCopyPath) != os.path.abspath(pszOutputPath):
            shutil.copy2(pszOutputPath, pszCopyPath)
    return pszOutputPath


def _apply_pj_summary_sales_cost_sg_admin_margin_borders(
    objSheet,
    iLastRow: int,
    iLastColumn: int,
) -> None:
    if iLastRow <= 0 or iLastColumn <= 0:
        return

    objThickSide = Side(style="medium", color="000000")
    objSolidSide = Side(style="thin", color="000000")
    objDottedSide = Side(style="dotted")

    def set_border(
        iRowIndex: int,
        iColumnIndex: int,
        objLeft: Optional[Side] = None,
        objRight: Optional[Side] = None,
        objTop: Optional[Side] = None,
        objBottom: Optional[Side] = None,
    ) -> None:
        objCell = objSheet.cell(row=iRowIndex, column=iColumnIndex)
        objBorder = copy(objCell.border)
        if objLeft is not None:
            objBorder.left = objLeft
        if objRight is not None:
            objBorder.right = objRight
        if objTop is not None:
            objBorder.top = objTop
        if objBottom is not None:
            objBorder.bottom = objBottom
        objCell.border = Border(
            left=objBorder.left,
            right=objBorder.right,
            top=objBorder.top,
            bottom=objBorder.bottom,
            diagonal=objBorder.diagonal,
            diagonal_direction=objBorder.diagonal_direction,
            outline=objBorder.outline,
            vertical=objBorder.vertical,
            horizontal=objBorder.horizontal,
        )

    for iRowIndex in range(1, iLastRow + 1):
        for iColumnIndex in range(1, iLastColumn + 1):
            objLeft: Optional[Side] = None
            objRight: Optional[Side] = None
            objTop: Optional[Side] = None
            objBottom: Optional[Side] = None

            if iRowIndex == 1:
                objTop = objThickSide
            if iRowIndex == 2:
                objBottom = objThickSide
            elif 3 <= iRowIndex < iLastRow:
                objBottom = objDottedSide
            if iRowIndex == iLastRow:
                objBottom = objThickSide

            if iColumnIndex == 1:
                objLeft = objThickSide
                objRight = objSolidSide
            elif iColumnIndex == 2:
                objRight = objSolidSide
            elif iColumnIndex == 3:
                objRight = objThickSide
            elif iColumnIndex >= 4:
                if (iColumnIndex - 4) % 2 == 0:
                    objLeft = objThickSide
                    objRight = objSolidSide
                else:
                    objLeft = objSolidSide
                    objRight = objThickSide

            if iColumnIndex == iLastColumn:
                objRight = objThickSide

            set_border(
                iRowIndex,
                iColumnIndex,
                objLeft=objLeft,
                objRight=objRight,
                objTop=objTop,
                objBottom=objBottom,
            )


def _clear_sheet_borders_below_last_row(
    objSheet,
    iStartRow: int,
    iLastColumn: int,
) -> None:
    if iStartRow <= 0 or iLastColumn <= 0:
        return

    iSheetMaxRow: int = objSheet.max_row
    if iStartRow > iSheetMaxRow:
        return

    for iRowIndex in range(iStartRow, iSheetMaxRow + 1):
        for iColumnIndex in range(1, iLastColumn + 1):
            objCell = objSheet.cell(row=iRowIndex, column=iColumnIndex)
            objCell.border = Border()


def _clear_sheet_values(objSheet) -> None:
    iMaxRow: int = objSheet.max_row
    iMaxColumn: int = objSheet.max_column
    if iMaxRow <= 0 or iMaxColumn <= 0:
        return
    for objRow in objSheet.iter_rows(
        min_row=1,
        max_row=iMaxRow,
        min_col=1,
        max_col=iMaxColumn,
    ):
        for objCell in objRow:
            objCell.value = None


def create_pj_summary_sales_cost_sg_admin_margin_excel(pszDirectory: str) -> Optional[str]:
    objCandidates: List[str] = []
    objPattern = re.compile(r"^0001_PJサマリ_step0009_.*_単月・累計_損益計算書\.tsv$")
    objSheetNamePattern = re.compile(
        r"^0001_PJサマリ_step0009_(.+)_単月・累計_損益計算書\.tsv$"
    )
    for pszName in os.listdir(pszDirectory):
        if objPattern.match(pszName):
            objCandidates.append(pszName)
    if not objCandidates:
        return None
    objCandidates.sort()
    pszCountPath: str = os.path.join(
        pszDirectory,
        "0001_PJサマリ_step0009_count.txt",
    )
    with open(pszCountPath, "w", encoding="utf-8") as objCountFile:
        objCountFile.write(f"count={len(objCandidates)}\n")
        for pszCandidate in objCandidates:
            objCountFile.write(pszCandidate + "\n")
    pszTemplatePath: str = os.path.join(
        os.path.dirname(__file__),
        "TEMPLATE_PJサマリ_PJ別_売上・売上原価・販管費・利益率.xlsx",
    )
    if not os.path.isfile(pszTemplatePath):
        return None
    objWorkbook = load_workbook(pszTemplatePath)
    objTemplateSheet = objWorkbook.worksheets[0]
    for objSheetToRemove in objWorkbook.worksheets[1:]:
        objWorkbook.remove(objSheetToRemove)
    for pszInputName in objCandidates:
        objSheet = objWorkbook.copy_worksheet(objTemplateSheet)
        objSheetNameMatch = objSheetNamePattern.match(pszInputName)
        if objSheetNameMatch:
            objSheet.title = f"Div実績_{objSheetNameMatch.group(1)}"
        _clear_sheet_values(objSheet)
        objRows = read_tsv_rows(os.path.join(pszDirectory, pszInputName))
        iFormatRowIndex: int = 2 if objSheet.max_row >= 2 else 1
        iLastColumn: int = max((len(objRow) for objRow in objRows), default=0)
        for iRowIndex, objRow in enumerate(objRows, start=1):
            for iColumnIndex, pszValue in enumerate(objRow, start=1):
                objCellValue = parse_tsv_value_for_excel(pszValue)
                objCell = objSheet.cell(
                    row=iRowIndex,
                    column=iColumnIndex,
                    value=objCellValue,
                )
                if iRowIndex >= 2:
                    objFormatCell = objSheet.cell(row=iFormatRowIndex, column=iColumnIndex)
                    if objFormatCell.number_format:
                        objCell.number_format = objFormatCell.number_format
        _apply_pj_summary_sales_cost_sg_admin_margin_borders(
            objSheet,
            len(objRows),
            iLastColumn,
        )
        _clear_sheet_borders_below_last_row(
            objSheet,
            len(objRows) + 1,
            max(iLastColumn, objSheet.max_column),
        )
    if objTemplateSheet in objWorkbook.worksheets:
        objWorkbook.remove(objTemplateSheet)
    pszTargetDirectory: str = os.path.join(pszDirectory, "PJサマリ")
    os.makedirs(pszTargetDirectory, exist_ok=True)
    pszOutputPath: str = os.path.join(
        pszTargetDirectory,
        "PJサマリ_PJ別_売上・売上原価・販管費・利益率.xlsx",
    )
    objWorkbook.save(pszOutputPath)
    if EXECUTION_ROOT_DIRECTORY:
        pszCompanyResultsDirectory = os.path.join(
            EXECUTION_ROOT_DIRECTORY,
            "カンパニー実績",
        )
        os.makedirs(pszCompanyResultsDirectory, exist_ok=True)
        shutil.copy2(
            pszOutputPath,
            os.path.join(pszCompanyResultsDirectory, os.path.basename(pszOutputPath)),
        )
    return pszOutputPath


def create_pj_summary_pl_cr_manhour_excel(
    pszDirectory: str,
    pszProjectName: str,
    pszInputPath: str,
) -> Optional[str]:
    def parse_h_mm_ss_to_excel_serial(pszTimeText: str) -> Optional[float]:
        objMatch = re.fullmatch(r"(\d+):(\d{2}):(\d{2})", (pszTimeText or "").strip())
        if objMatch is None:
            return None
        iHours: int = int(objMatch.group(1))
        iMinutes: int = int(objMatch.group(2))
        iSeconds: int = int(objMatch.group(3))
        return (iHours * 3600 + iMinutes * 60 + iSeconds) / 86400.0

    objSheetNamePattern = re.compile(r"^(P\d{5}|[A-OQ-Z]\d{3})")
    if not os.path.isfile(pszInputPath):
        return None
    pszTemplatePath: str = os.path.join(
        os.path.dirname(__file__),
        "TEMPLATE_PJサマリ_単月・累計_損益計算書・製造原価報告書・工数.xlsx",
    )
    if not os.path.isfile(pszTemplatePath):
        return None
    objWorkbook = load_workbook(pszTemplatePath)
    objSheet = objWorkbook.worksheets[0]
    objSheetNameMatch = objSheetNamePattern.match(pszProjectName)
    if objSheetNameMatch:
        objSheet.title = objSheetNameMatch.group(1)
    objRows = read_tsv_rows(pszInputPath)
    objValueColumnIndices = {2, 6}
    objCurrencyTargetRows = {"工数1時間当たり純売上高", "工数1時間当たり営業利益"}
    for iRowIndex, objRow in enumerate(objRows, start=1):
        pszRowLabel: str = objRow[0] if len(objRow) >= 1 else ""
        for iColumnIndex, pszValue in enumerate(objRow, start=1):
            objCellValue = parse_tsv_value_for_excel(pszValue)
            objCell = objSheet.cell(
                row=iRowIndex,
                column=iColumnIndex,
                value=objCellValue,
            )
            if pszRowLabel == "工数行(h:mm:ss)" and iColumnIndex in objValueColumnIndices:
                objExcelTimeSerial = parse_h_mm_ss_to_excel_serial(pszValue)
                if objExcelTimeSerial is not None:
                    objCell.value = objExcelTimeSerial
                    objCell.number_format = "[h]:mm:ss"
            elif pszRowLabel in objCurrencyTargetRows and iColumnIndex in objValueColumnIndices:
                objCell.number_format = "#,##0;[Red]-#,##0"
    pszTargetDirectory: str = os.path.join(
        pszDirectory,
        "PJサマリ",
        "PJ別_損益計算書・製造原価報告書・工数",
    )
    os.makedirs(pszTargetDirectory, exist_ok=True)
    pszOutputPath: str = os.path.join(
        pszTargetDirectory,
        f"PJサマリ_単・累計_{pszProjectName}.xlsx",
    )
    objWorkbook.save(pszOutputPath)
    if EXECUTION_ROOT_DIRECTORY:
        pszProjectProfitDirectory = os.path.join(
            EXECUTION_ROOT_DIRECTORY,
            "プロジェクト損益",
        )
        os.makedirs(pszProjectProfitDirectory, exist_ok=True)
        shutil.copy2(
            pszOutputPath,
            os.path.join(pszProjectProfitDirectory, os.path.basename(pszOutputPath)),
        )
    return pszOutputPath


def create_pj_summary_pl_cr_manhour_all_project_excel(
    pszDirectory: str,
    objProjectInputs: List[Tuple[str, str]],
) -> Optional[str]:
    def parse_h_mm_ss_to_excel_serial(pszTimeText: str) -> Optional[float]:
        objMatch = re.fullmatch(r"(\d+):(\d{2}):(\d{2})", (pszTimeText or "").strip())
        if objMatch is None:
            return None
        iHours: int = int(objMatch.group(1))
        iMinutes: int = int(objMatch.group(2))
        iSeconds: int = int(objMatch.group(3))
        return (iHours * 3600 + iMinutes * 60 + iSeconds) / 86400.0

    objSheetNamePattern = re.compile(r"^(P\d{5}|[A-OQ-Z]\d{3})")

    objValidInputs: List[Tuple[str, str]] = [
        (pszProjectName, pszInputPath)
        for pszProjectName, pszInputPath in objProjectInputs
        if os.path.isfile(pszInputPath)
    ]
    if not objValidInputs:
        return None

    pszTemplatePath: str = os.path.join(
        os.path.dirname(__file__),
        "TEMPLATE_PJサマリ_単月・累計_損益計算書・製造原価報告書・工数.xlsx",
    )
    if not os.path.isfile(pszTemplatePath):
        return None

    objWorkbook = load_workbook(pszTemplatePath)
    objTemplateSheet = objWorkbook.worksheets[0]

    for iIndex, objProjectInput in enumerate(objValidInputs):
        pszProjectName, pszInputPath = objProjectInput
        if iIndex == 0:
            objSheet = objTemplateSheet
        else:
            objSheet = objWorkbook.copy_worksheet(objTemplateSheet)
        objSheetNameMatch = objSheetNamePattern.match(pszProjectName)
        if objSheetNameMatch:
            objSheet.title = objSheetNameMatch.group(1)
        else:
            objSheet.title = pszProjectName
        objRows = read_tsv_rows(pszInputPath)
        objValueColumnIndices = {2, 6}
        objCurrencyTargetRows = {"工数1時間当たり純売上高", "工数1時間当たり営業利益"}
        for iRowIndex, objRow in enumerate(objRows, start=1):
            pszRowLabel: str = objRow[0] if len(objRow) >= 1 else ""
            for iColumnIndex, pszValue in enumerate(objRow, start=1):
                objCellValue = parse_tsv_value_for_excel(pszValue)
                objCell = objSheet.cell(
                    row=iRowIndex,
                    column=iColumnIndex,
                    value=objCellValue,
                )
                if pszRowLabel == "工数行(h:mm:ss)" and iColumnIndex in objValueColumnIndices:
                    objExcelTimeSerial = parse_h_mm_ss_to_excel_serial(pszValue)
                    if objExcelTimeSerial is not None:
                        objCell.value = objExcelTimeSerial
                        objCell.number_format = "[h]:mm:ss"
                elif pszRowLabel in objCurrencyTargetRows and iColumnIndex in objValueColumnIndices:
                    objCell.number_format = "#,##0;[Red]-#,##0"

    pszTargetDirectory: str = os.path.join(
        pszDirectory,
        "PJサマリ",
        "PJ別_損益計算書・製造原価報告書・工数",
    )
    os.makedirs(pszTargetDirectory, exist_ok=True)
    pszOutputPath: str = os.path.join(
        pszTargetDirectory,
        "PJサマリ_単・累計_AllProject.xlsx",
    )
    objWorkbook.save(pszOutputPath)
    if EXECUTION_ROOT_DIRECTORY:
        pszProjectProfitDirectory = os.path.join(
            EXECUTION_ROOT_DIRECTORY,
            "プロジェクト損益",
        )
        os.makedirs(pszProjectProfitDirectory, exist_ok=True)
        shutil.copy2(
            pszOutputPath,
            os.path.join(pszProjectProfitDirectory, os.path.basename(pszOutputPath)),
        )
    return pszOutputPath


def _apply_step0010_income_statement_borders(
    objSheet,
    iLastRow: int,
    iLastColumn: int,
) -> None:
    if iLastRow <= 0 or iLastColumn <= 0:
        return

    objThickSide = Side(style="medium", color="000000")
    objSolidSide = Side(style="thin", color="000000")
    objDottedSide = Side(style="dotted")

    def set_border(
        iRowIndex: int,
        iColumnIndex: int,
        objLeft: Optional[Side] = None,
        objRight: Optional[Side] = None,
        objTop: Optional[Side] = None,
        objBottom: Optional[Side] = None,
    ) -> None:
        objCell = objSheet.cell(row=iRowIndex, column=iColumnIndex)
        objBorder = copy(objCell.border)
        if objLeft is not None:
            objBorder.left = objLeft
        if objRight is not None:
            objBorder.right = objRight
        if objTop is not None:
            objBorder.top = objTop
        if objBottom is not None:
            objBorder.bottom = objBottom
        objCell.border = Border(
            left=objBorder.left,
            right=objBorder.right,
            top=objBorder.top,
            bottom=objBorder.bottom,
            diagonal=objBorder.diagonal,
            diagonal_direction=objBorder.diagonal_direction,
            outline=objBorder.outline,
            vertical=objBorder.vertical,
            horizontal=objBorder.horizontal,
        )

    for iRowIndex in range(1, iLastRow + 1):
        for iColumnIndex in range(1, iLastColumn + 1):
            objLeft: Optional[Side] = None
            objRight: Optional[Side] = None
            objTop: Optional[Side] = None
            objBottom: Optional[Side] = None

            if iColumnIndex == 1:
                objLeft = objThickSide
            else:
                objLeft = objSolidSide

            if iColumnIndex < iLastColumn:
                objRight = objSolidSide
            else:
                objRight = objThickSide

            if iRowIndex == 1:
                objBottom = objThickSide
                objTop = objThickSide
            elif iRowIndex < iLastRow:
                objBottom = objDottedSide
            else:
                objBottom = objThickSide

            set_border(
                iRowIndex,
                iColumnIndex,
                objLeft=objLeft,
                objRight=objRight,
                objTop=objTop,
                objBottom=objBottom,
            )


def _clear_step0010_income_statement_borders_outside_data(
    objSheet,
    iLastRow: int,
    iLastColumn: int,
) -> None:
    iSheetMaxRow: int = objSheet.max_row
    iSheetMaxColumn: int = objSheet.max_column

    if iLastRow > 0 and iLastColumn > 0:
        for iRowIndex in range(1, min(iLastRow, iSheetMaxRow) + 1):
            for iColumnIndex in range(iLastColumn + 1, iSheetMaxColumn + 1):
                objSheet.cell(row=iRowIndex, column=iColumnIndex).border = Border()

    if iLastRow < iSheetMaxRow:
        for iRowIndex in range(max(iLastRow + 1, 1), iSheetMaxRow + 1):
            for iColumnIndex in range(1, iSheetMaxColumn + 1):
                objSheet.cell(row=iRowIndex, column=iColumnIndex).border = Border()


def _normalize_step0010_manhour_label_fonts(
    objSheet,
    iLastRow: int,
) -> None:
    if iLastRow <= 0:
        return

    objTargetNames: List[str] = [
        "1Cカンパニー販管費の工数",
        "2Cカンパニー販管費の工数",
    ]
    objSourceNamePriority: List[str] = [
        "3Cカンパニー販管費の工数",
        "4Cカンパニー販管費の工数",
        "事業開発カンパニー販管費の工数",
    ]

    objNameToRowIndex: Dict[str, int] = {}
    for iRowIndex in range(1, iLastRow + 1):
        pszLabel = str(objSheet.cell(row=iRowIndex, column=1).value or "")
        if pszLabel:
            objNameToRowIndex[pszLabel] = iRowIndex

    iSourceRowIndex: Optional[int] = None
    for pszSourceName in objSourceNamePriority:
        if pszSourceName in objNameToRowIndex:
            iSourceRowIndex = objNameToRowIndex[pszSourceName]
            break
    if iSourceRowIndex is None:
        return

    objSourceFont = copy(objSheet.cell(row=iSourceRowIndex, column=1).font)
    for pszTargetName in objTargetNames:
        iTargetRowIndex = objNameToRowIndex.get(pszTargetName)
        if iTargetRowIndex is None:
            continue
        objSheet.cell(row=iTargetRowIndex, column=1).font = copy(objSourceFont)


def create_step0010_pj_income_statement_excel_from_tsv(
    pszStep0010Path: str,
) -> Optional[str]:
    pszBaseName = os.path.basename(pszStep0010Path)
    objMatch = re.fullmatch(
        r"損益計算書_販管費配賦_step0010_(\d{4}年\d{2}月)_A∪B_プロジェクト名_C∪D\.tsv",
        pszBaseName,
    )
    if objMatch is None:
        return None
    pszYearMonth = objMatch.group(1)
    pszScriptDirectory: str = os.path.dirname(__file__)
    pszTemplatePath: str = os.path.join(
        pszScriptDirectory,
        f"TEMPLATE_販管費配賦後_損益計算書_{pszYearMonth}_A∪B_プロジェクト名_C∪D.xlsx",
    )
    if not os.path.isfile(pszTemplatePath):
        pszTemplatePath = os.path.join(
            pszScriptDirectory,
            "TEMPLATE_販管費配賦後_損益計算書_YYYY年MM月_A∪B_プロジェクト名_C∪D.xlsx",
        )
    if not os.path.isfile(pszTemplatePath):
        return None

    objWorkbook = load_workbook(pszTemplatePath)
    objSheet = objWorkbook.worksheets[0]
    objSheet.title = f"PJ別損益計算書_{pszYearMonth}"
    objRows = read_tsv_rows(pszStep0010Path)
    iLastColumn: int = max((len(objRow) for objRow in objRows), default=0)
    for iRowIndex, objRow in enumerate(objRows, start=1):
        for iColumnIndex, pszValue in enumerate(objRow, start=1):
            objCellValue = parse_tsv_value_for_excel(pszValue)
            objSheet.cell(row=iRowIndex, column=iColumnIndex, value=objCellValue)

    _apply_step0010_income_statement_borders(objSheet, len(objRows), iLastColumn)
    _clear_step0010_income_statement_borders_outside_data(
        objSheet,
        len(objRows),
        iLastColumn,
    )
    _normalize_step0010_manhour_label_fonts(objSheet, len(objRows))

    pszOutputPath: str = os.path.join(
        os.path.dirname(pszStep0010Path),
        f"販管費配賦後_損益計算書_{pszYearMonth}_A∪B_プロジェクト名_C∪D.xlsx",
    )
    objWorkbook.save(pszOutputPath)
    if EXECUTION_ROOT_DIRECTORY:
        pszTargetDirectory = os.path.join(EXECUTION_ROOT_DIRECTORY, "PJ別損益計算書")
        os.makedirs(pszTargetDirectory, exist_ok=True)
        shutil.copy2(
            pszOutputPath,
            os.path.join(pszTargetDirectory, os.path.basename(pszOutputPath)),
        )
    return pszOutputPath


def create_step0010_pj_income_statement_vertical_excel_from_tsv(
    pszStep0010VerticalPath: str,
) -> Optional[str]:
    pszBaseName = os.path.basename(pszStep0010VerticalPath)
    objMatch = re.fullmatch(
        r"損益計算書_販管費配賦_step0010_(\d{4}年\d{2}月)_A∪B_プロジェクト名_C∪D_vertical\.tsv",
        pszBaseName,
    )
    if objMatch is None:
        return None
    pszYearMonth = objMatch.group(1)
    pszScriptDirectory: str = os.path.dirname(__file__)
    pszTemplatePath: str = os.path.join(
        pszScriptDirectory,
        f"TEMPLATE_販管費配賦後_損益計算書_{pszYearMonth}_A∪B_プロジェクト名_C∪D_vertical.xlsx",
    )
    if not os.path.isfile(pszTemplatePath):
        pszTemplatePath = os.path.join(
            pszScriptDirectory,
            "TEMPLATE_販管費配賦後_損益計算書_YYYY年MM月_A∪B_プロジェクト名_C∪D_vertical.xlsx",
        )
    if not os.path.isfile(pszTemplatePath):
        return None

    objWorkbook = load_workbook(pszTemplatePath)
    objSheet = objWorkbook.worksheets[0]
    objSheet.title = f"PJ別損益計算書_{pszYearMonth}_vertical"
    objRows = read_tsv_rows(pszStep0010VerticalPath)
    iLastColumn: int = max((len(objRow) for objRow in objRows), default=0)
    for iRowIndex, objRow in enumerate(objRows, start=1):
        for iColumnIndex, pszValue in enumerate(objRow, start=1):
            objCellValue = parse_tsv_value_for_excel(pszValue)
            objSheet.cell(row=iRowIndex, column=iColumnIndex, value=objCellValue)

    _apply_step0010_income_statement_borders(objSheet, len(objRows), iLastColumn)
    _clear_step0010_income_statement_borders_outside_data(
        objSheet,
        len(objRows),
        iLastColumn,
    )
    _normalize_step0010_manhour_label_fonts(objSheet, len(objRows))

    pszOutputPath: str = os.path.join(
        os.path.dirname(pszStep0010VerticalPath),
        f"販管費配賦後_損益計算書_{pszYearMonth}_A∪B_プロジェクト名_C∪D_vertical.xlsx",
    )
    objWorkbook.save(pszOutputPath)
    if EXECUTION_ROOT_DIRECTORY:
        pszTargetDirectory = os.path.join(EXECUTION_ROOT_DIRECTORY, "PJ別損益計算書")
        os.makedirs(pszTargetDirectory, exist_ok=True)
        shutil.copy2(
            pszOutputPath,
            os.path.join(pszTargetDirectory, os.path.basename(pszOutputPath)),
        )
    return pszOutputPath


def copy_excel_sheet_contents(objSourceSheet, objDestinationSheet) -> None:
    objDestinationSheet.sheet_format = copy(objSourceSheet.sheet_format)
    objDestinationSheet.sheet_properties = copy(objSourceSheet.sheet_properties)
    objDestinationSheet.page_margins = copy(objSourceSheet.page_margins)
    objDestinationSheet.page_setup = copy(objSourceSheet.page_setup)
    objDestinationSheet.print_options = copy(objSourceSheet.print_options)

    for iRowIndex, objDimension in objSourceSheet.row_dimensions.items():
        objDestinationSheet.row_dimensions[iRowIndex] = copy(objDimension)

    for pszColumnName, objDimension in objSourceSheet.column_dimensions.items():
        objDestinationSheet.column_dimensions[pszColumnName] = copy(objDimension)

    for objRow in objSourceSheet.iter_rows():
        for objCell in objRow:
            objDestinationCell = objDestinationSheet.cell(
                row=objCell.row,
                column=objCell.column,
                value=objCell.value,
            )
            if objCell.has_style:
                objDestinationCell._style = copy(objCell._style)
            if objCell.number_format is not None:
                objDestinationCell.number_format = objCell.number_format
            if objCell.protection is not None:
                objDestinationCell.protection = copy(objCell.protection)
            if objCell.alignment is not None:
                objDestinationCell.alignment = copy(objCell.alignment)
            if objCell.fill is not None:
                objDestinationCell.fill = copy(objCell.fill)
            if objCell.font is not None:
                objDestinationCell.font = copy(objCell.font)
            if objCell.border is not None:
                objDestinationCell.border = copy(objCell.border)

    for objMergedCellRange in objSourceSheet.merged_cells.ranges:
        objDestinationSheet.merge_cells(str(objMergedCellRange))


def create_step0010_pj_income_statement_both_excel(
    pszNormalExcelPath: str,
    pszVerticalExcelPath: str,
) -> Optional[str]:
    pszNormalName: str = os.path.basename(pszNormalExcelPath)
    objMatch = re.fullmatch(
        r"販管費配賦後_損益計算書_(\d{4}年\d{2}月)_A∪B_プロジェクト名_C∪D\.xlsx",
        pszNormalName,
    )
    if objMatch is None:
        return None
    pszYearMonth = objMatch.group(1)

    if not os.path.isfile(pszNormalExcelPath) or not os.path.isfile(pszVerticalExcelPath):
        return None

    pszOutputPath: str = os.path.join(
        os.path.dirname(pszNormalExcelPath),
        f"販管費配賦後_損益計算書_{pszYearMonth}_A∪B_プロジェクト名_C∪D_両方.xlsx",
    )
    shutil.copy2(pszNormalExcelPath, pszOutputPath)

    objBothWorkbook = load_workbook(pszOutputPath)
    objVerticalWorkbook = load_workbook(pszVerticalExcelPath)
    if not objBothWorkbook.worksheets or not objVerticalWorkbook.worksheets:
        return None

    objVerticalSourceSheet = objVerticalWorkbook.worksheets[0]
    pszVerticalSheetTitle: str = objVerticalSourceSheet.title
    if pszVerticalSheetTitle in objBothWorkbook.sheetnames:
        iSuffix: int = 2
        while f"{pszVerticalSheetTitle}_{iSuffix}" in objBothWorkbook.sheetnames:
            iSuffix += 1
        pszVerticalSheetTitle = f"{pszVerticalSheetTitle}_{iSuffix}"

    objVerticalSheet = objBothWorkbook.create_sheet(title=pszVerticalSheetTitle)
    copy_excel_sheet_contents(objVerticalSourceSheet, objVerticalSheet)

    objBothWorkbook.save(pszOutputPath)
    if EXECUTION_ROOT_DIRECTORY:
        pszTargetDirectory = os.path.join(EXECUTION_ROOT_DIRECTORY, "PJ別損益計算書")
        os.makedirs(pszTargetDirectory, exist_ok=True)
        shutil.copy2(
            pszOutputPath,
            os.path.join(pszTargetDirectory, os.path.basename(pszOutputPath)),
        )
    return pszOutputPath


def create_step0010_pj_income_statement_range_excel_from_tsvs(
    pszDirectory: str,
    objMonthlyPaths: List[str],
    bVertical: bool,
) -> Optional[str]:
    if not objMonthlyPaths:
        return None

    objPathPairs: List[Tuple[Tuple[int, int], str]] = []
    pszPattern = (
        r"損益計算書_販管費配賦_step0010_(\d{4})年(\d{2})月_A∪B_プロジェクト名_C∪D_vertical\.tsv"
        if bVertical
        else r"損益計算書_販管費配賦_step0010_(\d{4})年(\d{2})月_A∪B_プロジェクト名_C∪D\.tsv"
    )
    for pszPath in objMonthlyPaths:
        pszName: str = os.path.basename(pszPath)
        objMatch = re.fullmatch(pszPattern, pszName)
        if objMatch is None:
            continue
        iYear: int = int(objMatch.group(1))
        iMonth: int = int(objMatch.group(2))
        objPathPairs.append(((iYear, iMonth), pszPath))

    if not objPathPairs:
        return None

    objPathPairs.sort(key=lambda objItem: objItem[0])
    objStart = objPathPairs[0][0]
    objEnd = objPathPairs[-1][0]
    pszStartLabel: str = f"{objStart[0]}年{objStart[1]:02d}月"
    pszEndLabel: str = f"{objEnd[0]}年{objEnd[1]:02d}月"

    pszScriptDirectory: str = os.path.dirname(__file__)
    pszTemplateName: str = (
        "TEMPLATE_販管費配賦後_損益計算書_YYYY年MM月_A∪B_プロジェクト名_C∪D_vertical.xlsx"
        if bVertical
        else "TEMPLATE_販管費配賦後_損益計算書_YYYY年MM月_A∪B_プロジェクト名_C∪D.xlsx"
    )
    pszTemplatePath: str = os.path.join(pszScriptDirectory, pszTemplateName)
    if not os.path.isfile(pszTemplatePath):
        pszEndYearMonth: str = f"{objEnd[0]}年{objEnd[1]:02d}月"
        pszMonthlyTemplateName: str = (
            f"TEMPLATE_販管費配賦後_損益計算書_{pszEndYearMonth}_A∪B_プロジェクト名_C∪D_vertical.xlsx"
            if bVertical
            else f"TEMPLATE_販管費配賦後_損益計算書_{pszEndYearMonth}_A∪B_プロジェクト名_C∪D.xlsx"
        )
        pszTemplatePath = os.path.join(pszScriptDirectory, pszMonthlyTemplateName)
    if not os.path.isfile(pszTemplatePath):
        return None

    objWorkbook = load_workbook(pszTemplatePath)
    objTemplateSheet = objWorkbook.worksheets[0]
    for objSheetToRemove in objWorkbook.worksheets[1:]:
        objWorkbook.remove(objSheetToRemove)

    for objYearMonth, pszPath in objPathPairs:
        pszSheetName: str = f"{objYearMonth[0]}年{objYearMonth[1]:02d}月"
        if bVertical:
            pszSheetName = f"{pszSheetName}_vertical"

        objSheet = objWorkbook.copy_worksheet(objTemplateSheet)
        objSheet.title = pszSheetName
        _clear_sheet_values(objSheet)

        objRows = read_tsv_rows(pszPath)
        iLastColumn: int = max((len(objRow) for objRow in objRows), default=0)
        for iRowIndex, objRow in enumerate(objRows, start=1):
            for iColumnIndex, pszValue in enumerate(objRow, start=1):
                objCellValue = parse_tsv_value_for_excel(pszValue)
                objSheet.cell(row=iRowIndex, column=iColumnIndex, value=objCellValue)

        _apply_step0010_income_statement_borders(objSheet, len(objRows), iLastColumn)
        _clear_step0010_income_statement_borders_outside_data(
            objSheet,
            len(objRows),
            iLastColumn,
        )
        _normalize_step0010_manhour_label_fonts(objSheet, len(objRows))
    if objTemplateSheet in objWorkbook.worksheets:
        objWorkbook.remove(objTemplateSheet)

    pszOutputName: str = (
        f"販管費配賦後_損益計算書_{pszStartLabel}-{pszEndLabel}_A∪B_プロジェクト名_C∪D_vertical.xlsx"
        if bVertical
        else f"販管費配賦後_損益計算書_{pszStartLabel}-{pszEndLabel}_A∪B_プロジェクト名_C∪D.xlsx"
    )
    pszOutputPath: str = os.path.join(pszDirectory, pszOutputName)
    objWorkbook.save(pszOutputPath)
    if EXECUTION_ROOT_DIRECTORY:
        pszTargetDirectory = os.path.join(EXECUTION_ROOT_DIRECTORY, "PJ別損益計算書")
        os.makedirs(pszTargetDirectory, exist_ok=True)
        shutil.copy2(
            pszOutputPath,
            os.path.join(pszTargetDirectory, os.path.basename(pszOutputPath)),
        )
    return pszOutputPath


def create_step0010_pj_income_statement_excels(pszDirectory: str) -> List[str]:
    objOutputs: List[str] = []
    objNormalOutputByYearMonth: Dict[Tuple[int, int], str] = {}
    objVerticalOutputByYearMonth: Dict[Tuple[int, int], str] = {}
    objSelectedRangePath: Optional[str] = find_selected_range_path(pszDirectory)
    objSelectedRange = (
        parse_selected_range(objSelectedRangePath)
        if objSelectedRangePath is not None
        else None
    )
    objTargetYearMonth: Optional[Tuple[int, int]] = (
        objSelectedRange[1] if objSelectedRange is not None else None
    )
    objMonthlyNormalPaths: List[str] = []
    objMonthlyVerticalPaths: List[str] = []

    for pszName in sorted(os.listdir(pszDirectory)):
        pszPath = os.path.join(pszDirectory, pszName)
        if not os.path.isfile(pszPath):
            continue
        if objTargetYearMonth is not None:
            objYearMonth = extract_year_month_from_path(pszPath)
            if objYearMonth != objTargetYearMonth:
                continue
        if re.fullmatch(
            r"損益計算書_販管費配賦_step0010_\d{4}年\d{2}月_A∪B_プロジェクト名_C∪D\.tsv",
            pszName,
        ) is not None:
            objMonthlyNormalPaths.append(pszPath)
            if objTargetYearMonth is not None:
                objYearMonth = extract_year_month_from_path(pszPath)
                if objYearMonth != objTargetYearMonth:
                    continue
            pszOutput = create_step0010_pj_income_statement_excel_from_tsv(pszPath)
            if pszOutput is not None:
                objOutputs.append(pszOutput)
                objYearMonth = extract_year_month_from_path(pszPath)
                if objYearMonth is not None:
                    objNormalOutputByYearMonth[objYearMonth] = pszOutput
            continue
        if re.fullmatch(
            r"損益計算書_販管費配賦_step0010_\d{4}年\d{2}月_A∪B_プロジェクト名_C∪D_vertical\.tsv",
            pszName,
        ) is not None:
            objMonthlyVerticalPaths.append(pszPath)
            if objTargetYearMonth is not None:
                objYearMonth = extract_year_month_from_path(pszPath)
                if objYearMonth != objTargetYearMonth:
                    continue
            pszOutput = create_step0010_pj_income_statement_vertical_excel_from_tsv(pszPath)
            if pszOutput is not None:
                objOutputs.append(pszOutput)
                objYearMonth = extract_year_month_from_path(pszPath)
                if objYearMonth is not None:
                    objVerticalOutputByYearMonth[objYearMonth] = pszOutput

    pszRangeOutput = create_step0010_pj_income_statement_range_excel_from_tsvs(
        pszDirectory,
        objMonthlyNormalPaths,
        False,
    )
    if pszRangeOutput is not None:
        objOutputs.append(pszRangeOutput)

    pszRangeVerticalOutput = create_step0010_pj_income_statement_range_excel_from_tsvs(
        pszDirectory,
        objMonthlyVerticalPaths,
        True,
    )
    if pszRangeVerticalOutput is not None:
        objOutputs.append(pszRangeVerticalOutput)

    objBothYearMonths = sorted(
        set(objNormalOutputByYearMonth.keys()) & set(objVerticalOutputByYearMonth.keys())
    )
    for objYearMonth in objBothYearMonths:
        pszBothOutput = create_step0010_pj_income_statement_both_excel(
            objNormalOutputByYearMonth[objYearMonth],
            objVerticalOutputByYearMonth[objYearMonth],
        )
        if pszBothOutput is not None:
            objOutputs.append(pszBothOutput)

    return objOutputs


def reorder_cp_step0006_rows(objRows: List[List[str]]) -> List[List[str]]:
    if not objRows:
        return objRows
    objOrder = [
        "第一インキュ",
        "第二インキュ",
        "第三インキュ",
        "第四インキュ",
        "事業開発",
        "子会社",
        "投資先",
        "本部",
    ]
    objOrderSet = set(objOrder)
    objHeader = objRows[0]
    objBody = objRows[1:]
    objOrderedRows: List[List[str]] = []
    for pszCompany in objOrder:
        for objRow in objBody:
            if objRow and objRow[0] == pszCompany:
                objOrderedRows.append(objRow)
    for objRow in objBody:
        if not objRow or objRow[0] not in objOrderSet:
            objOrderedRows.append(objRow)
    return [objHeader] + objOrderedRows


def build_step0007_rows_for_cp(
    objRows: List[List[str]],
    objPriorMap: Dict[str, str],
    pszPriorLabel: str,
    pszCurrentLabel: str,
    pszPriorRowLabel: str,
    pszPrefix: str,
) -> List[List[str]]:
    if not objRows:
        return []
    objInsertedRows: List[List[str]] = []
    for objRow in objRows:
        pszLabel = objRow[0] if objRow else ""
        pszValue = objRow[1] if len(objRow) > 1 else ""
        objInsertedRows.append([pszLabel, "", "", pszValue])
    pszCompanyLabel = objInsertedRows[0][3] if len(objInsertedRows[0]) > 3 else ""
    objInsertedRows.insert(0, [pszCompanyLabel, "", "", ""])
    objInsertedRows[0][1] = pszPriorLabel
    objInsertedRows[0][2] = pszCurrentLabel
    objInsertedRows[0][3] = pszCurrentLabel
    if len(objInsertedRows) > 1:
        objInsertedRows[1][1] = pszPriorRowLabel
        objInsertedRows[1][2] = "計画"
        objInsertedRows[1][3] = "実績"
    for iRowIndex, objRow in enumerate(objInsertedRows):
        if len(objRow) < 4:
            objRow.extend([""] * (4 - len(objRow)))
        objRow.extend(["", ""])
        if iRowIndex == 1:
            objRow[4] = "前年比"
            objRow[5] = "計画比"
    bAllPriorValuesZero: bool = True
    bHasPriorNumericValue: bool = False
    iCurrentPeriodMonthCount: int = 0
    iPriorPeriodMonthCount: int = 0
    if pszPriorRowLabel == "前期":
        iCurrentPeriodMonthCount = parse_period_month_count(pszCurrentLabel)
        iPriorPeriodMonthCount = parse_period_month_count(pszPriorLabel)

    for objRow in objInsertedRows[2:]:
        pszLabel = objRow[0] if objRow else ""
        if not pszLabel:
            continue
        if pszLabel not in objPriorMap:
            bAllPriorValuesZero = False
            break
        pszPriorValue = objPriorMap[pszLabel]
        pszTrimmedPriorValue = (pszPriorValue or "").strip()
        if pszTrimmedPriorValue == "":
            bAllPriorValuesZero = False
            break
        try:
            fPriorValue = float(pszTrimmedPriorValue)
        except ValueError:
            bAllPriorValuesZero = False
            break
        if abs(fPriorValue) < 0.0000001:
            bHasPriorNumericValue = True
            continue
        bAllPriorValuesZero = False
        break

    for objRow in objInsertedRows[2:]:
        pszLabel = objRow[0] if objRow else ""
        if not pszLabel:
            continue
        if pszLabel in objPriorMap:
            pszPriorValue = objPriorMap[pszLabel]
            pszTrimmedPriorValue = (pszPriorValue or "").strip()
            if pszTrimmedPriorValue == "":
                objRow[1] = pszPriorValue
            else:
                try:
                    fPriorValue = float(pszTrimmedPriorValue)
                except ValueError:
                    objRow[1] = pszPriorValue
                else:
                    if bAllPriorValuesZero and bHasPriorNumericValue and abs(fPriorValue) < 0.0000001:
                        objRow[1] = "'－"
                    else:
                        objRow[1] = pszPriorValue
        else:
            objRow[1] = "'－"
        pszPriorValueForYoY: str = objRow[1] if len(objRow) > 1 else ""
        pszActualValueForYoY: str = objRow[3] if len(objRow) > 3 else ""
        pszTrimmedPriorForYoY: str = (pszPriorValueForYoY or "").strip()
        pszTrimmedActualForYoY: str = (pszActualValueForYoY or "").strip()
        if pszTrimmedPriorForYoY == "" or pszTrimmedPriorForYoY == "'－":
            continue
        if pszTrimmedActualForYoY == "":
            continue
        try:
            fPriorValueForYoY: float = float(pszTrimmedPriorForYoY)
            fActualValueForYoY: float = float(pszTrimmedActualForYoY)
        except ValueError:
            continue
        fPriorValueDenominator: float = fPriorValueForYoY
        if pszPriorRowLabel == "前期" and iCurrentPeriodMonthCount > 0 and iPriorPeriodMonthCount > 0:
            fPeriodRatio: float = float(iCurrentPeriodMonthCount) / float(iPriorPeriodMonthCount)
            fPriorValueDenominator = fPriorValueForYoY * fPeriodRatio

        if abs(fPriorValueDenominator) < 0.0000001:
            if fActualValueForYoY > 0:
                objRow[4] = "＋∞"
            elif fActualValueForYoY < 0:
                objRow[4] = "－∞"
            continue
        objRow[4] = "{0:.4f}".format(fActualValueForYoY / fPriorValueDenominator)
    apply_cp_company_plan_values(objInsertedRows, pszCurrentLabel, pszPrefix)
    apply_cp_group_plan_values(objInsertedRows, pszCurrentLabel, pszPrefix)

    for objRow in objInsertedRows[2:]:
        pszPlanValue: str = objRow[2] if len(objRow) > 2 else ""
        pszActualValue: str = objRow[3] if len(objRow) > 3 else ""
        fPlanValue = parse_plan_numeric_value(pszPlanValue)
        fActualValue = parse_plan_numeric_value(pszActualValue)
        if fPlanValue is None or fActualValue is None:
            continue
        if abs(fPlanValue) < 0.0000001:
            if fActualValue > 0:
                objRow[5] = "＋∞"
            elif fActualValue < 0:
                objRow[5] = "－∞"
            else:
                objRow[5] = ""
            continue
        objRow[5] = "{0:.4f}".format(fActualValue / fPlanValue)

    return objInsertedRows


CP_COMPANY_ALLOWED_NAMES: List[str] = [
    "テクノロジーインキュベーション",
    "コンテンツビジネス",
    "スタートアップサイド",
    "スタートアップコミュニティ",
    "スタートアップグロース",
    "経営管理",
    "第一インキュ",
    "第二インキュ",
    "第三インキュ",
    "第四インキュ",
    "事業開発",
    "子会社",
    "投資先",
    "本部",
    "合計",
]


CP_GROUP_ALLOWED_NAMES: List[str] = [
    "受託事業-施設運営",
    "受託事業-その他",
    "自社-施設運営",
    "自社-その他",
    "合計",
]


def parse_japanese_year_month_label(pszLabel: str) -> Optional[Tuple[int, int]]:
    objMatch = re.match(r"^(\d{4})年(\d{1,2})月$", (pszLabel or "").strip())
    if objMatch is None:
        return None
    iYear: int = int(objMatch.group(1))
    iMonth: int = int(objMatch.group(2))
    if iMonth < 1 or iMonth > 12:
        return None
    return iYear, iMonth


def parse_current_period_months_for_cp(pszCurrentLabel: str) -> List[Tuple[int, int]]:
    objSingle = parse_japanese_year_month_label(pszCurrentLabel)
    if objSingle is not None:
        return [objSingle]
    objRangeMatch = re.match(
        r"^(\d{4})年(\d{2})月-(\d{4})年(\d{2})月$",
        (pszCurrentLabel or "").strip(),
    )
    if objRangeMatch is None:
        return []
    iStartYear: int = int(objRangeMatch.group(1))
    iStartMonth: int = int(objRangeMatch.group(2))
    iEndYear: int = int(objRangeMatch.group(3))
    iEndMonth: int = int(objRangeMatch.group(4))
    return build_month_sequence((iStartYear, iStartMonth), (iEndYear, iEndMonth))


def parse_plan_numeric_value(pszValue: str) -> Optional[float]:
    pszText: str = (pszValue or "").strip()
    if pszText == "":
        return None
    pszNormalized: str = pszText.replace(",", "")
    if pszNormalized.endswith("%"):
        pszNormalized = pszNormalized[:-1]
    if pszNormalized == "":
        return None
    try:
        return float(pszNormalized)
    except ValueError:
        return None


CP_COMPANY_PLAN_CACHE: Optional[Dict[Tuple[str, str], Dict[Tuple[int, int], str]]] = None
CP_GROUP_PLAN_CACHE: Optional[Dict[Tuple[str, str], Dict[Tuple[int, int], str]]] = None


def read_cp_company_plan_map() -> Dict[Tuple[str, str], Dict[Tuple[int, int], str]]:
    global CP_COMPANY_PLAN_CACHE
    if CP_COMPANY_PLAN_CACHE is not None:
        return CP_COMPANY_PLAN_CACHE
    pszPlanPath: str = os.path.join(get_script_base_directory(), "計画.csv")
    if not os.path.isfile(pszPlanPath):
        CP_COMPANY_PLAN_CACHE = {}
        return CP_COMPANY_PLAN_CACHE

    objRows: List[List[str]] = []
    with open(pszPlanPath, "r", encoding="utf-8-sig", newline="") as objFile:
        objSniffer = csv.Sniffer()
        pszSample: str = objFile.read(4096)
        objFile.seek(0)
        try:
            objDialect = objSniffer.sniff(pszSample, delimiters=",\t")
            pszDelimiter = objDialect.delimiter
        except csv.Error:
            pszDelimiter = "\t"
        objReader = csv.reader(objFile, delimiter=pszDelimiter)
        for objRow in objReader:
            objRows.append(list(objRow))

    if not objRows:
        CP_COMPANY_PLAN_CACHE = {}
        return CP_COMPANY_PLAN_CACHE

    objMonthColumns: Dict[int, Tuple[int, int]] = {}
    for iColumnIndex, pszLabel in enumerate(objRows[0]):
        objMonth = parse_japanese_year_month_label(pszLabel)
        if objMonth is not None:
            objMonthColumns[iColumnIndex] = objMonth

    objAllowedCompanySet = set(CP_COMPANY_ALLOWED_NAMES)
    objGroupStartNames = {
        "受託事業-施設運営",
        "受託事業-その他",
        "自社-施設運営",
        "自社-その他",
    }
    pszCurrentCompany: str = ""
    objPlanMap: Dict[Tuple[str, str], Dict[Tuple[int, int], str]] = {}

    for objRow in objRows[1:]:
        pszCompanyCell: str = objRow[0].strip() if len(objRow) > 0 else ""
        if pszCompanyCell in objGroupStartNames:
            break
        if pszCompanyCell != "":
            pszCurrentCompany = pszCompanyCell
        if pszCurrentCompany not in objAllowedCompanySet:
            continue
        pszSubject: str = objRow[1].strip() if len(objRow) > 1 else ""
        if pszSubject == "":
            continue
        objKey = (pszCurrentCompany, pszSubject)
        objMonthMap: Dict[Tuple[int, int], str] = objPlanMap.setdefault(objKey, {})
        for iColumnIndex, objMonth in objMonthColumns.items():
            if iColumnIndex >= len(objRow):
                continue
            objMonthMap[objMonth] = (objRow[iColumnIndex] or "").strip()

    CP_COMPANY_PLAN_CACHE = objPlanMap
    return CP_COMPANY_PLAN_CACHE



def read_cp_group_plan_map() -> Dict[Tuple[str, str], Dict[Tuple[int, int], str]]:
    global CP_GROUP_PLAN_CACHE
    if CP_GROUP_PLAN_CACHE is not None:
        return CP_GROUP_PLAN_CACHE
    pszPlanPath: str = os.path.join(get_script_base_directory(), "計画.csv")
    if not os.path.isfile(pszPlanPath):
        CP_GROUP_PLAN_CACHE = {}
        return CP_GROUP_PLAN_CACHE

    objRows: List[List[str]] = []
    with open(pszPlanPath, "r", encoding="utf-8-sig", newline="") as objFile:
        objSniffer = csv.Sniffer()
        pszSample: str = objFile.read(4096)
        objFile.seek(0)
        try:
            objDialect = objSniffer.sniff(pszSample, delimiters=",	")
            pszDelimiter = objDialect.delimiter
        except csv.Error:
            pszDelimiter = "	"
        objReader = csv.reader(objFile, delimiter=pszDelimiter)
        for objRow in objReader:
            objRows.append(list(objRow))

    if not objRows:
        CP_GROUP_PLAN_CACHE = {}
        return CP_GROUP_PLAN_CACHE

    objMonthColumns: Dict[int, Tuple[int, int]] = {}
    for iColumnIndex, pszLabel in enumerate(objRows[0]):
        objMonth = parse_japanese_year_month_label(pszLabel)
        if objMonth is not None:
            objMonthColumns[iColumnIndex] = objMonth

    objAllowedGroupSet = set(CP_GROUP_ALLOWED_NAMES)
    objGroupStartNames = {
        "受託事業-施設運営",
        "受託事業-その他",
        "自社-施設運営",
        "自社-その他",
    }
    pszCurrentGroup: str = ""
    bInGroupSection: bool = False
    objPlanMap: Dict[Tuple[str, str], Dict[Tuple[int, int], str]] = {}

    for objRow in objRows[1:]:
        pszGroupCell: str = objRow[0].strip() if len(objRow) > 0 else ""
        if pszGroupCell in objGroupStartNames:
            bInGroupSection = True
        if not bInGroupSection:
            continue
        if pszGroupCell != "":
            pszCurrentGroup = pszGroupCell
        if pszCurrentGroup not in objAllowedGroupSet:
            continue
        pszSubject: str = objRow[1].strip() if len(objRow) > 1 else ""
        if pszSubject == "":
            continue
        objKey = (pszCurrentGroup, pszSubject)
        objMonthMap: Dict[Tuple[int, int], str] = objPlanMap.setdefault(objKey, {})
        for iColumnIndex, objMonth in objMonthColumns.items():
            if iColumnIndex >= len(objRow):
                continue
            objMonthMap[objMonth] = (objRow[iColumnIndex] or "").strip()

    CP_GROUP_PLAN_CACHE = objPlanMap
    return CP_GROUP_PLAN_CACHE

def apply_cp_company_plan_values(
    objInsertedRows: List[List[str]],
    pszCurrentLabel: str,
    pszPrefix: str,
) -> None:
    if pszPrefix != "0001_CP別":
        return
    if len(objInsertedRows) < 3:
        return
    pszCompany: str = (objInsertedRows[0][0] if objInsertedRows[0] else "").strip()
    if pszCompany not in set(CP_COMPANY_ALLOWED_NAMES):
        return
    objMonths: List[Tuple[int, int]] = parse_current_period_months_for_cp(pszCurrentLabel)
    if not objMonths:
        return
    objPlanMap = read_cp_company_plan_map()
    if not objPlanMap:
        return

    def get_monthly_plan_values(pszSubject: str) -> List[str]:
        objMonthMap = objPlanMap.get((pszCompany, pszSubject), {})
        return [objMonthMap.get(objMonth, "") for objMonth in objMonths]

    def compute_sum_value_text(pszSubject: str) -> str:
        objValues: List[str] = get_monthly_plan_values(pszSubject)
        fTotal: float = 0.0
        bHasNumeric: bool = False
        for pszValue in objValues:
            fParsed = parse_plan_numeric_value(pszValue)
            if fParsed is None:
                continue
            bHasNumeric = True
            fTotal += fParsed
        if not bHasNumeric:
            return ""
        if abs(fTotal - round(fTotal)) < 0.0000001:
            return str(int(round(fTotal)))
        return ("{0:.10f}".format(fTotal)).rstrip("0").rstrip(".")

    def compute_sum_numeric(pszSubject: str) -> Optional[float]:
        pszText = compute_sum_value_text(pszSubject)
        return parse_plan_numeric_value(pszText)

    bIsRange: bool = len(objMonths) > 1
    fSalesTotal: Optional[float] = None
    fGrossTotal: Optional[float] = None
    fOperatingTotal: Optional[float] = None
    if bIsRange:
        fSalesTotal = compute_sum_numeric("純売上高")
        fGrossTotal = compute_sum_numeric("売上総利益")
        fOperatingTotal = compute_sum_numeric("営業利益")

    for objRow in objInsertedRows[2:]:
        pszSubject: str = (objRow[0] if objRow else "").strip()
        if pszSubject == "":
            continue
        if bIsRange and pszSubject == "売上総利益率":
            if fSalesTotal is None or abs(fSalesTotal) < 0.0000001 or fGrossTotal is None:
                objRow[2] = ""
            else:
                objRow[2] = "{0:.2f}".format((fGrossTotal / fSalesTotal) * 100.0)
            continue
        if bIsRange and pszSubject == "営業利益率":
            if fSalesTotal is None or abs(fSalesTotal) < 0.0000001 or fOperatingTotal is None:
                objRow[2] = ""
            else:
                objRow[2] = "{0:.2f}".format((fOperatingTotal / fSalesTotal) * 100.0)
            continue
        if bIsRange:
            objRow[2] = compute_sum_value_text(pszSubject)
            continue
        objValues: List[str] = get_monthly_plan_values(pszSubject)
        objRow[2] = objValues[0] if objValues else ""


def apply_cp_group_plan_values(
    objInsertedRows: List[List[str]],
    pszCurrentLabel: str,
    pszPrefix: str,
) -> None:
    if pszPrefix != "0002_CP別":
        return
    if len(objInsertedRows) < 3:
        return
    pszGroup: str = (objInsertedRows[0][0] if objInsertedRows[0] else "").strip()
    if pszGroup not in set(CP_GROUP_ALLOWED_NAMES):
        return
    objMonths: List[Tuple[int, int]] = parse_current_period_months_for_cp(pszCurrentLabel)
    if not objMonths:
        return
    objPlanMap = read_cp_group_plan_map()
    if not objPlanMap:
        return

    def get_monthly_plan_values(pszSubject: str) -> List[str]:
        objMonthMap = objPlanMap.get((pszGroup, pszSubject), {})
        return [objMonthMap.get(objMonth, "") for objMonth in objMonths]

    def compute_sum_value_text(pszSubject: str) -> str:
        objValues: List[str] = get_monthly_plan_values(pszSubject)
        fTotal: float = 0.0
        bHasNumeric: bool = False
        for pszValue in objValues:
            fParsed = parse_plan_numeric_value(pszValue)
            if fParsed is None:
                continue
            bHasNumeric = True
            fTotal += fParsed
        if not bHasNumeric:
            return ""
        if abs(fTotal - round(fTotal)) < 0.0000001:
            return str(int(round(fTotal)))
        return ("{0:.10f}".format(fTotal)).rstrip("0").rstrip(".")

    def compute_sum_numeric(pszSubject: str) -> Optional[float]:
        pszText = compute_sum_value_text(pszSubject)
        return parse_plan_numeric_value(pszText)

    bIsRange: bool = len(objMonths) > 1
    fSalesTotal: Optional[float] = None
    fGrossTotal: Optional[float] = None
    fOperatingTotal: Optional[float] = None
    if bIsRange:
        fSalesTotal = compute_sum_numeric("純売上高")
        fGrossTotal = compute_sum_numeric("売上総利益")
        fOperatingTotal = compute_sum_numeric("営業利益")

    for objRow in objInsertedRows[2:]:
        pszSubject: str = (objRow[0] if objRow else "").strip()
        if pszSubject == "":
            continue
        if bIsRange and pszSubject == "売上総利益率":
            if fSalesTotal is None or abs(fSalesTotal) < 0.0000001 or fGrossTotal is None:
                objRow[2] = ""
            else:
                objRow[2] = "{0:.2f}".format((fGrossTotal / fSalesTotal) * 100.0)
            continue
        if bIsRange and pszSubject == "営業利益率":
            if fSalesTotal is None or abs(fSalesTotal) < 0.0000001 or fOperatingTotal is None:
                objRow[2] = ""
            else:
                objRow[2] = "{0:.2f}".format((fOperatingTotal / fSalesTotal) * 100.0)
            continue
        if bIsRange:
            objRow[2] = compute_sum_value_text(pszSubject)
            continue
        objValues: List[str] = get_monthly_plan_values(pszSubject)
        objRow[2] = objValues[0] if objValues else ""


def parse_period_month_count(pszLabel: str) -> int:
    objRangeMatch = re.match(r"^(\d{4})年(\d{2})月-(\d{4})年(\d{2})月$", (pszLabel or "").strip())
    if objRangeMatch is not None:
        iStartYear = int(objRangeMatch.group(1))
        iStartMonth = int(objRangeMatch.group(2))
        iEndYear = int(objRangeMatch.group(3))
        iEndMonth = int(objRangeMatch.group(4))
        return (iEndYear - iStartYear) * 12 + (iEndMonth - iStartMonth) + 1

    objSingleMatch = re.match(r"^(\d{4})年(\d{2})月$", (pszLabel or "").strip())
    if objSingleMatch is not None:
        return 1

    return 0


def build_step0006_prior_map(pszPriorPath: str) -> Dict[str, str]:
    if not os.path.isfile(pszPriorPath):
        return {}
    objPriorRows = read_tsv_rows(pszPriorPath)
    if not objPriorRows:
        return {}
    objPriorMap: Dict[str, str] = {}
    for objRow in objPriorRows[1:]:
        if not objRow:
            continue
        pszLabel = objRow[0]
        pszValue = objRow[1] if len(objRow) > 1 else ""
        objPriorMap[pszLabel] = pszValue
    return objPriorMap


def build_prior_range_for_cumulative(
    objStart: Tuple[int, int],
    objEnd: Tuple[int, int],
) -> Optional[Tuple[Tuple[int, int], Tuple[int, int]]]:
    iStartYear, iStartMonth = objStart
    iEndYear, iEndMonth = objEnd
    if iStartYear <= 0 or iEndYear <= 0:
        return None
    return (iStartYear - 1, iStartMonth), (iEndYear - 1, iEndMonth)


def build_cp_group_step0008_vertical(
    pszDirectory: str,
    pszPeriodLabel: str,
    pszTimeLabel: str,
) -> Optional[str]:
    objGroupOrder: List[str] = [
        "受託事業-施設運営",
        "受託事業-その他",
        "自社-施設運営",
        "自社-その他",
        "合計",
    ]
    pszPrefix: str = (
        f"0002_CP別_step0007_{pszPeriodLabel}_損益計算書_{pszTimeLabel}_"
    )
    objRows: List[List[str]] = []
    for pszGroup in objGroupOrder:
        pszInputPath: str = os.path.join(
            pszDirectory,
            f"{pszPrefix}{pszGroup}_vertical.tsv",
        )
        if not os.path.isfile(pszInputPath):
            return None
        objRows.extend(read_tsv_rows(pszInputPath))

    pszOutputPath: str = os.path.join(
        pszDirectory,
        f"0002_CP別_step0008_{pszPeriodLabel}_損益計算書_{pszTimeLabel}_計上グループ_vertical.tsv",
    )
    write_tsv_rows(pszOutputPath, objRows)

    pszScriptDirectory: str = os.path.dirname(__file__)
    pszTargetDirectory: str = os.path.join(pszScriptDirectory, "0002_CP別_step0008")
    os.makedirs(pszTargetDirectory, exist_ok=True)
    pszTargetPath: str = os.path.join(pszTargetDirectory, os.path.basename(pszOutputPath))
    shutil.copy2(pszOutputPath, pszTargetPath)
    return pszOutputPath


def build_cp_company_step0008_vertical(
    pszDirectory: str,
    pszPeriodLabel: str,
    pszTimeLabel: str,
    pszOrgMode: str,
) -> Optional[str]:
    if pszOrgMode == "new":
        objCompanyOrder: List[str] = [
            "コンテンツビジネス",
            "スタートアップコミュニティ",
            "スタートアップグロース",
            "テクノロジーインキュベーション",
            "経営管理",
            "事業開発",
            "スタートアップサイド",
            "子会社",
            "投資先",
            "本部",
            "合計",
        ]
    else:
        objCompanyOrder = [
            "第一インキュ",
            "第二インキュ",
            "第三インキュ",
            "第四インキュ",
            "事業開発",
            "子会社",
            "投資先",
            "本部",
            "合計",
        ]
    pszPrefix: str = (
        f"0001_CP別_step0007_{pszPeriodLabel}_損益計算書_{pszTimeLabel}_"
    )
    objRows: List[List[str]] = []
    for pszCompany in objCompanyOrder:
        pszInputPath: str = os.path.join(
            pszDirectory,
            f"{pszPrefix}{pszCompany}_vertical.tsv",
        )
        if not os.path.isfile(pszInputPath):
            return None
        objRows.extend(read_tsv_rows(pszInputPath))

    pszOutputPath: str = os.path.join(
        pszDirectory,
        f"0001_CP別_step0008_{pszPeriodLabel}_損益計算書_{pszTimeLabel}_計上カンパニー_vertical.tsv",
    )
    write_tsv_rows(pszOutputPath, objRows)
    pszScriptDirectory: str = os.path.dirname(__file__)
    pszTargetDirectory: str = os.path.join(pszScriptDirectory, "0001_CP別_step0008")
    os.makedirs(pszTargetDirectory, exist_ok=True)
    pszTargetPath: str = os.path.join(pszTargetDirectory, os.path.basename(pszOutputPath))
    shutil.copy2(pszOutputPath, pszTargetPath)
    return pszOutputPath


def try_create_cp_company_step0008_vertical(
    pszStep0007Path: str,
    pszOrgMode: Optional[str] = None,
) -> Optional[str]:
    def resolve_org_mode(pszHint: Optional[str]) -> Optional[str]:
        if pszHint in ("legacy", "new"):
            return pszHint
        if EXECUTION_ROOT_DIRECTORY is None:
            return None
        pszModePath: str = os.path.join(
            EXECUTION_ROOT_DIRECTORY,
            "company_or_division.txt",
        )
        if not os.path.isfile(pszModePath):
            return None
        try:
            with open(pszModePath, "r", encoding="utf-8", newline="") as objModeFile:
                pszModeLabel: str = objModeFile.read().strip().lower()
        except OSError:
            return None
        if pszModeLabel == "company":
            return "legacy"
        if pszModeLabel == "division":
            return "new"
        return None

    pszResolvedOrgMode = resolve_org_mode(pszOrgMode)
    if pszResolvedOrgMode is None:
        return None

    pszFileName = os.path.basename(pszStep0007Path)
    objMatch = re.match(
        r"0001_CP別_step0007_(単月|累計)_損益計算書_(.+?)_(.+)_vertical\.tsv",
        pszFileName,
    )
    if objMatch is None:
        return None
    pszPeriodLabel: str = objMatch.group(1)
    pszTimeLabel: str = objMatch.group(2)
    pszCompanyLabel: str = objMatch.group(3)
    if pszResolvedOrgMode == "new":
        objAllowedCompanies: set[str] = {
            "テクノロジーインキュベーション",
            "コンテンツビジネス",
            "スタートアップサイド",
            "スタートアップコミュニティ",
            "スタートアップグロース",
            "経営管理",
            "事業開発",
            "子会社",
            "投資先",
            "本部",
            "合計",
        }
    else:
        objAllowedCompanies = {
            "第一インキュ",
            "第二インキュ",
            "第三インキュ",
            "第四インキュ",
            "事業開発",
            "子会社",
            "投資先",
            "本部",
            "合計",
        }
    if pszCompanyLabel not in objAllowedCompanies:
        return None
    pszDirectory: str = os.path.dirname(pszStep0007Path)
    return build_cp_company_step0008_vertical(
        pszDirectory,
        pszPeriodLabel,
        pszTimeLabel,
        pszResolvedOrgMode,
    )


def try_create_cp_group_step0008_vertical(pszStep0007Path: str) -> Optional[str]:
    pszFileName = os.path.basename(pszStep0007Path)
    objMatch = re.match(
        r"0002_CP別_step0007_(単月|累計)_損益計算書_(.+?)_(.+)_vertical\.tsv",
        pszFileName,
    )
    if objMatch is None:
        return None
    pszPeriodLabel: str = objMatch.group(1)
    pszTimeLabel: str = objMatch.group(2)
    pszGroupLabel: str = objMatch.group(3)
    objAllowedGroups: set[str] = {
        "受託事業-施設運営",
        "受託事業-その他",
        "自社-施設運営",
        "自社-その他",
        "合計",
    }
    if pszGroupLabel not in objAllowedGroups:
        return None
    pszDirectory: str = os.path.dirname(pszStep0007Path)
    return build_cp_group_step0008_vertical(
        pszDirectory,
        pszPeriodLabel,
        pszTimeLabel,
    )


def build_cp_company_step0008_single_path(
    pszDirectory: str,
    objMonth: Tuple[int, int],
    pszPrefix: str,
) -> str:
    iYear, iMonth = objMonth
    pszMonth = f"{iMonth:02d}"
    return os.path.join(
        pszDirectory,
        f"{pszPrefix}_CP別_step0008_単月_損益計算書_{iYear}年{pszMonth}月_計上カンパニー_vertical.tsv",
    )


def build_cp_group_step0008_single_path(
    pszDirectory: str,
    objMonth: Tuple[int, int],
    pszPrefix: str,
) -> str:
    iYear, iMonth = objMonth
    pszMonth = f"{iMonth:02d}"
    return os.path.join(
        pszDirectory,
        f"{pszPrefix}_CP別_step0008_単月_損益計算書_{iYear}年{pszMonth}月_計上グループ_vertical.tsv",
    )


def build_cp_company_step0008_cumulative_path(
    pszDirectory: str,
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
    pszPrefix: str,
) -> str:
    (iStartYear, iStartMonth), (iEndYear, iEndMonth) = objRange
    pszStartMonth = f"{iStartMonth:02d}"
    pszEndMonth = f"{iEndMonth:02d}"
    return os.path.join(
        pszDirectory,
        (
            f"{pszPrefix}_CP別_step0008_累計_損益計算書_"
            f"{iStartYear}年{pszStartMonth}月-{iEndYear}年{pszEndMonth}月_"
            "計上カンパニー_vertical.tsv"
        ),
    )


def build_cp_group_step0008_cumulative_path(
    pszDirectory: str,
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
    pszPrefix: str,
) -> str:
    (iStartYear, iStartMonth), (iEndYear, iEndMonth) = objRange
    pszStartMonth = f"{iStartMonth:02d}"
    pszEndMonth = f"{iEndMonth:02d}"
    return os.path.join(
        pszDirectory,
        (
            f"{pszPrefix}_CP別_step0008_累計_損益計算書_"
            f"{iStartYear}年{pszStartMonth}月-{iEndYear}年{pszEndMonth}月_"
            "計上グループ_vertical.tsv"
        ),
    )


def build_cp_company_step0009_cumulative_path(
    pszDirectory: str,
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
) -> str:
    (iStartYear, iStartMonth), (iEndYear, iEndMonth) = objRange
    pszStartMonth = f"{iStartMonth:02d}"
    pszEndMonth = f"{iEndMonth:02d}"
    return os.path.join(
        pszDirectory,
        (
            "0001_CP別_step0009_累計_損益計算書_"
            f"{iStartYear}年{pszStartMonth}月-{iEndYear}年{pszEndMonth}月_"
            "計上カンパニー_vertical.tsv"
        ),
    )


def build_cp_company_step0009_single_path(
    pszDirectory: str,
    objMonth: Tuple[int, int],
) -> str:
    iYear, iMonth = objMonth
    pszMonth = f"{iMonth:02d}"
    return os.path.join(
        pszDirectory,
        f"0001_CP別_step0009_単月_損益計算書_{iYear}年{pszMonth}月_計上カンパニー_vertical.tsv",
    )


def find_cp_company_step0009_vertical_paths(
    pszDirectory: str,
) -> List[Tuple[str, str]]:
    objMatches: List[Tuple[str, str]] = []
    for pszFileName in os.listdir(pszDirectory):
        objMatch = re.match(
            (
                r"0001_CP別_step0009_累計_損益計算書_"
                r"(\d{4}年\d{2}月-\d{4}年\d{2}月)_計上カンパニー_vertical\.tsv"
            ),
            pszFileName,
        )
        if objMatch is None:
            continue
        pszPeriodLabel: str = objMatch.group(1)
        objMatches.append((pszPeriodLabel, os.path.join(pszDirectory, pszFileName)))
    objMatches.sort(key=lambda objItem: objItem[0])
    return objMatches


def build_cp_group_step0009_cumulative_path(
    pszDirectory: str,
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
) -> str:
    (iStartYear, iStartMonth), (iEndYear, iEndMonth) = objRange
    pszStartMonth = f"{iStartMonth:02d}"
    pszEndMonth = f"{iEndMonth:02d}"
    return os.path.join(
        pszDirectory,
        (
            "0002_CP別_step0009_累計_損益計算書_"
            f"{iStartYear}年{pszStartMonth}月-{iEndYear}年{pszEndMonth}月_"
            "計上グループ_vertical.tsv"
        ),
    )


def find_cp_group_step0009_vertical_paths(
    pszDirectory: str,
) -> List[Tuple[str, str]]:
    objMatches: List[Tuple[str, str]] = []
    for pszFileName in os.listdir(pszDirectory):
        objMatch = re.match(
            (
                r"0002_CP別_step0009_累計_損益計算書_"
                r"(\d{4}年\d{2}月-\d{4}年\d{2}月)_計上グループ_vertical\.tsv"
            ),
            pszFileName,
        )
        if objMatch is None:
            continue
        pszPeriodLabel: str = objMatch.group(1)
        objMatches.append((pszPeriodLabel, os.path.join(pszDirectory, pszFileName)))
    objMatches.sort(key=lambda objItem: objItem[0])
    return objMatches


def parse_tsv_value_for_excel(pszValue: str) -> Optional[object]:
    pszText: str = (pszValue or "").strip()
    if pszText == "":
        return None
    if pszText == "'－":
        return "－"
    if pszText == "'－∞":
        return "－∞"
    if pszText == "'＋∞":
        return "＋∞"
    pszNormalized = pszText
    if pszNormalized.startswith("'"):
        pszNormalized = pszNormalized[1:]
    pszNormalized = pszNormalized.replace("－", "-").replace("＋", "+")
    pszNormalized = pszNormalized.replace(",", "")
    if re.fullmatch(r"[+-]?\d+", pszNormalized):
        return int(pszNormalized)
    if re.fullmatch(r"[+-]?\d+\.\d+", pszNormalized):
        return float(pszNormalized)
    return pszText


def create_cp_company_step0009_excel(pszScriptDirectory: str) -> Optional[str]:
    def read_company_or_division_mode_label() -> Optional[str]:
        if not EXECUTION_ROOT_DIRECTORY:
            return None
        pszModePath: str = os.path.join(
            EXECUTION_ROOT_DIRECTORY,
            "company_or_division.txt",
        )
        if not os.path.isfile(pszModePath):
            return None
        try:
            with open(pszModePath, "r", encoding="utf-8", newline="") as objModeFile:
                pszModeLabel: str = objModeFile.read().strip().lower()
        except OSError:
            return None
        if pszModeLabel in ("company", "division"):
            return pszModeLabel
        return None

    pszTargetDirectory: str = os.path.join(pszScriptDirectory, "0001_CP別_step0009")
    if not os.path.isdir(pszTargetDirectory):
        return None
    objRangePath = find_selected_range_path(pszTargetDirectory)
    objSelectedRange = parse_selected_range(objRangePath) if objRangePath else None

    objTsvPaths: List[Tuple[str, str]] = []
    if objSelectedRange is not None:
        objRequiredPeriodRanges = build_cp_period_ranges_from_selected_range(objSelectedRange)
        _, objSelectedEnd = objSelectedRange
        objRequiredCurrentPeriodRanges: List[Tuple[Tuple[int, int], Tuple[int, int]]] = [
            objPeriodRange
            for objPeriodRange in objRequiredPeriodRanges
            if objPeriodRange[1] == objSelectedEnd
        ]
        objRequiredCheckRanges = objRequiredCurrentPeriodRanges or objRequiredPeriodRanges
        objMissingPeriodItems: List[Tuple[str, str]] = []
        for objPeriodRange in objRequiredCheckRanges:
            pszInputPath = build_cp_company_step0009_cumulative_path(pszTargetDirectory, objPeriodRange)
            (iStartYear, iStartMonth), (iEndYear, iEndMonth) = objPeriodRange
            pszPeriodLabel = f"{iStartYear}年{iStartMonth:02d}月-{iEndYear}年{iEndMonth:02d}月"
            if not os.path.isfile(pszInputPath):
                objMissingPeriodItems.append((pszPeriodLabel, os.path.basename(pszInputPath)))
                continue
            objTsvPaths.append((pszPeriodLabel, pszInputPath))
        (iStartYear, iStartMonth), (iEndYear, iEndMonth) = objSelectedRange
        pszSelectedStartLabel = f"{iStartYear}年{iStartMonth:02d}月"
        pszSelectedEndLabel = f"{iEndYear}年{iEndMonth:02d}月"
        objExistingTsvPaths = find_cp_company_step0009_vertical_paths(pszTargetDirectory)
        pszMissingReportPath: str = os.path.join(
            pszScriptDirectory,
            "0001_CP別_step0009_不足期間一覧.txt",
        )
        objLines: List[str] = []
        objLines.append(f"生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        objLines.append(f"選択範囲: {pszSelectedStartLabel}〜{pszSelectedEndLabel}")
        objLines.append("")
        objLines.append("必要期間一覧:")
        for objRangeItem in objRequiredCheckRanges:
            (iReqStartYear, iReqStartMonth), (iReqEndYear, iReqEndMonth) = objRangeItem
            pszRequiredPeriodLabel = (
                f"{iReqStartYear}年{iReqStartMonth:02d}月-"
                f"{iReqEndYear}年{iReqEndMonth:02d}月"
            )
            objLines.append(f"- {pszRequiredPeriodLabel}")
        objLines.append("")
        objLines.append("存在しているTSV一覧:")
        if objExistingTsvPaths:
            for _, pszExistingPath in objExistingTsvPaths:
                objLines.append(f"- {os.path.basename(pszExistingPath)}")
        else:
            objLines.append("- なし")
        objLines.append("")
        objLines.append("不足期間一覧:")
        if objMissingPeriodItems:
            for pszMissingPeriodLabel, pszMissingFileName in objMissingPeriodItems:
                objLines.append(f"- 期間ラベル: {pszMissingPeriodLabel}")
                objLines.append(f"  想定ファイル名（フル名）: {pszMissingFileName}")
            objLines.append("")
            objLines.append("判定結果: Excel未作成: 必要TSV不足")
        else:
            objLines.append("- 不足なし。")
            objLines.append("")
            objLines.append("判定結果: 必要TSV不足なし")
        with open(pszMissingReportPath, "w", encoding="utf-8", newline="\n") as objReportFile:
            objReportFile.write("\n".join(objLines) + "\n")
        if objMissingPeriodItems:
            return None
    else:
        objTsvPaths = find_cp_company_step0009_vertical_paths(pszTargetDirectory)
        if not objTsvPaths:
            return None

    pszModeLabel = read_company_or_division_mode_label()
    if pszModeLabel == "division":
        pszTemplateFileName = "TEMPLATE_CP別経営管理_計上div_累計.xlsx"
        pszOutputFileNamePrefix = "CP別経営管理_計上div_累計"
    else:
        pszTemplateFileName = "TEMPLATE_CP別経営管理_計上カンパニー_累計.xlsx"
        pszOutputFileNamePrefix = "CP別経営管理_計上カンパニー_累計"
    pszTemplatePath: str = os.path.join(
        pszScriptDirectory,
        pszTemplateFileName,
    )
    if not os.path.isfile(pszTemplatePath):
        return None

    objWorkbook = load_workbook(pszTemplatePath)
    objTemplateSheet = objWorkbook.worksheets[0]
    for objSheetToRemove in objWorkbook.worksheets[1:]:
        objWorkbook.remove(objSheetToRemove)
    for pszPeriodLabel, pszInputPath in objTsvPaths:
        pszSheetTitle = (
            f"経営管理_計上div_{pszPeriodLabel}"
            if pszModeLabel == "division"
            else pszPeriodLabel
        )
        if pszSheetTitle in objWorkbook.sheetnames:
            objWorkbook.remove(objWorkbook[pszSheetTitle])
        objSheet = objWorkbook.copy_worksheet(objTemplateSheet)
        objSheet.title = pszSheetTitle
        _clear_sheet_values(objSheet)
        objRows = read_tsv_rows(pszInputPath)
        for iRowIndex, objRow in enumerate(objRows, start=1):
            for iColumnIndex, pszValue in enumerate(objRow, start=1):
                objCellValue = parse_tsv_value_for_excel(pszValue)
                objSheet.cell(
                    row=iRowIndex,
                    column=iColumnIndex,
                    value=objCellValue,
                )
    if objTemplateSheet in objWorkbook.worksheets:
        objWorkbook.remove(objTemplateSheet)

    if objSelectedRange is not None:
        (iStartYear, iStartMonth), (iEndYear, iEndMonth) = objSelectedRange
        pszStartLabel = f"{iStartYear}年{iStartMonth:02d}月"
        pszEndLabel = f"{iEndYear}年{iEndMonth:02d}月"
        pszOutputFileName = (
            f"{pszOutputFileNamePrefix}_{pszStartLabel}-{pszEndLabel}.xlsx"
        )
    else:
        pszOutputFileName = f"{pszOutputFileNamePrefix}.xlsx"
    pszOutputPath: str = os.path.join(
        pszTargetDirectory,
        pszOutputFileName,
    )
    objWorkbook.save(pszOutputPath)
    return pszOutputPath


def create_cp_group_step0009_excel(pszScriptDirectory: str) -> Optional[str]:
    pszTargetDirectory: str = os.path.join(pszScriptDirectory, "0002_CP別_step0009")
    if not os.path.isdir(pszTargetDirectory):
        return None
    objRangePath = find_selected_range_path(pszTargetDirectory)
    objSelectedRange = parse_selected_range(objRangePath) if objRangePath else None

    objTsvPaths: List[Tuple[str, str]] = []
    if objSelectedRange is not None:
        for objPeriodRange in build_cp_period_ranges_from_selected_range(objSelectedRange):
            pszInputPath = build_cp_group_step0009_cumulative_path(pszTargetDirectory, objPeriodRange)
            if not os.path.isfile(pszInputPath):
                return None
            (iStartYear, iStartMonth), (iEndYear, iEndMonth) = objPeriodRange
            pszPeriodLabel = f"{iStartYear}年{iStartMonth:02d}月-{iEndYear}年{iEndMonth:02d}月"
            objTsvPaths.append((pszPeriodLabel, pszInputPath))
    else:
        objTsvPaths = find_cp_group_step0009_vertical_paths(pszTargetDirectory)
        if not objTsvPaths:
            return None

    pszTemplatePath: str = os.path.join(
        pszScriptDirectory,
        "TEMPLATE_CP別経営管理_計上グループ_累計.xlsx",
    )
    if not os.path.isfile(pszTemplatePath):
        return None

    objWorkbook = load_workbook(pszTemplatePath)
    objTemplateSheet = objWorkbook.worksheets[0]
    for pszPeriodLabel, pszInputPath in objTsvPaths:
        pszSheetTitle = f"経営管理_計上グループ_{pszPeriodLabel}"
        if pszSheetTitle in objWorkbook.sheetnames:
            objWorkbook.remove(objWorkbook[pszSheetTitle])
        objSheet = objWorkbook.copy_worksheet(objTemplateSheet)
        objSheet.title = pszSheetTitle
        objRows = read_tsv_rows(pszInputPath)
        for iRowIndex, objRow in enumerate(objRows, start=1):
            for iColumnIndex, pszValue in enumerate(objRow, start=1):
                objCellValue = parse_tsv_value_for_excel(pszValue)
                objSheet.cell(
                    row=iRowIndex,
                    column=iColumnIndex,
                    value=objCellValue,
                )
    if objTemplateSheet in objWorkbook.worksheets:
        objWorkbook.remove(objTemplateSheet)

    if objSelectedRange is not None:
        (iStartYear, iStartMonth), (iEndYear, iEndMonth) = objSelectedRange
        pszStartLabel = f"{iStartYear}年{iStartMonth:02d}月"
        pszEndLabel = f"{iEndYear}年{iEndMonth:02d}月"
        pszOutputFileName = (
            f"CP別経営管理_計上グループ_累計_{pszStartLabel}-{pszEndLabel}.xlsx"
        )
    else:
        pszOutputFileName = "CP別経営管理_計上グループ_累計.xlsx"
    pszOutputPath: str = os.path.join(
        pszTargetDirectory,
        pszOutputFileName,
    )
    objWorkbook.save(pszOutputPath)
    return pszOutputPath


def append_vertical_columns(
    objBaseRows: List[List[str]],
    objAppendRows: List[List[str]],
) -> List[List[str]]:
    if not objBaseRows:
        return [list(objRow) for objRow in objAppendRows]
    iAppendColumnCount: int = max((len(objRow) for objRow in objAppendRows), default=1) - 1
    if iAppendColumnCount < 0:
        iAppendColumnCount = 0

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objBaseRows):
        objAppendRow: List[str] = (
            objAppendRows[iRowIndex] if iRowIndex < len(objAppendRows) else []
        )
        objAppendValues = list(objAppendRow[1:]) if objAppendRow else []
        if len(objAppendValues) < iAppendColumnCount:
            objAppendValues.extend([""] * (iAppendColumnCount - len(objAppendValues)))
        if not objRow:
            objOutputRows.append([""] * (len(objRow) + iAppendColumnCount))
            continue
        objOutputRows.append(list(objRow) + objAppendValues)
    return objOutputRows


def build_cp_company_step0009_single_for_month(
    pszDirectory: str,
    objMonth: Tuple[int, int],
) -> Optional[str]:
    pszInputPath = build_cp_company_step0008_single_path(
        pszDirectory,
        objMonth,
        "0001",
    )
    if not os.path.isfile(pszInputPath):
        return None
    objRows = read_tsv_rows(pszInputPath)
    pszOutputPath = build_cp_company_step0009_single_path(pszDirectory, objMonth)
    write_tsv_rows(pszOutputPath, objRows)
    return pszOutputPath


def build_cp_step0009_vertical_for_range(
    pszDirectory: str,
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
) -> Optional[str]:
    pszCumulativePath = build_cp_company_step0008_cumulative_path(
        pszDirectory,
        objRange,
        "0001",
    )
    if not os.path.isfile(pszCumulativePath):
        return None
    objOutputRows: List[List[str]] = read_tsv_rows(pszCumulativePath)
    objMonths = build_month_sequence(objRange[0], objRange[1])
    for objMonth in objMonths:
        pszSinglePath = build_cp_company_step0008_single_path(
            pszDirectory,
            objMonth,
            "0001",
        )
        if not os.path.isfile(pszSinglePath):
            return None
        objSingleRows = read_tsv_rows(pszSinglePath)
        objOutputRows = append_vertical_columns(objOutputRows, objSingleRows)

    pszOutputPath = build_cp_company_step0009_cumulative_path(pszDirectory, objRange)
    write_tsv_rows(pszOutputPath, objOutputRows)

    pszScriptDirectory: str = os.path.dirname(__file__)
    pszTargetDirectory: str = os.path.join(pszScriptDirectory, "0001_CP別_step0009")
    os.makedirs(pszTargetDirectory, exist_ok=True)
    pszTargetPath: str = os.path.join(pszTargetDirectory, os.path.basename(pszOutputPath))
    shutil.copy2(pszOutputPath, pszTargetPath)
    return pszOutputPath


def build_cp_group_step0009_vertical_for_range(
    pszDirectory: str,
    objRange: Tuple[Tuple[int, int], Tuple[int, int]],
) -> Optional[str]:
    pszCumulativePath = build_cp_group_step0008_cumulative_path(
        pszDirectory,
        objRange,
        "0002",
    )
    if not os.path.isfile(pszCumulativePath):
        return None
    objOutputRows: List[List[str]] = read_tsv_rows(pszCumulativePath)
    objMonths = build_month_sequence(objRange[0], objRange[1])
    for objMonth in objMonths:
        pszSinglePath = build_cp_group_step0008_single_path(
            pszDirectory,
            objMonth,
            "0002",
        )
        if not os.path.isfile(pszSinglePath):
            return None
        objSingleRows = read_tsv_rows(pszSinglePath)
        objOutputRows = append_vertical_columns(objOutputRows, objSingleRows)

    pszOutputPath = build_cp_group_step0009_cumulative_path(pszDirectory, objRange)
    write_tsv_rows(pszOutputPath, objOutputRows)

    pszScriptDirectory: str = os.path.dirname(__file__)
    pszTargetDirectory: str = os.path.join(pszScriptDirectory, "0002_CP別_step0009")
    os.makedirs(pszTargetDirectory, exist_ok=True)
    pszTargetPath: str = os.path.join(pszTargetDirectory, os.path.basename(pszOutputPath))
    shutil.copy2(pszOutputPath, pszTargetPath)
    return pszOutputPath


def copy_cp_management_excels(
    pszCompanyPath: Optional[str],
    pszGroupPath: Optional[str],
) -> None:
    if not pszCompanyPath or not pszGroupPath:
        return
    if not os.path.isfile(pszCompanyPath) or not os.path.isfile(pszGroupPath):
        return
    if not EXECUTION_ROOT_DIRECTORY:
        return
    pszCompanyDirectory: str = os.path.join(
        EXECUTION_ROOT_DIRECTORY,
        "CP別経営管理表_計上カンパニー",
    )
    pszGroupDirectory: str = os.path.join(
        EXECUTION_ROOT_DIRECTORY,
        "CP別経営管理表_計上グループ",
    )
    os.makedirs(pszCompanyDirectory, exist_ok=True)
    os.makedirs(pszGroupDirectory, exist_ok=True)
    pszCompanyTargetPath: str = os.path.join(
        pszCompanyDirectory,
        os.path.basename(pszCompanyPath),
    )
    pszGroupTargetPath: str = os.path.join(
        pszGroupDirectory,
        os.path.basename(pszGroupPath),
    )
    shutil.copy2(pszCompanyPath, pszCompanyTargetPath)
    shutil.copy2(pszGroupPath, pszGroupTargetPath)


def try_create_cp_step0009_vertical(pszDirectory: str) -> Optional[str]:
    pszRangePath: Optional[str] = find_selected_range_path(pszDirectory)
    if pszRangePath is None:
        return None
    objRange = parse_selected_range(pszRangePath)
    if objRange is None:
        return None

    objTargetRanges: List[Tuple[Tuple[int, int], Tuple[int, int]]] = build_cp_period_ranges_from_selected_range(objRange)

    for objRangeItem in objTargetRanges:
        build_cp_step0009_vertical_for_range(pszDirectory, objRangeItem)
    return create_cp_company_step0009_excel(os.path.dirname(__file__))


def try_create_cp_group_step0009_vertical(pszDirectory: str) -> Optional[str]:
    pszRangePath: Optional[str] = find_selected_range_path(pszDirectory)
    if pszRangePath is None:
        return None
    objRange = parse_selected_range(pszRangePath)
    if objRange is None:
        return None

    objTargetRanges: List[Tuple[Tuple[int, int], Tuple[int, int]]] = build_cp_period_ranges_from_selected_range(objRange)

    for objRangeItem in objTargetRanges:
        pszCumulativePath = build_cp_group_step0008_cumulative_path(
            pszDirectory,
            objRangeItem,
            "0002",
        )
        if not os.path.isfile(pszCumulativePath):
            return None

    for objRangeItem in objTargetRanges:
        build_cp_group_step0009_vertical_for_range(pszDirectory, objRangeItem)
    return create_cp_group_step0009_excel(os.path.dirname(__file__))


def create_cp_step0007_file_company(pszStep0006Path: str, pszPrefix: str) -> None:
    pszFileName = os.path.basename(pszStep0006Path)
    pszDirectory = os.path.dirname(pszStep0006Path)
    objSingleMatch = re.match(
        rf"{pszPrefix}_step0006_単月_損益計算書_(\d{{4}})年(\d{{2}})月_(.+)_vertical\.tsv",
        pszFileName,
    )
    objCumulativeMatch = re.match(
        rf"{pszPrefix}_step0006_累計_損益計算書_(\d{{4}})年(\d{{2}})月-(\d{{4}})年(\d{{2}})月_(.+)_vertical\.tsv",
        pszFileName,
    )
    pszPriorLabel = ""
    pszCurrentLabel = ""
    pszPriorPath = ""
    pszOutputPath = ""
    pszPriorRowLabel = "前年"
    if objSingleMatch:
        iYear = int(objSingleMatch.group(1))
        iMonth = int(objSingleMatch.group(2))
        pszCompany = objSingleMatch.group(3)
        pszPriorLabel = f"{iYear - 1}年{iMonth:02d}月"
        pszCurrentLabel = f"{iYear}年{iMonth:02d}月"
        iPriorYear = iYear - 1
        pszPriorPath = os.path.join(
            pszDirectory,
            (
                f"{pszPrefix}_step0006_単月_損益計算書_"
                f"{iPriorYear}年{iMonth:02d}月_{pszCompany}_vertical.tsv"
            ),
        )
        pszOutputPath = os.path.join(
            pszDirectory,
            (
                f"{pszPrefix}_step0007_単月_損益計算書_"
                f"{iYear}年{iMonth:02d}月_{pszCompany}_vertical.tsv"
            ),
        )
    elif objCumulativeMatch:
        iStartYear = int(objCumulativeMatch.group(1))
        iStartMonth = int(objCumulativeMatch.group(2))
        iEndYear = int(objCumulativeMatch.group(3))
        iEndMonth = int(objCumulativeMatch.group(4))
        pszCompany = objCumulativeMatch.group(5)
        pszCurrentLabel = (
            f"{iStartYear}年{iStartMonth:02d}月-"
            f"{iEndYear}年{iEndMonth:02d}月"
        )
        objPriorRange = build_prior_range_for_cumulative(
            (iStartYear, iStartMonth),
            (iEndYear, iEndMonth),
        )
        if objPriorRange is not None:
            (iPriorStartYear, iPriorStartMonth), (iPriorEndYear, iPriorEndMonth) = objPriorRange
            pszPriorLabel = (
                f"{iPriorStartYear}年{iPriorStartMonth:02d}月-"
                f"{iPriorEndYear}年{iPriorEndMonth:02d}月"
            )
            pszPriorPath = os.path.join(
                pszDirectory,
                (
                    f"{pszPrefix}_step0006_累計_損益計算書_"
                    f"{pszPriorLabel}_{pszCompany}_vertical.tsv"
                ),
            )
        pszPriorRowLabel = "前期"
        pszOutputPath = os.path.join(
            pszDirectory,
            (
                f"{pszPrefix}_step0007_累計_損益計算書_"
                f"{pszCurrentLabel}_{pszCompany}_vertical.tsv"
            ),
        )
    else:
        return

    objRows = read_tsv_rows(pszStep0006Path)
    if not objRows:
        return
    objRows = reorder_cp_step0006_rows(objRows)
    objPriorMap = build_step0006_prior_map(pszPriorPath)
    objOutputRows = build_step0007_rows_for_cp(
        objRows,
        objPriorMap,
        pszPriorLabel,
        pszCurrentLabel,
        pszPriorRowLabel,
        pszPrefix,
    )
    write_tsv_rows(pszOutputPath, objOutputRows)
    pszTargetDirectory = os.path.join(get_script_base_directory(), f"{pszPrefix}_step0007")
    os.makedirs(pszTargetDirectory, exist_ok=True)
    pszTargetPath = os.path.join(pszTargetDirectory, os.path.basename(pszOutputPath))
    shutil.copy2(pszOutputPath, pszTargetPath)


def create_cp_step0007_file_0001(
    pszStep0006Path: str,
    pszOrgMode: Optional[str] = None,
) -> None:
    create_cp_step0007_file_company(pszStep0006Path, "0001_CP別")
    pszOutputPath = os.path.join(
        get_script_base_directory(),
        os.path.basename(pszStep0006Path).replace("_step0006_", "_step0007_"),
    )
    if os.path.isfile(pszOutputPath):
        try_create_cp_company_step0008_vertical(
            pszOutputPath,
            pszOrgMode=pszOrgMode,
        )


def create_cp_step0007_file_0002(pszStep0006Path: str) -> None:
    create_cp_step0007_file_company(pszStep0006Path, "0002_CP別")
    pszTargetDirectory = os.path.join(os.path.dirname(__file__), "0002_CP別_step0007")
    os.makedirs(pszTargetDirectory, exist_ok=True)
    pszOutputPath = os.path.join(
        os.path.dirname(pszStep0006Path),
        os.path.basename(pszStep0006Path).replace("_step0006_", "_step0007_"),
    )
    if os.path.isfile(pszOutputPath):
        pszTargetPath = os.path.join(pszTargetDirectory, os.path.basename(pszOutputPath))
        shutil.copy2(pszOutputPath, pszTargetPath)
        try_create_cp_group_step0008_vertical(pszOutputPath)
        try_create_cp_group_step0008_vertical(pszTargetPath)


def create_empty_previous_fiscal_cp_step0005_vertical(
    pszDirectory: str,
    objStart: Tuple[int, int],
    objEnd: Tuple[int, int],
    pszCpPrefix: str,
) -> Optional[str]:
    iStartYear, iStartMonth = objStart
    iEndYear, iEndMonth = objEnd
    if iStartMonth == 4:
        iPriorStartYear = iStartYear - 1
        iPriorStartMonth = 4
        iPriorEndYear = iStartYear
        iPriorEndMonth = 3
    elif iStartMonth == 9:
        iPriorStartYear = iStartYear - 1
        iPriorStartMonth = 9
        iPriorEndYear = iStartYear
        iPriorEndMonth = 8
    else:
        return None

    pszTemplatePath: str = os.path.join(
        pszDirectory,
        (
            f"{pszCpPrefix}_step0005_累計_損益計算書_"
            f"{iStartYear}年{iStartMonth:02d}月-"
            f"{iEndYear}年{iEndMonth:02d}月_vertical.tsv"
        ),
    )
    pszTargetPath: str = os.path.join(
        pszDirectory,
        (
            f"{pszCpPrefix}_step0005_累計_損益計算書_"
            f"{iPriorStartYear}年{iPriorStartMonth:02d}月-"
            f"{iPriorEndYear}年{iPriorEndMonth:02d}月_vertical.tsv"
        ),
    )
    if not os.path.isfile(pszTemplatePath):
        return None
    if os.path.isfile(pszTargetPath):
        return pszTargetPath

    objRows = read_tsv_rows(pszTemplatePath)
    if not objRows:
        return None
    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objRows):
        if iRowIndex == 0:
            objOutputRows.append(list(objRow))
            continue
        if not objRow:
            objOutputRows.append([])
            continue
        pszLabel: str = objRow[0]
        if pszLabel in ("売上総利益率", "営業利益率"):
            objOutputRows.append([pszLabel] + ["0.0"] * (len(objRow) - 1))
        else:
            objOutputRows.append([pszLabel] + ["0"] * (len(objRow) - 1))
    write_tsv_rows(pszTargetPath, objOutputRows)
    return pszTargetPath


def main(argv: list[str]) -> int:
    if len(argv) < 3:
        print_usage()
        return 1

    objCsvInputs: List[str] = [pszPath for pszPath in argv[1:] if pszPath.lower().endswith(".csv")]
    objTsvInputs: List[str] = [pszPath for pszPath in argv[1:] if pszPath.lower().endswith(".tsv")]

    if objCsvInputs and objTsvInputs:
        print(
            "Error: CSV と TSV を混在させて実行することはできません。"
            " CSV は CSV だけでドラッグ＆ドロップしてください。"
            " TSV は TSV だけでドラッグ＆ドロップしてください。"
        )
        return 1

    if objCsvInputs and not objTsvInputs:
        print(
            "Error: 本スクリプトは TSV 専用です。CSV は CSV だけでドラッグ＆ドロップしてください。"
            " TSV を扱う場合は TSV のみを指定してください。"
        )
        print_usage()
        return 1

    objArgv: list[str] = [argv[0]] + (objTsvInputs if objTsvInputs else argv[1:])

    if len(objArgv) < 3:
        print_usage()
        return 1

    create_execution_folders()

    if len(objArgv) == 4:
        objPairs: List[List[str]] = [[objArgv[1], objArgv[2], objArgv[3]]]
    else:
        iArgCount: int = len(objArgv) - 1
        if iArgCount % 2 != 0:
            print_usage()
            return 1
        objPairs = []
        objManhourCandidates: List[str] = []
        objPlCandidates: List[str] = []
        bGroupedOrder: bool = True
        bSeenPl: bool = False
        for pszCandidate in objArgv[1:]:
            pszBaseName: str = os.path.basename(pszCandidate)
            if pszBaseName.startswith("工数_"):
                if bSeenPl:
                    bGroupedOrder = False
                objManhourCandidates.append(pszCandidate)
                continue
            if pszBaseName.startswith("損益計算書_"):
                bSeenPl = True
                objPlCandidates.append(pszCandidate)
                continue
            bGroupedOrder = False
            break

        bSplitByGroup: bool = (
            bGroupedOrder
            and len(objManhourCandidates) == len(objPlCandidates)
            and len(objManhourCandidates) + len(objPlCandidates) == iArgCount
        )

        if bSplitByGroup:
            for iIndex in range(len(objManhourCandidates)):
                objPairs.append([objManhourCandidates[iIndex], objPlCandidates[iIndex]])
        else:
            for iIndex in range(1, len(objArgv), 2):
                if iIndex + 1 >= len(objArgv):
                    print_usage()
                    return 1
                objPairs.append([objArgv[iIndex], objArgv[iIndex + 1]])

    objPairsWithMonths: List[Tuple[str, str, Optional[Tuple[int, int]]]] = []
    objParsedMonths: List[Tuple[int, int]] = []
    for objPair in objPairs:
        pszManhourPath: str = objPair[0]
        pszPlPath: str = objPair[1]
        objMonthPl: Optional[Tuple[int, int]] = extract_year_month_from_path(pszPlPath)
        objMonthManhour: Optional[Tuple[int, int]] = extract_year_month_from_path(pszManhourPath)
        objMonth: Optional[Tuple[int, int]] = objMonthPl if objMonthPl is not None else objMonthManhour
        if objMonthPl is not None and objMonthManhour is not None and objMonthPl != objMonthManhour:
            objMonth = None
        objPairsWithMonths.append((pszManhourPath, pszPlPath, objMonth))
        if objMonth is not None:
            objParsedMonths.append(objMonth)

    objSelectedRange: Optional[Tuple[Tuple[int, int], Tuple[int, int]]] = None
    if objParsedMonths:
        objSelectedRange = find_best_continuous_range(objParsedMonths)

    if objSelectedRange is None:
        pszBaseDirectory: str = os.path.dirname(objPairs[0][1]) if objPairs else os.getcwd()
        pszRangePath: Optional[str] = find_selected_range_path(pszBaseDirectory)
        if pszRangePath is not None:
            objSelectedRange = parse_selected_range(pszRangePath)
        if objSelectedRange is None:
            print("Error: 採用範囲を取得できませんでした。", file=sys.stderr)
            return 1

    objSelectedPairs: List[List[str]]
    if objParsedMonths:
        objSelectedPairs = [
            [pszManhourPath, pszPlPath]
            for pszManhourPath, pszPlPath, objMonth in objPairsWithMonths
            if objMonth is not None and is_month_in_range(objMonth, objSelectedRange)
        ]
        objSelectedPairs.sort(
            key=lambda objPair: extract_year_month_from_path(objPair[1])
            or extract_year_month_from_path(objPair[0])
            or (0, 0),
        )
    else:
        objSelectedPairs = [objPair[:2] for objPair in objPairsWithMonths]

    if not objSelectedPairs:
        print("Error: 採用範囲に合致する入力がありません。", file=sys.stderr)
        return 1

    pszRangeFileDirectory: str = get_script_base_directory()
    pszRangePathSelected: str = ensure_selected_range_file(pszRangeFileDirectory, objSelectedRange)
    record_created_file(pszRangePathSelected)

    objPairs = objSelectedPairs

    for objPair in objPairs:
        pszManhourPath: str = objPair[0]
        pszPlPath: str = objPair[1]
        pszOutputPath: str
        if len(objPair) == 3:
            pszOutputPath = objPair[2]
        else:
            pszOutputPath = build_default_output_path(pszPlPath)
        pszOutputFinalPath: str = build_output_path_with_step(pszPlPath, "販管費配賦_")
        pszOutputStep0001Path: str = build_output_path_with_step(pszPlPath, "販管費配賦_step0001_")
        pszOutputStep0002Path: str = build_output_path_with_step(pszPlPath, "販管費配賦_step0002_")
        pszOutputStep0003ZeroPath: str = build_output_path_with_step(pszPlPath, "販管費配賦_step0003_")
        pszOutputStep0003Path: str = build_output_path_with_step(pszPlPath, "販管費配賦_step0007_")
        pszOutputStep0004Path: str = build_output_path_with_step(pszPlPath, "販管費配賦_step0008_")
        pszOutputStep0009Path: str = build_output_path_with_step(pszPlPath, "販管費配賦_step0009_")
        pszOutputStep0005Path: str = build_output_path_with_step(pszPlPath, "販管費配賦_step0005_")
        pszOutputStep0006Path: str = build_output_path_with_step(pszPlPath, "販管費配賦_step0006_")
        pszOutputStep0010Path: str = build_output_path_with_step(pszPlPath, "販管費配賦_step0010_")

        if not os.path.exists(pszManhourPath):
            print(f"Input file not found: {pszManhourPath}")
            return 1
        if not os.path.exists(pszPlPath):
            print(f"Input file not found: {pszPlPath}")
            return 1

        objManhourMap: Dict[str, str] = load_manhour_map(pszManhourPath)
        objCompanyMap: Dict[str, str] = load_company_map(pszManhourPath)
        process_pl_tsv(
            pszPlPath,
            pszOutputPath,
            pszOutputStep0001Path,
            pszOutputStep0002Path,
            pszOutputStep0003ZeroPath,
            pszOutputStep0003Path,
            pszOutputStep0004Path,
            pszOutputStep0009Path,
            pszOutputStep0005Path,
            pszOutputStep0006Path,
            pszOutputStep0010Path,
            pszOutputFinalPath,
            objManhourMap,
            objCompanyMap,
        )

        print(f"Output: {pszOutputStep0001Path}")
        print(f"Output: {pszOutputStep0002Path}")
        print(f"Output: {pszOutputStep0003ZeroPath}")
        print(f"Output: {pszOutputStep0003Path}")
        print(f"Output: {pszOutputStep0004Path}")
        print(f"Output: {pszOutputStep0009Path}")
        print(f"Output: {pszOutputStep0005Path}")
        print(f"Output: {pszOutputStep0006Path}")
        print(f"Output: {pszOutputStep0010Path}")
        print(f"Output: {pszOutputFinalPath}")

    if objPairs:
        create_step0010_pj_income_statement_excels(get_script_base_directory())
        create_cumulative_reports(objPairs[0][1])
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
