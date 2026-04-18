from __future__ import annotations

import argparse
import calendar
import csv
import re
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation, ROUND_FLOOR, ROUND_HALF_UP
from pathlib import Path
from typing import List


INVALID_FILE_CHARS_PATTERN: re.Pattern[str] = re.compile(r'[\\/:*?"<>|]')
YEAR_MONTH_PATTERN: re.Pattern[str] = re.compile(r"(\d{2})\.(\d{1,2})月")
DURATION_TEXT_PATTERN: re.Pattern[str] = re.compile(r"^\s*(\d+)\s+day(?:s)?,\s*(\d+):(\d{2}):(\d{2})\s*$")
TIME_TEXT_PATTERN: re.Pattern[str] = re.compile(r"^\d+:\d{2}:\d{2}$")
SALARY_PAYMENT_STEP0001_FILE_PATTERN: re.Pattern[str] = re.compile(r"^支給・控除等一覧表_給与_step0001_.+\.tsv$")
NEW_RAWDATA_STEP0001_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0001_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0002_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0002_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0003_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0003_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0004_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0004_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0005_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0005_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0006_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0006_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0007_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0007_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0008_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0008_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0009_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0009_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0010_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0010_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0011_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0011_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0012_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0012_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0013_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0013_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0012_PREPAYED_COMMUTE_FILE_PATTERN: re.Pattern[str] = re.compile(
    r"^新_ローデータ_シート_step0012_\d{4}年(?:04-09月|10-03月)_\d{2}月_前払通勤交通費按分表\.tsv$"
)
NEW_RAWDATA_STEP0013_PREPAYED_COMMUTE_FILE_PATTERN: re.Pattern[str] = re.compile(
    r"^新_ローデータ_シート_step0013_\d{4}年(?:04-09月|10-03月)_\d{2}月_前払通勤交通費按分表\.tsv$"
)
NEW_RAWDATA_STEP0014_PREPAYED_COMMUTE_FILE_PATTERN: re.Pattern[str] = re.compile(
    r"^新_ローデータ_シート_step0014_\d{4}年(?:04-09月|10-03月)_\d{2}月_前払通勤交通費按分表\.tsv$"
)
NEW_RAWDATA_STEP0015_PREPAYED_COMMUTE_FILE_PATTERN: re.Pattern[str] = re.compile(
    r"^新_ローデータ_シート_step0015_\d{4}年(?:04-09月|10-03月)_\d{2}月_前払通勤交通費按分表\.tsv$"
)
NEW_RAWDATA_STEP0014_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0014_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0015_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0015_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0016_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0016_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0017_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0017_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0018_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0018_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0019_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0019_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0020_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0020_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0021_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0021_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0022_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0022_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0023_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0023_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0024_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0024_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0025_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0025_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0026_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0026_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0027_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0027_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0028_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0028_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0029_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0029_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0013_NONTAX_COMMUTE_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0013_非課税通勤手当_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0013_STATUTORY_WELFARE_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0013_法定福利費_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0014_STATUTORY_WELFARE_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0014_法定福利費_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0015_STATUTORY_WELFARE_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0015_法定福利費_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0016_STATUTORY_WELFARE_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0016_法定福利費_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0017_STATUTORY_WELFARE_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0017_法定福利費_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0018_STATUTORY_WELFARE_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0018_法定福利費_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0019_STATUTORY_WELFARE_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0019_法定福利費_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0020_STATUTORY_WELFARE_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0020_法定福利費_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0021_STATUTORY_WELFARE_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0021_法定福利費_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0022_STATUTORY_WELFARE_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0022_法定福利費_\d{4}年\d{2}月\.tsv$")
NEW_RAWDATA_STEP0023_STATUTORY_WELFARE_FILE_PATTERN: re.Pattern[str] = re.compile(r"^新_ローデータ_シート_step0023_法定福利費_\d{4}年\d{2}月\.tsv$")
PREPAYED_COMMUTE_STEP0006_MONTHLY_FILE_PATTERN: re.Pattern[str] = re.compile(
    r"^新_ローデータ_シート_step0006_(\d{4}年(?:04-09月|10-03月))_(\d{2})月_前払通勤交通費按分表\.tsv$"
)
PREPAYED_COMMUTE_STEP0023_FILE_PATTERN: re.Pattern[str] = re.compile(
    r"^新_ローデータ_シート_step0023_(\d{4})年(?:04-09月|10-03月)_(\d{2})月_前払通勤交通費按分表\.tsv$"
)
SALARY_PAYMENT_DEDUCTION_REQUIRED_HEADERS: tuple[str, ...] = (
    "従業員名",
    "スタッフコード",
    "基本給",
    "課税通勤手当",
    "非課税通勤手当",
    "残業手当",
    "残業時間(60時間以上)",
    "深夜労働手当",
    "休日労働手当",
    "固定残業代",
    "赴任手当",
    "テレワーク手当",
    "プロジェクトリーダー手当",
    "その他支給",
    "欠勤控除",
    "遅刻早退控除",
    "立替経費",
    "その他手当",
    "その他控除",
    "健保事業主負担",
    "介護事業主負担",
    "厚年事業主負担",
    "雇保事業主負担",
    "労災保険料",
    "一般拠出金",
    "子育拠出金",
)
MANAGEMENT_ACCOUNTING_MANHOUR_REQUIRED_HEADERS: tuple[str, ...] = (
    "日時",
    "スタッフコード",
    "姓 名",
    "所属グループ名",
    "スタッフ種別",
    "総労働時間",
    "プロジェクトコード",
    "プロジェクト名",
    "タスクコード",
    "タスク名",
    "工数",
)


def build_candidate_paths(pszInputPath: str) -> List[Path]:
    objInputPath: Path = Path(pszInputPath)
    objScriptDirectoryPath: Path = Path(__file__).resolve().parent
    objInputDirectoryPath: Path = Path.cwd() / "input"
    return [
        objInputPath,
        objScriptDirectoryPath / pszInputPath,
        objInputDirectoryPath / pszInputPath,
    ]


def resolve_existing_input_path(pszInputPath: str) -> Path:
    for objCandidatePath in build_candidate_paths(pszInputPath):
        if objCandidatePath.exists():
            return objCandidatePath
    raise FileNotFoundError(f"Input file not found: {pszInputPath}")


def sanitize_sheet_name_for_file_name(pszSheetName: str) -> str:
    pszSanitized: str = INVALID_FILE_CHARS_PATTERN.sub("_", pszSheetName).strip()
    if pszSanitized == "":
        return "Sheet"
    return pszSanitized


def build_unique_output_path(
    objBaseDirectoryPath: Path,
    pszExcelStem: str,
    pszSanitizedSheetName: str,
    objUsedPaths: set[Path],
) -> Path:
    objOutputPath: Path = objBaseDirectoryPath / f"{pszExcelStem}_{pszSanitizedSheetName}.tsv"
    if objOutputPath not in objUsedPaths:
        objUsedPaths.add(objOutputPath)
        return objOutputPath

    iSuffix: int = 2
    while True:
        objCandidatePath: Path = (
            objBaseDirectoryPath / f"{pszExcelStem}_{pszSanitizedSheetName}_{iSuffix}.tsv"
        )
        if objCandidatePath not in objUsedPaths:
            objUsedPaths.add(objCandidatePath)
            return objCandidatePath
        iSuffix += 1


def format_timedelta_as_h_mm_ss(objDuration: timedelta) -> str:
    iTotalSeconds: int = int(objDuration.total_seconds())
    iSign: int = -1 if iTotalSeconds < 0 else 1
    iAbsTotalSeconds: int = abs(iTotalSeconds)
    iHours: int = iAbsTotalSeconds // 3600
    iMinutes: int = (iAbsTotalSeconds % 3600) // 60
    iSeconds: int = iAbsTotalSeconds % 60
    pszPrefix: str = "-" if iSign < 0 else ""
    return f"{pszPrefix}{iHours}:{iMinutes:02d}:{iSeconds:02d}"


def normalize_duration_text_if_needed(pszText: str) -> str:
    objMatch = DURATION_TEXT_PATTERN.match(pszText)
    if objMatch is None:
        return pszText
    iDays: int = int(objMatch.group(1))
    iHours: int = int(objMatch.group(2))
    iMinutes: int = int(objMatch.group(3))
    iSeconds: int = int(objMatch.group(4))
    iTotalHours: int = iDays * 24 + iHours
    return f"{iTotalHours}:{iMinutes:02d}:{iSeconds:02d}"


def normalize_cell_value(objValue: object) -> str:
    if objValue is None:
        return ""
    if isinstance(objValue, timedelta):
        return format_timedelta_as_h_mm_ss(objValue)
    pszText: str = str(objValue)
    pszText = normalize_duration_text_if_needed(pszText)
    return pszText.replace("\t", "_")


def write_sheet_to_tsv(objOutputPath: Path, objRows: List[List[object]]) -> None:
    with open(objOutputPath, mode="w", encoding="utf-8", newline="") as objFile:
        objWriter: csv.writer = csv.writer(objFile, delimiter="\t", lineterminator="\n")
        for objRow in objRows:
            objWriter.writerow([normalize_cell_value(objValue) for objValue in objRow])


def convert_csv_rows_to_tsv_file(objOutputPath: Path, objRows: List[List[str]]) -> None:
    write_sheet_to_tsv(objOutputPath, objRows)


def format_xlsx_cell_value_for_tsv(objValue: object) -> object:
    if isinstance(objValue, datetime):
        if (
            objValue.hour == 0
            and objValue.minute == 0
            and objValue.second == 0
            and objValue.microsecond == 0
        ):
            return objValue.strftime("%Y/%m/%d")
        return objValue.strftime("%Y/%m/%d %H:%M:%S")

    if isinstance(objValue, date):
        return objValue.strftime("%Y/%m/%d")

    if isinstance(objValue, time):
        if objValue.second == 0 and objValue.microsecond == 0:
            return f"{objValue.hour}:{objValue.minute:02d}"
        return f"{objValue.hour}:{objValue.minute:02d}:{objValue.second:02d}"

    if isinstance(objValue, timedelta):
        pszText: str = format_timedelta_as_h_mm_ss(objValue)
        return re.sub(r"^(\d+):(\d{2}):00$", r"\1:\2", pszText)

    return objValue


def convert_xlsx_rows_to_tsv_file(objOutputPath: Path, objRows: List[List[object]]) -> None:
    objNormalizedRows: List[List[object]] = []
    for objRow in objRows:
        objNormalizedRows.append([
            format_xlsx_cell_value_for_tsv(objValue) for objValue in objRow
        ])
    write_sheet_to_tsv(objOutputPath, objNormalizedRows)


def read_tsv_rows(objInputPath: Path) -> List[List[str]]:
    objRows: List[List[str]] = []
    with open(objInputPath, mode="r", encoding="utf-8-sig", newline="") as objFile:
        objReader = csv.reader(objFile, delimiter="\t")
        for objRow in objReader:
            objRows.append(list(objRow))
    return objRows


def is_blank_text(pszText: str) -> bool:
    return (pszText or "").strip().replace("\u3000", "") == ""


def get_effective_column_count(objRow: List[str]) -> int:
    for iIndex in range(len(objRow) - 1, -1, -1):
        if not is_blank_text(objRow[iIndex]):
            return iIndex + 1
    return 0


def is_jobcan_long_format_tsv(objRows: List[List[str]]) -> bool:
    objNonEmptyRows: List[List[str]] = [
        objRow for objRow in objRows if any(not is_blank_text(pszCell) for pszCell in objRow)
    ]
    if not objNonEmptyRows:
        return False

    iTotal: int = len(objNonEmptyRows)
    iFourColumnsLike: int = 0
    iTimeTextRows: int = 0
    iProjectCodeRows: int = 0
    for objRow in objNonEmptyRows:
        iEffectiveColumns: int = get_effective_column_count(objRow)
        if 3 <= iEffectiveColumns <= 5:
            iFourColumnsLike += 1

        if len(objRow) >= 4:
            pszTimeText: str = (objRow[3] or "").strip()
            if TIME_TEXT_PATTERN.match(pszTimeText) is not None or DURATION_TEXT_PATTERN.match(pszTimeText) is not None:
                iTimeTextRows += 1

        if len(objRow) >= 2:
            pszProjectText: str = (objRow[1] or "").strip()
            if re.match(r"^(P\d{5}|[A-OQ-Z]\d{3})", pszProjectText) is not None:
                iProjectCodeRows += 1

    return (
        iFourColumnsLike / iTotal >= 0.7
        and iTimeTextRows / iTotal >= 0.5
        and iProjectCodeRows / iTotal >= 0.5
    )


def is_salary_payment_deduction_list_tsv(objRows: List[List[str]]) -> bool:
    if len(objRows) < 2:
        return False

    objHeaderRow: List[str] = objRows[0]
    objHeaderSet: set[str] = {
        (pszCell or "").strip()
        for pszCell in objHeaderRow
        if (pszCell or "").strip() != ""
    }
    if not all(pszRequiredHeader in objHeaderSet for pszRequiredHeader in SALARY_PAYMENT_DEDUCTION_REQUIRED_HEADERS):
        return False

    iStaffCodeIndex: int = objHeaderRow.index("スタッフコード")
    bHasStaffCodeValue: bool = False
    for objRow in objRows[1:]:
        if iStaffCodeIndex >= len(objRow):
            continue
        pszStaffCode: str = (objRow[iStaffCodeIndex] or "").strip()
        if re.match(r"^\d+$", pszStaffCode) is not None:
            bHasStaffCodeValue = True
            break
    return bHasStaffCodeValue


def is_management_accounting_manhour_csv(objRows: List[List[str]]) -> bool:
    if len(objRows) < 2:
        return False

    objHeaderRow: List[str] = objRows[0]
    objHeaderSet: set[str] = {
        (pszCell or "").strip()
        for pszCell in objHeaderRow
        if (pszCell or "").strip() != ""
    }
    if not all(
        pszRequiredHeader in objHeaderSet
        for pszRequiredHeader in MANAGEMENT_ACCOUNTING_MANHOUR_REQUIRED_HEADERS
    ):
        return False

    iStaffCodeIndex: int = objHeaderRow.index("スタッフコード")
    iManhourIndex: int = objHeaderRow.index("工数")

    bHasStaffCode: bool = False
    bHasManhour: bool = False
    for objRow in objRows[1:]:
        if iStaffCodeIndex < len(objRow):
            pszStaffCode: str = (objRow[iStaffCodeIndex] or "").strip()
            if re.match(r"^\d+$", pszStaffCode) is not None:
                bHasStaffCode = True
        if iManhourIndex < len(objRow):
            pszManhour: str = (objRow[iManhourIndex] or "").strip()
            if re.match(r"^\d+:\d{2}(?::\d{2})?$", pszManhour) is not None:
                bHasManhour = True
        if bHasStaffCode and bHasManhour:
            return True

    return False


def is_management_accounting_manhour_tsv(objRows: List[List[str]]) -> bool:
    return is_management_accounting_manhour_csv(objRows)


def is_prepaid_commute_allocation_table_tsv(objResolvedInputPath: Path, objRows: List[List[str]]) -> bool:
    if "前払通勤交通費按分表" not in objResolvedInputPath.name:
        return False
    if not objRows:
        return False
    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    return "スタッフコード" in objHeaderRow


def is_management_accounting_manhour_xlsx_sheet(objRows: List[List[object]]) -> bool:
    objStringRows: List[List[str]] = []
    for objRow in objRows:
        objStringRows.append([
            ("" if objValue is None else str(format_xlsx_cell_value_for_tsv(objValue))).strip()
            for objValue in objRow
        ])
    return is_management_accounting_manhour_csv(objStringRows)


def build_staff_code_by_name_from_management_accounting_rows(
    objRows: List[List[str]],
) -> dict[str, str]:
    if not objRows:
        return {}

    objHeaderRow: List[str] = [(pszCell or "").strip() for pszCell in objRows[0]]
    if "スタッフコード" not in objHeaderRow or "姓 名" not in objHeaderRow:
        return {}

    iStaffCodeIndex: int = objHeaderRow.index("スタッフコード")
    iStaffNameIndex: int = objHeaderRow.index("姓 名")

    objStaffCodeByName: dict[str, str] = {}
    for objRow in objRows[1:]:
        if iStaffCodeIndex >= len(objRow) or iStaffNameIndex >= len(objRow):
            continue
        pszStaffCode: str = (objRow[iStaffCodeIndex] or "").strip()
        pszStaffName: str = (objRow[iStaffNameIndex] or "").strip()
        if pszStaffName == "" or pszStaffCode == "":
            continue
        if re.match(r"^\d+$", pszStaffCode) is None:
            continue
        if pszStaffName not in objStaffCodeByName:
            objStaffCodeByName[pszStaffName] = pszStaffCode

    return objStaffCodeByName


def load_staff_code_by_name_from_management_accounting_file(
    objManagementAccountingPath: Path,
) -> dict[str, str]:
    pszSuffix: str = objManagementAccountingPath.suffix.lower()

    if pszSuffix == ".tsv":
        objRows: List[List[str]] = read_tsv_rows(objManagementAccountingPath)
        if not is_management_accounting_manhour_tsv(objRows):
            raise ValueError(f"Not management accounting manhour TSV: {objManagementAccountingPath}")
        return build_staff_code_by_name_from_management_accounting_rows(objRows)

    if pszSuffix == ".csv":
        objRows = []
        with open(objManagementAccountingPath, mode="r", encoding="utf-8-sig", newline="") as objFile:
            objReader = csv.reader(objFile)
            for objRow in objReader:
                objRows.append(list(objRow))
        if not is_management_accounting_manhour_csv(objRows):
            raise ValueError(f"Not management accounting manhour CSV: {objManagementAccountingPath}")
        return build_staff_code_by_name_from_management_accounting_rows(objRows)

    if pszSuffix == ".xlsx":
        try:
            import openpyxl
        except Exception as objException:
            raise RuntimeError(f"Failed to import openpyxl: {objException}") from objException

        objWorkbook = openpyxl.load_workbook(
            filename=objManagementAccountingPath,
            read_only=True,
            data_only=True,
        )
        try:
            for objWorksheet in objWorkbook.worksheets:
                objRowsXlsx: List[List[object]] = [
                    list(objRow)
                    for objRow in objWorksheet.iter_rows(values_only=True)
                ]
                if not is_management_accounting_manhour_xlsx_sheet(objRowsXlsx):
                    continue

                objRowsString: List[List[str]] = []
                for objRow in objRowsXlsx:
                    objRowsString.append([
                        "" if objValue is None else str(format_xlsx_cell_value_for_tsv(objValue)).strip()
                        for objValue in objRow
                    ])
                return build_staff_code_by_name_from_management_accounting_rows(objRowsString)
        finally:
            objWorkbook.close()

        raise ValueError(f"No management accounting manhour sheet found in XLSX: {objManagementAccountingPath}")

    raise ValueError(f"Unsupported management accounting extension: {objManagementAccountingPath}")


def build_new_rawdata_step0003_output_path_from_step0002(objStep0002Path: Path) -> Path:
    pszFileName: str = objStep0002Path.name
    if "_step0002_" not in pszFileName:
        raise ValueError(f"Input is not step0002 file: {objStep0002Path}")
    pszOutputFileName: str = pszFileName.replace("_step0002_", "_step0003_", 1)
    return objStep0002Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0003_error_path_from_step0002(objStep0002Path: Path) -> Path:
    objStep0003Path: Path = build_new_rawdata_step0003_output_path_from_step0002(objStep0002Path)
    return objStep0003Path.resolve().parent / f"{objStep0003Path.stem}_error.txt"


def write_new_rawdata_step0003_error_file_for_prepayed_commute(
    objStep0002Path: Path,
    objException: Exception,
) -> None:
    objErrorPath: Path = build_new_rawdata_step0003_error_path_from_step0002(objStep0002Path)
    pszMessage: str = (
        "step0003 作成エラー\n"
        f"対象ファイル: {objStep0002Path}\n"
        "理由: step0002 の読み込みまたは処理に失敗したため、step0003 を作成できませんでした。\n"
        f"詳細: {objException}\n"
    )
    with open(objErrorPath, mode="w", encoding="utf-8", newline="\n") as objFile:
        objFile.write(pszMessage)


def build_new_rawdata_step0004_output_path_from_step0003(objStep0003Path: Path) -> Path:
    pszFileName: str = objStep0003Path.name
    if "_step0003_" not in pszFileName:
        raise ValueError(f"Input is not step0003 file: {objStep0003Path}")
    pszOutputFileName: str = pszFileName.replace("_step0003_", "_step0004_", 1)
    return objStep0003Path.resolve().parent / pszOutputFileName


def _apply_prepayed_commute_hamilton_allocation_to_step0004_rows(objRows: List[List[str]]) -> None:
    if not objRows:
        return

    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    objRequiredHeaders: List[str] = ["前払支給分", "申請の有無", "等分"]
    if any(pszHeader not in objHeaderRow for pszHeader in objRequiredHeaders):
        return

    try:
        pszPeriodLabel: str = detect_prepayed_commute_period_label(objRows)
    except Exception:
        return

    objPeriodMonths: List[int] = [4, 5, 6, 7, 8, 9] if pszPeriodLabel == "04-09月" else [10, 11, 12, 1, 2, 3]
    objMonthHeaders: List[str] = [f"{iMonth}月" for iMonth in objPeriodMonths]
    if any(pszHeader not in objHeaderRow for pszHeader in objMonthHeaders):
        return

    iAdvanceIndex: int = objHeaderRow.index("前払支給分")
    iApplicationIndex: int = objHeaderRow.index("申請の有無")
    iEqualIndex: int = objHeaderRow.index("等分")
    iStaffCodeIndex: int = objHeaderRow.index("スタッフコード") if "スタッフコード" in objHeaderRow else -1
    objMonthIndicesByMonth: dict[int, int] = {iMonth: objHeaderRow.index(f"{iMonth}月") for iMonth in objPeriodMonths}

    for iDataRowIndex, objRow in enumerate(objRows[1:], start=2):
        while len(objRow) <= max(iEqualIndex, max(objMonthIndicesByMonth.values())):
            objRow.append("")

        pszAdvanceText: str = (objRow[iAdvanceIndex] or "").strip() if iAdvanceIndex < len(objRow) else ""
        if pszAdvanceText == "":
            continue
        pszApplicationText: str = (objRow[iApplicationIndex] or "").strip() if iApplicationIndex < len(objRow) else ""
        if pszApplicationText == "":
            continue
        pszEqualText: str = (objRow[iEqualIndex] or "").strip() if iEqualIndex < len(objRow) else ""
        if pszEqualText == "":
            continue

        try:
            objAdvanceValue: Decimal = Decimal(pszAdvanceText)
            objEqualValue: Decimal = Decimal(pszEqualText)
        except (InvalidOperation, ValueError):
            pszStaffCode: str = (objRow[iStaffCodeIndex] or "").strip() if iStaffCodeIndex >= 0 and iStaffCodeIndex < len(objRow) else ""
            print(
                "Warning: step0004 prepaid decimal parse failed at row={0}, staff_code={1}, 前払支給分='{2}', 等分='{3}', 申請の有無='{4}'".format(
                    iDataRowIndex,
                    pszStaffCode,
                    pszAdvanceText,
                    pszEqualText,
                    pszApplicationText,
                )
            )
            continue

        objMatch = re.match(r"^(\d{1,2})/\d{1,2}支給$", pszApplicationText)
        if objMatch is None:
            pszStaffCode = (objRow[iStaffCodeIndex] or "").strip() if iStaffCodeIndex >= 0 and iStaffCodeIndex < len(objRow) else ""
            print(
                "Warning: step0004 prepaid invalid 申請の有無 format at row={0}, staff_code={1}, 申請の有無='{2}'".format(
                    iDataRowIndex,
                    pszStaffCode,
                    pszApplicationText,
                )
            )
            continue

        iPaymentMonth: int = int(objMatch.group(1))
        iStartMonth: int = iPaymentMonth
        if pszPeriodLabel == "04-09月" and iPaymentMonth in (3, 4):
            iStartMonth = 4
        if pszPeriodLabel == "10-03月" and iPaymentMonth in (9, 10):
            iStartMonth = 10
        if iStartMonth not in objPeriodMonths:
            pszStaffCode = (objRow[iStaffCodeIndex] or "").strip() if iStaffCodeIndex >= 0 and iStaffCodeIndex < len(objRow) else ""
            print(
                "Warning: step0004 prepaid payment month out of period at row={0}, staff_code={1}, period={2}, 支給月={3}, 申請の有無='{4}'".format(
                    iDataRowIndex,
                    pszStaffCode,
                    pszPeriodLabel,
                    iPaymentMonth,
                    pszApplicationText,
                )
            )
            continue

        iStartIndex: int = objPeriodMonths.index(iStartMonth)
        objTargetMonths: List[int] = objPeriodMonths[iStartIndex:]
        if not objTargetMonths:
            continue

        for iMonth in objPeriodMonths:
            iMonthIndex: int = objMonthIndicesByMonth[iMonth]
            objRow[iMonthIndex] = ""

        if objEqualValue == objEqualValue.to_integral_value():
            pszAssignedValue: str = format_decimal_for_tsv_cell(objEqualValue)
            for iMonth in objTargetMonths:
                objRow[objMonthIndicesByMonth[iMonth]] = pszAssignedValue
            continue

        iScaleDigits: int = 0
        iScale: int = 1
        objScaledAdvance: Decimal = objAdvanceValue * Decimal(iScale)
        try:
            iTargetScaledUnits = int(objScaledAdvance.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        except InvalidOperation:
            pszStaffCode = (objRow[iStaffCodeIndex] or "").strip() if iStaffCodeIndex >= 0 and iStaffCodeIndex < len(objRow) else ""
            print(
                "Warning: step0004 prepaid quantize failed at row={0}, staff_code={1}, 前払支給分='{2}', 等分='{3}', 申請の有無='{4}', advance={5}, equal={6}, advance_is_finite={7}, equal_is_finite={8}, scale_digits={9}, scaled_advance={10}".format(
                    iDataRowIndex,
                    pszStaffCode,
                    pszAdvanceText,
                    pszEqualText,
                    pszApplicationText,
                    objAdvanceValue,
                    objEqualValue,
                    objAdvanceValue.is_finite(),
                    objEqualValue.is_finite(),
                    iScaleDigits,
                    objScaledAdvance,
                )
            )
            continue

        iMonthCount: int = len(objTargetMonths)
        objRaw: Decimal = Decimal(iTargetScaledUnits) / Decimal(iMonthCount)
        objFloorValue: Decimal = objRaw.to_integral_value(rounding=ROUND_FLOOR)
        iBaseValue: int = int(objFloorValue)
        iRemain: int = iTargetScaledUnits - (iBaseValue * iMonthCount)

        objAllocatedUnitsByMonth: dict[int, int] = {iMonth: iBaseValue for iMonth in objTargetMonths}
        if iRemain > 0:
            for iOffset in range(iRemain):
                objAllocatedUnitsByMonth[objTargetMonths[iOffset]] += 1
        elif iRemain < 0:
            for iOffset in range(-iRemain):
                objAllocatedUnitsByMonth[objTargetMonths[-(iOffset + 1)]] -= 1

        for iMonth in objTargetMonths:
            iMonthIndex: int = objMonthIndicesByMonth[iMonth]
            objRow[iMonthIndex] = format_scaled_units(objAllocatedUnitsByMonth[iMonth], iScaleDigits)


def process_new_rawdata_step0004_from_step0003(
    objNewRawdataStep0003Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0003Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0003Path}")

    if "前払通勤交通費按分表" in objNewRawdataStep0003Path.name:
        objOutputRows: List[List[str]] = [list(objRow) for objRow in objInputRows]
        _apply_prepayed_commute_hamilton_allocation_to_step0004_rows(objOutputRows)
    else:
        objOutputRows = []
        objSeenStaffCodes: set[str] = set()
        for objRow in objInputRows:
            objOriginalRow: List[str] = list(objRow)

            pszDisplayStaffCode: str = ""
            if objOriginalRow:
                pszStaffCodeCell: str = (objOriginalRow[0] or "").strip()
                if pszStaffCodeCell != "":
                    if pszStaffCodeCell not in objSeenStaffCodes:
                        pszDisplayStaffCode = pszStaffCodeCell
                        objSeenStaffCodes.add(pszStaffCodeCell)

            objOutputRows.append([pszDisplayStaffCode] + objOriginalRow)

    objOutputPath: Path = build_new_rawdata_step0004_output_path_from_step0003(objNewRawdataStep0003Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def build_new_rawdata_step0005_output_path_from_step0004(objStep0004Path: Path) -> Path:
    pszFileName: str = objStep0004Path.name
    if "_step0004_" not in pszFileName:
        raise ValueError(f"Input is not step0004 file: {objStep0004Path}")
    pszOutputFileName: str = pszFileName.replace("_step0004_", "_step0005_", 1)
    return objStep0004Path.resolve().parent / pszOutputFileName


def parse_numeric_text(pszText: str) -> float | None:
    pszValue: str = (pszText or "").strip()
    if pszValue == "":
        return None
    if re.match(r"^-?\d+(?:\.\d+)?$", pszValue) is None:
        return None
    try:
        return float(pszValue)
    except Exception:
        return None


def _build_new_rawdata_step0005_monthly_output_path_from_step0004(
    objStep0004Path: Path,
    iTargetMonth: int,
) -> Path:
    objBaseStep0005Path: Path = build_new_rawdata_step0005_output_path_from_step0004(objStep0004Path)
    pszBaseName: str = objBaseStep0005Path.name
    pszSuffix: str = "_前払通勤交通費按分表.tsv"
    if not pszBaseName.endswith(pszSuffix):
        raise ValueError(f"Input is not prepaid commute step0004 file: {objStep0004Path}")
    pszMonthLabel: str = f"{iTargetMonth:02d}月"
    objTenToThreeMatch = re.match(
        r"^新_ローデータ_シート_step0005_(\d{4})年(10-03月)_前払通勤交通費按分表\.tsv$",
        pszBaseName,
    )
    if objTenToThreeMatch is not None:
        iBaseYear: int = int(objTenToThreeMatch.group(1))
        pszPeriodLabel: str = objTenToThreeMatch.group(2)
        iOutputYear: int = iBaseYear - 1 if iTargetMonth in (10, 11, 12) else iBaseYear
        pszOutputName = (
            f"新_ローデータ_シート_step0005_{iOutputYear}年{pszPeriodLabel}_{pszMonthLabel}_前払通勤交通費按分表.tsv"
        )
    else:
        pszOutputName = pszBaseName[: -len(pszSuffix)] + f"_{pszMonthLabel}" + pszSuffix
    return objBaseStep0005Path.resolve().parent / pszOutputName


def _parse_decimal_text(pszText: str) -> Decimal | None:
    pszValue: str = (pszText or "").strip()
    if pszValue == "":
        return None
    try:
        return Decimal(pszValue)
    except (InvalidOperation, ValueError):
        return None


def _build_prepayed_commute_step0005_rows_for_target_month(
    objInputRows: List[List[str]],
    iTargetMonth: int,
) -> List[List[str]]:
    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objInputRows[0]]
    iAdvanceIndex: int = objHeaderRow.index("前払支給分")
    iApplicationIndex: int = objHeaderRow.index("申請の有無")
    iTotalIndex: int = objHeaderRow.index("合計") if "合計" in objHeaderRow else -1
    iRemainingIndex: int = objHeaderRow.index("残り") if "残り" in objHeaderRow else -1

    pszPeriodLabel: str = detect_prepayed_commute_period_label(objInputRows)
    objPeriodMonths: List[int] = [4, 5, 6, 7, 8, 9] if pszPeriodLabel == "04-09月" else [10, 11, 12, 1, 2, 3]
    objMonthIndicesByMonth: dict[int, int] = {iMonth: objHeaderRow.index(f"{iMonth}月") for iMonth in objPeriodMonths}
    objMonthOrderIndexByMonth: dict[int, int] = {
        iMonth: iOrderIndex for iOrderIndex, iMonth in enumerate(objPeriodMonths)
    }
    if iTargetMonth not in objMonthOrderIndexByMonth:
        raise ValueError(f"Target month {iTargetMonth} is out of period months {objPeriodMonths}")
    iTargetOrderIndex: int = objMonthOrderIndexByMonth[iTargetMonth]

    objOutputRows: List[List[str]] = [list(objInputRows[0])]
    iRequiredLength: int = max(
        iAdvanceIndex,
        iApplicationIndex,
        iTotalIndex if iTotalIndex >= 0 else 0,
        iRemainingIndex if iRemainingIndex >= 0 else 0,
        max(objMonthIndicesByMonth.values()),
    )
    for objRow in objInputRows[1:]:
        objNewRow: List[str] = list(objRow)
        while len(objNewRow) <= iRequiredLength:
            objNewRow.append("")

        pszAdvanceText: str = (objNewRow[iAdvanceIndex] or "").strip()
        if pszAdvanceText == "":
            objOutputRows.append(objNewRow)
            continue

        pszApplicationText: str = (objNewRow[iApplicationIndex] or "").strip()
        objMatch = re.match(r"^(\d{1,2})/\d{1,2}支給$", pszApplicationText)
        if objMatch is None:
            objNewRow[iApplicationIndex] = "?/??支給_不明"
            objOutputRows.append(objNewRow)
            continue

        iPaymentMonth: int = int(objMatch.group(1))
        iStartMonth: int = iPaymentMonth
        if pszPeriodLabel == "04-09月" and iPaymentMonth in (3, 4):
            iStartMonth = 4
        if pszPeriodLabel == "10-03月" and iPaymentMonth in (9, 10):
            iStartMonth = 10
        if iStartMonth not in objMonthOrderIndexByMonth:
            continue
        if iTargetOrderIndex < objMonthOrderIndexByMonth[iStartMonth]:
            continue

        objMonthSum: Decimal = Decimal("0")
        for iMonth in objPeriodMonths:
            iMonthIndex: int = objMonthIndicesByMonth[iMonth]
            pszMonthText: str = (objNewRow[iMonthIndex] or "").strip()
            if objMonthOrderIndexByMonth[iMonth] > iTargetOrderIndex:
                objNewRow[iMonthIndex] = ""
                continue
            objMonthValue: Decimal | None = _parse_decimal_text(pszMonthText)
            if objMonthValue is not None:
                objMonthSum += objMonthValue

        if iTotalIndex >= 0:
            objNewRow[iTotalIndex] = format_decimal_for_tsv_cell(objMonthSum)
        if iRemainingIndex >= 0:
            objAdvanceValue: Decimal | None = _parse_decimal_text(pszAdvanceText)
            if objAdvanceValue is not None:
                objNewRow[iRemainingIndex] = format_decimal_for_tsv_cell(objAdvanceValue - objMonthSum)

        objOutputRows.append(objNewRow)

    return objOutputRows


def _process_new_rawdata_step0005_monthly_prepayed_commute_from_step0004(
    objNewRawdataStep0004Path: Path,
    objInputRows: List[List[str]],
) -> int:
    pszPeriodLabel: str = detect_prepayed_commute_period_label(objInputRows)
    objPeriodMonths: List[int] = [4, 5, 6, 7, 8, 9] if pszPeriodLabel == "04-09月" else [10, 11, 12, 1, 2, 3]
    for iTargetMonth in objPeriodMonths:
        objOutputRows: List[List[str]] = _build_prepayed_commute_step0005_rows_for_target_month(
            objInputRows,
            iTargetMonth,
        )
        objOutputPath: Path = _build_new_rawdata_step0005_monthly_output_path_from_step0004(
            objNewRawdataStep0004Path,
            iTargetMonth,
        )
        write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def _extract_target_month_from_prepayed_step_file_name(objPath: Path) -> int | None:
    objMatch = re.search(r"_([01]\d)月_前払通勤交通費按分表\.tsv$", objPath.name)
    if objMatch is None:
        return None
    iMonth: int = int(objMatch.group(1))
    if 1 <= iMonth <= 12:
        return iMonth
    return None


def _detect_name_column_index_for_prepayed_step_rows(
    objHeaderRow: List[str],
    iStaffCodeIndex: int,
    iMonthIndex: int,
) -> int:
    for pszHeader in ("氏名", "姓名", "姓 名", "名前"):
        if pszHeader in objHeaderRow:
            return objHeaderRow.index(pszHeader)
    if "申請の有無" in objHeaderRow:
        iCandidate: int = objHeaderRow.index("申請の有無") + 1
        if 0 <= iCandidate < len(objHeaderRow):
            return iCandidate
    iFallback: int = iStaffCodeIndex + 2
    if iFallback == iMonthIndex:
        iFallback = iStaffCodeIndex + 1
    if 0 <= iFallback < len(objHeaderRow):
        return iFallback
    return min(len(objHeaderRow) - 1, iStaffCodeIndex + 1)


def _process_new_rawdata_step0006_monthly_prepayed_commute_from_step0005(
    objNewRawdataStep0005Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0005Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0005Path}")

    iTargetMonth: int | None = _extract_target_month_from_prepayed_step_file_name(objNewRawdataStep0005Path)
    if iTargetMonth is None:
        raise ValueError(f"Could not detect target month from step0005 prepaid file name: {objNewRawdataStep0005Path}")

    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objInputRows[0]]
    if "スタッフコード" not in objHeaderRow:
        raise ValueError(f"Missing header: スタッフコード ({objNewRawdataStep0005Path})")
    pszMonthHeader: str = f"{iTargetMonth}月"
    if pszMonthHeader not in objHeaderRow:
        raise ValueError(f"Missing header: {pszMonthHeader} ({objNewRawdataStep0005Path})")

    iStaffCodeIndex: int = objHeaderRow.index("スタッフコード")
    iMonthIndex: int = objHeaderRow.index(pszMonthHeader)
    iNameIndex: int = _detect_name_column_index_for_prepayed_step_rows(objHeaderRow, iStaffCodeIndex, iMonthIndex)

    pszNameHeader: str = objHeaderRow[iNameIndex] if 0 <= iNameIndex < len(objHeaderRow) else ""
    objOutputRows: List[List[str]] = [["スタッフコード", pszNameHeader, pszMonthHeader]]
    iRequiredLength: int = max(iStaffCodeIndex, iNameIndex, iMonthIndex)
    for objRow in objInputRows[1:]:
        objNewRow: List[str] = list(objRow)
        while len(objNewRow) <= iRequiredLength:
            objNewRow.append("")
        objOutputRows.append([
            objNewRow[iStaffCodeIndex],
            objNewRow[iNameIndex],
            objNewRow[iMonthIndex],
        ])

    objOutputPath: Path = build_new_rawdata_step0006_output_path_from_step0005(objNewRawdataStep0005Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_new_rawdata_step0006_monthly_prepayed_commute_from_step0004(
    objNewRawdataStep0004Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0004Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0004Path}")
    pszPeriodLabel: str = detect_prepayed_commute_period_label(objInputRows)
    objPeriodMonths: List[int] = [4, 5, 6, 7, 8, 9] if pszPeriodLabel == "04-09月" else [10, 11, 12, 1, 2, 3]
    for iTargetMonth in objPeriodMonths:
        objStep0005Path: Path = _build_new_rawdata_step0005_monthly_output_path_from_step0004(
            objNewRawdataStep0004Path,
            iTargetMonth,
        )
        if not objStep0005Path.exists():
            continue
        _process_new_rawdata_step0006_monthly_prepayed_commute_from_step0005(objStep0005Path)
    return 0


def process_new_rawdata_step0005_from_step0004(
    objNewRawdataStep0004Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0004Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0004Path}")
    if "前払通勤交通費按分表" in objNewRawdataStep0004Path.name:
        return _process_new_rawdata_step0005_monthly_prepayed_commute_from_step0004(
            objNewRawdataStep0004Path,
            objInputRows,
        )

    objRankTargets: List[float] = []
    for objRow in objInputRows:
        if not objRow:
            continue
        fValue = parse_numeric_text(objRow[0])
        if fValue is not None:
            objRankTargets.append(fValue)

    objOutputRows: List[List[str]] = []
    for objRow in objInputRows:
        objNewRow: List[str] = list(objRow)
        pszRankText: str = ""
        if objNewRow:
            fValue = parse_numeric_text(objNewRow[0])
            if fValue is not None:
                iRank: int = 1 + sum(1 for fTarget in objRankTargets if fTarget < fValue)
                pszRankText = str(iRank)
        objOutputRows.append([pszRankText] + objNewRow)

    objOutputPath: Path = build_new_rawdata_step0005_output_path_from_step0004(objNewRawdataStep0004Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def build_new_rawdata_step0006_output_path_from_step0005(objStep0005Path: Path) -> Path:
    pszFileName: str = objStep0005Path.name
    if "_step0005_" not in pszFileName:
        raise ValueError(f"Input is not step0005 file: {objStep0005Path}")
    pszOutputFileName = pszFileName.replace("_step0005_", "_step0006_", 1)
    return objStep0005Path.resolve().parent / pszOutputFileName


def process_new_rawdata_step0006_from_step0005(
    objNewRawdataStep0005Path: Path,
) -> int:
    if "前払通勤交通費按分表" in objNewRawdataStep0005Path.name:
        iTargetMonth: int | None = _extract_target_month_from_prepayed_step_file_name(objNewRawdataStep0005Path)
        if iTargetMonth is not None:
            return _process_new_rawdata_step0006_monthly_prepayed_commute_from_step0005(
                objNewRawdataStep0005Path,
            )

    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0005Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0005Path}")

    objOutputRows: List[List[str]] = [list(objRow) for objRow in objInputRows]

    iCurrentStaffCode: str = ""
    iRowIndex: int = 0
    while iRowIndex < len(objOutputRows):
        objRow: List[str] = objOutputRows[iRowIndex]
        if len(objRow) < 4:
            if iCurrentStaffCode != "" and len(objRow) >= 3 and (objRow[2] or "").strip() == "":
                objRow[2] = iCurrentStaffCode
            iRowIndex += 1
            continue

        pszStaffName: str = (objRow[3] or "").strip()
        if pszStaffName == "":
            if iCurrentStaffCode != "" and len(objRow) >= 3 and (objRow[2] or "").strip() == "":
                objRow[2] = iCurrentStaffCode
            iRowIndex += 1
            continue

        pszStaffCode: str = (objRow[2] or "").strip() if len(objRow) >= 3 else ""
        pszProjectName: str = objRow[4] if len(objRow) >= 5 else ""
        pszManhour: str = objRow[5] if len(objRow) >= 6 else ""

        while len(objRow) < 6:
            objRow.append("")

        objRow[4] = "合計"
        objRow[5] = ""

        objNewDetailRow: List[str] = [""] * max(len(objRow), 6)
        objNewDetailRow[2] = pszStaffCode
        objNewDetailRow[4] = pszProjectName
        objNewDetailRow[5] = pszManhour
        objOutputRows.insert(iRowIndex + 1, objNewDetailRow)

        iCurrentStaffCode = pszStaffCode
        iRowIndex += 2

    objOutputPath: Path = build_new_rawdata_step0006_output_path_from_step0005(objNewRawdataStep0005Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def build_new_rawdata_step0007_output_path_from_step0006(objStep0006Path: Path) -> Path:
    pszFileName: str = objStep0006Path.name
    if "_step0006_" not in pszFileName:
        raise ValueError(f"Input is not step0006 file: {objStep0006Path}")
    pszOutputFileName: str = pszFileName.replace("_step0006_", "_step0007_", 1)
    return objStep0006Path.resolve().parent / pszOutputFileName


def parse_time_text_to_seconds(pszTimeText: str) -> int:
    objParts: List[str] = (pszTimeText or "").strip().split(":")
    if len(objParts) != 3:
        raise ValueError(f"Invalid time format: {pszTimeText}")
    iHours: int = int(objParts[0])
    iMinutes: int = int(objParts[1])
    iSeconds: int = int(objParts[2])
    return iHours * 3600 + iMinutes * 60 + iSeconds


def process_new_rawdata_step0007_from_step0006(
    objNewRawdataStep0006Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0006Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0006Path}")

    objOutputRows: List[List[str]] = [list(objRow) for objRow in objInputRows]

    iRowIndex: int = 0
    while iRowIndex < len(objOutputRows):
        objRow: List[str] = objOutputRows[iRowIndex]
        pszStaffName: str = (objRow[3] or "").strip() if len(objRow) >= 4 else ""
        pszProject: str = (objRow[4] or "").strip() if len(objRow) >= 5 else ""
        if pszStaffName == "" or pszProject != "合計":
            iRowIndex += 1
            continue

        while len(objRow) < 6:
            objRow.append("")

        iTotalSeconds: int = 0
        iDetailIndex: int = iRowIndex + 1
        while iDetailIndex < len(objOutputRows):
            objDetailRow: List[str] = objOutputRows[iDetailIndex]
            pszDetailStaffName: str = (objDetailRow[3] or "").strip() if len(objDetailRow) >= 4 else ""
            pszDetailProject: str = (objDetailRow[4] or "").strip() if len(objDetailRow) >= 5 else ""
            if pszDetailStaffName != "" and pszDetailProject == "合計":
                break

            pszManhour: str = (objDetailRow[5] or "").strip() if len(objDetailRow) >= 6 else ""
            if pszManhour != "":
                iTotalSeconds += parse_time_text_to_seconds(pszManhour)
            iDetailIndex += 1

        objRow[5] = format_timedelta_as_h_mm_ss(timedelta(seconds=iTotalSeconds))
        iRowIndex = iDetailIndex

    objOutputPath: Path = build_new_rawdata_step0007_output_path_from_step0006(objNewRawdataStep0006Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0




def build_new_rawdata_step0008_output_path_from_step0007(objStep0007Path: Path) -> Path:
    pszFileName: str = objStep0007Path.name
    if "_step0007_" not in pszFileName:
        raise ValueError(f"Input is not step0007 file: {objStep0007Path}")
    pszOutputFileName: str = pszFileName.replace("_step0007_", "_step0008_", 1)
    return objStep0007Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0007_prepayed_commute_error_path(objStep0007Path: Path) -> Path:
    pszSuffix: str = "_前払通勤交通費按分表.tsv"
    if not objStep0007Path.name.endswith(pszSuffix):
        raise ValueError(f"Input is not prepaid commute step0007 file: {objStep0007Path}")
    return objStep0007Path.resolve().parent / f"{objStep0007Path.stem}_error.txt"


def build_new_rawdata_step0007_prepayed_commute_no_manhour_staff_path(objStep0007Path: Path) -> Path:
    pszSuffix: str = "_前払通勤交通費按分表.tsv"
    if not objStep0007Path.name.endswith(pszSuffix):
        raise ValueError(f"Input is not prepaid commute step0007 file: {objStep0007Path}")
    return objStep0007Path.resolve().parent / f"{objStep0007Path.stem}_工数なしスタッフ.tsv"


def _extract_staff_amounts_from_prepayed_step0006_rows(
    objStep0006Rows: List[List[str]],
    objStep0006Path: Path,
) -> Tuple[Dict[str, str], Dict[str, str], str]:
    if not objStep0006Rows:
        raise ValueError(f"Input TSV has no rows: {objStep0006Path}")
    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objStep0006Rows[0]]
    if "スタッフコード" not in objHeaderRow:
        raise ValueError(f"Missing header: スタッフコード ({objStep0006Path})")
    iStaffCodeIndex: int = objHeaderRow.index("スタッフコード")
    iAmountIndex: int = len(objHeaderRow) - 1
    iNameIndex: int = 1 if len(objHeaderRow) >= 2 else -1
    pszAmountHeader: str = objHeaderRow[iAmountIndex] if 0 <= iAmountIndex < len(objHeaderRow) else "金額"
    objAmountsByStaffCode: Dict[str, str] = {}
    objNameByStaffCode: Dict[str, str] = {}
    for objRow in objStep0006Rows[1:]:
        pszStaffCode: str = (objRow[iStaffCodeIndex] or "").strip() if iStaffCodeIndex < len(objRow) else ""
        if pszStaffCode == "":
            continue
        pszStaffName: str = (objRow[iNameIndex] or "").strip() if 0 <= iNameIndex < len(objRow) else ""
        pszAmount: str = (objRow[iAmountIndex] or "").strip() if iAmountIndex < len(objRow) else ""
        if pszAmount == "":
            continue
        objAmountsByStaffCode[pszStaffCode] = pszAmount
        objNameByStaffCode[pszStaffCode] = pszStaffName
    return objAmountsByStaffCode, objNameByStaffCode, pszAmountHeader


def _collect_staff_blocks_from_prepayed_step0007_rows(
    objStep0007Rows: List[List[str]],
    objStep0007Path: Path,
) -> Dict[str, List[List[str]]]:
    objBlocksByStaffCode: Dict[str, List[List[str]]] = {}
    iRowIndex: int = 0
    while iRowIndex < len(objStep0007Rows):
        objRow: List[str] = list(objStep0007Rows[iRowIndex])
        pszProject: str = (objRow[4] or "").strip() if len(objRow) >= 5 else ""
        pszStaffCode: str = (objRow[2] or "").strip() if len(objRow) >= 3 else ""
        if pszProject != "合計" or pszStaffCode == "":
            iRowIndex += 1
            continue
        iBlockStart: int = iRowIndex
        iBlockEnd: int = iBlockStart + 1
        while iBlockEnd < len(objStep0007Rows):
            objNextRow: List[str] = objStep0007Rows[iBlockEnd]
            pszNextProject: str = (objNextRow[4] or "").strip() if len(objNextRow) >= 5 else ""
            pszNextStaffCode: str = (objNextRow[2] or "").strip() if len(objNextRow) >= 3 else ""
            if pszNextProject == "合計" and pszNextStaffCode != "":
                break
            iBlockEnd += 1
        if pszStaffCode in objBlocksByStaffCode:
            raise ValueError(
                f"Duplicate staff block in step0007 prepaid file: staff_code={pszStaffCode} ({objStep0007Path})"
            )
        objBlocksByStaffCode[pszStaffCode] = [list(objItem) for objItem in objStep0007Rows[iBlockStart:iBlockEnd]]
        iRowIndex = iBlockEnd
    return objBlocksByStaffCode


def _build_prepayed_step0008_rows_from_step0006_and_step0007(
    objStep0006Rows: List[List[str]],
    objStep0007Rows: List[List[str]],
    objStep0006Path: Path,
    objStep0007Path: Path,
) -> Tuple[List[List[str]], List[List[str]]]:
    objAmountsByStaffCode, objNameByStaffCode, pszAmountHeader = _extract_staff_amounts_from_prepayed_step0006_rows(
        objStep0006Rows,
        objStep0006Path,
    )
    objBlocksByStaffCode: Dict[str, List[List[str]]] = _collect_staff_blocks_from_prepayed_step0007_rows(
        objStep0007Rows,
        objStep0007Path,
    )
    objOutputRows: List[List[str]] = [[
        "スタッフ昇順",
        "スタッフコード(先頭)",
        "スタッフコード",
        "氏名",
        "プロジェクト名",
        "工数",
        "前払通勤交通費按分",
    ]]
    objNoManhourRows: List[List[str]] = [["スタッフコード", "", pszAmountHeader]]
    for pszStaffCode, pszAmount in objAmountsByStaffCode.items():
        if pszStaffCode not in objBlocksByStaffCode:
            objNoManhourRows.append([pszStaffCode, objNameByStaffCode.get(pszStaffCode, ""), pszAmount])
            continue
        objBlockRows: List[List[str]] = [list(objRow) for objRow in objBlocksByStaffCode[pszStaffCode]]
        if not objBlockRows:
            continue
        objSummaryRow: List[str] = list(objBlockRows[0])
        while len(objSummaryRow) < 6:
            objSummaryRow.append("")
        objSummaryRow.append(pszAmount)
        objOutputRows.append(objSummaryRow)
        for objDetailRow in objBlockRows[1:]:
            objOutputRows.append(list(objDetailRow))
    return objOutputRows, objNoManhourRows


def _write_prepayed_step0007_error_file(
    objStep0007Path: Path,
    objException: Exception,
) -> None:
    objErrorPath: Path = build_new_rawdata_step0007_prepayed_commute_error_path(objStep0007Path)
    with open(objErrorPath, mode="w", encoding="utf-8", newline="") as objFile:
        objFile.write(str(objException))
        objFile.write("\n")


def process_new_rawdata_step0008_mainstream_from_step0007(
    objNewRawdataStep0007Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0007Path)
    objOutputRows: List[List[str]] = [[
        "スタッフ昇順",
        "スタッフコード(先頭)",
        "スタッフコード",
        "氏名",
        "プロジェクト名",
        "工数",
    ]]
    objOutputRows.extend([list(objRow) for objRow in objInputRows])

    objOutputPath: Path = build_new_rawdata_step0008_output_path_from_step0007(objNewRawdataStep0007Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_new_rawdata_step0008_prepayed_commute_from_step0007(
    objNewRawdataStep0007Path: Path,
) -> int:
    objStep0006Path: Path = objNewRawdataStep0007Path.resolve().parent / objNewRawdataStep0007Path.name.replace(
        "_step0007_",
        "_step0006_",
        1,
    )
    try:
        objStep0006Rows: List[List[str]] = read_tsv_rows(objStep0006Path)
        objStep0007Rows: List[List[str]] = read_tsv_rows(objNewRawdataStep0007Path)
        objOutputRows, objNoManhourRows = _build_prepayed_step0008_rows_from_step0006_and_step0007(
            objStep0006Rows,
            objStep0007Rows,
            objStep0006Path,
            objNewRawdataStep0007Path,
        )
        objOutputPath: Path = build_new_rawdata_step0008_output_path_from_step0007(objNewRawdataStep0007Path)
        write_sheet_to_tsv(objOutputPath, objOutputRows)
        if len(objNoManhourRows) >= 2:
            objNoManhourPath: Path = build_new_rawdata_step0007_prepayed_commute_no_manhour_staff_path(
                objNewRawdataStep0007Path
            )
            write_sheet_to_tsv(objNoManhourPath, objNoManhourRows)
        return 0
    except Exception as objException:
        _write_prepayed_step0007_error_file(objNewRawdataStep0007Path, objException)
        raise


def process_new_rawdata_step0008_from_step0007(
    objNewRawdataStep0007Path: Path,
) -> int:
    if objNewRawdataStep0007Path.name.endswith("_前払通勤交通費按分表.tsv"):
        return process_new_rawdata_step0008_prepayed_commute_from_step0007(objNewRawdataStep0007Path)
    return process_new_rawdata_step0008_mainstream_from_step0007(objNewRawdataStep0007Path)


def process_prepayed_commute_step0007_from_mainstream_step0007(
    objMainstreamStep0007Path: Path,
) -> int:
    objMainMatch = re.match(
        r"^新_ローデータ_シート_step0007_(\d{4})年(\d{2})月\.tsv$",
        objMainstreamStep0007Path.name,
    )
    if objMainMatch is None:
        return 0
    iTargetMonth: int = int(objMainMatch.group(2))
    pszTargetMonthLabel: str = f"{iTargetMonth:02d}月"

    objPrepaidStep0006Paths: List[Path] = sorted(
        objMainstreamStep0007Path.resolve().parent.glob(
            "新_ローデータ_シート_step0006_*_前払通勤交通費按分表.tsv"
        )
    )
    pszTargetPeriodLabel: str = ""
    for objStep0006Path in objPrepaidStep0006Paths:
        objMatch = PREPAYED_COMMUTE_STEP0006_MONTHLY_FILE_PATTERN.match(objStep0006Path.name)
        if objMatch is None:
            continue
        pszPeriodLabel: str = objMatch.group(1)
        pszMonthLabel: str = f"{int(objMatch.group(2)):02d}月"
        if pszMonthLabel == pszTargetMonthLabel:
            pszTargetPeriodLabel = pszPeriodLabel
            break

    if pszTargetPeriodLabel == "":
        return 0

    objRows: List[List[str]] = read_tsv_rows(objMainstreamStep0007Path)
    if not objRows:
        raise ValueError(f"Input TSV has no rows: {objMainstreamStep0007Path}")

    objOutputPath: Path = (
        objMainstreamStep0007Path.resolve().parent
        / f"新_ローデータ_シート_step0007_{pszTargetPeriodLabel}_{pszTargetMonthLabel}_前払通勤交通費按分表.tsv"
    )
    write_sheet_to_tsv(objOutputPath, objRows)
    return 0



def build_new_rawdata_step0009_output_path_from_step0008(objStep0008Path: Path) -> Path:
    pszFileName: str = objStep0008Path.name
    if "_step0008_" not in pszFileName:
        raise ValueError(f"Input is not step0008 file: {objStep0008Path}")
    pszOutputFileName: str = pszFileName.replace("_step0008_", "_step0009_", 1)
    return objStep0008Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0010_output_path_from_prepayed_step0009(objStep0009Path: Path) -> Path:
    pszFileName: str = objStep0009Path.name
    if "_step0009_" not in pszFileName:
        raise ValueError(f"Input is not step0009 file: {objStep0009Path}")
    pszOutputFileName: str = pszFileName.replace("_step0009_", "_step0010_", 1)
    return objStep0009Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0009_5_output_path_from_step0009(objStep0009Path: Path) -> Path:
    pszFileName: str = objStep0009Path.name
    if "_step0009_" not in pszFileName:
        raise ValueError(f"Input is not step0009 file: {objStep0009Path}")
    pszOutputFileName: str = pszFileName.replace("_step0009_", "_step0009.5_", 1)
    return objStep0009Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0009_5_error_path_from_step0009_5(objStep0009_5Path: Path) -> Path:
    return objStep0009_5Path.resolve().parent / f"{objStep0009_5Path.stem}_error.txt"


def process_new_rawdata_step0009_from_step0008_and_salary_step0001(
    objNewRawdataStep0008Path: Path,
    objSalaryStep0001Path: Path,
) -> int:
    objStep0008Rows: List[List[str]] = read_tsv_rows(objNewRawdataStep0008Path)
    if not objStep0008Rows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0008Path}")

    objSalaryRows: List[List[str]] = read_tsv_rows(objSalaryStep0001Path)
    if not objSalaryRows:
        raise ValueError(f"Input TSV has no rows: {objSalaryStep0001Path}")

    objSalaryHeader: List[str] = list(objSalaryRows[0])
    objAdditionalHeaders: List[str] = objSalaryHeader[2:] if len(objSalaryHeader) >= 2 else []

    objOutputRows: List[List[str]] = [list(objRow) for objRow in objStep0008Rows]
    objOutputRows[0].extend(objAdditionalHeaders)

    objOutputPath: Path = build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_new_rawdata_step0009_5_from_step0009_with_summary_row_merge(
    objStep0009Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0009Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0009Path}")

    objOutputRows: List[List[str]] = []
    if objInputRows:
        objOutputRows.append(list(objInputRows[0]))

    objErrorLines: List[str] = []
    iRowIndex: int = 1
    while iRowIndex < len(objInputRows):
        objRow: List[str] = list(objInputRows[iRowIndex])
        pszStaffName: str = (objRow[3] or "").strip() if len(objRow) >= 4 else ""
        pszProjectName: str = (objRow[4] or "").strip() if len(objRow) >= 5 else ""
        if pszStaffName != "" and pszProjectName == "合計":
            if iRowIndex + 1 >= len(objInputRows):
                objOutputRows.append(objRow)
                objErrorLines.append(
                    f"{iRowIndex + 1}行目: 合計行の直下行が存在しないため、合計行をそのまま出力しました。"
                )
                iRowIndex += 1
                continue

            objNextRow: List[str] = list(objInputRows[iRowIndex + 1])
            pszNextStaffName: str = (objNextRow[3] or "").strip() if len(objNextRow) >= 4 else ""
            if pszNextStaffName != "":
                objOutputRows.append(objRow)
                objErrorLines.append(
                    f"{iRowIndex + 1}行目: 合計行の直下行が不正なため、合計行をそのまま出力しました。"
                )
                iRowIndex += 1
                continue

            objMergedRow: List[str] = list(objRow[:4]) + list(objNextRow[4:])
            if len(objMergedRow) < len(objRow):
                objMergedRow.extend([""] * (len(objRow) - len(objMergedRow)))
            objOutputRows.append(objMergedRow)
            iRowIndex += 2
            continue

        objOutputRows.append(objRow)
        iRowIndex += 1

    objOutputPath: Path = build_new_rawdata_step0009_5_output_path_from_step0009(objStep0009Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)

    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0009_5_error_path_from_step0009_5(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    return 0


def process_new_rawdata_step0009_prepayed_commute_from_step0008(
    objNewRawdataStep0008Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0008Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0008Path}")

    objOutputRows: List[List[str]] = [list(objRow) for objRow in objInputRows]

    iRowIndex: int = 0
    while iRowIndex < len(objOutputRows):
        objSummaryRow: List[str] = objOutputRows[iRowIndex]
        pszStaffName: str = (objSummaryRow[3] or "").strip() if len(objSummaryRow) >= 4 else ""
        pszProject: str = (objSummaryRow[4] or "").strip() if len(objSummaryRow) >= 5 else ""
        if pszStaffName == "" or pszProject != "合計":
            iRowIndex += 1
            continue

        iNextSummaryIndex: int = iRowIndex + 1
        objDetailIndices: List[int] = []
        while iNextSummaryIndex < len(objOutputRows):
            objCandidateRow: List[str] = objOutputRows[iNextSummaryIndex]
            pszCandidateName: str = (objCandidateRow[3] or "").strip() if len(objCandidateRow) >= 4 else ""
            pszCandidateProject: str = (objCandidateRow[4] or "").strip() if len(objCandidateRow) >= 5 else ""
            if pszCandidateName != "" and pszCandidateProject == "合計":
                break
            if pszCandidateName == "":
                objDetailIndices.append(iNextSummaryIndex)
            iNextSummaryIndex += 1

        if objDetailIndices and len(objSummaryRow) > 6:
            objWeights: List[int] = []
            for iDetailIndex in objDetailIndices:
                objDetailRow: List[str] = objOutputRows[iDetailIndex]
                pszManhour: str = (objDetailRow[5] or "").strip() if len(objDetailRow) >= 6 else ""
                if pszManhour == "":
                    objWeights.append(0)
                    continue
                try:
                    objWeights.append(parse_time_text_to_seconds(pszManhour))
                except Exception:
                    objWeights.append(0)

            iWeightTotal: int = sum(objWeights)
            pszTotalText: str = objSummaryRow[6] if 6 < len(objSummaryRow) else ""
            objTotalValue: Decimal | None = parse_decimal_text(pszTotalText)
            if objTotalValue is not None:
                iScaleDigits: int = count_decimal_places(pszTotalText)
                iScale: int = 10 ** iScaleDigits
                objAbsTotalScaled: Decimal = (abs(objTotalValue) * Decimal(iScale)).quantize(Decimal("1"))
                iTotalScaledUnits: int = int(objAbsTotalScaled)

                objAllocatedUnits: List[int] = [0] * len(objDetailIndices)
                if iWeightTotal > 0 and iTotalScaledUnits > 0:
                    objFloors: List[int] = []
                    objRemainders: List[tuple[int, Decimal]] = []
                    for iIndex, iWeight in enumerate(objWeights):
                        if iWeight <= 0:
                            objFloors.append(0)
                            objRemainders.append((iIndex, Decimal("-1")))
                            continue
                        objRaw: Decimal = Decimal(iTotalScaledUnits) * Decimal(iWeight) / Decimal(iWeightTotal)
                        objFloorValue: Decimal = objRaw.to_integral_value(rounding=ROUND_FLOOR)
                        iFloor: int = int(objFloorValue)
                        objFloors.append(iFloor)
                        objRemainders.append((iIndex, objRaw - objFloorValue))

                    iFloorSum: int = sum(objFloors)
                    iRemaining: int = iTotalScaledUnits - iFloorSum
                    objRemainders.sort(key=lambda objItem: (-objItem[1], objItem[0]))
                    for iIndex, _ in objRemainders[:iRemaining]:
                        objFloors[iIndex] += 1
                    objAllocatedUnits = objFloors

                iSign: int = -1 if objTotalValue < 0 else 1
                for iIndex, iDetailIndex in enumerate(objDetailIndices):
                    objDetailRow: List[str] = objOutputRows[iDetailIndex]
                    while len(objDetailRow) <= 6:
                        objDetailRow.append("")
                    objDetailRow[6] = format_scaled_units(iSign * objAllocatedUnits[iIndex], iScaleDigits)

        iRowIndex = iNextSummaryIndex

    objOutputPath: Path = build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0



def build_new_rawdata_step0010_output_path_from_step0009(objStep0009Path: Path) -> Path:
    pszFileName: str = objStep0009Path.name
    if "_step0009.5_" in pszFileName:
        pszOutputFileName: str = pszFileName.replace("_step0009.5_", "_step0010_", 1)
    elif "_step0009_" in pszFileName:
        pszOutputFileName = pszFileName.replace("_step0009_", "_step0010_", 1)
    else:
        raise ValueError(f"Input is not step0009/step0009.5 file: {objStep0009Path}")
    return objStep0009Path.resolve().parent / pszOutputFileName


def process_new_rawdata_step0010_from_step0009_and_salary_step0001(
    objNewRawdataStep0009Path: Path,
    objSalaryStep0001Path: Path,
) -> int:
    objStep0009Rows: List[List[str]] = read_tsv_rows(objNewRawdataStep0009Path)
    if not objStep0009Rows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0009Path}")

    objSalaryRows: List[List[str]] = read_tsv_rows(objSalaryStep0001Path)
    if len(objSalaryRows) < 2:
        raise ValueError(f"Input TSV has no data rows: {objSalaryStep0001Path}")

    objSalaryRowByStaffCode: dict[str, List[str]] = {}
    for objSalaryRow in objSalaryRows[1:]:
        if len(objSalaryRow) < 2:
            continue
        pszStaffCode: str = (objSalaryRow[1] or "").strip()
        if pszStaffCode == "":
            continue
        if pszStaffCode not in objSalaryRowByStaffCode:
            objSalaryRowByStaffCode[pszStaffCode] = list(objSalaryRow)

    objOutputRows: List[List[str]] = []
    for objRow in objStep0009Rows:
        objNewRow: List[str] = list(objRow)
        pszStaffName: str = (objNewRow[3] or "").strip() if len(objNewRow) >= 4 else ""
        pszProject: str = (objNewRow[4] or "").strip() if len(objNewRow) >= 5 else ""
        pszStaffCode: str = (objNewRow[2] or "").strip() if len(objNewRow) >= 3 else ""
        if pszStaffName != "" and pszProject == "合計" and pszStaffCode != "":
            objSalaryRow: List[str] | None = objSalaryRowByStaffCode.get(pszStaffCode)
            if objSalaryRow is not None:
                objNewRow.extend(objSalaryRow[2:] if len(objSalaryRow) >= 2 else [])
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0010_output_path_from_step0009(objNewRawdataStep0009Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_new_rawdata_step0010_prepayed_commute_from_step0009(
    objStep0009Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0009Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0009Path}")
    objOutputRows: List[List[str]] = remove_columns_by_1_based_indices(objInputRows, {1, 2, 3, 4})

    objOutputPath: Path = build_new_rawdata_step0010_output_path_from_step0009(objStep0009Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0011_from_step0010(objOutputPath)
    objStep0011Path: Path = build_new_rawdata_step0011_output_path_from_step0010(objOutputPath)
    process_new_rawdata_step0012_from_step0011(objStep0011Path)
    process_new_rawdata_step0013_prepayed_commute_from_step0012(
        build_new_rawdata_step0012_output_path_from_step0011(objStep0011Path)
    )
    return 0



def build_new_rawdata_step0011_output_path_from_step0010(objStep0010Path: Path) -> Path:
    pszFileName: str = objStep0010Path.name
    if "_step0010_" not in pszFileName:
        raise ValueError(f"Input is not step0010 file: {objStep0010Path}")
    pszOutputFileName: str = pszFileName.replace("_step0010_", "_step0011_", 1)
    return objStep0010Path.resolve().parent / pszOutputFileName


def parse_decimal_text(pszText: str) -> Decimal | None:
    pszValue: str = (pszText or "").strip()
    if pszValue == "":
        return None
    try:
        return Decimal(pszValue)
    except InvalidOperation:
        return None


def count_decimal_places(pszText: str) -> int:
    pszValue: str = (pszText or "").strip()
    if "." not in pszValue:
        return 0
    return len(pszValue.split(".", 1)[1])


def format_scaled_units(iUnits: int, iScaleDigits: int) -> str:
    if iScaleDigits <= 0:
        return str(iUnits)
    iSign: str = "-" if iUnits < 0 else ""
    iAbsUnits: int = abs(iUnits)
    iScale: int = 10 ** iScaleDigits
    iIntegerPart: int = iAbsUnits // iScale
    iFractionPart: int = iAbsUnits % iScale
    pszFraction: str = f"{iFractionPart:0{iScaleDigits}d}".rstrip("0")
    if pszFraction == "":
        return f"{iSign}{iIntegerPart}"
    return f"{iSign}{iIntegerPart}.{pszFraction}"


def process_new_rawdata_step0011_from_step0010(
    objNewRawdataStep0010Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0010Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0010Path}")

    objOutputRows: List[List[str]] = [list(objRow) for objRow in objInputRows]

    iRowIndex: int = 0
    while iRowIndex < len(objOutputRows):
        objSummaryRow: List[str] = objOutputRows[iRowIndex]
        pszStaffName: str = (objSummaryRow[3] or "").strip() if len(objSummaryRow) >= 4 else ""
        pszProject: str = (objSummaryRow[4] or "").strip() if len(objSummaryRow) >= 5 else ""
        if pszStaffName == "" or pszProject != "合計":
            iRowIndex += 1
            continue

        iNextSummaryIndex: int = iRowIndex + 1
        objDetailIndices: List[int] = []
        while iNextSummaryIndex < len(objOutputRows):
            objCandidateRow: List[str] = objOutputRows[iNextSummaryIndex]
            pszCandidateName: str = (objCandidateRow[3] or "").strip() if len(objCandidateRow) >= 4 else ""
            pszCandidateProject: str = (objCandidateRow[4] or "").strip() if len(objCandidateRow) >= 5 else ""
            if pszCandidateName != "" and pszCandidateProject == "合計":
                break
            if pszCandidateName == "":
                objDetailIndices.append(iNextSummaryIndex)
            iNextSummaryIndex += 1

        if objDetailIndices and len(objSummaryRow) > 6:
            objWeights: List[int] = []
            for iDetailIndex in objDetailIndices:
                objDetailRow: List[str] = objOutputRows[iDetailIndex]
                pszManhour: str = (objDetailRow[5] or "").strip() if len(objDetailRow) >= 6 else ""
                if pszManhour == "":
                    objWeights.append(0)
                    continue
                try:
                    objWeights.append(parse_time_text_to_seconds(pszManhour))
                except Exception:
                    objWeights.append(0)

            iWeightTotal: int = sum(objWeights)
            for iColumnIndex in range(6, len(objSummaryRow)):
                pszTotalText: str = objSummaryRow[iColumnIndex] if iColumnIndex < len(objSummaryRow) else ""
                objTotalValue: Decimal | None = parse_decimal_text(pszTotalText)
                if objTotalValue is None:
                    continue

                iScaleDigits: int = count_decimal_places(pszTotalText)
                iScale: int = 10 ** iScaleDigits
                objAbsTotalScaled: Decimal = (abs(objTotalValue) * Decimal(iScale)).quantize(Decimal("1"))
                iTotalScaledUnits: int = int(objAbsTotalScaled)

                objAllocatedUnits: List[int] = [0] * len(objDetailIndices)
                if iWeightTotal > 0 and iTotalScaledUnits > 0:
                    objFloors: List[int] = []
                    objRemainders: List[tuple[int, Decimal]] = []
                    for iIndex, iWeight in enumerate(objWeights):
                        if iWeight <= 0:
                            objFloors.append(0)
                            objRemainders.append((iIndex, Decimal("-1")))
                            continue
                        objRaw: Decimal = Decimal(iTotalScaledUnits) * Decimal(iWeight) / Decimal(iWeightTotal)
                        objFloorValue: Decimal = objRaw.to_integral_value(rounding=ROUND_FLOOR)
                        iFloor: int = int(objFloorValue)
                        objFloors.append(iFloor)
                        objRemainders.append((iIndex, objRaw - objFloorValue))

                    iFloorSum: int = sum(objFloors)
                    iRemaining: int = iTotalScaledUnits - iFloorSum
                    objRemainders.sort(key=lambda objItem: (-objItem[1], objItem[0]))
                    for iIndex, _ in objRemainders[:iRemaining]:
                        objFloors[iIndex] += 1
                    objAllocatedUnits = objFloors

                iSign: int = -1 if objTotalValue < 0 else 1
                for iIndex, iDetailIndex in enumerate(objDetailIndices):
                    objDetailRow: List[str] = objOutputRows[iDetailIndex]
                    while len(objDetailRow) <= iColumnIndex:
                        objDetailRow.append("")
                    objDetailRow[iColumnIndex] = format_scaled_units(iSign * objAllocatedUnits[iIndex], iScaleDigits)

        iRowIndex = iNextSummaryIndex

    objOutputPath: Path = build_new_rawdata_step0011_output_path_from_step0010(objNewRawdataStep0010Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0

def build_new_rawdata_step0012_output_path_from_step0011(objStep0011Path: Path) -> Path:
    pszFileName: str = objStep0011Path.name
    if "_step0011_" not in pszFileName:
        raise ValueError(f"Input is not step0011 file: {objStep0011Path}")
    pszOutputFileName: str = pszFileName.replace("_step0011_", "_step0012_", 1)
    return objStep0011Path.resolve().parent / pszOutputFileName


def process_new_rawdata_step0012_from_step0011(
    objNewRawdataStep0011Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0011Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0011Path}")

    objPrefixRows: List[List[str]] = []
    iRowIndex: int = 0
    while iRowIndex < len(objInputRows):
        objRow: List[str] = objInputRows[iRowIndex]
        pszStaffName: str = (objRow[3] or "").strip() if len(objRow) >= 4 else ""
        pszProject: str = (objRow[4] or "").strip() if len(objRow) >= 5 else ""
        if pszStaffName != "" and pszProject == "合計":
            break
        objPrefixRows.append(list(objRow))
        iRowIndex += 1

    objBlocks: List[tuple[tuple[int, float | str], int, List[List[str]]]] = []
    iBlockOrder: int = 0
    while iRowIndex < len(objInputRows):
        objRow: List[str] = objInputRows[iRowIndex]
        pszStaffName: str = (objRow[3] or "").strip() if len(objRow) >= 4 else ""
        pszProject: str = (objRow[4] or "").strip() if len(objRow) >= 5 else ""
        if not (pszStaffName != "" and pszProject == "合計"):
            objPrefixRows.append(list(objRow))
            iRowIndex += 1
            continue

        iBlockEnd: int = iRowIndex + 1
        while iBlockEnd < len(objInputRows):
            objNextRow: List[str] = objInputRows[iBlockEnd]
            pszNextStaffName: str = (objNextRow[3] or "").strip() if len(objNextRow) >= 4 else ""
            pszNextProject: str = (objNextRow[4] or "").strip() if len(objNextRow) >= 5 else ""
            if pszNextStaffName != "" and pszNextProject == "合計":
                break
            iBlockEnd += 1

        objBlockRows: List[List[str]] = [list(objBlockRow) for objBlockRow in objInputRows[iRowIndex:iBlockEnd]]
        pszStaffCodeHead: str = (objRow[1] or "").strip() if len(objRow) >= 2 else ""
        fSortNumber: float | None = parse_numeric_text(pszStaffCodeHead)
        if fSortNumber is None:
            objSortKey: tuple[int, float | str] = (1, pszStaffCodeHead)
        else:
            objSortKey = (0, fSortNumber)
        objBlocks.append((objSortKey, iBlockOrder, objBlockRows))
        iBlockOrder += 1
        iRowIndex = iBlockEnd

    objBlocks.sort(key=lambda objItem: (objItem[0], objItem[1]))

    objOutputRows: List[List[str]] = list(objPrefixRows)
    for _, _, objBlockRows in objBlocks:
        objOutputRows.extend(objBlockRows)

    objOutputPath: Path = build_new_rawdata_step0012_output_path_from_step0011(objNewRawdataStep0011Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def build_new_rawdata_step0013_output_path_from_step0012(objStep0012Path: Path) -> Path:
    pszFileName: str = objStep0012Path.name
    if "_step0012_" not in pszFileName:
        raise ValueError(f"Input is not step0012 file: {objStep0012Path}")
    pszOutputFileName: str = pszFileName.replace("_step0012_", "_step0013_", 1)
    return objStep0012Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0013_prepayed_commute_output_path_from_step0012(objStep0012Path: Path) -> Path:
    pszFileName: str = objStep0012Path.name
    if NEW_RAWDATA_STEP0012_PREPAYED_COMMUTE_FILE_PATTERN.match(pszFileName) is None:
        raise ValueError(f"Input is not prepaid commute step0012 file: {objStep0012Path}")
    pszOutputFileName: str = pszFileName.replace("_step0012_", "_step0013_", 1)
    if NEW_RAWDATA_STEP0013_PREPAYED_COMMUTE_FILE_PATTERN.match(pszOutputFileName) is None:
        raise ValueError(f"Could not build prepaid commute step0013 file name: {objStep0012Path}")
    return objStep0012Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0014_prepayed_commute_output_path_from_step0013(objStep0013Path: Path) -> Path:
    pszFileName: str = objStep0013Path.name
    if NEW_RAWDATA_STEP0013_PREPAYED_COMMUTE_FILE_PATTERN.match(pszFileName) is None:
        raise ValueError(f"Input is not prepaid commute step0013 file: {objStep0013Path}")
    pszOutputFileName: str = pszFileName.replace("_step0013_", "_step0014_", 1)
    if NEW_RAWDATA_STEP0014_PREPAYED_COMMUTE_FILE_PATTERN.match(pszOutputFileName) is None:
        raise ValueError(f"Could not build prepaid commute step0014 file name: {objStep0013Path}")
    return objStep0013Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0015_prepayed_commute_output_path_from_step0014(objStep0014Path: Path) -> Path:
    pszFileName: str = objStep0014Path.name
    if NEW_RAWDATA_STEP0014_PREPAYED_COMMUTE_FILE_PATTERN.match(pszFileName) is None:
        raise ValueError(f"Input is not prepaid commute step0014 file: {objStep0014Path}")
    pszOutputFileName: str = pszFileName.replace("_step0014_", "_step0015_", 1)
    if NEW_RAWDATA_STEP0015_PREPAYED_COMMUTE_FILE_PATTERN.match(pszOutputFileName) is None:
        raise ValueError(f"Could not build prepaid commute step0015 file name: {objStep0014Path}")
    return objStep0014Path.resolve().parent / pszOutputFileName


def build_debit_project_code_from_project_name_for_prepayed_commute(pszProjectName: str) -> str:
    iUnderscoreIndex: int = pszProjectName.find("_")
    if iUnderscoreIndex == -1:
        return ""
    pszProjectCodePrefix: str = pszProjectName[:iUnderscoreIndex]
    if re.match(r"^A\d{3}$", pszProjectCodePrefix) is not None:
        return "999"
    if re.match(r"^H\d{3}$", pszProjectCodePrefix) is not None:
        return "999"
    return pszProjectCodePrefix


def step0004_normalize_project_name_for_org_table(pszProjectName: str) -> str:
    pszNormalized: str = pszProjectName or ""
    pszNormalized = re.sub(
        r'"([^"]*)"',
        lambda objMatch: objMatch.group(1).replace("\t", "_"),
        pszNormalized,
    )
    pszNormalized = pszNormalized.replace('"', "")
    pszNormalized = re.sub(
        r"((?:P\d{5}|[A-OQ-Z]\d{3}))[\u0020\u3000]+",
        r"\1_",
        pszNormalized,
    )
    if pszNormalized.startswith("【"):
        objMatchBracket: re.Match[str] | None = re.search(
            r"(P\d{5}|[A-OQ-Z]\d{3})",
            pszNormalized,
        )
        if objMatchBracket is not None:
            pszCodeBracket: str = objMatchBracket.group(1)
            pszRestBracket: str = (
                pszNormalized[: objMatchBracket.start()]
                + pszNormalized[objMatchBracket.end() :]
            )
            return pszCodeBracket + "_" + pszRestBracket
    objMatchP: re.Match[str] | None = re.match(r"^(P\d{5})(.*)$", pszNormalized)
    if objMatchP is not None:
        pszCode: str = objMatchP.group(1)
        pszRest: str = objMatchP.group(2)
        if pszRest.startswith("【"):
            pszNormalized = pszCode + "_" + pszRest
    else:
        objMatchOther: re.Match[str] | None = re.match(r"^([A-OQ-Z]\d{3})(.*)$", pszNormalized)
        if objMatchOther is not None:
            pszCodeOther: str = objMatchOther.group(1)
            pszRestOther: str = objMatchOther.group(2)
            if pszRestOther.startswith("【"):
                pszNormalized = pszCodeOther + "_" + pszRestOther
    return pszNormalized


def normalize_org_table_project_code_step0004_for_prepayed_commute(pszProjectCode: str) -> str:
    pszNormalized: str = step0004_normalize_project_name_for_org_table(pszProjectCode or "")
    return re.sub(r"[ \u3000]+", "_", pszNormalized)


def write_org_table_tsv_from_csv_for_prepayed_commute(objBaseDirectoryPath: Path) -> int:
    objScriptDirectoryPath: Path = Path(__file__).resolve().parent
    objOrgTableCsvPath: Path = objScriptDirectoryPath / "管轄PJ表_定期代.csv"
    if not objOrgTableCsvPath.exists():
        objOrgTableCsvPath = objBaseDirectoryPath / "管轄PJ表_定期代.csv"

    objOrgTableTsvPath: Path = objBaseDirectoryPath / "管轄PJ表_定期代.tsv"
    if not objOrgTableCsvPath.exists():
        raise FileNotFoundError(f"管轄PJ表_定期代.csv が見つかりません。Path = {objOrgTableCsvPath}")

    objRows: List[List[str]] = []
    arrEncodings: List[str] = ["utf-8-sig", "cp932"]
    objLastDecodeError: Exception | None = None
    for pszEncoding in arrEncodings:
        try:
            with open(
                objOrgTableCsvPath,
                mode="r",
                encoding=pszEncoding,
                newline="",
            ) as objInputFile:
                objReader = csv.reader(objInputFile)
                for objRow in objReader:
                    objRows.append(list(objRow))
            objLastDecodeError = None
            break
        except UnicodeDecodeError as objError:
            objLastDecodeError = objError
            objRows = []
    if objLastDecodeError is not None:
        raise ValueError(f"unexpected exception while reading 管轄PJ表_定期代.csv. Detail = {objLastDecodeError}")

    for iRowIndex, objRow in enumerate(objRows):
        if len(objRow) > 1:
            objRow[1] = normalize_org_table_project_code_step0004_for_prepayed_commute(objRow[1])
        objRows[iRowIndex] = objRow

    objOrgTableTsvPath.parent.mkdir(parents=True, exist_ok=True)
    with open(objOrgTableTsvPath, mode="w", encoding="utf-8", newline="") as objOutputFile:
        objWriter: csv.writer = csv.writer(objOutputFile, delimiter="\t")
        for objRow in objRows:
            objWriter.writerow(objRow)
    return 0


def load_org_table_debit_project_codes_for_prepayed_commute(
    objOrgTableTsvPath: Path,
) -> Dict[str, List[str]]:
    if not objOrgTableTsvPath.exists():
        raise FileNotFoundError(f"管轄PJ表_定期代.tsv が見つかりません。Path = {objOrgTableTsvPath}")
    objRows: List[List[str]] = read_tsv_rows(objOrgTableTsvPath)
    if not objRows:
        raise ValueError(f"管轄PJ表_定期代.tsv has no rows: {objOrgTableTsvPath}")

    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    iProjectCodeIndex: int = objHeaderRow.index("PJコード") if "PJコード" in objHeaderRow else 1
    iDebitProjectCodeIndex: int = (
        objHeaderRow.index("借方プロジェクトコード")
        if "借方プロジェクトコード" in objHeaderRow
        else 5
    )
    iStartIndex: int = 1 if "PJコード" in objHeaderRow or "借方プロジェクトコード" in objHeaderRow else 0

    objCodeToDebits: Dict[str, List[str]] = {}
    for objRow in objRows[iStartIndex:]:
        if iProjectCodeIndex >= len(objRow):
            continue
        pszOrgProjectCodeText: str = (objRow[iProjectCodeIndex] or "").strip()
        if pszOrgProjectCodeText == "":
            continue
        objCodeMatch: re.Match[str] | None = re.match(r"^(P\d{5}|[A-OQ-Z]\d{3})", pszOrgProjectCodeText)
        if objCodeMatch is None:
            continue
        pszCode: str = objCodeMatch.group(1)
        pszDebitCode: str = (objRow[iDebitProjectCodeIndex] or "").strip() if iDebitProjectCodeIndex < len(objRow) else ""
        objCodeToDebits.setdefault(pszCode, []).append(pszDebitCode)
    return objCodeToDebits


def build_debit_project_code_from_project_name_and_org_table_for_prepayed_commute(
    pszProjectName: str,
    objCodeToDebits: Dict[str, List[str]],
    objErrorLines: List[str],
) -> str:
    iUnderscoreIndex: int = pszProjectName.find("_")
    if iUnderscoreIndex == -1:
        return ""
    pszProjectCodePrefix: str = pszProjectName[:iUnderscoreIndex]
    if re.match(r"^A\d{3}$", pszProjectCodePrefix) is not None:
        return "999"
    if re.match(r"^H\d{3}$", pszProjectCodePrefix) is not None:
        return "999"

    objCandidates: List[str] = objCodeToDebits.get(pszProjectCodePrefix, [])
    if len(objCandidates) == 0:
        objErrorLines.append(
            f"管轄PJ表に該当コードがありません。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
        )
        return pszProjectCodePrefix
    if len(objCandidates) >= 2:
        objErrorLines.append(
            f"管轄PJ表に一致候補が複数あります。先頭一致を採用しました。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
        )
    return objCandidates[0]


def process_new_rawdata_step0015_prepayed_commute_with_org_table_from_step0014(
    objStep0014Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0014Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0014Path}")
    objOrgTablePath: Path = objStep0014Path.resolve().parent / "管轄PJ表_定期代.tsv"
    objCodeToDebits: Dict[str, List[str]] = load_org_table_debit_project_codes_for_prepayed_commute(objOrgTablePath)

    objHeaderRow: List[str] = list(objInputRows[0])
    iProjectNameIndex: int = objHeaderRow.index("プロジェクト名") if "プロジェクト名" in objHeaderRow else 0
    iInsertIndex: int = iProjectNameIndex

    objOutputRows: List[List[str]] = []
    objErrorLines: List[str] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex == 0:
            pszDebitProjectCode: str = "借方プロジェクトコード"
        else:
            pszProjectName: str = objNewRow[iProjectNameIndex] if len(objNewRow) > iProjectNameIndex else ""
            pszDebitProjectCode = build_debit_project_code_from_project_name_and_org_table_for_prepayed_commute(
                pszProjectName,
                objCodeToDebits,
                objErrorLines,
            )
        objNewRow.insert(iInsertIndex, pszDebitProjectCode)
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0015_prepayed_commute_output_path_from_step0014(objStep0014Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0015_error_path_from_step0015(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    process_new_rawdata_step0016_prepayed_commute_from_step0015(objOutputPath)
    return 0


def process_new_rawdata_step0016_prepayed_commute_from_step0015(
    objStep0015Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0015Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0015Path}")

    objHeaderRow: List[str] = list(objInputRows[0])
    iDebitProjectCodeIndex: int = (
        objHeaderRow.index("借方プロジェクトコード")
        if "借方プロジェクトコード" in objHeaderRow
        else 0
    )
    iInsertIndex: int = iDebitProjectCodeIndex
    objInsertedHeaders: List[str] = [
        "借方税区分コード",
        "借方税率種別コード",
        "借方税率",
        "借方取引先コード",
    ]

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        objInsertedValues: List[str] = list(objInsertedHeaders) if iRowIndex == 0 else ["", "", "", ""]
        objNewRow[iInsertIndex:iInsertIndex] = objInsertedValues
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0016_output_path_from_step0015(objStep0015Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0017_prepayed_commute_from_step0016_with_org_table(objOutputPath)
    return 0


def load_org_table_subaccount_codes_for_prepayed_commute(
    objOrgTableTsvPath: Path,
) -> tuple[Dict[str, List[str]], List[str]]:
    if not objOrgTableTsvPath.exists():
        raise FileNotFoundError(f"管轄PJ表_定期代.tsv が見つかりません。Path = {objOrgTableTsvPath}")
    objRows: List[List[str]] = read_tsv_rows(objOrgTableTsvPath)
    if not objRows:
        raise ValueError(f"管轄PJ表_定期代.tsv has no rows: {objOrgTableTsvPath}")

    objErrors: List[str] = []
    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    if "PJコード" not in objHeaderRow or "補助科目コード" not in objHeaderRow:
        objErrors.append("管轄PJ表_定期代.tsv に必須列（PJコード / 補助科目コード）がありません。")
        return {}, objErrors
    iProjectCodeIndex: int = objHeaderRow.index("PJコード")
    iSubaccountCodeIndex: int = objHeaderRow.index("補助科目コード")

    objCodeToSubaccounts: Dict[str, List[str]] = {}
    for objRow in objRows[1:]:
        if iProjectCodeIndex >= len(objRow):
            continue
        pszOrgProjectCodeText: str = (objRow[iProjectCodeIndex] or "").strip()
        if pszOrgProjectCodeText == "":
            continue
        objCodeMatch: re.Match[str] | None = re.match(r"^(P\d{5}|[A-OQ-Z]\d{3})", pszOrgProjectCodeText)
        if objCodeMatch is None:
            continue
        pszCode: str = objCodeMatch.group(1)
        pszSubaccountCode: str = (
            (objRow[iSubaccountCodeIndex] or "").strip() if iSubaccountCodeIndex < len(objRow) else ""
        )
        objCodeToSubaccounts.setdefault(pszCode, []).append(pszSubaccountCode)
    return objCodeToSubaccounts, objErrors


def build_new_rawdata_step0017_error_path_from_step0017(objStep0017Path: Path) -> Path:
    return objStep0017Path.resolve().parent / f"{objStep0017Path.stem}_error.txt"


def process_new_rawdata_step0017_prepayed_commute_from_step0016_with_org_table(
    objStep0016Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0016Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0016Path}")

    objOrgTablePath: Path = objStep0016Path.resolve().parent / "管轄PJ表_定期代.tsv"
    objCodeToSubaccounts, objErrorLines = load_org_table_subaccount_codes_for_prepayed_commute(objOrgTablePath)

    objHeaderRow: List[str] = list(objInputRows[0])
    iDebitTaxCategoryCodeIndex: int = (
        objHeaderRow.index("借方税区分コード")
        if "借方税区分コード" in objHeaderRow
        else 0
    )
    iProjectNameIndex: int = objHeaderRow.index("プロジェクト名") if "プロジェクト名" in objHeaderRow else -1
    iInsertIndex: int = iDebitTaxCategoryCodeIndex

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex == 0:
            pszSubaccountCode: str = "借方補助科目コード"
        else:
            pszSubaccountCode = ""
            if iProjectNameIndex == -1:
                objErrorLines.append("step0016 にプロジェクト名列がありません。")
            else:
                pszProjectName: str = objNewRow[iProjectNameIndex] if len(objNewRow) > iProjectNameIndex else ""
                pszProjectCodePrefix: str = extract_project_code_prefix_step0017(pszProjectName)
                objCandidates: List[str] = objCodeToSubaccounts.get(pszProjectCodePrefix, [])
                if len(objCandidates) == 0:
                    objErrorLines.append(
                        f"管轄PJ表_定期代.tsv に該当する補助科目コードがありません。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
                    )
                elif len(objCandidates) >= 2:
                    pszSubaccountCode = objCandidates[0]
                    objErrorLines.append(
                        f"管轄PJ表_定期代.tsv に一致候補が複数あります。先頭一致を採用しました。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
                    )
                else:
                    pszSubaccountCode = objCandidates[0]
        objNewRow.insert(iInsertIndex, pszSubaccountCode)
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0017_output_path_from_step0016(objStep0016Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0017_error_path_from_step0017(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    process_new_rawdata_step0018_prepayed_commute_from_step0017_with_org_table(objOutputPath)
    return 0


def load_org_table_account_codes_for_prepayed_commute(
    objOrgTableTsvPath: Path,
) -> tuple[Dict[str, List[str]], List[str]]:
    if not objOrgTableTsvPath.exists():
        raise FileNotFoundError(f"管轄PJ表_定期代.tsv が見つかりません。Path = {objOrgTableTsvPath}")
    objRows: List[List[str]] = read_tsv_rows(objOrgTableTsvPath)
    if not objRows:
        raise ValueError(f"管轄PJ表_定期代.tsv has no rows: {objOrgTableTsvPath}")

    objErrors: List[str] = []
    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    if "PJコード" not in objHeaderRow or "勘定科目コード" not in objHeaderRow:
        objErrors.append("管轄PJ表_定期代.tsv に必須列（PJコード / 勘定科目コード）がありません。")
        return {}, objErrors
    iProjectCodeIndex: int = objHeaderRow.index("PJコード")
    iAccountCodeIndex: int = objHeaderRow.index("勘定科目コード")

    objCodeToAccounts: Dict[str, List[str]] = {}
    for objRow in objRows[1:]:
        if iProjectCodeIndex >= len(objRow):
            continue
        pszOrgProjectCodeText: str = (objRow[iProjectCodeIndex] or "").strip()
        if pszOrgProjectCodeText == "":
            continue
        objCodeMatch: re.Match[str] | None = re.match(r"^(P\d{5}|[A-OQ-Z]\d{3})", pszOrgProjectCodeText)
        if objCodeMatch is None:
            continue
        pszCode: str = objCodeMatch.group(1)
        pszAccountCode: str = (
            (objRow[iAccountCodeIndex] or "").strip() if iAccountCodeIndex < len(objRow) else ""
        )
        objCodeToAccounts.setdefault(pszCode, []).append(pszAccountCode)
    return objCodeToAccounts, objErrors


def build_new_rawdata_step0018_error_path_from_step0018(objStep0018Path: Path) -> Path:
    return objStep0018Path.resolve().parent / f"{objStep0018Path.stem}_error.txt"


def process_new_rawdata_step0018_prepayed_commute_from_step0017_with_org_table(
    objStep0017Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0017Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0017Path}")

    objOrgTablePath: Path = objStep0017Path.resolve().parent / "管轄PJ表_定期代.tsv"
    objCodeToAccounts, objErrorLines = load_org_table_account_codes_for_prepayed_commute(objOrgTablePath)

    objHeaderRow: List[str] = list(objInputRows[0])
    iDebitSubaccountCodeIndex: int = (
        objHeaderRow.index("借方補助科目コード")
        if "借方補助科目コード" in objHeaderRow
        else 0
    )
    iProjectNameIndex: int = objHeaderRow.index("プロジェクト名") if "プロジェクト名" in objHeaderRow else -1
    iInsertIndex: int = iDebitSubaccountCodeIndex

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex == 0:
            pszAccountCode: str = "借方勘定科目コード"
        else:
            pszAccountCode = ""
            if iProjectNameIndex == -1:
                objErrorLines.append("step0017 にプロジェクト名列がありません。")
            else:
                pszProjectName: str = objNewRow[iProjectNameIndex] if len(objNewRow) > iProjectNameIndex else ""
                pszProjectCodePrefix: str = extract_project_code_prefix_step0017(pszProjectName)
                objCandidates: List[str] = objCodeToAccounts.get(pszProjectCodePrefix, [])
                if len(objCandidates) == 0:
                    objErrorLines.append(
                        f"管轄PJ表_定期代.tsv に該当する勘定科目コードがありません。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
                    )
                elif len(objCandidates) >= 2:
                    pszAccountCode = objCandidates[0]
                    objErrorLines.append(
                        f"管轄PJ表_定期代.tsv に一致候補が複数あります。先頭一致を採用しました。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
                    )
                else:
                    pszAccountCode = objCandidates[0]
        objNewRow.insert(iInsertIndex, pszAccountCode)
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0018_output_path_from_step0017(objStep0017Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0018_error_path_from_step0018(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    process_new_rawdata_step0019_prepayed_commute_from_step0018_with_org_table(objOutputPath)
    return 0


def load_org_table_department_codes_for_prepayed_commute(
    objOrgTableTsvPath: Path,
) -> tuple[Dict[str, List[str]], List[str]]:
    if not objOrgTableTsvPath.exists():
        raise FileNotFoundError(f"管轄PJ表_定期代.tsv が見つかりません。Path = {objOrgTableTsvPath}")
    objRows: List[List[str]] = read_tsv_rows(objOrgTableTsvPath)
    if not objRows:
        raise ValueError(f"管轄PJ表_定期代.tsv has no rows: {objOrgTableTsvPath}")

    objErrors: List[str] = []
    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    if "PJコード" not in objHeaderRow or "部門コード" not in objHeaderRow:
        objErrors.append("管轄PJ表_定期代.tsv に必須列（PJコード / 部門コード）がありません。")
        return {}, objErrors
    iProjectCodeIndex: int = objHeaderRow.index("PJコード")
    iDepartmentCodeIndex: int = objHeaderRow.index("部門コード")

    objCodeToDepartments: Dict[str, List[str]] = {}
    for objRow in objRows[1:]:
        if iProjectCodeIndex >= len(objRow):
            continue
        pszOrgProjectCodeText: str = (objRow[iProjectCodeIndex] or "").strip()
        if pszOrgProjectCodeText == "":
            continue
        objCodeMatch: re.Match[str] | None = re.match(r"^(P\d{5}|[A-OQ-Z]\d{3})", pszOrgProjectCodeText)
        if objCodeMatch is None:
            continue
        pszCode: str = objCodeMatch.group(1)
        pszDepartmentCode: str = (
            (objRow[iDepartmentCodeIndex] or "").strip() if iDepartmentCodeIndex < len(objRow) else ""
        )
        objCodeToDepartments.setdefault(pszCode, []).append(pszDepartmentCode)
    return objCodeToDepartments, objErrors


def build_new_rawdata_step0019_error_path_from_step0019(objStep0019Path: Path) -> Path:
    return objStep0019Path.resolve().parent / f"{objStep0019Path.stem}_error.txt"


def process_new_rawdata_step0019_prepayed_commute_from_step0018_with_org_table(
    objStep0018Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0018Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0018Path}")

    objOrgTablePath: Path = objStep0018Path.resolve().parent / "管轄PJ表_定期代.tsv"
    objCodeToDepartments, objErrorLines = load_org_table_department_codes_for_prepayed_commute(objOrgTablePath)

    objHeaderRow: List[str] = list(objInputRows[0])
    iDebitAccountCodeIndex: int = (
        objHeaderRow.index("借方勘定科目コード")
        if "借方勘定科目コード" in objHeaderRow
        else 0
    )
    iProjectNameIndex: int = objHeaderRow.index("プロジェクト名") if "プロジェクト名" in objHeaderRow else -1
    iInsertIndex: int = iDebitAccountCodeIndex

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex == 0:
            pszDepartmentCode: str = "借方部門コード"
        else:
            pszDepartmentCode = ""
            if iProjectNameIndex == -1:
                objErrorLines.append("step0018 にプロジェクト名列がありません。")
            else:
                pszProjectName: str = objNewRow[iProjectNameIndex] if len(objNewRow) > iProjectNameIndex else ""
                pszProjectCodePrefix: str = extract_project_code_prefix_step0017(pszProjectName)
                objCandidates: List[str] = objCodeToDepartments.get(pszProjectCodePrefix, [])
                if len(objCandidates) == 0:
                    objErrorLines.append(
                        f"管轄PJ表_定期代.tsv に該当する部門コードがありません。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
                    )
                elif len(objCandidates) >= 2:
                    pszDepartmentCode = objCandidates[0]
                    objErrorLines.append(
                        f"管轄PJ表_定期代.tsv に一致候補が複数あります。先頭一致を採用しました。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
                    )
                else:
                    pszDepartmentCode = objCandidates[0]
        objNewRow.insert(iInsertIndex, pszDepartmentCode)
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0019_output_path_from_step0018(objStep0018Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0019_error_path_from_step0019(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    process_new_rawdata_step0020_prepayed_commute_from_step0019(objOutputPath)
    return 0


def build_new_rawdata_step0020_output_path_from_step0019(objStep0019Path: Path) -> Path:
    pszFileName: str = objStep0019Path.name
    if "_step0019_" not in pszFileName:
        raise ValueError(f"Input is not step0019 file: {objStep0019Path}")
    pszOutputFileName: str = pszFileName.replace("_step0019_", "_step0020_", 1)
    return objStep0019Path.resolve().parent / pszOutputFileName


def build_prepayed_commute_voucher_date_from_step0019_path(objStep0019Path: Path) -> str:
    objMatch: re.Match[str] | None = re.match(
        r"^新_ローデータ_シート_step0019_(\d{4})年(04-09月|10-03月)_(\d{2})月_前払通勤交通費按分表\.tsv$",
        objStep0019Path.name,
    )
    if objMatch is None:
        raise ValueError(f"Input is not prepaid commute step0019 file: {objStep0019Path}")
    iBaseYear: int = int(objMatch.group(1))
    pszPeriodLabel: str = objMatch.group(2)
    iMonth: int = int(objMatch.group(3))
    iTargetYear: int = iBaseYear
    if pszPeriodLabel == "10-03月" and 1 <= iMonth <= 3:
        iTargetYear = iBaseYear + 1
    iLastDay: int = calendar.monthrange(iTargetYear, iMonth)[1]
    return f"{iTargetYear}/{iMonth}/{iLastDay}"


def process_new_rawdata_step0020_prepayed_commute_from_step0019(
    objStep0019Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0019Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0019Path}")
    objOrgTablePath: Path = objStep0019Path.resolve().parent / "管轄PJ表_定期代.tsv"
    _ = read_tsv_rows(objOrgTablePath)

    pszVoucherDate: str = build_prepayed_commute_voucher_date_from_step0019_path(objStep0019Path)
    objHeaderRow: List[str] = list(objInputRows[0])
    iDebitDepartmentCodeIndex: int = (
        objHeaderRow.index("借方部門コード")
        if "借方部門コード" in objHeaderRow
        else 0
    )
    iInsertIndex: int = iDebitDepartmentCodeIndex

    objOutputRows: List[List[str]] = []
    objInsertedHeaders: List[str] = ["区切", "伝票区分コード", "日付"]
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        objInsertedValues: List[str] = list(objInsertedHeaders) if iRowIndex == 0 else ["", "", ""]
        objNewRow[iInsertIndex:iInsertIndex] = objInsertedValues
        objOutputRows.append(objNewRow)

    if len(objOutputRows) >= 2:
        objOutputRows[1][iInsertIndex] = "*"
        objOutputRows[1][iInsertIndex + 1] = "0"
    for objRow in objOutputRows[1:]:
        objRow[iInsertIndex + 2] = pszVoucherDate

    objOutputPath: Path = build_new_rawdata_step0020_output_path_from_step0019(objStep0019Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0021_prepayed_commute_from_step0020(objOutputPath)
    return 0


def build_new_rawdata_step0021_output_path_from_step0020(objStep0020Path: Path) -> Path:
    pszFileName: str = objStep0020Path.name
    if "_step0020_" not in pszFileName:
        raise ValueError(f"Input is not step0020 file: {objStep0020Path}")
    pszOutputFileName: str = pszFileName.replace("_step0020_", "_step0021_", 1)
    return objStep0020Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0021_error_path_from_step0021(objStep0021Path: Path) -> Path:
    return objStep0021Path.resolve().parent / f"{objStep0021Path.stem}_error.txt"


def process_new_rawdata_step0021_prepayed_commute_from_step0020(
    objStep0020Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0020Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0020Path}")

    objHeaderRow: List[str] = list(objInputRows[0])
    objErrorLines: List[str] = []
    objOutputPath: Path = build_new_rawdata_step0021_output_path_from_step0020(objStep0020Path)

    if "前払通勤交通費按分" not in objHeaderRow:
        objErrorLines.append("step0020 に前払通勤交通費按分列がありません。")
        objErrorPath: Path = build_new_rawdata_step0021_error_path_from_step0021(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
        raise ValueError(f"Missing required header: 前払通勤交通費按分. Input = {objStep0020Path}")
    iAmountIndex: int = objHeaderRow.index("前払通勤交通費按分")

    objDeleteColumnNames: List[str] = ["プロジェクト名", "工数"]
    objDeleteIndices: List[int] = []
    for pszColumnName in objDeleteColumnNames:
        if pszColumnName in objHeaderRow:
            objDeleteIndices.append(objHeaderRow.index(pszColumnName))
        else:
            objErrorLines.append(f"step0020 に{pszColumnName}列がありません。")

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex == 0 and iAmountIndex < len(objNewRow):
            objNewRow[iAmountIndex] = "借方本体金額"
        for iDeleteIndex in sorted(set(objDeleteIndices), reverse=True):
            if iDeleteIndex < len(objNewRow):
                del objNewRow[iDeleteIndex]
        objOutputRows.append(objNewRow)

    write_sheet_to_tsv(objOutputPath, objOutputRows)
    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0021_error_path_from_step0021(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    process_new_rawdata_step0022_prepayed_commute_from_step0021(objOutputPath)
    return 0


def process_new_rawdata_step0022_from_step0021(
    objStep0021Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0021Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0021Path}")

    objHeaderRow: List[str] = list(objInputRows[0])
    iDebitProjectCodeIndex: int = (
        objHeaderRow.index("借方プロジェクトコード")
        if "借方プロジェクトコード" in objHeaderRow
        else 0
    )
    iInsertIndex: int = iDebitProjectCodeIndex
    objInsertedHeaders: List[str] = [
        "借方税区分コード",
        "借方税率種別コード",
        "借方税率",
        "借方取引先コード",
    ]

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        objInsertedValues: List[str] = list(objInsertedHeaders) if iRowIndex == 0 else ["", "", "", ""]
        objNewRow[iInsertIndex:iInsertIndex] = objInsertedValues
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0022_output_path_from_step0021(objStep0021Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0023_from_step0022_with_org_table(objOutputPath)
    return 0


def load_org_table_subaccount_codes_for_salary(
    objOrgTableTsvPath: Path,
) -> tuple[Dict[str, List[str]], List[str]]:
    if not objOrgTableTsvPath.exists():
        raise FileNotFoundError(f"管轄PJ表_給与.tsv が見つかりません。Path = {objOrgTableTsvPath}")
    objRows: List[List[str]] = read_tsv_rows(objOrgTableTsvPath)
    if not objRows:
        raise ValueError(f"管轄PJ表_給与.tsv has no rows: {objOrgTableTsvPath}")

    objErrors: List[str] = []
    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    if "PJコード" not in objHeaderRow or "補助科目コード" not in objHeaderRow:
        objErrors.append("管轄PJ表_給与.tsv に必須列（PJコード / 補助科目コード）がありません。")
        return {}, objErrors
    iProjectCodeIndex: int = objHeaderRow.index("PJコード")
    iSubaccountCodeIndex: int = objHeaderRow.index("補助科目コード")

    objCodeToSubaccounts: Dict[str, List[str]] = {}
    for objRow in objRows[1:]:
        if iProjectCodeIndex >= len(objRow):
            continue
        pszOrgProjectCodeText: str = (objRow[iProjectCodeIndex] or "").strip()
        if pszOrgProjectCodeText == "":
            continue
        objCodeMatch: re.Match[str] | None = re.match(r"^(P\d{5}|[A-OQ-Z]\d{3})", pszOrgProjectCodeText)
        if objCodeMatch is None:
            continue
        pszCode: str = objCodeMatch.group(1)
        pszSubaccountCode: str = (
            (objRow[iSubaccountCodeIndex] or "").strip() if iSubaccountCodeIndex < len(objRow) else ""
        )
        objCodeToSubaccounts.setdefault(pszCode, []).append(pszSubaccountCode)
    return objCodeToSubaccounts, objErrors


def process_new_rawdata_step0023_from_step0022_with_org_table(
    objStep0022Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0022Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0022Path}")

    objOrgTablePath: Path = objStep0022Path.resolve().parent / "管轄PJ表_給与.tsv"
    objCodeToSubaccounts, objErrorLines = load_org_table_subaccount_codes_for_salary(objOrgTablePath)

    objHeaderRow: List[str] = list(objInputRows[0])
    iDebitTaxCategoryCodeIndex: int = (
        objHeaderRow.index("借方税区分コード")
        if "借方税区分コード" in objHeaderRow
        else 0
    )
    iProjectNameIndex: int = objHeaderRow.index("プロジェクト名") if "プロジェクト名" in objHeaderRow else -1
    iInsertIndex: int = iDebitTaxCategoryCodeIndex

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex == 0:
            pszSubaccountCode: str = "借方補助科目コード"
        else:
            pszSubaccountCode = ""
            if iProjectNameIndex == -1:
                objErrorLines.append("step0022 にプロジェクト名列がありません。")
            else:
                pszProjectName: str = objNewRow[iProjectNameIndex] if len(objNewRow) > iProjectNameIndex else ""
                pszProjectCodePrefix: str = extract_project_code_prefix_step0017(pszProjectName)
                objCandidates: List[str] = objCodeToSubaccounts.get(pszProjectCodePrefix, [])
                if len(objCandidates) == 0:
                    objErrorLines.append(
                        f"管轄PJ表_給与.tsv に該当する補助科目コードがありません。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
                    )
                elif len(objCandidates) >= 2:
                    pszSubaccountCode = objCandidates[0]
                    objErrorLines.append(
                        f"管轄PJ表_給与.tsv に一致候補が複数あります。先頭一致を採用しました。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
                    )
                else:
                    pszSubaccountCode = objCandidates[0]
        objNewRow.insert(iInsertIndex, pszSubaccountCode)
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0023_output_path_from_step0022(objStep0022Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0023_error_path_from_step0023(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    process_new_rawdata_step0024_from_step0023_with_org_table(objOutputPath)
    return 0


def build_new_rawdata_step0024_output_path_from_step0023(objStep0023Path: Path) -> Path:
    pszFileName: str = objStep0023Path.name
    if "_step0023_" not in pszFileName:
        raise ValueError(f"Input is not step0023 file: {objStep0023Path}")
    pszOutputFileName: str = pszFileName.replace("_step0023_", "_step0024_", 1)
    return objStep0023Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0024_error_path_from_step0024(objStep0024Path: Path) -> Path:
    return objStep0024Path.resolve().parent / f"{objStep0024Path.stem}_error.txt"


def load_org_table_account_codes_for_salary(
    objOrgTableTsvPath: Path,
) -> tuple[Dict[str, List[str]], List[str]]:
    if not objOrgTableTsvPath.exists():
        raise FileNotFoundError(f"管轄PJ表_給与.tsv が見つかりません。Path = {objOrgTableTsvPath}")
    objRows: List[List[str]] = read_tsv_rows(objOrgTableTsvPath)
    if not objRows:
        raise ValueError(f"管轄PJ表_給与.tsv has no rows: {objOrgTableTsvPath}")

    objErrors: List[str] = []
    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    if "PJコード" not in objHeaderRow or "勘定科目コード" not in objHeaderRow:
        objErrors.append("管轄PJ表_給与.tsv に必須列（PJコード / 勘定科目コード）がありません。")
        return {}, objErrors
    iProjectCodeIndex: int = objHeaderRow.index("PJコード")
    iAccountCodeIndex: int = objHeaderRow.index("勘定科目コード")

    objCodeToAccounts: Dict[str, List[str]] = {}
    for objRow in objRows[1:]:
        if iProjectCodeIndex >= len(objRow):
            continue
        pszOrgProjectCodeText: str = (objRow[iProjectCodeIndex] or "").strip()
        if pszOrgProjectCodeText == "":
            continue
        objCodeMatch: re.Match[str] | None = re.match(r"^(P\d{5}|[A-OQ-Z]\d{3})", pszOrgProjectCodeText)
        if objCodeMatch is None:
            continue
        pszCode: str = objCodeMatch.group(1)
        pszAccountCode: str = (
            (objRow[iAccountCodeIndex] or "").strip() if iAccountCodeIndex < len(objRow) else ""
        )
        objCodeToAccounts.setdefault(pszCode, []).append(pszAccountCode)
    return objCodeToAccounts, objErrors


def process_new_rawdata_step0024_from_step0023_with_org_table(
    objStep0023Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0023Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0023Path}")

    objOrgTablePath: Path = objStep0023Path.resolve().parent / "管轄PJ表_給与.tsv"
    objCodeToAccounts, objErrorLines = load_org_table_account_codes_for_salary(objOrgTablePath)

    objHeaderRow: List[str] = list(objInputRows[0])
    iDebitSubaccountCodeIndex: int = (
        objHeaderRow.index("借方補助科目コード")
        if "借方補助科目コード" in objHeaderRow
        else 0
    )
    iProjectNameIndex: int = objHeaderRow.index("プロジェクト名") if "プロジェクト名" in objHeaderRow else -1
    iInsertIndex: int = iDebitSubaccountCodeIndex

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex == 0:
            pszAccountCode: str = "借方勘定科目コード"
        else:
            pszAccountCode = ""
            if iProjectNameIndex == -1:
                objErrorLines.append("step0023 にプロジェクト名列がありません。")
            else:
                pszProjectName: str = objNewRow[iProjectNameIndex] if len(objNewRow) > iProjectNameIndex else ""
                pszProjectCodePrefix: str = extract_project_code_prefix_step0017(pszProjectName)
                objCandidates: List[str] = objCodeToAccounts.get(pszProjectCodePrefix, [])
                if len(objCandidates) == 0:
                    objErrorLines.append(
                        f"管轄PJ表_給与.tsv に該当する勘定科目コードがありません。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
                    )
                elif len(objCandidates) >= 2:
                    pszAccountCode = objCandidates[0]
                    objErrorLines.append(
                        f"管轄PJ表_給与.tsv に一致候補が複数あります。先頭一致を採用しました。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
                    )
                else:
                    pszAccountCode = objCandidates[0]
        objNewRow.insert(iInsertIndex, pszAccountCode)
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0024_output_path_from_step0023(objStep0023Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0024_error_path_from_step0024(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    process_new_rawdata_step0025_from_step0024_with_org_table(objOutputPath)
    return 0


def build_new_rawdata_step0025_output_path_from_step0024(objStep0024Path: Path) -> Path:
    pszFileName: str = objStep0024Path.name
    if "_step0024_" not in pszFileName:
        raise ValueError(f"Input is not step0024 file: {objStep0024Path}")
    pszOutputFileName: str = pszFileName.replace("_step0024_", "_step0025_", 1)
    return objStep0024Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0025_error_path_from_step0025(objStep0025Path: Path) -> Path:
    return objStep0025Path.resolve().parent / f"{objStep0025Path.stem}_error.txt"


def build_new_rawdata_step0026_output_path_from_step0025(objStep0025Path: Path) -> Path:
    pszFileName: str = objStep0025Path.name
    if "_step0025_" not in pszFileName:
        raise ValueError(f"Input is not step0025 file: {objStep0025Path}")
    pszOutputFileName: str = pszFileName.replace("_step0025_", "_step0026_", 1)
    return objStep0025Path.resolve().parent / pszOutputFileName


def build_salary_voucher_date_from_step0025_path(objStep0025Path: Path) -> str:
    objMatch: re.Match[str] | None = re.match(
        r"^新_ローデータ_シート_step0025_(\d{4})年(\d{2})月\.tsv$",
        objStep0025Path.name,
    )
    if objMatch is None:
        raise ValueError(f"Input is not salary step0025 file: {objStep0025Path}")
    iYear: int = int(objMatch.group(1))
    iMonth: int = int(objMatch.group(2))
    iLastDay: int = calendar.monthrange(iYear, iMonth)[1]
    return f"{iYear}/{iMonth}/{iLastDay}"


def build_new_rawdata_step0027_output_path_from_step0026(objStep0026Path: Path) -> Path:
    pszFileName: str = objStep0026Path.name
    if "_step0026_" not in pszFileName:
        raise ValueError(f"Input is not step0026 file: {objStep0026Path}")
    pszOutputFileName: str = pszFileName.replace("_step0026_", "_step0027_", 1)
    return objStep0026Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0027_error_path_from_step0027(objStep0027Path: Path) -> Path:
    return objStep0027Path.resolve().parent / f"{objStep0027Path.stem}_error.txt"


def load_org_table_department_codes_for_salary(
    objOrgTableTsvPath: Path,
) -> tuple[Dict[str, List[str]], List[str]]:
    if not objOrgTableTsvPath.exists():
        raise FileNotFoundError(f"管轄PJ表_給与.tsv が見つかりません。Path = {objOrgTableTsvPath}")
    objRows: List[List[str]] = read_tsv_rows(objOrgTableTsvPath)
    if not objRows:
        raise ValueError(f"管轄PJ表_給与.tsv has no rows: {objOrgTableTsvPath}")

    objErrors: List[str] = []
    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    if "PJコード" not in objHeaderRow or "部門コード" not in objHeaderRow:
        objErrors.append("管轄PJ表_給与.tsv に必須列（PJコード / 部門コード）がありません。")
        return {}, objErrors
    iProjectCodeIndex: int = objHeaderRow.index("PJコード")
    iDepartmentCodeIndex: int = objHeaderRow.index("部門コード")

    objCodeToDepartments: Dict[str, List[str]] = {}
    for objRow in objRows[1:]:
        if iProjectCodeIndex >= len(objRow):
            continue
        pszOrgProjectCodeText: str = (objRow[iProjectCodeIndex] or "").strip()
        if pszOrgProjectCodeText == "":
            continue
        objCodeMatch: re.Match[str] | None = re.match(r"^(P\d{5}|[A-OQ-Z]\d{3})", pszOrgProjectCodeText)
        if objCodeMatch is None:
            continue
        pszCode: str = objCodeMatch.group(1)
        pszDepartmentCode: str = (
            (objRow[iDepartmentCodeIndex] or "").strip() if iDepartmentCodeIndex < len(objRow) else ""
        )
        objCodeToDepartments.setdefault(pszCode, []).append(pszDepartmentCode)
    return objCodeToDepartments, objErrors


def process_new_rawdata_step0025_from_step0024_with_org_table(
    objStep0024Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0024Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0024Path}")

    objOrgTablePath: Path = objStep0024Path.resolve().parent / "管轄PJ表_給与.tsv"
    objCodeToDepartments, objErrorLines = load_org_table_department_codes_for_salary(objOrgTablePath)

    objHeaderRow: List[str] = list(objInputRows[0])
    iDebitAccountCodeIndex: int = (
        objHeaderRow.index("借方勘定科目コード")
        if "借方勘定科目コード" in objHeaderRow
        else 0
    )
    iProjectNameIndex: int = objHeaderRow.index("プロジェクト名") if "プロジェクト名" in objHeaderRow else -1
    iInsertIndex: int = iDebitAccountCodeIndex

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex == 0:
            pszDepartmentCode: str = "借方部門コード"
        else:
            pszDepartmentCode = ""
            if iProjectNameIndex == -1:
                objErrorLines.append("step0024 にプロジェクト名列がありません。")
            else:
                pszProjectName: str = objNewRow[iProjectNameIndex] if len(objNewRow) > iProjectNameIndex else ""
                pszProjectCodePrefix: str = extract_project_code_prefix_step0017(pszProjectName)
                objCandidates: List[str] = objCodeToDepartments.get(pszProjectCodePrefix, [])
                if len(objCandidates) == 0:
                    objErrorLines.append(
                        f"管轄PJ表_給与.tsv に該当する部門コードがありません。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
                    )
                elif len(objCandidates) >= 2:
                    pszDepartmentCode = objCandidates[0]
                    objErrorLines.append(
                        f"管轄PJ表_給与.tsv に一致候補が複数あります。先頭一致を採用しました。プロジェクト名={pszProjectName} コード={pszProjectCodePrefix}"
                    )
                else:
                    pszDepartmentCode = objCandidates[0]
        objNewRow.insert(iInsertIndex, pszDepartmentCode)
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0025_output_path_from_step0024(objStep0024Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0025_error_path_from_step0025(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    process_new_rawdata_step0026_from_step0025(objOutputPath)
    return 0


def process_new_rawdata_step0026_from_step0025(
    objStep0025Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0025Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0025Path}")

    pszVoucherDate: str = build_salary_voucher_date_from_step0025_path(objStep0025Path)
    objHeaderRow: List[str] = list(objInputRows[0])
    iDebitDepartmentCodeIndex: int = (
        objHeaderRow.index("借方部門コード")
        if "借方部門コード" in objHeaderRow
        else 0
    )
    iInsertIndex: int = iDebitDepartmentCodeIndex
    objInsertedHeaders: List[str] = ["区切", "伝票区分コード", "日付"]

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        objInsertedValues: List[str] = list(objInsertedHeaders) if iRowIndex == 0 else ["", "", ""]
        objNewRow[iInsertIndex:iInsertIndex] = objInsertedValues
        objOutputRows.append(objNewRow)

    if len(objOutputRows) >= 2:
        objOutputRows[1][iInsertIndex] = "*"
        objOutputRows[1][iInsertIndex + 1] = "0"
    for objRow in objOutputRows[1:]:
        objRow[iInsertIndex + 2] = pszVoucherDate

    objOutputPath: Path = build_new_rawdata_step0026_output_path_from_step0025(objStep0025Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0027_from_step0026(objOutputPath)
    return 0


def process_new_rawdata_step0027_from_step0026(
    objStep0026Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0026Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0026Path}")

    objHeaderRow: List[str] = list(objInputRows[0])
    objErrorLines: List[str] = []
    objOutputPath: Path = build_new_rawdata_step0027_output_path_from_step0026(objStep0026Path)

    if "給与合計" not in objHeaderRow:
        objErrorLines.append("step0026 に給与合計列がありません。")
        objErrorPath: Path = build_new_rawdata_step0027_error_path_from_step0027(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
        raise ValueError(f"Missing required header: 給与合計. Input = {objStep0026Path}")
    iAmountIndex: int = objHeaderRow.index("給与合計")

    objDeleteColumnName: str = "プロジェクト名"
    objDeleteIndices: List[int] = []
    if objDeleteColumnName in objHeaderRow:
        objDeleteIndices.append(objHeaderRow.index(objDeleteColumnName))
    else:
        objErrorLines.append(f"step0026 に{objDeleteColumnName}列がありません。")

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex == 0 and iAmountIndex < len(objNewRow):
            objNewRow[iAmountIndex] = "借方本体金額"
        for iDeleteIndex in sorted(set(objDeleteIndices), reverse=True):
            if iDeleteIndex < len(objNewRow):
                del objNewRow[iDeleteIndex]
        objOutputRows.append(objNewRow)

    write_sheet_to_tsv(objOutputPath, objOutputRows)
    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0027_error_path_from_step0027(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    process_new_rawdata_step0028_from_step0027(objOutputPath)
    return 0


def build_new_rawdata_step0028_output_path_from_step0027(objStep0027Path: Path) -> Path:
    pszFileName: str = objStep0027Path.name
    if "_step0027_" not in pszFileName:
        raise ValueError(f"Input is not step0027 file: {objStep0027Path}")
    pszOutputFileName: str = pszFileName.replace("_step0027_", "_step0028_", 1)
    return objStep0027Path.resolve().parent / pszOutputFileName


def process_new_rawdata_step0028_from_step0027(
    objStep0027Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0027Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0027Path}")

    objHeaderRow: List[str] = list(objInputRows[0])
    iDebitAmountIndex: int = objHeaderRow.index("借方本体金額") if "借方本体金額" in objHeaderRow else len(objHeaderRow)
    iInsertIndex: int = iDebitAmountIndex + 1
    objInsertedHeaders: List[str] = [
        "貸方部門コード",
        "貸方勘定科目コード",
        "貸方補助科目コード",
        "貸方税区分コード",
        "貸方税率種別コード",
        "貸方税率",
        "貸方取引先コード",
        "貸方プロジェクトコード",
        "貸方本体金額",
        "摘要",
    ]

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        objInsertedValues: List[str] = list(objInsertedHeaders) if iRowIndex == 0 else [""] * len(objInsertedHeaders)
        objNewRow[iInsertIndex:iInsertIndex] = objInsertedValues
        objOutputRows.append(objNewRow)

    iRequiredColumns: int = len(objOutputRows[0]) if objOutputRows else iInsertIndex + len(objInsertedHeaders)
    if len(objOutputRows) < 2:
        objOutputRows.append([""] * iRequiredColumns)
    if len(objOutputRows[1]) < iRequiredColumns:
        objOutputRows[1].extend([""] * (iRequiredColumns - len(objOutputRows[1])))
    objOutputRows[1][iInsertIndex + 0] = "0"
    objOutputRows[1][iInsertIndex + 1] = "151"
    objOutputRows[1][iInsertIndex + 2] = "6"
    objOutputRows[1][iInsertIndex + 9] = "社員通勤定期代振替"

    objOutputPath: Path = build_new_rawdata_step0028_output_path_from_step0027(objStep0027Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0029_from_step0028(objOutputPath)
    return 0


def build_new_rawdata_step0029_output_path_from_step0028(objStep0028Path: Path) -> Path:
    pszFileName: str = objStep0028Path.name
    if "_step0028_" not in pszFileName:
        raise ValueError(f"Input is not step0028 file: {objStep0028Path}")
    pszOutputFileName: str = pszFileName.replace("_step0028_", "_step0029_", 1)
    return objStep0028Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0029_error_path_from_step0029(objStep0029Path: Path) -> Path:
    return objStep0029Path.resolve().parent / f"{objStep0029Path.stem}_error.txt"


def process_new_rawdata_step0029_from_step0028(
    objStep0028Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0028Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0028Path}")

    objOutputRows: List[List[str]] = [list(objRow) for objRow in objInputRows]
    objHeaderRow: List[str] = list(objOutputRows[0])
    objErrorLines: List[str] = []
    objOutputPath: Path = build_new_rawdata_step0029_output_path_from_step0028(objStep0028Path)

    if "借方本体金額" not in objHeaderRow or "貸方本体金額" not in objHeaderRow:
        if "借方本体金額" not in objHeaderRow:
            objErrorLines.append("step0028 に借方本体金額列がありません。")
        if "貸方本体金額" not in objHeaderRow:
            objErrorLines.append("step0028 に貸方本体金額列がありません。")
        objErrorPath: Path = build_new_rawdata_step0029_error_path_from_step0029(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
        raise ValueError(f"Missing required headers. Input = {objStep0028Path}")

    iDebitAmountIndex: int = objHeaderRow.index("借方本体金額")
    iCreditAmountIndex: int = objHeaderRow.index("貸方本体金額")
    iRequiredColumns: int = len(objHeaderRow)
    if len(objOutputRows) < 2:
        objOutputRows.append([""] * iRequiredColumns)
    if len(objOutputRows[1]) < iRequiredColumns:
        objOutputRows[1].extend([""] * (iRequiredColumns - len(objOutputRows[1])))

    iDebitTotal: int = 0
    for iRowIndex, objRow in enumerate(objOutputRows[1:], start=2):
        if iDebitAmountIndex >= len(objRow):
            objErrorLines.append(f"{iRowIndex}行目: 借方本体金額列が不足しているため0扱いでスキップしました。")
            continue
        pszDebitAmountRaw: str = (objRow[iDebitAmountIndex] or "").strip()
        if pszDebitAmountRaw == "":
            objErrorLines.append(f"{iRowIndex}行目: 借方本体金額が空欄のため0扱いでスキップしました。")
            continue
        pszDebitAmountNormalized: str = pszDebitAmountRaw.replace(",", "")
        try:
            iDebitTotal += int(Decimal(pszDebitAmountNormalized))
        except (InvalidOperation, ValueError):
            objErrorLines.append(
                f"{iRowIndex}行目: 借方本体金額[{pszDebitAmountRaw}]は数値化できないため0扱いでスキップしました。"
            )
            continue

    objOutputRows[1][iCreditAmountIndex] = str(iDebitTotal)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0029_error_path_from_step0029(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    return 0


def build_new_rawdata_step0022_output_path_from_step0021(objStep0021Path: Path) -> Path:
    pszFileName: str = objStep0021Path.name
    if "_step0021_" not in pszFileName:
        raise ValueError(f"Input is not step0021 file: {objStep0021Path}")
    pszOutputFileName: str = pszFileName.replace("_step0021_", "_step0022_", 1)
    return objStep0021Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0023_output_path_from_step0022(objStep0022Path: Path) -> Path:
    pszFileName: str = objStep0022Path.name
    if "_step0022_" not in pszFileName:
        raise ValueError(f"Input is not step0022 file: {objStep0022Path}")
    pszOutputFileName: str = pszFileName.replace("_step0022_", "_step0023_", 1)
    return objStep0022Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0023_error_path_from_step0023(objStep0023Path: Path) -> Path:
    return objStep0023Path.resolve().parent / f"{objStep0023Path.stem}_error.txt"


def resolve_accounting_csv_year_month_from_step0023_file_name(objStep0023Path: Path) -> tuple[int, int]:
    objMatch = PREPAYED_COMMUTE_STEP0023_FILE_PATTERN.match(objStep0023Path.name)
    if objMatch is None:
        raise ValueError(f"Input is not prepaid commute step0023 file: {objStep0023Path}")

    iFiscalStartYear: int = int(objMatch.group(1))
    iMonth: int = int(objMatch.group(2))
    if iMonth in (10, 11, 12):
        return iFiscalStartYear, iMonth
    if iMonth in (1, 2, 3):
        return iFiscalStartYear + 1, iMonth
    raise ValueError(f"Unsupported month in step0023 file: {objStep0023Path}")


def build_salary_project_transfer_csv_output_paths_from_step0023(objStep0023Path: Path) -> tuple[Path, Path]:
    iYear: int
    iMonth: int
    iYear, iMonth = resolve_accounting_csv_year_month_from_step0023_file_name(objStep0023Path)
    pszYearMonthLabel: str = f"{iYear % 100:02d}.{iMonth:02d}月"
    objBaseDirectory: Path = objStep0023Path.resolve().parent
    objUtf8Path: Path = objBaseDirectory / f"通勤費プロジェクト振替_{pszYearMonthLabel}_勘定奉行用.csv"
    objSjisPath: Path = objBaseDirectory / f"通勤費プロジェクト振替_{pszYearMonthLabel}_勘定奉行用_sjis.csv"
    return objUtf8Path, objSjisPath


def write_rows_to_csv(objOutputPath: Path, objRows: List[List[str]], pszEncoding: str) -> None:
    with open(objOutputPath, mode="w", encoding=pszEncoding, newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter=",", lineterminator="\n")
        for objRow in objRows:
            objWriter.writerow(objRow)


def process_salary_project_transfer_csv_from_step0023(objStep0023Path: Path) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0023Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0023Path}")

    objUtf8Path: Path
    objSjisPath: Path
    objUtf8Path, objSjisPath = build_salary_project_transfer_csv_output_paths_from_step0023(objStep0023Path)
    write_rows_to_csv(objUtf8Path, objInputRows, "utf-8")
    write_rows_to_csv(objSjisPath, objInputRows, "cp932")
    return 0


def process_new_rawdata_step0022_prepayed_commute_from_step0021(
    objStep0021Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0021Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0021Path}")

    objHeaderRow: List[str] = list(objInputRows[0])
    iDebitAmountIndex: int = objHeaderRow.index("借方本体金額") if "借方本体金額" in objHeaderRow else len(objHeaderRow)
    iInsertIndex: int = iDebitAmountIndex + 1
    objInsertedHeaders: List[str] = [
        "貸方部門コード",
        "貸方勘定科目コード",
        "貸方補助科目コード",
        "貸方税区分コード",
        "貸方税率種別コード",
        "貸方税率",
        "貸方取引先コード",
        "貸方プロジェクトコード",
        "貸方本体金額",
        "摘要",
    ]

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        objInsertedValues: List[str] = list(objInsertedHeaders) if iRowIndex == 0 else [""] * len(objInsertedHeaders)
        objNewRow[iInsertIndex:iInsertIndex] = objInsertedValues
        objOutputRows.append(objNewRow)

    iRequiredColumns: int = len(objOutputRows[0]) if objOutputRows else iInsertIndex + len(objInsertedHeaders)
    if len(objOutputRows) < 2:
        objOutputRows.append([""] * iRequiredColumns)
    if len(objOutputRows[1]) < iRequiredColumns:
        objOutputRows[1].extend([""] * (iRequiredColumns - len(objOutputRows[1])))
    objOutputRows[1][iInsertIndex + 0] = "0"
    objOutputRows[1][iInsertIndex + 1] = "151"
    objOutputRows[1][iInsertIndex + 2] = "6"
    objOutputRows[1][iInsertIndex + 9] = "社員通勤定期代振替"

    objOutputPath: Path = build_new_rawdata_step0022_output_path_from_step0021(objStep0021Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0023_prepayed_commute_from_step0022(objOutputPath)
    return 0


def process_new_rawdata_step0023_prepayed_commute_from_step0022(
    objStep0022Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0022Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0022Path}")

    objOutputRows: List[List[str]] = [list(objRow) for objRow in objInputRows]
    objHeaderRow: List[str] = list(objOutputRows[0])
    objErrorLines: List[str] = []
    objOutputPath: Path = build_new_rawdata_step0023_output_path_from_step0022(objStep0022Path)

    if "借方本体金額" not in objHeaderRow or "貸方本体金額" not in objHeaderRow:
        if "借方本体金額" not in objHeaderRow:
            objErrorLines.append("step0022 に借方本体金額列がありません。")
        if "貸方本体金額" not in objHeaderRow:
            objErrorLines.append("step0022 に貸方本体金額列がありません。")
        objErrorPath: Path = build_new_rawdata_step0023_error_path_from_step0023(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
        raise ValueError(f"Missing required headers. Input = {objStep0022Path}")

    iDebitAmountIndex: int = objHeaderRow.index("借方本体金額")
    iCreditAmountIndex: int = objHeaderRow.index("貸方本体金額")
    iRequiredColumns: int = len(objHeaderRow)
    if len(objOutputRows) < 2:
        objOutputRows.append([""] * iRequiredColumns)
    if len(objOutputRows[1]) < iRequiredColumns:
        objOutputRows[1].extend([""] * (iRequiredColumns - len(objOutputRows[1])))

    iDebitTotal: int = 0
    for iRowIndex, objRow in enumerate(objOutputRows[1:], start=2):
        if iDebitAmountIndex >= len(objRow):
            objErrorLines.append(f"{iRowIndex}行目: 借方本体金額列が不足しているため0扱いでスキップしました。")
            continue
        pszDebitAmountRaw: str = (objRow[iDebitAmountIndex] or "").strip()
        if pszDebitAmountRaw == "":
            objErrorLines.append(f"{iRowIndex}行目: 借方本体金額が空欄のため0扱いでスキップしました。")
            continue
        pszDebitAmountNormalized: str = pszDebitAmountRaw.replace(",", "")
        try:
            iDebitTotal += int(Decimal(pszDebitAmountNormalized))
        except (InvalidOperation, ValueError):
            objErrorLines.append(
                f"{iRowIndex}行目: 借方本体金額[{pszDebitAmountRaw}]は数値化できないため0扱いでスキップしました。"
            )
            continue

    objOutputRows[1][iCreditAmountIndex] = str(iDebitTotal)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_salary_project_transfer_csv_from_step0023(objOutputPath)
    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0023_error_path_from_step0023(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    return 0


def build_new_rawdata_step0013_nontax_commute_output_path_from_step0012(objStep0012Path: Path) -> Path:
    pszBasePath: Path = build_new_rawdata_step0013_output_path_from_step0012(objStep0012Path)
    pszFileName: str = pszBasePath.name
    if "_step0013_" not in pszFileName:
        raise ValueError(f"Input is not step0013 file: {pszBasePath}")
    pszOutputFileName: str = pszFileName.replace("_step0013_", "_step0013_非課税通勤手当_", 1)
    return pszBasePath.resolve().parent / pszOutputFileName


def build_new_rawdata_step0013_statutory_welfare_output_path_from_step0012(objStep0012Path: Path) -> Path:
    pszBasePath: Path = build_new_rawdata_step0013_output_path_from_step0012(objStep0012Path)
    pszFileName: str = pszBasePath.name
    if "_step0013_" not in pszFileName:
        raise ValueError(f"Input is not step0013 file: {pszBasePath}")
    pszOutputFileName: str = pszFileName.replace("_step0013_", "_step0013_法定福利費_", 1)
    return pszBasePath.resolve().parent / pszOutputFileName


def select_columns_by_1_based_indices(objRows: List[List[str]], objIndices: List[int]) -> List[List[str]]:
    objOutputRows: List[List[str]] = []
    for objRow in objRows:
        objOutputRows.append([
            objRow[iIndex - 1] if iIndex - 1 < len(objRow) else ""
            for iIndex in objIndices
        ])
    return objOutputRows


def build_header_index_map(objHeaderRow: List[str]) -> dict[str, int]:
    objIndexMap: dict[str, int] = {}
    for iIndex, pszHeader in enumerate(objHeaderRow):
        pszKey: str = (pszHeader or "").strip()
        if pszKey == "":
            continue
        if pszKey not in objIndexMap:
            objIndexMap[pszKey] = iIndex
    return objIndexMap


def get_required_header_indices(objHeaderRow: List[str], objRequiredHeaders: List[str]) -> List[int]:
    objIndexMap: dict[str, int] = build_header_index_map(objHeaderRow)
    objMissingHeaders: List[str] = [pszHeader for pszHeader in objRequiredHeaders if pszHeader not in objIndexMap]
    if objMissingHeaders:
        raise ValueError(f"Missing required headers: {', '.join(objMissingHeaders)}")
    return [objIndexMap[pszHeader] for pszHeader in objRequiredHeaders]


def remove_columns_by_1_based_indices(objRows: List[List[str]], objExcludedIndices: set[int]) -> List[List[str]]:
    objOutputRows: List[List[str]] = []
    for objRow in objRows:
        objOutputRows.append([
            pszCell
            for iIndex, pszCell in enumerate(objRow, start=1)
            if iIndex not in objExcludedIndices
        ])
    return objOutputRows


def process_new_rawdata_step0013_from_step0012(
    objNewRawdataStep0012Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0012Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0012Path}")

    objNontaxCommuteColumns: List[int] = [1, 2, 3, 4, 5, 9]
    objExcludedColumns: set[int] = set([8, 9, 21, 23] + list(range(24, 31)))
    objStep0012HeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objInputRows[0]]
    objStatutoryWelfareHeaderNames: List[str] = [
        "スタッフ昇順",
        "スタッフコード(先頭)",
        "スタッフコード",
        "氏名",
        "プロジェクト名",
        "工数",
        "健保事業主負担",
        "介護事業主負担",
        "厚年事業主負担",
        "雇保事業主負担",
        "労災保険料",
        "一般拠出金",
        "子育拠出金",
    ]
    objStatutoryWelfareIndices0: List[int] = get_required_header_indices(
        objStep0012HeaderRow,
        objStatutoryWelfareHeaderNames,
    )
    objStatutoryWelfareColumns: List[int] = [iIndex + 1 for iIndex in objStatutoryWelfareIndices0]

    objStep0013Rows: List[List[str]] = remove_columns_by_1_based_indices(objInputRows, objExcludedColumns)
    objStep0013NontaxCommuteRows: List[List[str]] = select_columns_by_1_based_indices(
        objInputRows,
        objNontaxCommuteColumns,
    )
    objStep0013StatutoryWelfareRows: List[List[str]] = select_columns_by_1_based_indices(
        objInputRows,
        objStatutoryWelfareColumns,
    )

    objStep0013Path: Path = build_new_rawdata_step0013_output_path_from_step0012(objNewRawdataStep0012Path)
    objStep0013NontaxCommutePath: Path = build_new_rawdata_step0013_nontax_commute_output_path_from_step0012(
        objNewRawdataStep0012Path
    )
    objStep0013StatutoryWelfarePath: Path = build_new_rawdata_step0013_statutory_welfare_output_path_from_step0012(
        objNewRawdataStep0012Path
    )

    write_sheet_to_tsv(objStep0013Path, objStep0013Rows)
    process_new_rawdata_step0014_from_step0013(objStep0013Path)
    write_sheet_to_tsv(objStep0013StatutoryWelfarePath, objStep0013StatutoryWelfareRows)
    process_new_rawdata_step0014_statutory_welfare_from_step0013_statutory_welfare(objStep0013StatutoryWelfarePath)
    return 0


def process_new_rawdata_step0013_prepayed_commute_from_step0012(
    objStep0012Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0012Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0012Path}")

    objHeaderRow: List[str] = list(objInputRows[0])
    bHasHeader: bool = len(objHeaderRow) >= 2 and objHeaderRow[0].strip() == "プロジェクト名"
    iDataStartIndex: int = 1 if bHasHeader else 0

    objProjectOrder: List[str] = []
    objProjectSeconds: Dict[str, int] = {}
    objProjectNumericTotals: Dict[str, List[Decimal]] = {}
    objProjectNumericScales: Dict[str, List[int]] = {}
    iMaxNumericColumns: int = 0

    for objRow in objInputRows[iDataStartIndex:]:
        if not objRow:
            continue
        pszProjectName: str = (objRow[0] if len(objRow) >= 1 else "").strip()
        if pszProjectName == "":
            continue

        if pszProjectName not in objProjectSeconds:
            objProjectSeconds[pszProjectName] = 0
            objProjectNumericTotals[pszProjectName] = []
            objProjectNumericScales[pszProjectName] = []
            objProjectOrder.append(pszProjectName)

        pszManhour: str = objRow[1] if len(objRow) >= 2 else ""
        objProjectSeconds[pszProjectName] += parse_time_text_to_seconds(pszManhour)

        iNumericColumns: int = max(0, len(objRow) - 2)
        iMaxNumericColumns = max(iMaxNumericColumns, iNumericColumns)
        while len(objProjectNumericTotals[pszProjectName]) < iNumericColumns:
            objProjectNumericTotals[pszProjectName].append(Decimal("0"))
            objProjectNumericScales[pszProjectName].append(0)
        for iOffset in range(iNumericColumns):
            pszCell: str = objRow[iOffset + 2]
            objValue: Decimal | None = parse_decimal_text(pszCell)
            if objValue is not None:
                objProjectNumericTotals[pszProjectName][iOffset] += objValue
                objProjectNumericScales[pszProjectName][iOffset] = max(
                    objProjectNumericScales[pszProjectName][iOffset],
                    count_decimal_places(pszCell),
                )

    objOutputRows: List[List[str]] = []
    if bHasHeader:
        objOutputRows.append(objHeaderRow)
    for pszProjectName in objProjectOrder:
        objOutputRow: List[str] = [
            pszProjectName,
            format_timedelta_as_h_mm_ss(timedelta(seconds=objProjectSeconds[pszProjectName])),
        ]
        objTotals: List[Decimal] = objProjectNumericTotals[pszProjectName]
        objScales: List[int] = objProjectNumericScales[pszProjectName]
        for iOffset in range(iMaxNumericColumns):
            objTotal: Decimal = objTotals[iOffset] if iOffset < len(objTotals) else Decimal("0")
            iScaleDigits: int = objScales[iOffset] if iOffset < len(objScales) else 0
            objOutputRow.append(format_decimal_for_tsv(objTotal, iScaleDigits))
        objOutputRows.append(objOutputRow)

    objOutputPath: Path = build_new_rawdata_step0013_prepayed_commute_output_path_from_step0012(objStep0012Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0014_prepayed_commute_from_step0013_sorted(objOutputPath)
    return 0


def process_new_rawdata_step0014_prepayed_commute_from_step0013_sorted(
    objStep0013Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0013Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0013Path}")

    objHeaderRow: List[str] = list(objInputRows[0])
    bHasHeader: bool = len(objHeaderRow) >= 1 and objHeaderRow[0].strip() == "プロジェクト名"
    iDataStartIndex: int = 1 if bHasHeader else 0

    objMatchedRows: List[Tuple[str, int, List[str]]] = []
    objUnmatchedRows: List[Tuple[int, List[str]]] = []
    for iRowIndex, objRow in enumerate(objInputRows[iDataStartIndex:], start=iDataStartIndex):
        pszProjectName: str = (objRow[0] if len(objRow) >= 1 else "").strip()
        pszProjectCodePrefix: str = extract_project_code_prefix_step0017(pszProjectName)
        if is_supported_project_code_prefix_step0017(pszProjectCodePrefix):
            objMatchedRows.append((pszProjectCodePrefix, iRowIndex, list(objRow)))
        else:
            objUnmatchedRows.append((iRowIndex, list(objRow)))

    objMatchedRows.sort(key=lambda objItem: (objItem[0], objItem[1]))

    objOutputRows: List[List[str]] = []
    if bHasHeader:
        objOutputRows.append(objHeaderRow)
    objOutputRows.extend([objRow for _, _, objRow in objMatchedRows])
    objOutputRows.extend([objRow for _, objRow in objUnmatchedRows])

    objOutputPath: Path = build_new_rawdata_step0014_prepayed_commute_output_path_from_step0013(objStep0013Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    write_org_table_tsv_from_csv_for_prepayed_commute(objOutputPath.resolve().parent)
    process_new_rawdata_step0015_prepayed_commute_with_org_table_from_step0014(objOutputPath)
    return 0


def process_new_rawdata_step0015_prepayed_commute_from_step0014(
    objStep0014Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0014Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0014Path}")

    objHeaderRow: List[str] = list(objInputRows[0])
    iProjectNameIndex: int = objHeaderRow.index("プロジェクト名") if "プロジェクト名" in objHeaderRow else 0
    iInsertIndex: int = iProjectNameIndex

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex == 0:
            pszDebitProjectCode: str = "借方プロジェクトコード"
        else:
            pszProjectName: str = objNewRow[iProjectNameIndex] if len(objNewRow) > iProjectNameIndex else ""
            pszDebitProjectCode = build_debit_project_code_from_project_name_for_prepayed_commute(pszProjectName)
        objNewRow.insert(iInsertIndex, pszDebitProjectCode)
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0015_prepayed_commute_output_path_from_step0014(objStep0014Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0016_prepayed_commute_from_step0015(objOutputPath)
    return 0


def build_new_rawdata_step0014_statutory_welfare_output_path_from_step0013_statutory_welfare(
    objStep0013StatutoryWelfarePath: Path,
) -> Path:
    pszFileName: str = objStep0013StatutoryWelfarePath.name
    if "_step0013_法定福利費_" not in pszFileName:
        raise ValueError(f"Input is not step0013 statutory welfare file: {objStep0013StatutoryWelfarePath}")
    pszOutputFileName: str = pszFileName.replace("_step0013_法定福利費_", "_step0014_法定福利費_", 1)
    return objStep0013StatutoryWelfarePath.resolve().parent / pszOutputFileName


def process_new_rawdata_step0014_statutory_welfare_from_step0013_statutory_welfare(
    objStep0013StatutoryWelfarePath: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0013StatutoryWelfarePath)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0013StatutoryWelfarePath}")

    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objInputRows[0]]
    objWelfareHeaderNames: List[str] = [
        "健保事業主負担",
        "介護事業主負担",
        "厚年事業主負担",
        "雇保事業主負担",
        "労災保険料",
        "一般拠出金",
        "子育拠出金",
    ]
    objRequiredHeaderNames: List[str] = ["氏名", "プロジェクト名"] + objWelfareHeaderNames
    objRequiredIndices: List[int] = get_required_header_indices(objHeaderRow, objRequiredHeaderNames)
    iStaffNameIndex: int = objRequiredIndices[0]
    iProjectNameIndex: int = objRequiredIndices[1]
    objWelfareIndices: List[int] = objRequiredIndices[2:]
    iChildcareContributionIndex: int = objWelfareIndices[-1]
    iInsertIndex: int = iChildcareContributionIndex + 1

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        while len(objNewRow) < iInsertIndex:
            objNewRow.append("")

        pszTotalLegalWelfare: str = ""
        if iRowIndex == 0:
            pszTotalLegalWelfare = "法定福利費"
        else:
            pszStaffName: str = (objNewRow[iStaffNameIndex] or "").strip() if iStaffNameIndex < len(objNewRow) else ""
            pszProjectName: str = (objNewRow[iProjectNameIndex] or "").strip() if iProjectNameIndex < len(objNewRow) else ""
            if pszStaffName != "" and pszProjectName == "合計":
                objTotal: Decimal = Decimal("0")
                for iColumnIndex in objWelfareIndices:
                    objValue: Decimal | None = parse_decimal_text(objNewRow[iColumnIndex] if iColumnIndex < len(objNewRow) else "")
                    if objValue is not None:
                        objTotal += objValue
                iFlooredTotal: int = int(objTotal.to_integral_value(rounding=ROUND_FLOOR))
                pszTotalLegalWelfare = str(iFlooredTotal)

        objNewRow.insert(iInsertIndex, pszTotalLegalWelfare)
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0014_statutory_welfare_output_path_from_step0013_statutory_welfare(
        objStep0013StatutoryWelfarePath
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0015_statutory_welfare_from_step0014_statutory_welfare(objOutputPath)
    return 0


def build_new_rawdata_step0015_statutory_welfare_output_path_from_step0014_statutory_welfare(
    objStep0014StatutoryWelfarePath: Path,
) -> Path:
    pszFileName: str = objStep0014StatutoryWelfarePath.name
    if "_step0014_法定福利費_" not in pszFileName:
        raise ValueError(f"Input is not step0014 statutory welfare file: {objStep0014StatutoryWelfarePath}")
    pszOutputFileName: str = pszFileName.replace("_step0014_法定福利費_", "_step0015_法定福利費_", 1)
    return objStep0014StatutoryWelfarePath.resolve().parent / pszOutputFileName


def build_new_rawdata_step0014_output_path_from_step0013(objStep0013Path: Path) -> Path:
    pszFileName: str = objStep0013Path.name
    if "_step0013_" not in pszFileName:
        raise ValueError(f"Input is not step0013 file: {objStep0013Path}")
    pszOutputFileName: str = pszFileName.replace("_step0013_", "_step0014_", 1)
    return objStep0013Path.resolve().parent / pszOutputFileName


def format_decimal_for_tsv(objValue: Decimal, iScaleDigits: int) -> str:
    iScale: int = 10 ** iScaleDigits
    iSign: int = -1 if objValue < 0 else 1
    objAbsScaled: Decimal = (abs(objValue) * Decimal(iScale)).quantize(Decimal("1"))
    iUnits: int = int(objAbsScaled) * iSign
    return format_scaled_units(iUnits, iScaleDigits)


def process_new_rawdata_step0014_from_step0013(
    objStep0013Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0013Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0013Path}")

    iOutputColumnIndex: int = 19  # T列(0-based index 19)
    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        while len(objNewRow) <= iOutputColumnIndex:
            objNewRow.append("")

        pszSalaryTotal: str = ""
        if iRowIndex == 0:
            pszSalaryTotal = "給与合計"
        else:
            iScaleDigits: int = 0
            objAddTotal: Decimal = Decimal("0")
            for iColumnIndex in range(6, 16):
                pszCell: str = objNewRow[iColumnIndex] if iColumnIndex < len(objNewRow) else ""
                objValue: Decimal | None = parse_decimal_text(pszCell)
                if objValue is not None:
                    objAddTotal += objValue
                    iScaleDigits = max(iScaleDigits, count_decimal_places(pszCell))

            objSubTotal: Decimal = Decimal("0")
            for iColumnIndex in range(16, 18):
                pszCell = objNewRow[iColumnIndex] if iColumnIndex < len(objNewRow) else ""
                objValue = parse_decimal_text(pszCell)
                if objValue is not None:
                    objSubTotal += objValue
                    iScaleDigits = max(iScaleDigits, count_decimal_places(pszCell))

            pszOtherAllowance: str = objNewRow[18] if 18 < len(objNewRow) else ""
            objOtherAllowance: Decimal = parse_decimal_text(pszOtherAllowance) or Decimal("0")
            iScaleDigits = max(iScaleDigits, count_decimal_places(pszOtherAllowance))

            objTotal: Decimal = objAddTotal - objSubTotal + objOtherAllowance
            pszSalaryTotal = format_decimal_for_tsv(objTotal, iScaleDigits)

        objNewRow[iOutputColumnIndex] = pszSalaryTotal
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0014_output_path_from_step0013(objStep0013Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0015_from_step0014(objOutputPath)
    return 0


def build_new_rawdata_step0015_output_path_from_step0014(objStep0014Path: Path) -> Path:
    pszFileName: str = objStep0014Path.name
    if "_step0014_" not in pszFileName:
        raise ValueError(f"Input is not step0014 file: {objStep0014Path}")
    pszOutputFileName: str = pszFileName.replace("_step0014_", "_step0015_", 1)
    return objStep0014Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0015_error_path_from_step0015(objStep0015Path: Path) -> Path:
    return objStep0015Path.resolve().parent / f"{objStep0015Path.stem}_error.txt"


def build_new_rawdata_step0016_output_path_from_step0015(objStep0015Path: Path) -> Path:
    pszFileName: str = objStep0015Path.name
    if "_step0015_" not in pszFileName:
        raise ValueError(f"Input is not step0015 file: {objStep0015Path}")
    pszOutputFileName: str = pszFileName.replace("_step0015_", "_step0016_", 1)
    return objStep0015Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0017_output_path_from_step0016(objStep0016Path: Path) -> Path:
    pszFileName: str = objStep0016Path.name
    if "_step0016_" not in pszFileName:
        raise ValueError(f"Input is not step0016 file: {objStep0016Path}")
    pszOutputFileName: str = pszFileName.replace("_step0016_", "_step0017_", 1)
    return objStep0016Path.resolve().parent / pszOutputFileName


def extract_project_code_prefix_step0017(pszProjectName: str) -> str:
    iUnderscoreIndex: int = pszProjectName.find("_")
    if iUnderscoreIndex == -1:
        return pszProjectName
    return pszProjectName[:iUnderscoreIndex]


def is_supported_project_code_prefix_step0017(pszProjectCodePrefix: str) -> bool:
    if re.match(r"^P\d{5}$", pszProjectCodePrefix) is not None:
        return True
    if re.match(r"^[A-OQ-Z]\d{3}$", pszProjectCodePrefix) is not None:
        return True
    return False


def build_new_rawdata_step0018_output_path_from_step0017(objStep0017Path: Path) -> Path:
    pszFileName: str = objStep0017Path.name
    if "_step0017_" not in pszFileName:
        raise ValueError(f"Input is not step0017 file: {objStep0017Path}")
    pszOutputFileName: str = pszFileName.replace("_step0017_", "_step0018_", 1)
    return objStep0017Path.resolve().parent / pszOutputFileName


def build_new_rawdata_step0019_output_path_from_step0018(objStep0018Path: Path) -> Path:
    pszFileName: str = objStep0018Path.name
    if "_step0018_" not in pszFileName:
        raise ValueError(f"Input is not step0018 file: {objStep0018Path}")
    pszOutputFileName: str = pszFileName.replace("_step0018_", "_step0019_", 1)
    return objStep0018Path.resolve().parent / pszOutputFileName


def process_new_rawdata_step0019_from_step0018(
    objStep0018Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0018Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0018Path}")

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        if iRowIndex == 0:
            objOutputRows.append(["No"] + list(objRow))
        else:
            objOutputRows.append([str(iRowIndex)] + list(objRow))

    objOutputPath: Path = build_new_rawdata_step0019_output_path_from_step0018(objStep0018Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0020_from_step0019(objOutputPath)
    return 0


def process_new_rawdata_step0020_from_step0019(
    objStep0019Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0019Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0019Path}")

    objHeaderRow: List[str] = list(objInputRows[0])
    objRequiredHeaderNames: List[str] = ["プロジェクト名", "給与合計"]
    objRequiredIndices: List[int] = get_required_header_indices(objHeaderRow, objRequiredHeaderNames)

    objOutputRows: List[List[str]] = [objRequiredHeaderNames]
    for objRow in objInputRows[1:]:
        objOutputRows.append([
            objRow[iColumnIndex] if iColumnIndex < len(objRow) else ""
            for iColumnIndex in objRequiredIndices
        ])

    objOutputPath: Path = build_new_rawdata_step0020_output_path_from_step0019(objStep0019Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0021_from_step0020_with_org_table(objOutputPath)
    return 0


def load_org_table_debit_project_codes_for_salary(
    objOrgTableTsvPath: Path,
) -> Dict[str, List[str]]:
    if not objOrgTableTsvPath.exists():
        raise FileNotFoundError(f"管轄PJ表_給与.tsv が見つかりません。Path = {objOrgTableTsvPath}")
    objRows: List[List[str]] = read_tsv_rows(objOrgTableTsvPath)
    if not objRows:
        raise ValueError(f"管轄PJ表_給与.tsv has no rows: {objOrgTableTsvPath}")

    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    iProjectCodeIndex: int = objHeaderRow.index("PJコード") if "PJコード" in objHeaderRow else 1
    iDebitProjectCodeIndex: int = (
        objHeaderRow.index("借方プロジェクトコード")
        if "借方プロジェクトコード" in objHeaderRow
        else 5
    )
    iStartIndex: int = 1 if "PJコード" in objHeaderRow or "借方プロジェクトコード" in objHeaderRow else 0

    objCodeToDebits: Dict[str, List[str]] = {}
    for objRow in objRows[iStartIndex:]:
        if iProjectCodeIndex >= len(objRow):
            continue
        pszOrgProjectCodeText: str = (objRow[iProjectCodeIndex] or "").strip()
        if pszOrgProjectCodeText == "":
            continue
        objCodeMatch: re.Match[str] | None = re.match(r"^(P\d{5}|[A-OQ-Z]\d{3})", pszOrgProjectCodeText)
        if objCodeMatch is None:
            continue
        pszCode: str = objCodeMatch.group(1)
        pszDebitCode: str = (objRow[iDebitProjectCodeIndex] or "").strip() if iDebitProjectCodeIndex < len(objRow) else ""
        objCodeToDebits.setdefault(pszCode, []).append(pszDebitCode)
    return objCodeToDebits


def write_org_table_salary_tsv_from_csv(
    objBaseDirectoryPath: Path,
) -> int:
    objScriptDirectoryPath: Path = Path(__file__).resolve().parent
    objOrgTableCsvPath: Path = objScriptDirectoryPath / "管轄PJ表_給与.csv"
    if not objOrgTableCsvPath.exists():
        objOrgTableCsvPath = objBaseDirectoryPath / "管轄PJ表_給与.csv"

    objOrgTableTsvPath: Path = objBaseDirectoryPath / "管轄PJ表_給与.tsv"
    if not objOrgTableCsvPath.exists():
        raise FileNotFoundError(f"管轄PJ表_給与.csv が見つかりません。Path = {objOrgTableCsvPath}")

    objRows: List[List[str]] = []
    arrEncodings: List[str] = ["utf-8-sig", "cp932"]
    objLastDecodeError: Exception | None = None
    for pszEncoding in arrEncodings:
        try:
            with open(
                objOrgTableCsvPath,
                mode="r",
                encoding=pszEncoding,
                newline="",
            ) as objInputFile:
                objReader = csv.reader(objInputFile)
                for objRow in objReader:
                    objRows.append(list(objRow))
            objLastDecodeError = None
            break
        except UnicodeDecodeError as objError:
            objLastDecodeError = objError
            objRows = []
    if objLastDecodeError is not None:
        raise ValueError(f"unexpected exception while reading 管轄PJ表_給与.csv. Detail = {objLastDecodeError}")

    objOrgTableTsvPath.parent.mkdir(parents=True, exist_ok=True)
    with open(objOrgTableTsvPath, mode="w", encoding="utf-8", newline="") as objOutputFile:
        objWriter: csv.writer = csv.writer(objOutputFile, delimiter="\t")
        for objRow in objRows:
            objWriter.writerow(objRow)
    return 0


def load_org_table_debit_project_codes_for_statutory_welfare(
    objOrgTableTsvPath: Path,
) -> Dict[str, List[str]]:
    if not objOrgTableTsvPath.exists():
        raise FileNotFoundError(f"管轄PJ表_法定福利.tsv が見つかりません。Path = {objOrgTableTsvPath}")
    objRows: List[List[str]] = read_tsv_rows(objOrgTableTsvPath)
    if not objRows:
        raise ValueError(f"管轄PJ表_法定福利.tsv has no rows: {objOrgTableTsvPath}")

    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    iProjectCodeIndex: int = objHeaderRow.index("PJコード") if "PJコード" in objHeaderRow else 1
    iDebitProjectCodeIndex: int = (
        objHeaderRow.index("借方プロジェクトコード")
        if "借方プロジェクトコード" in objHeaderRow
        else 5
    )
    iStartIndex: int = 1 if "PJコード" in objHeaderRow or "借方プロジェクトコード" in objHeaderRow else 0

    objCodeToDebits: Dict[str, List[str]] = {}
    for objRow in objRows[iStartIndex:]:
        if iProjectCodeIndex >= len(objRow):
            continue
        pszOrgProjectCodeText: str = (objRow[iProjectCodeIndex] or "").strip()
        if pszOrgProjectCodeText == "":
            continue
        objCodeMatch: re.Match[str] | None = re.match(r"^(P\d{5}|[A-OQ-Z]\d{3})", pszOrgProjectCodeText)
        if objCodeMatch is None:
            continue
        pszCode: str = objCodeMatch.group(1)
        pszDebitCode: str = (objRow[iDebitProjectCodeIndex] or "").strip() if iDebitProjectCodeIndex < len(objRow) else ""
        objCodeToDebits.setdefault(pszCode, []).append(pszDebitCode)
    return objCodeToDebits


def write_org_table_statutory_welfare_tsv_from_csv(
    objBaseDirectoryPath: Path,
) -> int:
    objScriptDirectoryPath: Path = Path(__file__).resolve().parent
    objOrgTableCsvPath: Path = objScriptDirectoryPath / "管轄PJ表_法定福利.csv"
    if not objOrgTableCsvPath.exists():
        objOrgTableCsvPath = objBaseDirectoryPath / "管轄PJ表_法定福利.csv"

    objOrgTableTsvPath: Path = objBaseDirectoryPath / "管轄PJ表_法定福利.tsv"
    if not objOrgTableCsvPath.exists():
        raise FileNotFoundError(f"管轄PJ表_法定福利.csv が見つかりません。Path = {objOrgTableCsvPath}")

    objRows: List[List[str]] = []
    arrEncodings: List[str] = ["utf-8-sig", "cp932"]
    objLastDecodeError: Exception | None = None
    for pszEncoding in arrEncodings:
        try:
            with open(
                objOrgTableCsvPath,
                mode="r",
                encoding=pszEncoding,
                newline="",
            ) as objInputFile:
                objReader = csv.reader(objInputFile)
                for objRow in objReader:
                    objRows.append(list(objRow))
            objLastDecodeError = None
            break
        except UnicodeDecodeError as objError:
            objLastDecodeError = objError
            objRows = []
    if objLastDecodeError is not None:
        raise ValueError(f"unexpected exception while reading 管轄PJ表_法定福利.csv. Detail = {objLastDecodeError}")

    objOrgTableTsvPath.parent.mkdir(parents=True, exist_ok=True)
    with open(objOrgTableTsvPath, mode="w", encoding="utf-8", newline="") as objOutputFile:
        objWriter: csv.writer = csv.writer(objOutputFile, delimiter="\t")
        for objRow in objRows:
            objWriter.writerow(objRow)
    return 0


def process_new_rawdata_step0021_from_step0020_with_org_table(
    objStep0020Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0020Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0020Path}")
    objBaseDirectoryPath: Path = objStep0020Path.resolve().parent
    write_org_table_salary_tsv_from_csv(objBaseDirectoryPath)
    objOrgTablePath: Path = objBaseDirectoryPath / "管轄PJ表_給与.tsv"
    objCodeToDebits: Dict[str, List[str]] = load_org_table_debit_project_codes_for_salary(objOrgTablePath)

    objHeaderRow: List[str] = list(objInputRows[0])
    iProjectNameIndex: int = objHeaderRow.index("プロジェクト名") if "プロジェクト名" in objHeaderRow else 0
    iInsertIndex: int = iProjectNameIndex

    objOutputRows: List[List[str]] = []
    objErrorLines: List[str] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex == 0:
            pszDebitProjectCode: str = "借方プロジェクトコード"
        else:
            pszProjectName: str = objNewRow[iProjectNameIndex] if len(objNewRow) > iProjectNameIndex else ""
            pszDebitProjectCode = build_debit_project_code_from_project_name_and_org_table_for_prepayed_commute(
                pszProjectName,
                objCodeToDebits,
                objErrorLines,
            )
        objNewRow.insert(iInsertIndex, pszDebitProjectCode)
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0021_output_path_from_step0020(objStep0020Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0021_error_path_from_step0021(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    process_new_rawdata_step0022_from_step0021(objOutputPath)
    return 0


def process_new_rawdata_step0018_from_step0017(
    objStep0017Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0017Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0017Path}")

    objHeaderRow: List[str] = list(objInputRows[0])
    bHasHeader: bool = len(objHeaderRow) >= 1 and objHeaderRow[0].strip() == "プロジェクト名"
    iDataStartIndex: int = 1 if bHasHeader else 0

    objMatchedRows: List[Tuple[str, int, List[str]]] = []
    objUnmatchedRows: List[Tuple[int, List[str]]] = []
    for iRowIndex, objRow in enumerate(objInputRows[iDataStartIndex:], start=iDataStartIndex):
        pszProjectName: str = (objRow[0] if len(objRow) >= 1 else "").strip()
        pszProjectCodePrefix: str = extract_project_code_prefix_step0017(pszProjectName)
        if is_supported_project_code_prefix_step0017(pszProjectCodePrefix):
            objMatchedRows.append((pszProjectCodePrefix, iRowIndex, list(objRow)))
        else:
            objUnmatchedRows.append((iRowIndex, list(objRow)))

    objMatchedRows.sort(key=lambda objItem: (objItem[0], objItem[1]))

    objOutputRows: List[List[str]] = []
    if bHasHeader:
        objOutputRows.append(objHeaderRow)
    objOutputRows.extend([objRow for _, _, objRow in objMatchedRows])
    objOutputRows.extend([objRow for _, objRow in objUnmatchedRows])

    objOutputPath: Path = build_new_rawdata_step0018_output_path_from_step0017(objStep0017Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0019_from_step0018(objOutputPath)
    return 0


def process_new_rawdata_step0017_from_step0016(
    objStep0016Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0016Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0016Path}")

    objHeaderRow: List[str] = list(objInputRows[0])
    bHasHeader: bool = len(objHeaderRow) >= 2 and objHeaderRow[0].strip() == "プロジェクト名"
    iDataStartIndex: int = 1 if bHasHeader else 0

    objProjectOrder: List[str] = []
    objProjectSeconds: Dict[str, int] = {}
    objProjectNumericTotals: Dict[str, List[Decimal]] = {}
    objProjectNumericScales: Dict[str, List[int]] = {}
    iMaxNumericColumns: int = 0

    for objRow in objInputRows[iDataStartIndex:]:
        if not objRow:
            continue
        pszProjectName: str = (objRow[0] if len(objRow) >= 1 else "").strip()
        if pszProjectName == "":
            continue

        pszProjectCodePrefix: str = extract_project_code_prefix_step0017(pszProjectName)
        _ = is_supported_project_code_prefix_step0017(pszProjectCodePrefix)

        if pszProjectName not in objProjectSeconds:
            objProjectSeconds[pszProjectName] = 0
            objProjectNumericTotals[pszProjectName] = []
            objProjectNumericScales[pszProjectName] = []
            objProjectOrder.append(pszProjectName)

        pszManhour: str = objRow[1] if len(objRow) >= 2 else ""
        objProjectSeconds[pszProjectName] += parse_time_text_to_seconds(pszManhour)

        iNumericColumns: int = max(0, len(objRow) - 2)
        iMaxNumericColumns = max(iMaxNumericColumns, iNumericColumns)
        while len(objProjectNumericTotals[pszProjectName]) < iNumericColumns:
            objProjectNumericTotals[pszProjectName].append(Decimal("0"))
            objProjectNumericScales[pszProjectName].append(0)
        for iOffset in range(iNumericColumns):
            pszCell: str = objRow[iOffset + 2]
            objValue: Decimal | None = parse_decimal_text(pszCell)
            if objValue is not None:
                objProjectNumericTotals[pszProjectName][iOffset] += objValue
                objProjectNumericScales[pszProjectName][iOffset] = max(
                    objProjectNumericScales[pszProjectName][iOffset],
                    count_decimal_places(pszCell),
                )

    objOutputRows: List[List[str]] = []
    if bHasHeader:
        objOutputRows.append(objHeaderRow)
    for pszProjectName in objProjectOrder:
        objOutputRow: List[str] = [
            pszProjectName,
            format_timedelta_as_h_mm_ss(timedelta(seconds=objProjectSeconds[pszProjectName])),
        ]
        objTotals: List[Decimal] = objProjectNumericTotals[pszProjectName]
        objScales: List[int] = objProjectNumericScales[pszProjectName]
        for iOffset in range(iMaxNumericColumns):
            objTotal: Decimal = objTotals[iOffset] if iOffset < len(objTotals) else Decimal("0")
            iScaleDigits: int = objScales[iOffset] if iOffset < len(objScales) else 0
            objOutputRow.append(format_decimal_for_tsv(objTotal, iScaleDigits))
        objOutputRows.append(objOutputRow)

    objOutputPath: Path = build_new_rawdata_step0017_output_path_from_step0016(objStep0016Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0018_from_step0017(objOutputPath)
    return 0


def process_new_rawdata_step0016_from_step0015(
    objStep0015Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0015Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0015Path}")
    objOutputRows: List[List[str]] = remove_columns_by_1_based_indices(objInputRows, {1, 2, 3, 4})
    objOutputPath: Path = build_new_rawdata_step0016_output_path_from_step0015(objStep0015Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0017_from_step0016(objOutputPath)
    return 0


def process_new_rawdata_step0015_from_step0014(
    objStep0014Path: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0014Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0014Path}")

    objOutputRows: List[List[str]] = []
    if objInputRows:
        objOutputRows.append(list(objInputRows[0]))

    objErrorLines: List[str] = []
    iRowIndex: int = 1
    while iRowIndex < len(objInputRows):
        objRow: List[str] = list(objInputRows[iRowIndex])
        pszStaffName: str = (objRow[3] or "").strip() if len(objRow) >= 4 else ""
        pszProjectName: str = (objRow[4] or "").strip() if len(objRow) >= 5 else ""
        if pszStaffName != "" and pszProjectName == "合計":
            if iRowIndex + 1 >= len(objInputRows):
                objOutputRows.append(objRow)
                objErrorLines.append(
                    f"{iRowIndex + 1}行目: 合計行の直下行が存在しないため、合計行をそのまま出力しました。"
                )
                iRowIndex += 1
                continue

            objNextRow: List[str] = list(objInputRows[iRowIndex + 1])
            pszNextStaffName: str = (objNextRow[3] or "").strip() if len(objNextRow) >= 4 else ""
            if pszNextStaffName != "":
                objOutputRows.append(objRow)
                objErrorLines.append(
                    f"{iRowIndex + 1}行目: 合計行の直下行が不正なため、合計行をそのまま出力しました。"
                )
                iRowIndex += 1
                continue

            objMergedRow: List[str] = list(objRow[:4]) + list(objNextRow[4:])
            if len(objMergedRow) < len(objRow):
                objMergedRow.extend([""] * (len(objRow) - len(objMergedRow)))
            objOutputRows.append(objMergedRow)
            iRowIndex += 2
            continue

        objOutputRows.append(objRow)
        iRowIndex += 1

    objOutputPath: Path = build_new_rawdata_step0015_output_path_from_step0014(objStep0014Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)

    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0015_error_path_from_step0015(objOutputPath)
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")

    process_new_rawdata_step0016_from_step0015(objOutputPath)
    return 0


def process_new_rawdata_step0015_statutory_welfare_from_step0014_statutory_welfare(
    objStep0014StatutoryWelfarePath: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0014StatutoryWelfarePath)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0014StatutoryWelfarePath}")

    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objInputRows[0]]
    objRequiredHeaderNames: List[str] = ["氏名", "プロジェクト名", "工数", "法定福利費"]
    objRequiredIndices: List[int] = get_required_header_indices(objHeaderRow, objRequiredHeaderNames)
    iStaffNameIndex: int = objRequiredIndices[0]
    iProjectNameIndex: int = objRequiredIndices[1]
    iManhourIndex: int = objRequiredIndices[2]
    iLegalWelfareIndex: int = objRequiredIndices[3]

    objOutputRows: List[List[str]] = [list(objRow) for objRow in objInputRows]
    iRowIndex: int = 1
    while iRowIndex < len(objOutputRows):
        objSummaryRow: List[str] = objOutputRows[iRowIndex]
        pszStaffName: str = (objSummaryRow[iStaffNameIndex] or "").strip() if iStaffNameIndex < len(objSummaryRow) else ""
        pszProjectName: str = (objSummaryRow[iProjectNameIndex] or "").strip() if iProjectNameIndex < len(objSummaryRow) else ""
        if pszStaffName == "" or pszProjectName != "合計":
            iRowIndex += 1
            continue

        iNextSummaryIndex: int = iRowIndex + 1
        objDetailIndices: List[int] = []
        while iNextSummaryIndex < len(objOutputRows):
            objCandidateRow: List[str] = objOutputRows[iNextSummaryIndex]
            pszCandidateName: str = (objCandidateRow[iStaffNameIndex] or "").strip() if iStaffNameIndex < len(objCandidateRow) else ""
            pszCandidateProject: str = (objCandidateRow[iProjectNameIndex] or "").strip() if iProjectNameIndex < len(objCandidateRow) else ""
            if pszCandidateName != "" and pszCandidateProject == "合計":
                break
            if pszCandidateName == "":
                objDetailIndices.append(iNextSummaryIndex)
            iNextSummaryIndex += 1

        if objDetailIndices:
            pszTotalText: str = objSummaryRow[iLegalWelfareIndex] if iLegalWelfareIndex < len(objSummaryRow) else ""
            objTotalValue: Decimal | None = parse_decimal_text(pszTotalText)
            if objTotalValue is not None:
                objWeights: List[int] = []
                for iDetailIndex in objDetailIndices:
                    objDetailRow: List[str] = objOutputRows[iDetailIndex]
                    pszManhour: str = (objDetailRow[iManhourIndex] or "").strip() if iManhourIndex < len(objDetailRow) else ""
                    if pszManhour == "":
                        objWeights.append(0)
                        continue
                    try:
                        objWeights.append(parse_time_text_to_seconds(pszManhour))
                    except Exception:
                        objWeights.append(0)

                iWeightTotal: int = sum(objWeights)
                iScaleDigits: int = count_decimal_places(pszTotalText)
                iScale: int = 10 ** iScaleDigits
                objAbsTotalScaled: Decimal = (abs(objTotalValue) * Decimal(iScale)).quantize(Decimal("1"))
                iTotalScaledUnits: int = int(objAbsTotalScaled)

                objAllocatedUnits: List[int] = [0] * len(objDetailIndices)
                if iWeightTotal > 0 and iTotalScaledUnits > 0:
                    objFloors: List[int] = []
                    objRemainders: List[tuple[int, Decimal]] = []
                    for iIndex, iWeight in enumerate(objWeights):
                        if iWeight <= 0:
                            objFloors.append(0)
                            objRemainders.append((iIndex, Decimal("-1")))
                            continue
                        objRaw: Decimal = Decimal(iTotalScaledUnits) * Decimal(iWeight) / Decimal(iWeightTotal)
                        objFloorValue: Decimal = objRaw.to_integral_value(rounding=ROUND_FLOOR)
                        iFloor: int = int(objFloorValue)
                        objFloors.append(iFloor)
                        objRemainders.append((iIndex, objRaw - objFloorValue))

                    iFloorSum: int = sum(objFloors)
                    iRemaining: int = iTotalScaledUnits - iFloorSum
                    objRemainders.sort(key=lambda objItem: (-objItem[1], objItem[0]))
                    for iIndex, _ in objRemainders[:iRemaining]:
                        objFloors[iIndex] += 1
                    objAllocatedUnits = objFloors

                iSign: int = -1 if objTotalValue < 0 else 1
                for iIndex, iDetailIndex in enumerate(objDetailIndices):
                    objDetailRow: List[str] = objOutputRows[iDetailIndex]
                    while len(objDetailRow) <= iLegalWelfareIndex:
                        objDetailRow.append("")
                    objDetailRow[iLegalWelfareIndex] = format_scaled_units(iSign * objAllocatedUnits[iIndex], iScaleDigits)

        iRowIndex = iNextSummaryIndex

    objOutputPath: Path = build_new_rawdata_step0015_statutory_welfare_output_path_from_step0014_statutory_welfare(
        objStep0014StatutoryWelfarePath
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0016_statutory_welfare_from_step0015_statutory_welfare(objOutputPath)
    return 0


def build_new_rawdata_step0016_statutory_welfare_output_path_from_step0015_statutory_welfare(
    objStep0015StatutoryWelfarePath: Path,
) -> Path:
    pszFileName: str = objStep0015StatutoryWelfarePath.name
    if "_step0015_法定福利費_" not in pszFileName:
        raise ValueError(f"Input is not step0015 statutory welfare file: {objStep0015StatutoryWelfarePath}")
    pszOutputFileName: str = pszFileName.replace("_step0015_法定福利費_", "_step0016_法定福利費_", 1)
    return objStep0015StatutoryWelfarePath.resolve().parent / pszOutputFileName


def build_new_rawdata_step0016_statutory_welfare_error_path_from_step0016_statutory_welfare(
    objStep0016StatutoryWelfarePath: Path,
) -> Path:
    return objStep0016StatutoryWelfarePath.resolve().parent / f"{objStep0016StatutoryWelfarePath.stem}_error.txt"


def process_new_rawdata_step0016_statutory_welfare_from_step0015_statutory_welfare(
    objStep0015StatutoryWelfarePath: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0015StatutoryWelfarePath)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0015StatutoryWelfarePath}")

    objOutputRows: List[List[str]] = []
    if objInputRows:
        objOutputRows.append(list(objInputRows[0]))

    objErrorLines: List[str] = []
    iRowIndex: int = 1
    while iRowIndex < len(objInputRows):
        objRow: List[str] = list(objInputRows[iRowIndex])
        pszStaffName: str = (objRow[3] or "").strip() if len(objRow) >= 4 else ""
        pszProjectName: str = (objRow[4] or "").strip() if len(objRow) >= 5 else ""
        if pszStaffName != "" and pszProjectName == "合計":
            if iRowIndex + 1 >= len(objInputRows):
                objOutputRows.append(objRow)
                objErrorLines.append(
                    f"{iRowIndex + 1}行目: 合計行の直下行が存在しないため、合計行をそのまま出力しました。"
                )
                iRowIndex += 1
                continue

            objNextRow: List[str] = list(objInputRows[iRowIndex + 1])
            pszNextStaffName: str = (objNextRow[3] or "").strip() if len(objNextRow) >= 4 else ""
            if pszNextStaffName != "":
                objOutputRows.append(objRow)
                objErrorLines.append(
                    f"{iRowIndex + 1}行目: 合計行の直下行が不正なため、合計行をそのまま出力しました。"
                )
                iRowIndex += 1
                continue

            objMergedRow: List[str] = list(objRow[:4]) + list(objNextRow[4:])
            if len(objMergedRow) < len(objRow):
                objMergedRow.extend([""] * (len(objRow) - len(objMergedRow)))
            objOutputRows.append(objMergedRow)
            iRowIndex += 2
            continue

        objOutputRows.append(objRow)
        iRowIndex += 1

    objOutputPath: Path = build_new_rawdata_step0016_statutory_welfare_output_path_from_step0015_statutory_welfare(
        objStep0015StatutoryWelfarePath
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)

    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0016_statutory_welfare_error_path_from_step0016_statutory_welfare(
            objOutputPath
        )
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    process_new_rawdata_step0017_statutory_welfare_from_step0016_statutory_welfare(objOutputPath)
    return 0


def build_new_rawdata_step0017_statutory_welfare_output_path_from_step0016_statutory_welfare(
    objStep0016StatutoryWelfarePath: Path,
) -> Path:
    pszFileName: str = objStep0016StatutoryWelfarePath.name
    if "_step0016_法定福利費_" not in pszFileName:
        raise ValueError(f"Input is not step0016 statutory welfare file: {objStep0016StatutoryWelfarePath}")
    pszOutputFileName: str = pszFileName.replace("_step0016_法定福利費_", "_step0017_法定福利費_", 1)
    return objStep0016StatutoryWelfarePath.resolve().parent / pszOutputFileName


def process_new_rawdata_step0017_statutory_welfare_from_step0016_statutory_welfare(
    objStep0016StatutoryWelfarePath: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0016StatutoryWelfarePath)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0016StatutoryWelfarePath}")
    objOutputRows: List[List[str]] = remove_columns_by_1_based_indices(objInputRows, {1, 2, 3, 4})
    objOutputPath: Path = build_new_rawdata_step0017_statutory_welfare_output_path_from_step0016_statutory_welfare(
        objStep0016StatutoryWelfarePath
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0018_statutory_welfare_from_step0017_statutory_welfare(objOutputPath)
    return 0


def build_new_rawdata_step0018_statutory_welfare_output_path_from_step0017_statutory_welfare(
    objStep0017StatutoryWelfarePath: Path,
) -> Path:
    pszFileName: str = objStep0017StatutoryWelfarePath.name
    if "_step0017_法定福利費_" not in pszFileName:
        raise ValueError(f"Input is not step0017 statutory welfare file: {objStep0017StatutoryWelfarePath}")
    pszOutputFileName: str = pszFileName.replace("_step0017_法定福利費_", "_step0018_法定福利費_", 1)
    return objStep0017StatutoryWelfarePath.resolve().parent / pszOutputFileName


def process_new_rawdata_step0018_statutory_welfare_from_step0017_statutory_welfare(
    objStep0017StatutoryWelfarePath: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0017StatutoryWelfarePath)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0017StatutoryWelfarePath}")

    objHeaderRow: List[str] = list(objInputRows[0])
    bHasHeader: bool = len(objHeaderRow) >= 2 and objHeaderRow[0].strip() == "プロジェクト名"
    iDataStartIndex: int = 1 if bHasHeader else 0

    objProjectOrder: List[str] = []
    objProjectSeconds: Dict[str, int] = {}
    objProjectNumericTotals: Dict[str, List[Decimal]] = {}
    objProjectNumericScales: Dict[str, List[int]] = {}
    iMaxNumericColumns: int = 0

    for objRow in objInputRows[iDataStartIndex:]:
        if not objRow:
            continue
        pszProjectName: str = (objRow[0] if len(objRow) >= 1 else "").strip()
        if pszProjectName == "":
            continue

        pszProjectCodePrefix: str = extract_project_code_prefix_step0017(pszProjectName)
        _ = is_supported_project_code_prefix_step0017(pszProjectCodePrefix)

        if pszProjectName not in objProjectSeconds:
            objProjectSeconds[pszProjectName] = 0
            objProjectNumericTotals[pszProjectName] = []
            objProjectNumericScales[pszProjectName] = []
            objProjectOrder.append(pszProjectName)

        pszManhour: str = objRow[1] if len(objRow) >= 2 else ""
        objProjectSeconds[pszProjectName] += parse_time_text_to_seconds(pszManhour)

        iNumericColumns: int = max(0, len(objRow) - 2)
        iMaxNumericColumns = max(iMaxNumericColumns, iNumericColumns)
        while len(objProjectNumericTotals[pszProjectName]) < iNumericColumns:
            objProjectNumericTotals[pszProjectName].append(Decimal("0"))
            objProjectNumericScales[pszProjectName].append(0)
        for iOffset in range(iNumericColumns):
            pszCell: str = objRow[iOffset + 2]
            objValue: Decimal | None = parse_decimal_text(pszCell)
            if objValue is not None:
                objProjectNumericTotals[pszProjectName][iOffset] += objValue
                objProjectNumericScales[pszProjectName][iOffset] = max(
                    objProjectNumericScales[pszProjectName][iOffset],
                    count_decimal_places(pszCell),
                )

    objOutputRows: List[List[str]] = []
    if bHasHeader:
        objOutputRows.append(objHeaderRow)
    for pszProjectName in objProjectOrder:
        objOutputRow: List[str] = [
            pszProjectName,
            format_timedelta_as_h_mm_ss(timedelta(seconds=objProjectSeconds[pszProjectName])),
        ]
        objTotals: List[Decimal] = objProjectNumericTotals[pszProjectName]
        objScales: List[int] = objProjectNumericScales[pszProjectName]
        for iOffset in range(iMaxNumericColumns):
            objTotal: Decimal = objTotals[iOffset] if iOffset < len(objTotals) else Decimal("0")
            iScaleDigits: int = objScales[iOffset] if iOffset < len(objScales) else 0
            objOutputRow.append(format_decimal_for_tsv(objTotal, iScaleDigits))
        objOutputRows.append(objOutputRow)

    objOutputPath: Path = build_new_rawdata_step0018_statutory_welfare_output_path_from_step0017_statutory_welfare(
        objStep0017StatutoryWelfarePath
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0019_statutory_welfare_from_step0018_statutory_welfare(objOutputPath)
    return 0


def build_new_rawdata_step0019_statutory_welfare_output_path_from_step0018_statutory_welfare(
    objStep0018StatutoryWelfarePath: Path,
) -> Path:
    pszFileName: str = objStep0018StatutoryWelfarePath.name
    if "_step0018_法定福利費_" not in pszFileName:
        raise ValueError(f"Input is not step0018 statutory welfare file: {objStep0018StatutoryWelfarePath}")
    pszOutputFileName: str = pszFileName.replace("_step0018_法定福利費_", "_step0019_法定福利費_", 1)
    return objStep0018StatutoryWelfarePath.resolve().parent / pszOutputFileName


def process_new_rawdata_step0019_statutory_welfare_from_step0018_statutory_welfare(
    objStep0018StatutoryWelfarePath: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0018StatutoryWelfarePath)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0018StatutoryWelfarePath}")

    objHeaderRow: List[str] = list(objInputRows[0])
    bHasHeader: bool = len(objHeaderRow) >= 1 and objHeaderRow[0].strip() == "プロジェクト名"
    iDataStartIndex: int = 1 if bHasHeader else 0

    objMatchedRows: List[Tuple[str, int, List[str]]] = []
    objUnmatchedRows: List[Tuple[int, List[str]]] = []
    for iRowIndex, objRow in enumerate(objInputRows[iDataStartIndex:], start=iDataStartIndex):
        pszProjectName: str = (objRow[0] if len(objRow) >= 1 else "").strip()
        pszProjectCodePrefix: str = extract_project_code_prefix_step0017(pszProjectName)
        if is_supported_project_code_prefix_step0017(pszProjectCodePrefix):
            objMatchedRows.append((pszProjectCodePrefix, iRowIndex, list(objRow)))
        else:
            objUnmatchedRows.append((iRowIndex, list(objRow)))

    objMatchedRows.sort(key=lambda objItem: (objItem[0], objItem[1]))

    objOutputRows: List[List[str]] = []
    if bHasHeader:
        objOutputRows.append(objHeaderRow)
    objOutputRows.extend([objRow for _, _, objRow in objMatchedRows])
    objOutputRows.extend([objRow for _, objRow in objUnmatchedRows])

    objOutputPath: Path = build_new_rawdata_step0019_statutory_welfare_output_path_from_step0018_statutory_welfare(
        objStep0018StatutoryWelfarePath
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0020_statutory_welfare_from_step0019_statutory_welfare(objOutputPath)
    return 0


def build_new_rawdata_step0020_statutory_welfare_output_path_from_step0019_statutory_welfare(
    objStep0019StatutoryWelfarePath: Path,
) -> Path:
    pszFileName: str = objStep0019StatutoryWelfarePath.name
    if "_step0019_法定福利費_" not in pszFileName:
        raise ValueError(f"Input is not step0019 statutory welfare file: {objStep0019StatutoryWelfarePath}")
    pszOutputFileName: str = pszFileName.replace("_step0019_法定福利費_", "_step0020_法定福利費_", 1)
    return objStep0019StatutoryWelfarePath.resolve().parent / pszOutputFileName


def process_new_rawdata_step0020_statutory_welfare_from_step0019_statutory_welfare(
    objStep0019StatutoryWelfarePath: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0019StatutoryWelfarePath)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0019StatutoryWelfarePath}")

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        if iRowIndex == 0:
            objOutputRows.append(["No"] + list(objRow))
        else:
            objOutputRows.append([str(iRowIndex)] + list(objRow))

    objOutputPath: Path = build_new_rawdata_step0020_statutory_welfare_output_path_from_step0019_statutory_welfare(
        objStep0019StatutoryWelfarePath
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0021_statutory_welfare_from_step0020_statutory_welfare(objOutputPath)
    return 0


def build_new_rawdata_step0021_statutory_welfare_output_path_from_step0020_statutory_welfare(
    objStep0020StatutoryWelfarePath: Path,
) -> Path:
    pszFileName: str = objStep0020StatutoryWelfarePath.name
    if "_step0020_法定福利費_" not in pszFileName:
        raise ValueError(f"Input is not step0020 statutory welfare file: {objStep0020StatutoryWelfarePath}")
    pszOutputFileName: str = pszFileName.replace("_step0020_法定福利費_", "_step0021_法定福利費_", 1)
    return objStep0020StatutoryWelfarePath.resolve().parent / pszOutputFileName


def process_new_rawdata_step0021_statutory_welfare_from_step0020_statutory_welfare(
    objStep0020StatutoryWelfarePath: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0020StatutoryWelfarePath)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0020StatutoryWelfarePath}")

    objHeaderRow: List[str] = list(objInputRows[0])
    objRequiredHeaderNames: List[str] = ["プロジェクト名", "法定福利費"]
    objRequiredIndices: List[int] = get_required_header_indices(objHeaderRow, objRequiredHeaderNames)

    objOutputRows: List[List[str]] = [objRequiredHeaderNames]
    for objRow in objInputRows[1:]:
        objOutputRows.append([
            objRow[iColumnIndex] if iColumnIndex < len(objRow) else ""
            for iColumnIndex in objRequiredIndices
        ])

    objOutputPath: Path = build_new_rawdata_step0021_statutory_welfare_output_path_from_step0020_statutory_welfare(
        objStep0020StatutoryWelfarePath
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0022_statutory_welfare_from_step0021_statutory_welfare_with_org_table(objOutputPath)
    return 0


def build_new_rawdata_step0022_statutory_welfare_output_path_from_step0021_statutory_welfare(
    objStep0021StatutoryWelfarePath: Path,
) -> Path:
    pszFileName: str = objStep0021StatutoryWelfarePath.name
    if "_step0021_法定福利費_" not in pszFileName:
        raise ValueError(f"Input is not step0021 statutory welfare file: {objStep0021StatutoryWelfarePath}")
    pszOutputFileName: str = pszFileName.replace("_step0021_法定福利費_", "_step0022_法定福利費_", 1)
    return objStep0021StatutoryWelfarePath.resolve().parent / pszOutputFileName


def build_new_rawdata_step0022_statutory_welfare_error_path_from_step0022_statutory_welfare(
    objStep0022StatutoryWelfarePath: Path,
) -> Path:
    return objStep0022StatutoryWelfarePath.resolve().parent / f"{objStep0022StatutoryWelfarePath.stem}_error.txt"


def process_new_rawdata_step0022_statutory_welfare_from_step0021_statutory_welfare_with_org_table(
    objStep0021StatutoryWelfarePath: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0021StatutoryWelfarePath)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0021StatutoryWelfarePath}")
    objBaseDirectoryPath: Path = objStep0021StatutoryWelfarePath.resolve().parent
    write_org_table_statutory_welfare_tsv_from_csv(objBaseDirectoryPath)
    objOrgTablePath: Path = objBaseDirectoryPath / "管轄PJ表_法定福利.tsv"
    objCodeToDebits: Dict[str, List[str]] = load_org_table_debit_project_codes_for_statutory_welfare(objOrgTablePath)

    objHeaderRow: List[str] = list(objInputRows[0])
    iProjectNameIndex: int = objHeaderRow.index("プロジェクト名") if "プロジェクト名" in objHeaderRow else 0
    iInsertIndex: int = iProjectNameIndex

    objOutputRows: List[List[str]] = []
    objErrorLines: List[str] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        if iRowIndex == 0:
            pszDebitProjectCode: str = "借方プロジェクトコード"
        else:
            pszProjectName: str = objNewRow[iProjectNameIndex] if len(objNewRow) > iProjectNameIndex else ""
            pszDebitProjectCode = build_debit_project_code_from_project_name_and_org_table_for_prepayed_commute(
                pszProjectName,
                objCodeToDebits,
                objErrorLines,
            )
        objNewRow.insert(iInsertIndex, pszDebitProjectCode)
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0022_statutory_welfare_output_path_from_step0021_statutory_welfare(
        objStep0021StatutoryWelfarePath
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    if objErrorLines:
        objErrorPath: Path = build_new_rawdata_step0022_statutory_welfare_error_path_from_step0022_statutory_welfare(
            objOutputPath
        )
        objErrorPath.write_text("\n".join(objErrorLines) + "\n", encoding="utf-8")
    process_new_rawdata_step0023_statutory_welfare_from_step0022_statutory_welfare(objOutputPath)
    return 0


def build_new_rawdata_step0023_statutory_welfare_output_path_from_step0022_statutory_welfare(
    objStep0022StatutoryWelfarePath: Path,
) -> Path:
    pszFileName: str = objStep0022StatutoryWelfarePath.name
    if "_step0022_法定福利費_" not in pszFileName:
        raise ValueError(f"Input is not step0022 statutory welfare file: {objStep0022StatutoryWelfarePath}")
    pszOutputFileName: str = pszFileName.replace("_step0022_法定福利費_", "_step0023_法定福利費_", 1)
    return objStep0022StatutoryWelfarePath.resolve().parent / pszOutputFileName


def process_new_rawdata_step0023_statutory_welfare_from_step0022_statutory_welfare(
    objStep0022StatutoryWelfarePath: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0022StatutoryWelfarePath)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0022StatutoryWelfarePath}")

    objHeaderRow: List[str] = list(objInputRows[0])
    iDebitProjectCodeIndex: int = (
        objHeaderRow.index("借方プロジェクトコード")
        if "借方プロジェクトコード" in objHeaderRow
        else 0
    )
    iInsertIndex: int = iDebitProjectCodeIndex
    objInsertedHeaders: List[str] = [
        "借方税区分コード",
        "借方税率種別コード",
        "借方税率",
        "借方取引先コード",
    ]

    objOutputRows: List[List[str]] = []
    for iRowIndex, objRow in enumerate(objInputRows):
        objNewRow: List[str] = list(objRow)
        objInsertedValues: List[str] = list(objInsertedHeaders) if iRowIndex == 0 else ["", "", "", ""]
        objNewRow[iInsertIndex:iInsertIndex] = objInsertedValues
        objOutputRows.append(objNewRow)

    objOutputPath: Path = build_new_rawdata_step0023_statutory_welfare_output_path_from_step0022_statutory_welfare(
        objStep0022StatutoryWelfarePath
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_new_rawdata_step0012_and_step0013_from_step0011(
    objNewRawdataStep0011Path: Path,
) -> int:
    process_new_rawdata_step0012_from_step0011(objNewRawdataStep0011Path)
    process_new_rawdata_step0013_from_step0012(
        build_new_rawdata_step0012_output_path_from_step0011(objNewRawdataStep0011Path)
    )
    return 0


def process_new_rawdata_step0011_and_step0012_from_step0010(
    objNewRawdataStep0010Path: Path,
) -> int:
    process_new_rawdata_step0011_from_step0010(objNewRawdataStep0010Path)
    objNewRawdataStep0011Path: Path = build_new_rawdata_step0011_output_path_from_step0010(objNewRawdataStep0010Path)
    process_new_rawdata_step0012_from_step0011(objNewRawdataStep0011Path)
    process_new_rawdata_step0013_from_step0012(
        build_new_rawdata_step0012_output_path_from_step0011(objNewRawdataStep0011Path)
    )
    return 0



def fill_missing_staff_codes_in_new_rawdata_step0002_by_management_accounting(
    objNewRawdataStep0002Path: Path,
    objStaffCodeByName: dict[str, str],
) -> int:
    if not objStaffCodeByName:
        raise ValueError("No staff code mapping from management accounting file")

    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0002Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0002Path}")

    objOutputRows: List[List[str]] = []
    pszCurrentStaffName: str = ""
    for objRow in objInputRows:
        objNewRow: List[str] = list(objRow)
        if not objNewRow:
            objOutputRows.append(objNewRow)
            continue

        if len(objNewRow) >= 2:
            pszStaffNameCell: str = (objNewRow[1] or "").strip()
            if pszStaffNameCell != "":
                pszCurrentStaffName = pszStaffNameCell

        pszStaffCodeCell: str = (objNewRow[0] or "").strip()
        if pszStaffCodeCell == "" and pszCurrentStaffName != "":
            pszFilledCode: str = objStaffCodeByName.get(pszCurrentStaffName, "")
            if pszFilledCode != "":
                objNewRow[0] = pszFilledCode

        objOutputRows.append(objNewRow)

    apply_prepayed_commute_equal_allocation_to_step0003_rows(objOutputRows)

    objOutputPath: Path = build_new_rawdata_step0003_output_path_from_step0002(objNewRawdataStep0002Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_prepayed_commute_step0004_from_step0003_path(objStep0003Path: Path) -> Path:
    process_new_rawdata_step0004_from_step0003(objStep0003Path)
    return build_new_rawdata_step0004_output_path_from_step0003(objStep0003Path)


def process_new_rawdata_step0003_from_prepayed_commute_step0002(
    objNewRawdataStep0002Path: Path,
) -> int:
    objOutputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0002Path)
    if not objOutputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0002Path}")
    apply_prepayed_commute_equal_allocation_to_step0003_rows(objOutputRows)
    objOutputPath: Path = build_new_rawdata_step0003_output_path_from_step0002(objNewRawdataStep0002Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_prepayed_commute_step0004_from_step0003_path(objOutputPath)
    return 0


def format_decimal_for_tsv_cell(objValue: Decimal) -> str:
    if objValue == objValue.to_integral_value():
        return str(int(objValue))
    return format(objValue.normalize(), "f")


def resolve_prepayed_commute_divisor_from_application_text(
    pszApplicationText: str,
    pszPeriodLabel: str = "04-09月",
) -> int | None:
    objMatch = re.match(r"^(\d{1,2})/\d{1,2}支給$", (pszApplicationText or "").strip())
    if objMatch is None:
        return None
    iMonth: int = int(objMatch.group(1))
    if pszPeriodLabel == "10-03月":
        if iMonth in (9, 10):
            return 6
        if iMonth == 11:
            return 5
        if iMonth == 12:
            return 4
        if iMonth == 1:
            return 3
        if iMonth == 2:
            return 2
        if iMonth == 3:
            return 1
        return None

    if iMonth in (3, 4):
        return 6
    if iMonth == 5:
        return 5
    if iMonth == 6:
        return 4
    if iMonth == 7:
        return 3
    if iMonth == 8:
        return 2
    if iMonth == 9:
        return 1
    return None


def apply_prepayed_commute_equal_allocation_to_step0003_rows(objRows: List[List[str]]) -> None:
    if not objRows:
        return

    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    if "前払支給分" not in objHeaderRow or "申請の有無" not in objHeaderRow or "等分" not in objHeaderRow:
        return
    try:
        pszPeriodLabel: str = detect_prepayed_commute_period_label(objRows)
    except Exception:
        return
    objRequiredMonthHeaders: List[str] = (
        ["4月", "5月", "6月", "7月", "8月", "9月"]
        if pszPeriodLabel == "04-09月"
        else ["10月", "11月", "12月", "1月", "2月", "3月"]
    )
    if any(pszMonthHeader not in objHeaderRow for pszMonthHeader in objRequiredMonthHeaders):
        return

    iAdvanceIndex: int = objHeaderRow.index("前払支給分")
    iApplicationIndex: int = objHeaderRow.index("申請の有無")
    iEqualIndex: int = objHeaderRow.index("等分")

    for objRow in objRows[1:]:
        if iAdvanceIndex >= len(objRow):
            continue
        pszAdvanceText: str = (objRow[iAdvanceIndex] or "").strip()
        if pszAdvanceText == "":
            continue
        try:
            objAdvanceValue: Decimal = Decimal(pszAdvanceText)
        except (InvalidOperation, ValueError):
            continue

        pszApplicationText: str = ""
        if iApplicationIndex < len(objRow):
            pszApplicationText = (objRow[iApplicationIndex] or "").strip()
        iDivisor = resolve_prepayed_commute_divisor_from_application_text(
            pszApplicationText,
            pszPeriodLabel,
        )
        if iDivisor is None:
            continue

        objEqualValue: Decimal = objAdvanceValue / Decimal(iDivisor)
        while len(objRow) <= iEqualIndex:
            objRow.append("")
        objRow[iEqualIndex] = format_decimal_for_tsv_cell(objEqualValue)


def process_management_accounting_manhour_csv_input(
    objResolvedInputPath: Path,
    objRows: List[List[str]],
) -> int:
    objOutputPath: Path = objResolvedInputPath.resolve().with_suffix(".tsv")
    convert_csv_rows_to_tsv_file(objOutputPath, objRows)
    return 0


def extract_year_text_for_prepayed_commute_allocation(
    objInputPath: Path,
    pszPeriodLabel: str | None = None,
) -> str:
    objMatch = re.search(r"(20\d{2})", str(objInputPath))
    if objMatch is not None:
        return objMatch.group(1)
    try:
        pszYearMonthText: str = extract_year_month_text_from_path(objInputPath)
        return pszYearMonthText[:4]
    except Exception:
        iCurrentYear: int = date.today().year
        iCurrentMonth: int = date.today().month
        if pszPeriodLabel == "10-03月" and 1 <= iCurrentMonth <= 3:
            return str(iCurrentYear - 1)
        return str(iCurrentYear)


def detect_prepayed_commute_period_label(
    objRows: List[List[str]],
) -> str:
    if not objRows:
        raise ValueError("Input TSV has no rows")
    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    if "前払支給分" not in objHeaderRow:
        raise ValueError("Missing header: 前払支給分")
    iAdvanceIndex: int = objHeaderRow.index("前払支給分")
    objMonthValues: List[int] = []
    for iOffset in range(1, 7):
        iIndex: int = iAdvanceIndex + iOffset
        if iIndex >= len(objHeaderRow):
            break
        objMatch = re.match(r"^(\d{1,2})月$", objHeaderRow[iIndex])
        if objMatch is not None:
            objMonthValues.append(int(objMatch.group(1)))
    if objMonthValues == [4, 5, 6, 7, 8, 9]:
        return "04-09月"
    if objMonthValues == [10, 11, 12, 1, 2, 3]:
        return "10-03月"
    raise ValueError(f"Could not detect prepaid commute period from month headers: {objMonthValues}")


def collect_payment_months_from_prepayed_commute_rows(objRows: List[List[str]]) -> set[int]:
    if not objRows:
        return set()
    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    if "申請の有無" not in objHeaderRow:
        return set()
    iApplicationIndex: int = objHeaderRow.index("申請の有無")
    objMonths: set[int] = set()
    for objRow in objRows[1:]:
        if iApplicationIndex >= len(objRow):
            continue
        pszCell: str = (objRow[iApplicationIndex] or "").strip()
        objMatch = re.match(r"^(\d{1,2})/\d{1,2}支給$", pszCell)
        if objMatch is None:
            continue
        objMonths.add(int(objMatch.group(1)))
    return objMonths


def build_new_rawdata_step0001_output_path_for_prepayed_commute_allocation(
    objResolvedInputPath: Path,
    pszPeriodLabel: str,
) -> Path:
    pszYearText: str = extract_year_text_for_prepayed_commute_allocation(objResolvedInputPath, pszPeriodLabel)
    return (
        objResolvedInputPath.resolve().parent
        / f"新_ローデータ_シート_step0001_{pszYearText}年{pszPeriodLabel}_前払通勤交通費按分表.tsv"
    )


def process_prepayed_commute_allocation_tsv_input(
    objResolvedInputPath: Path,
    objRows: List[List[str]],
) -> int:
    pszPeriodLabel: str = detect_prepayed_commute_period_label(objRows)
    objPaymentMonths: set[int] = collect_payment_months_from_prepayed_commute_rows(objRows)
    if pszPeriodLabel == "04-09月" and objPaymentMonths and not objPaymentMonths.issubset({3, 4, 5, 6, 7, 8, 9}):
        print(
            "Warning: payment months {0} do not match expected months {{3,4,5,6,7,8,9}} for 4~9 period: {1}".format(
                sorted(objPaymentMonths),
                objResolvedInputPath,
            )
        )
    if pszPeriodLabel == "10-03月" and objPaymentMonths and not objPaymentMonths.issubset({9, 10, 11, 12, 1, 2, 3}):
        print(
            "Warning: payment months {0} do not match expected months {{9,10,11,12,1,2,3}} for 10~3 period: {1}".format(
                sorted(objPaymentMonths),
                objResolvedInputPath,
            )
        )

    objHeaderRow: List[str] = [("" if objCell is None else str(objCell)).strip() for objCell in objRows[0]]
    objStaffCodeIndices: List[int] = [
        iIndex for iIndex, pszCell in enumerate(objHeaderRow) if pszCell == "スタッフコード"
    ]
    objOutputRows: List[List[str]] = [list(objRow) for objRow in objRows]
    if len(objStaffCodeIndices) >= 2:
        iRemoveIndex: int = objStaffCodeIndices[0]
        objRemovedRows: List[List[str]] = []
        for objRow in objOutputRows:
            objNewRow: List[str] = list(objRow)
            if iRemoveIndex < len(objNewRow):
                objNewRow.pop(iRemoveIndex)
            objRemovedRows.append(objNewRow)
        objOutputRows = objRemovedRows

    objOutputPath: Path = build_new_rawdata_step0001_output_path_for_prepayed_commute_allocation(
        objResolvedInputPath,
        pszPeriodLabel,
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)

    objStep0002Path: Path = build_new_rawdata_step0002_output_path_from_step0001(objOutputPath)
    objStep0002Rows: List[List[str]] = [list(objRow) for objRow in objOutputRows]
    if objStep0002Rows:
        objStep0002HeaderRow: List[str] = [
            ("" if objCell is None else str(objCell)).strip() for objCell in objStep0002Rows[0]
        ]
        objTargetHeaders: List[str] = (
            ["4月", "5月", "6月", "7月", "8月", "9月", "合計", "残り", "等分"]
            if pszPeriodLabel == "04-09月"
            else ["10月", "11月", "12月", "1月", "2月", "3月", "合計", "残り", "等分"]
        )
        objTargetIndices: List[int] = []
        for pszHeader in objTargetHeaders:
            if pszHeader not in objStep0002HeaderRow:
                raise ValueError(f"Missing required header for prepaid commute step0002: {pszHeader}")
            objTargetIndices.append(objStep0002HeaderRow.index(pszHeader))

        for objRow in objStep0002Rows[1:]:
            for iTargetIndex in objTargetIndices:
                if iTargetIndex < len(objRow):
                    objRow[iTargetIndex] = ""
    write_sheet_to_tsv(objStep0002Path, objStep0002Rows)

    try:
        objStep0003Rows: List[List[str]] = [list(objRow) for objRow in objStep0002Rows]
        apply_prepayed_commute_equal_allocation_to_step0003_rows(objStep0003Rows)
        objStep0003Path: Path = build_new_rawdata_step0003_output_path_from_step0002(objStep0002Path)
        write_sheet_to_tsv(objStep0003Path, objStep0003Rows)
        objStep0004Path: Path = process_prepayed_commute_step0004_from_step0003_path(objStep0003Path)
        process_new_rawdata_step0005_from_step0004(objStep0004Path)
        process_new_rawdata_step0006_monthly_prepayed_commute_from_step0004(objStep0004Path)
        objBaseDirectory: Path = objStep0004Path.resolve().parent
        objMainstreamStep0007Paths: List[Path] = sorted(
            [
                objPath
                for objPath in objBaseDirectory.glob("新_ローデータ_シート_step0007_*.tsv")
                if NEW_RAWDATA_STEP0007_FILE_PATTERN.match(objPath.name) is not None
            ]
        )
        for objMainstreamStep0007Path in objMainstreamStep0007Paths:
            process_prepayed_commute_step0007_from_mainstream_step0007(objMainstreamStep0007Path)
        objPrepaidStep0007Paths: List[Path] = sorted(
            objBaseDirectory.glob("新_ローデータ_シート_step0007_*_前払通勤交通費按分表.tsv")
        )
        for objPrepaidStep0007Path in objPrepaidStep0007Paths:
            process_new_rawdata_step0008_prepayed_commute_from_step0007(objPrepaidStep0007Path)
            process_new_rawdata_step0009_prepayed_commute_from_step0008(
                build_new_rawdata_step0008_output_path_from_step0007(objPrepaidStep0007Path)
            )
            process_new_rawdata_step0009_5_from_step0009_with_summary_row_merge(
                build_new_rawdata_step0009_output_path_from_step0008(
                    build_new_rawdata_step0008_output_path_from_step0007(objPrepaidStep0007Path)
                )
            )
            process_new_rawdata_step0010_prepayed_commute_from_step0009(
                build_new_rawdata_step0009_5_output_path_from_step0009(
                    build_new_rawdata_step0009_output_path_from_step0008(
                        build_new_rawdata_step0008_output_path_from_step0007(objPrepaidStep0007Path)
                    )
                )
            )
    except Exception as objException:
        try:
            write_new_rawdata_step0003_error_file_for_prepayed_commute(
                objStep0002Path,
                objException,
            )
        except Exception:
            pass
        raise

    return 0


def extract_year_month_text_from_path(objInputPath: Path) -> str:
    objMatch = YEAR_MONTH_PATTERN.search(str(objInputPath))
    if objMatch is None:
        raise ValueError(f"Could not extract YY.MM月 from input path: {objInputPath}")
    iYear: int = 2000 + int(objMatch.group(1))
    iMonth: int = int(objMatch.group(2))
    return f"{iYear}年{iMonth:02d}月"


def normalize_project_name_for_jobcan_long_tsv(pszProjectName: str) -> str:
    pszNormalized: str = pszProjectName or ""
    pszNormalized = pszNormalized.replace("\t", "_")
    pszNormalized = re.sub(r"(P\d{5})(?![ _\t　【])", r"\1_", pszNormalized)
    pszNormalized = re.sub(r"([A-OQ-Z]\d{3})(?![ _\t　【])", r"\1_", pszNormalized)
    pszNormalized = re.sub(r"^([A-OQ-Z]\d{3}) +", r"\1_", pszNormalized)
    pszNormalized = re.sub(r"([A-OQ-Z]\d{3})[ 　]+", r"\1_", pszNormalized)
    pszNormalized = re.sub(r"(P\d{5})[ 　]+", r"\1_", pszNormalized)
    return pszNormalized


def process_jobcan_long_tsv_input_rawdata_sheet_step0001(
    objResolvedInputPath: Path,
    objRows: List[List[str]],
) -> int:
    pszYearMonthText: str = extract_year_month_text_from_path(objResolvedInputPath)

    objOutputRows: List[List[str]] = []
    pszCurrentStaffName: str = ""
    pszLastOutputStaffName: str = ""
    for objRow in objRows:
        if not any(not is_blank_text(pszCell) for pszCell in objRow):
            continue
        if len(objRow) < 4:
            continue

        pszStaffName: str = (objRow[0] or "").strip()
        if pszStaffName != "":
            pszCurrentStaffName = pszStaffName
        if pszCurrentStaffName == "":
            continue

        pszProjectName: str = normalize_project_name_for_jobcan_long_tsv((objRow[1] or "").strip())
        pszManhour: str = (objRow[3] or "").strip()
        if pszProjectName == "" and pszManhour == "":
            continue

        pszOutputStaffName: str = pszCurrentStaffName
        if pszCurrentStaffName == pszLastOutputStaffName:
            pszOutputStaffName = ""
        else:
            pszLastOutputStaffName = pszCurrentStaffName

        objOutputRows.append([pszOutputStaffName, pszProjectName, pszManhour])

    if not objOutputRows:
        raise ValueError("No output rows generated for Jobcan long-format TSV")

    objOutputPath: Path = (
        objResolvedInputPath.resolve().parent
        / f"ローデータ_シート_step0001_{pszYearMonthText}.tsv"
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_jobcan_long_tsv_input_new_rawdata_sheet_step0001(
    objResolvedInputPath: Path,
    objRows: List[List[str]],
) -> int:
    pszYearMonthText: str = extract_year_month_text_from_path(objResolvedInputPath)

    objOutputRows: List[List[str]] = []
    pszCurrentStaffName: str = ""
    pszLastOutputStaffName: str = ""
    for objRow in objRows:
        if not any(not is_blank_text(pszCell) for pszCell in objRow):
            continue
        if len(objRow) < 4:
            continue

        pszStaffName: str = (objRow[0] or "").strip()
        if pszStaffName != "":
            pszCurrentStaffName = pszStaffName
        if pszCurrentStaffName == "":
            continue

        pszProjectName: str = normalize_project_name_for_jobcan_long_tsv((objRow[1] or "").strip())
        pszManhour: str = (objRow[3] or "").strip()
        if pszProjectName == "" and pszManhour == "":
            continue

        pszOutputStaffName: str = pszCurrentStaffName
        if pszCurrentStaffName == pszLastOutputStaffName:
            pszOutputStaffName = ""
        else:
            pszLastOutputStaffName = pszCurrentStaffName

        objOutputRows.append([pszOutputStaffName, pszProjectName, pszManhour])

    if not objOutputRows:
        raise ValueError("No output rows generated for Jobcan long-format TSV")

    objOutputPath: Path = (
        objResolvedInputPath.resolve().parent
        / f"新_ローデータ_シート_step0001_{pszYearMonthText}.tsv"
    )
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_jobcan_long_tsv_input(objResolvedInputPath: Path, objRows: List[List[str]]) -> int:
    process_jobcan_long_tsv_input_rawdata_sheet_step0001(objResolvedInputPath, objRows)
    process_jobcan_long_tsv_input_new_rawdata_sheet_step0001(objResolvedInputPath, objRows)
    return 0


def build_new_rawdata_step0002_output_path_from_step0001(objStep0001Path: Path) -> Path:
    pszFileName: str = objStep0001Path.name
    if "_step0001_" not in pszFileName:
        raise ValueError(f"Input is not step0001 file: {objStep0001Path}")
    pszOutputFileName: str = pszFileName.replace("_step0001_", "_step0002_", 1)
    return objStep0001Path.resolve().parent / pszOutputFileName


def build_staff_code_by_name_from_salary_step0001(objSalaryStep0001Path: Path) -> dict[str, str]:
    objRows: List[List[str]] = read_tsv_rows(objSalaryStep0001Path)
    objStaffCodeByName: dict[str, str] = {}
    for iRowIndex, objRow in enumerate(objRows):
        if len(objRow) < 2:
            continue
        pszStaffName: str = (objRow[0] or "").strip()
        pszStaffCode: str = (objRow[1] or "").strip()
        if pszStaffName == "" or pszStaffCode == "":
            continue
        if iRowIndex == 0 and pszStaffName == "従業員名" and pszStaffCode == "スタッフコード":
            continue
        objStaffCodeByName[pszStaffName] = pszStaffCode
    if not objStaffCodeByName:
        raise ValueError(f"No staff codes found in salary step0001 TSV: {objSalaryStep0001Path}")
    return objStaffCodeByName


def build_staff_name_by_code_from_salary_step0001(objSalaryStep0001Path: Path) -> dict[str, str]:
    objRows: List[List[str]] = read_tsv_rows(objSalaryStep0001Path)
    objStaffNameByCode: dict[str, str] = {}
    for iRowIndex, objRow in enumerate(objRows):
        if len(objRow) < 2:
            continue
        pszStaffName: str = (objRow[0] or "").strip()
        pszStaffCode: str = (objRow[1] or "").strip()
        if pszStaffName == "" or pszStaffCode == "":
            continue
        if iRowIndex == 0 and pszStaffName == "従業員名" and pszStaffCode == "スタッフコード":
            continue
        if pszStaffCode not in objStaffNameByCode:
            objStaffNameByCode[pszStaffCode] = pszStaffName
    if not objStaffNameByCode:
        raise ValueError(f"No staff names found in salary step0001 TSV: {objSalaryStep0001Path}")
    return objStaffNameByCode


def build_new_rawdata_step0003_name_mapping_output_path(objStep0003Path: Path) -> Path:
    pszFileName: str = objStep0003Path.name
    if not NEW_RAWDATA_STEP0003_FILE_PATTERN.match(pszFileName):
        raise ValueError(f"Input is not step0003 file: {objStep0003Path}")
    pszStem: str = objStep0003Path.stem
    return objStep0003Path.resolve().parent / f"{pszStem}_工数の姓_給与の姓_対応表.tsv"


def build_new_rawdata_step0003_name_mapping_sorted_output_path(objStep0003NameMappingPath: Path) -> Path:
    pszFileName: str = objStep0003NameMappingPath.name
    pszSuffix: str = "_工数の姓_給与の姓_対応表.tsv"
    if not pszFileName.endswith(pszSuffix):
        raise ValueError(f"Input is not step0003 name mapping file: {objStep0003NameMappingPath}")
    pszOutputFileName: str = pszFileName[:-4] + "_昇順.tsv"
    return objStep0003NameMappingPath.resolve().parent / pszOutputFileName


def build_new_rawdata_step0003_old_current_name_mapping_output_path(objStep0003NameMappingPath: Path) -> Path:
    pszFileName: str = objStep0003NameMappingPath.name
    pszSuffix: str = "_工数の姓_給与の姓_対応表_昇順.tsv"
    if not pszFileName.endswith(pszSuffix):
        raise ValueError(f"Input is not step0003 name mapping sorted file: {objStep0003NameMappingPath}")
    pszOutputFileName: str = pszFileName[: -len(pszSuffix)] + "_旧姓_現在の姓_対応表_昇順.tsv"
    return objStep0003NameMappingPath.resolve().parent / pszOutputFileName


def extract_surname_from_full_name(pszFullName: str) -> str:
    pszName: str = (pszFullName or "").strip()
    if pszName == "":
        return ""
    objParts: List[str] = re.split(r"[\s　]+", pszName)
    if not objParts:
        return ""
    return objParts[0]


def process_new_rawdata_step0003_old_current_name_mapping(
    objStep0003NameMappingPath: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0003NameMappingPath)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0003NameMappingPath}")

    objOutputRows: List[List[str]] = [list(objInputRows[0])]
    for objRow in objInputRows[1:]:
        pszManhourName: str = (objRow[1] or "").strip() if len(objRow) >= 2 else ""
        pszSalaryName: str = (objRow[2] or "").strip() if len(objRow) >= 3 else ""
        pszManhourSurname: str = extract_surname_from_full_name(pszManhourName)
        pszSalarySurname: str = extract_surname_from_full_name(pszSalaryName)
        if pszManhourSurname == pszSalarySurname:
            continue
        objOutputRows.append(list(objRow))

    objOutputPath: Path = build_new_rawdata_step0003_old_current_name_mapping_output_path(objStep0003NameMappingPath)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_new_rawdata_step0003_name_mapping_sorted_by_staff_code(
    objStep0003NameMappingPath: Path,
) -> int:
    objInputRows: List[List[str]] = read_tsv_rows(objStep0003NameMappingPath)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objStep0003NameMappingPath}")

    objHeaderRow: List[str] = list(objInputRows[0])
    objDataRows: List[List[str]] = [list(objRow) for objRow in objInputRows[1:]]

    objDataRows.sort(key=lambda objRow: int((objRow[0] or "").strip()) if len(objRow) >= 1 and (objRow[0] or "").strip().isdigit() else 10 ** 18)

    objOutputRows: List[List[str]] = [objHeaderRow] + objDataRows
    objOutputPath: Path = build_new_rawdata_step0003_name_mapping_sorted_output_path(objStep0003NameMappingPath)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_salary_step0001_for_step0003_old_new_name_mapping(
    objNewRawdataStep0003Path: Path,
    objSalaryStep0001Path: Path,
) -> int:
    objStaffNameByCode: dict[str, str] = build_staff_name_by_code_from_salary_step0001(objSalaryStep0001Path)

    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0003Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0003Path}")

    objOutputRows: List[List[str]] = [["スタッフコード", "氏名", "氏名"]]
    objSeenStaffCodes: set[str] = set()
    for objRow in objInputRows:
        pszStaffCode: str = (objRow[0] or "").strip() if len(objRow) >= 1 else ""
        pszStep0003StaffName: str = (objRow[1] or "").strip() if len(objRow) >= 2 else ""
        if pszStep0003StaffName == "":
            continue
        if pszStaffCode == "":
            continue
        if pszStaffCode in objSeenStaffCodes:
            continue

        pszSalaryStaffName: str = objStaffNameByCode.get(pszStaffCode, "")
        objOutputRows.append([pszStaffCode, pszStep0003StaffName, pszSalaryStaffName])
        objSeenStaffCodes.add(pszStaffCode)

    objOutputPath: Path = build_new_rawdata_step0003_name_mapping_output_path(objNewRawdataStep0003Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    process_new_rawdata_step0003_name_mapping_sorted_by_staff_code(objOutputPath)
    objSortedOutputPath: Path = build_new_rawdata_step0003_name_mapping_sorted_output_path(objOutputPath)
    process_new_rawdata_step0003_old_current_name_mapping(objSortedOutputPath)
    return 0


def process_new_rawdata_step0002_from_salary_and_new_rawdata_step0001(
    objSalaryStep0001Path: Path,
    objNewRawdataStep0001Path: Path,
) -> int:
    objStaffCodeByName: dict[str, str] = build_staff_code_by_name_from_salary_step0001(objSalaryStep0001Path)

    objInputRows: List[List[str]] = read_tsv_rows(objNewRawdataStep0001Path)
    if not objInputRows:
        raise ValueError(f"Input TSV has no rows: {objNewRawdataStep0001Path}")

    objOutputRows: List[List[str]] = []
    pszCurrentStaffName: str = ""
    for objRow in objInputRows:
        objNewRow: List[str] = list(objRow)
        if objNewRow:
            pszStaffNameCell: str = (objNewRow[0] or "").strip()
            if pszStaffNameCell != "":
                pszCurrentStaffName = pszStaffNameCell
        pszStaffCode: str = objStaffCodeByName.get(pszCurrentStaffName, "") if pszCurrentStaffName != "" else ""
        objOutputRows.append([pszStaffCode] + objNewRow)

    objOutputPath: Path = build_new_rawdata_step0002_output_path_from_step0001(objNewRawdataStep0001Path)
    write_sheet_to_tsv(objOutputPath, objOutputRows)
    return 0


def process_tsv_input(objResolvedInputPath: Path) -> int:
    objRows: List[List[str]] = read_tsv_rows(objResolvedInputPath)
    if len(objRows) < 2:
        raise ValueError(f"Input TSV has too few rows: {objResolvedInputPath}")

    if is_prepaid_commute_allocation_table_tsv(objResolvedInputPath, objRows):
        return process_prepayed_commute_allocation_tsv_input(objResolvedInputPath, objRows)

    if is_jobcan_long_format_tsv(objRows):
        return process_jobcan_long_tsv_input(objResolvedInputPath, objRows)

    if is_salary_payment_deduction_list_tsv(objRows):
        raise ValueError(f"Salary payment/deduction list TSV is not supported yet: {objResolvedInputPath}")

    raise ValueError(f"Unsupported TSV format: {objResolvedInputPath}")


def build_salary_payment_deduction_step0001_output_path_from_csv(
    objResolvedInputPath: Path,
) -> Path:
    pszStem: str = objResolvedInputPath.stem
    pszStem = re.sub(r"^作成用データ：", "", pszStem)

    pszBaseName: str
    pszDateLabel: str
    pszBaseName, pszSeparator, pszDateLabel = pszStem.rpartition("_")
    if pszSeparator == "" or pszBaseName == "" or pszDateLabel == "":
        raise ValueError(f"Could not build salary step0001 output name from csv: {objResolvedInputPath}")

    pszOutputFileName: str = f"{pszBaseName}_step0001_{pszDateLabel}.tsv"
    return objResolvedInputPath.resolve().with_name(pszOutputFileName)


def process_csv_input(objResolvedInputPath: Path) -> int:
    objRows: List[List[str]] = []
    with open(objResolvedInputPath, mode="r", encoding="utf-8-sig", newline="") as objFile:
        objReader = csv.reader(objFile)
        for objRow in objReader:
            objRows.append(list(objRow))

    if is_management_accounting_manhour_csv(objRows):
        return process_management_accounting_manhour_csv_input(
            objResolvedInputPath,
            objRows,
        )

    objOutputPath: Path = objResolvedInputPath.resolve().with_suffix(".tsv")
    convert_csv_rows_to_tsv_file(objOutputPath, objRows)

    if is_salary_payment_deduction_list_tsv(objRows):
        objSalaryStep0001OutputPath: Path = build_salary_payment_deduction_step0001_output_path_from_csv(
            objResolvedInputPath
        )
        convert_csv_rows_to_tsv_file(objSalaryStep0001OutputPath, objRows)

    return 0


def process_single_input(pszInputXlsxPath: str) -> int:
    objResolvedInputPath: Path = resolve_existing_input_path(pszInputXlsxPath)
    pszSuffix: str = objResolvedInputPath.suffix.lower()

    if pszSuffix == ".tsv":
        return process_tsv_input(objResolvedInputPath)

    if pszSuffix == ".csv":
        return process_csv_input(objResolvedInputPath)

    if pszSuffix != ".xlsx":
        raise ValueError(f"Unsupported extension (only .xlsx/.tsv/.csv): {objResolvedInputPath}")

    objBaseDirectoryPath: Path = objResolvedInputPath.resolve().parent
    pszExcelStem: str = objResolvedInputPath.stem

    try:
        import openpyxl
    except Exception as objException:
        raise RuntimeError(f"Failed to import openpyxl: {objException}") from objException

    try:
        objWorkbook = openpyxl.load_workbook(
            filename=objResolvedInputPath,
            read_only=True,
            data_only=True,
        )
    except Exception as objException:
        raise RuntimeError(f"Failed to read workbook: {objResolvedInputPath}. Detail = {objException}") from objException

    objUsedPaths: set[Path] = set()
    try:
        for objWorksheet in objWorkbook.worksheets:
            pszSanitizedSheetName: str = sanitize_sheet_name_for_file_name(objWorksheet.title)
            objOutputPath: Path = build_unique_output_path(
                objBaseDirectoryPath,
                pszExcelStem,
                pszSanitizedSheetName,
                objUsedPaths,
            )
            objRows: List[List[object]] = [list(objRow) for objRow in objWorksheet.iter_rows(values_only=True)]
            convert_xlsx_rows_to_tsv_file(objOutputPath, objRows)
    finally:
        objWorkbook.close()

    return 0


def main() -> int:
    objParser: argparse.ArgumentParser = argparse.ArgumentParser()
    objParser.add_argument(
        "pszInputXlsxPaths",
        nargs="+",
        help="Input file paths (.xlsx or .tsv or .csv)",
    )
    objArgs: argparse.Namespace = objParser.parse_args()

    iExitCode: int = 0
    objHandledInputPaths: set[Path] = set()

    objSalaryStep0001Paths: List[Path] = []
    objNewRawdataStep0001Paths: List[Path] = []
    objNewRawdataStep0002Paths: List[Path] = []
    objNewRawdataStep0003Paths: List[Path] = []
    objNewRawdataStep0004Paths: List[Path] = []
    objNewRawdataStep0005Paths: List[Path] = []
    objNewRawdataStep0006Paths: List[Path] = []
    objNewRawdataStep0007Paths: List[Path] = []
    objNewRawdataStep0008Paths: List[Path] = []
    objNewRawdataStep0009Paths: List[Path] = []
    objNewRawdataStep0010Paths: List[Path] = []
    objNewRawdataStep0011Paths: List[Path] = []
    objNewRawdataStep0012Paths: List[Path] = []
    objNewRawdataStep0012PrepaidCommutePaths: List[Path] = []
    objNewRawdataStep0013Paths: List[Path] = []
    objNewRawdataStep0014Paths: List[Path] = []
    objNewRawdataStep0015Paths: List[Path] = []
    objNewRawdataStep0016Paths: List[Path] = []
    objNewRawdataStep0017Paths: List[Path] = []
    objNewRawdataStep0018Paths: List[Path] = []
    objNewRawdataStep0019Paths: List[Path] = []
    objNewRawdataStep0020Paths: List[Path] = []
    objNewRawdataStep0021Paths: List[Path] = []
    objNewRawdataStep0022Paths: List[Path] = []
    objNewRawdataStep0023Paths: List[Path] = []
    objNewRawdataStep0024Paths: List[Path] = []
    objNewRawdataStep0025Paths: List[Path] = []
    objNewRawdataStep0026Paths: List[Path] = []
    objNewRawdataStep0027Paths: List[Path] = []
    objNewRawdataStep0028Paths: List[Path] = []
    objNewRawdataStep0014StatutoryWelfarePaths: List[Path] = []
    objNewRawdataStep0015StatutoryWelfarePaths: List[Path] = []
    objNewRawdataStep0016StatutoryWelfarePaths: List[Path] = []
    objNewRawdataStep0017StatutoryWelfarePaths: List[Path] = []
    objNewRawdataStep0018StatutoryWelfarePaths: List[Path] = []
    objNewRawdataStep0019StatutoryWelfarePaths: List[Path] = []
    objNewRawdataStep0020StatutoryWelfarePaths: List[Path] = []
    objNewRawdataStep0021StatutoryWelfarePaths: List[Path] = []
    objNewRawdataStep0022StatutoryWelfarePaths: List[Path] = []
    objManagementAccountingCandidatePaths: List[Path] = []



    for pszInputXlsxPath in objArgs.pszInputXlsxPaths:
        try:
            objResolvedInputPath: Path = resolve_existing_input_path(pszInputXlsxPath)
        except Exception:
            continue

        if SALARY_PAYMENT_STEP0001_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objSalaryStep0001Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0001_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0001Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0002_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0002Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0003_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0003Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0004_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0004Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0005_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0005Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0006_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0006Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0007_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0007Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0008_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0008Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0009_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0009Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0010_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0010Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0011_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0011Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0012_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0012Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0012_PREPAYED_COMMUTE_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0012PrepaidCommutePaths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0013_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0013Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0014_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0014Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0015_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0015Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0016_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0016Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0017_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0017Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0018_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0018Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0019_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0019Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0020_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0020Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0021_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0021Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0022_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0022Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0023_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0023Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0024_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0024Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0025_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0025Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0026_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0026Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0027_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0027Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0028_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0028Paths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0014_STATUTORY_WELFARE_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0014StatutoryWelfarePaths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0015_STATUTORY_WELFARE_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0015StatutoryWelfarePaths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0016_STATUTORY_WELFARE_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0016StatutoryWelfarePaths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0017_STATUTORY_WELFARE_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0017StatutoryWelfarePaths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0018_STATUTORY_WELFARE_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0018StatutoryWelfarePaths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0019_STATUTORY_WELFARE_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0019StatutoryWelfarePaths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0020_STATUTORY_WELFARE_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0020StatutoryWelfarePaths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0021_STATUTORY_WELFARE_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0021StatutoryWelfarePaths.append(objResolvedInputPath)
        if NEW_RAWDATA_STEP0022_STATUTORY_WELFARE_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
            objNewRawdataStep0022StatutoryWelfarePaths.append(objResolvedInputPath)

        if objResolvedInputPath.suffix.lower() in (".tsv", ".csv", ".xlsx"):
            objManagementAccountingCandidatePaths.append(objResolvedInputPath)

    if objSalaryStep0001Paths:
        objSalaryStep0001Path: Path = objSalaryStep0001Paths[0]
        for objNewRawdataStep0001Path in objNewRawdataStep0001Paths:
            try:
                process_new_rawdata_step0002_from_salary_and_new_rawdata_step0001(
                    objSalaryStep0001Path,
                    objNewRawdataStep0001Path,
                )
                objHandledInputPaths.add(objSalaryStep0001Path.resolve())
                objHandledInputPaths.add(objNewRawdataStep0001Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0002 pair: {0} / {1}. Detail = {2}".format(
                        objSalaryStep0001Path,
                        objNewRawdataStep0001Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0002Paths:
        objPrepaidCommuteStep0002Paths: List[Path] = []
        objNormalStep0002Paths: List[Path] = []
        for objNewRawdataStep0002Path in objNewRawdataStep0002Paths:
            if "前払通勤交通費按分表" in objNewRawdataStep0002Path.name:
                objPrepaidCommuteStep0002Paths.append(objNewRawdataStep0002Path)
            else:
                objNormalStep0002Paths.append(objNewRawdataStep0002Path)

        for objPrepaidCommuteStep0002Path in objPrepaidCommuteStep0002Paths:
            try:
                process_new_rawdata_step0003_from_prepayed_commute_step0002(
                    objPrepaidCommuteStep0002Path
                )
                objNewRawdataStep0003Path: Path = build_new_rawdata_step0003_output_path_from_step0002(
                    objPrepaidCommuteStep0002Path
                )
                objHandledInputPaths.add(objPrepaidCommuteStep0002Path.resolve())
                objHandledInputPaths.add(objNewRawdataStep0003Path.resolve())
            except Exception as objException:
                try:
                    write_new_rawdata_step0003_error_file_for_prepayed_commute(
                        objPrepaidCommuteStep0002Path,
                        objException,
                    )
                except Exception:
                    pass
                print(
                    "Error: failed to process prepaid commute step0003 from step0002: {0}. Detail = {1}".format(
                        objPrepaidCommuteStep0002Path,
                        objException,
                    )
                )
                iExitCode = 1

        if objNormalStep0002Paths and not objSalaryStep0001Paths:
            print("Error: salary step0001 TSV is required to process step0003 from step0002")
            iExitCode = 1
        else:
            for objNewRawdataStep0002Path in objNormalStep0002Paths:
                for objManagementAccountingCandidatePath in objManagementAccountingCandidatePaths:
                    if objManagementAccountingCandidatePath.resolve() == objNewRawdataStep0002Path.resolve():
                        continue
                    try:
                        objStaffCodeByName: dict[str, str] = load_staff_code_by_name_from_management_accounting_file(
                            objManagementAccountingCandidatePath
                        )
                    except Exception:
                        continue

                    try:
                        fill_missing_staff_codes_in_new_rawdata_step0002_by_management_accounting(
                            objNewRawdataStep0002Path,
                            objStaffCodeByName,
                        )
                        objNewRawdataStep0003Path: Path = build_new_rawdata_step0003_output_path_from_step0002(
                            objNewRawdataStep0002Path
                        )
                        process_salary_step0001_for_step0003_old_new_name_mapping(
                            objNewRawdataStep0003Path,
                            objSalaryStep0001Paths[0],
                        )
                        process_new_rawdata_step0004_from_step0003(objNewRawdataStep0003Path)
                        objNewRawdataStep0004Path: Path = build_new_rawdata_step0004_output_path_from_step0003(
                            objNewRawdataStep0003Path
                        )
                        process_new_rawdata_step0005_from_step0004(objNewRawdataStep0004Path)
                        if "前払通勤交通費按分表" in objNewRawdataStep0004Path.name:
                            process_new_rawdata_step0006_monthly_prepayed_commute_from_step0004(
                                objNewRawdataStep0004Path
                            )
                            objNewRawdataStep0005Path = build_new_rawdata_step0005_output_path_from_step0004(
                                objNewRawdataStep0004Path
                            )
                            objNewRawdataStep0006Path = build_new_rawdata_step0006_output_path_from_step0005(
                                objNewRawdataStep0005Path
                            )
                        else:
                            objNewRawdataStep0005Path = build_new_rawdata_step0005_output_path_from_step0004(
                                objNewRawdataStep0004Path
                            )
                            process_new_rawdata_step0006_from_step0005(objNewRawdataStep0005Path)
                            objNewRawdataStep0006Path = build_new_rawdata_step0006_output_path_from_step0005(
                                objNewRawdataStep0005Path
                            )
                        process_new_rawdata_step0007_from_step0006(objNewRawdataStep0006Path)
                        objNewRawdataStep0007Path: Path = build_new_rawdata_step0007_output_path_from_step0006(
                            objNewRawdataStep0006Path
                        )
                        process_prepayed_commute_step0007_from_mainstream_step0007(objNewRawdataStep0007Path)
                        process_new_rawdata_step0008_mainstream_from_step0007(objNewRawdataStep0007Path)
                        if objSalaryStep0001Paths:
                            objNewRawdataStep0008Path: Path = build_new_rawdata_step0008_output_path_from_step0007(
                                objNewRawdataStep0007Path
                            )
                            process_new_rawdata_step0009_from_step0008_and_salary_step0001(
                                objNewRawdataStep0008Path,
                                objSalaryStep0001Paths[0],
                            )
                            process_new_rawdata_step0010_from_step0009_and_salary_step0001(
                                build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path),
                                objSalaryStep0001Paths[0],
                            )
                            process_new_rawdata_step0011_and_step0012_from_step0010(
                                build_new_rawdata_step0010_output_path_from_step0009(
                                    build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path)
                                )
                            )
                        objHandledInputPaths.add(objNewRawdataStep0005Path.resolve())
                        objHandledInputPaths.add(objNewRawdataStep0006Path.resolve())
                        objHandledInputPaths.add(objNewRawdataStep0002Path.resolve())
                        objHandledInputPaths.add(objNewRawdataStep0003Path.resolve())
                        objHandledInputPaths.add(objNewRawdataStep0004Path.resolve())
                        objHandledInputPaths.add(objManagementAccountingCandidatePath.resolve())
                    except Exception as objException:
                        if "前払通勤交通費按分表" in objNewRawdataStep0002Path.name:
                            try:
                                write_new_rawdata_step0003_error_file_for_prepayed_commute(
                                    objNewRawdataStep0002Path,
                                    objException,
                                )
                            except Exception:
                                pass
                        print(
                            "Error: failed to fill missing step0002 staff codes: {0} / {1}. Detail = {2}".format(
                                objNewRawdataStep0002Path,
                                objManagementAccountingCandidatePath,
                                objException,
                            )
                        )
                        iExitCode = 1
                    break
    if objNewRawdataStep0003Paths:
        for objNewRawdataStep0003Path in objNewRawdataStep0003Paths:
            if objNewRawdataStep0003Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0004_from_step0003(objNewRawdataStep0003Path)
                objNewRawdataStep0004Path: Path = build_new_rawdata_step0004_output_path_from_step0003(
                    objNewRawdataStep0003Path
                )
                process_new_rawdata_step0005_from_step0004(objNewRawdataStep0004Path)
                if "前払通勤交通費按分表" in objNewRawdataStep0004Path.name:
                    process_new_rawdata_step0006_monthly_prepayed_commute_from_step0004(
                        objNewRawdataStep0004Path
                    )
                    objHandledInputPaths.add(objNewRawdataStep0003Path.resolve())
                    objHandledInputPaths.add(objNewRawdataStep0004Path.resolve())
                    continue
                objNewRawdataStep0005Path = build_new_rawdata_step0005_output_path_from_step0004(
                    objNewRawdataStep0004Path
                )
                process_new_rawdata_step0006_from_step0005(objNewRawdataStep0005Path)
                objNewRawdataStep0006Path = build_new_rawdata_step0006_output_path_from_step0005(
                    objNewRawdataStep0005Path
                )
                process_new_rawdata_step0007_from_step0006(objNewRawdataStep0006Path)
                objNewRawdataStep0007Path: Path = build_new_rawdata_step0007_output_path_from_step0006(
                    objNewRawdataStep0006Path
                )
                process_prepayed_commute_step0007_from_mainstream_step0007(objNewRawdataStep0007Path)
                process_new_rawdata_step0008_mainstream_from_step0007(objNewRawdataStep0007Path)
                if objSalaryStep0001Paths:
                    objNewRawdataStep0008Path: Path = build_new_rawdata_step0008_output_path_from_step0007(
                        objNewRawdataStep0007Path
                    )
                    process_new_rawdata_step0009_from_step0008_and_salary_step0001(
                        objNewRawdataStep0008Path,
                        objSalaryStep0001Paths[0],
                    )
                    process_new_rawdata_step0010_from_step0009_and_salary_step0001(
                        build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path),
                        objSalaryStep0001Paths[0],
                    )
                    process_new_rawdata_step0011_and_step0012_from_step0010(
                        build_new_rawdata_step0010_output_path_from_step0009(
                            build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path)
                        )
                    )
                objHandledInputPaths.add(objNewRawdataStep0003Path.resolve())
                objHandledInputPaths.add(objNewRawdataStep0004Path.resolve())
                objHandledInputPaths.add(objNewRawdataStep0005Path.resolve())
                objHandledInputPaths.add(objNewRawdataStep0006Path.resolve())
                objHandledInputPaths.add(objNewRawdataStep0006Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0004 from step0003: {0}. Detail = {1}".format(
                        objNewRawdataStep0003Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0004Paths:
        for objNewRawdataStep0004Path in objNewRawdataStep0004Paths:
            if objNewRawdataStep0004Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0005_from_step0004(objNewRawdataStep0004Path)
                if "前払通勤交通費按分表" in objNewRawdataStep0004Path.name:
                    process_new_rawdata_step0006_monthly_prepayed_commute_from_step0004(
                        objNewRawdataStep0004Path
                    )
                    objHandledInputPaths.add(objNewRawdataStep0004Path.resolve())
                    continue
                objNewRawdataStep0005Path = build_new_rawdata_step0005_output_path_from_step0004(
                    objNewRawdataStep0004Path
                )
                process_new_rawdata_step0006_from_step0005(objNewRawdataStep0005Path)
                objNewRawdataStep0006Path = build_new_rawdata_step0006_output_path_from_step0005(
                    objNewRawdataStep0005Path
                )
                process_new_rawdata_step0007_from_step0006(objNewRawdataStep0006Path)
                objNewRawdataStep0007Path: Path = build_new_rawdata_step0007_output_path_from_step0006(
                    objNewRawdataStep0006Path
                )
                process_prepayed_commute_step0007_from_mainstream_step0007(objNewRawdataStep0007Path)
                process_new_rawdata_step0008_mainstream_from_step0007(objNewRawdataStep0007Path)
                if objSalaryStep0001Paths:
                    objNewRawdataStep0008Path: Path = build_new_rawdata_step0008_output_path_from_step0007(
                        objNewRawdataStep0007Path
                    )
                    process_new_rawdata_step0009_from_step0008_and_salary_step0001(
                        objNewRawdataStep0008Path,
                        objSalaryStep0001Paths[0],
                    )
                    process_new_rawdata_step0010_from_step0009_and_salary_step0001(
                        build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path),
                        objSalaryStep0001Paths[0],
                    )
                    process_new_rawdata_step0011_and_step0012_from_step0010(
                        build_new_rawdata_step0010_output_path_from_step0009(
                            build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path)
                        )
                    )
                objHandledInputPaths.add(objNewRawdataStep0004Path.resolve())
                objHandledInputPaths.add(objNewRawdataStep0005Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0005 from step0004: {0}. Detail = {1}".format(
                        objNewRawdataStep0004Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0005Paths:
        for objNewRawdataStep0005Path in objNewRawdataStep0005Paths:
            if objNewRawdataStep0005Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0006_from_step0005(objNewRawdataStep0005Path)
                objNewRawdataStep0006Path: Path = build_new_rawdata_step0006_output_path_from_step0005(
                    objNewRawdataStep0005Path
                )
                process_new_rawdata_step0007_from_step0006(objNewRawdataStep0006Path)
                objNewRawdataStep0007Path: Path = build_new_rawdata_step0007_output_path_from_step0006(
                    objNewRawdataStep0006Path
                )
                process_prepayed_commute_step0007_from_mainstream_step0007(objNewRawdataStep0007Path)
                process_new_rawdata_step0008_mainstream_from_step0007(objNewRawdataStep0007Path)
                if objSalaryStep0001Paths:
                    objNewRawdataStep0008Path: Path = build_new_rawdata_step0008_output_path_from_step0007(
                        objNewRawdataStep0007Path
                    )
                    process_new_rawdata_step0009_from_step0008_and_salary_step0001(
                        objNewRawdataStep0008Path,
                        objSalaryStep0001Paths[0],
                    )
                    process_new_rawdata_step0010_from_step0009_and_salary_step0001(
                        build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path),
                        objSalaryStep0001Paths[0],
                    )
                    process_new_rawdata_step0011_and_step0012_from_step0010(
                        build_new_rawdata_step0010_output_path_from_step0009(
                            build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path)
                        )
                    )
                objHandledInputPaths.add(objNewRawdataStep0005Path.resolve())
                objHandledInputPaths.add(objNewRawdataStep0006Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0006 from step0005: {0}. Detail = {1}".format(
                        objNewRawdataStep0005Path,
                        objException,
                    )
                )
                iExitCode = 1



    if objNewRawdataStep0006Paths:
        for objNewRawdataStep0006Path in objNewRawdataStep0006Paths:
            if objNewRawdataStep0006Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0007_from_step0006(objNewRawdataStep0006Path)
                objNewRawdataStep0007Path: Path = build_new_rawdata_step0007_output_path_from_step0006(
                    objNewRawdataStep0006Path
                )
                process_prepayed_commute_step0007_from_mainstream_step0007(objNewRawdataStep0007Path)
                process_new_rawdata_step0008_mainstream_from_step0007(objNewRawdataStep0007Path)
                if objSalaryStep0001Paths:
                    objNewRawdataStep0008Path: Path = build_new_rawdata_step0008_output_path_from_step0007(
                        objNewRawdataStep0007Path
                    )
                    process_new_rawdata_step0009_from_step0008_and_salary_step0001(
                        objNewRawdataStep0008Path,
                        objSalaryStep0001Paths[0],
                    )
                    process_new_rawdata_step0010_from_step0009_and_salary_step0001(
                        build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path),
                        objSalaryStep0001Paths[0],
                    )
                    process_new_rawdata_step0011_and_step0012_from_step0010(
                        build_new_rawdata_step0010_output_path_from_step0009(
                            build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path)
                        )
                    )
                objHandledInputPaths.add(objNewRawdataStep0006Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0007 from step0006: {0}. Detail = {1}".format(
                        objNewRawdataStep0006Path,
                        objException,
                    )
                )
                iExitCode = 1


    if objNewRawdataStep0007Paths:
        for objNewRawdataStep0007Path in objNewRawdataStep0007Paths:
            if objNewRawdataStep0007Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_prepayed_commute_step0007_from_mainstream_step0007(objNewRawdataStep0007Path)
                process_new_rawdata_step0008_mainstream_from_step0007(objNewRawdataStep0007Path)
                if objSalaryStep0001Paths:
                    objNewRawdataStep0008Path: Path = build_new_rawdata_step0008_output_path_from_step0007(
                        objNewRawdataStep0007Path
                    )
                    process_new_rawdata_step0009_from_step0008_and_salary_step0001(
                        objNewRawdataStep0008Path,
                        objSalaryStep0001Paths[0],
                    )
                    process_new_rawdata_step0010_from_step0009_and_salary_step0001(
                        build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path),
                        objSalaryStep0001Paths[0],
                    )
                    process_new_rawdata_step0011_and_step0012_from_step0010(
                        build_new_rawdata_step0010_output_path_from_step0009(
                            build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path)
                        )
                    )
                objHandledInputPaths.add(objNewRawdataStep0007Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0008 from step0007: {0}. Detail = {1}".format(
                        objNewRawdataStep0007Path,
                        objException,
                    )
                )
                iExitCode = 1


    if objNewRawdataStep0008Paths:
        objPrepaidStep0008Paths: List[Path] = []
        objNormalStep0008Paths: List[Path] = []
        for objNewRawdataStep0008Path in objNewRawdataStep0008Paths:
            if "前払通勤交通費按分表" in objNewRawdataStep0008Path.name:
                objPrepaidStep0008Paths.append(objNewRawdataStep0008Path)
            else:
                objNormalStep0008Paths.append(objNewRawdataStep0008Path)

        for objPrepaidStep0008Path in objPrepaidStep0008Paths:
            if objPrepaidStep0008Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0009_prepayed_commute_from_step0008(objPrepaidStep0008Path)
                process_new_rawdata_step0009_5_from_step0009_with_summary_row_merge(
                    build_new_rawdata_step0009_output_path_from_step0008(objPrepaidStep0008Path)
                )
                process_new_rawdata_step0010_prepayed_commute_from_step0009(
                    build_new_rawdata_step0009_5_output_path_from_step0009(
                        build_new_rawdata_step0009_output_path_from_step0008(objPrepaidStep0008Path)
                    )
                )
                objHandledInputPaths.add(objPrepaidStep0008Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process prepaid commute step0009 from step0008: {0}. Detail = {1}".format(
                        objPrepaidStep0008Path,
                        objException,
                    )
                )
                iExitCode = 1

        if objNormalStep0008Paths:
            if not objSalaryStep0001Paths:
                print("Error: salary step0001 TSV is required to process step0009 from step0008")
                iExitCode = 1
            else:
                for objNewRawdataStep0008Path in objNormalStep0008Paths:
                    if objNewRawdataStep0008Path.resolve() in objHandledInputPaths:
                        continue
                    try:
                        process_new_rawdata_step0009_from_step0008_and_salary_step0001(
                            objNewRawdataStep0008Path,
                            objSalaryStep0001Paths[0],
                        )
                        process_new_rawdata_step0010_from_step0009_and_salary_step0001(
                            build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path),
                            objSalaryStep0001Paths[0],
                        )
                        process_new_rawdata_step0011_and_step0012_from_step0010(
                            build_new_rawdata_step0010_output_path_from_step0009(
                                build_new_rawdata_step0009_output_path_from_step0008(objNewRawdataStep0008Path)
                            )
                        )
                        objHandledInputPaths.add(objNewRawdataStep0008Path.resolve())
                    except Exception as objException:
                        print(
                            "Error: failed to process step0009 from step0008: {0}. Detail = {1}".format(
                                objNewRawdataStep0008Path,
                                objException,
                            )
                        )
                        iExitCode = 1


    if objNewRawdataStep0009Paths:
        objPrepaidStep0009Paths: List[Path] = []
        objNormalStep0009Paths: List[Path] = []
        for objNewRawdataStep0009Path in objNewRawdataStep0009Paths:
            if "前払通勤交通費按分表" in objNewRawdataStep0009Path.name:
                objPrepaidStep0009Paths.append(objNewRawdataStep0009Path)
            else:
                objNormalStep0009Paths.append(objNewRawdataStep0009Path)

        for objPrepaidStep0009Path in objPrepaidStep0009Paths:
            if objPrepaidStep0009Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0009_5_from_step0009_with_summary_row_merge(objPrepaidStep0009Path)
                process_new_rawdata_step0010_prepayed_commute_from_step0009(
                    build_new_rawdata_step0009_5_output_path_from_step0009(objPrepaidStep0009Path)
                )
                objHandledInputPaths.add(objPrepaidStep0009Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process prepaid commute step0010 from step0009: {0}. Detail = {1}".format(
                        objPrepaidStep0009Path,
                        objException,
                    )
                )
                iExitCode = 1

        if objNormalStep0009Paths:
            if not objSalaryStep0001Paths:
                print("Error: salary step0001 TSV is required to process step0010 from step0009")
                iExitCode = 1
            else:
                for objNewRawdataStep0009Path in objNormalStep0009Paths:
                    if objNewRawdataStep0009Path.resolve() in objHandledInputPaths:
                        continue
                    try:
                        process_new_rawdata_step0010_from_step0009_and_salary_step0001(
                            objNewRawdataStep0009Path,
                            objSalaryStep0001Paths[0],
                        )
                        process_new_rawdata_step0011_and_step0012_from_step0010(
                            build_new_rawdata_step0010_output_path_from_step0009(objNewRawdataStep0009Path)
                        )
                        objHandledInputPaths.add(objNewRawdataStep0009Path.resolve())
                    except Exception as objException:
                        print(
                            "Error: failed to process step0010 from step0009: {0}. Detail = {1}".format(
                                objNewRawdataStep0009Path,
                                objException,
                            )
                        )
                        iExitCode = 1


    if objNewRawdataStep0010Paths:
        for objNewRawdataStep0010Path in objNewRawdataStep0010Paths:
            if objNewRawdataStep0010Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0011_and_step0012_from_step0010(objNewRawdataStep0010Path)
                objHandledInputPaths.add(objNewRawdataStep0010Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0011 from step0010: {0}. Detail = {1}".format(
                        objNewRawdataStep0010Path,
                        objException,
                    )
                )
                iExitCode = 1


    if objNewRawdataStep0011Paths:
        for objNewRawdataStep0011Path in objNewRawdataStep0011Paths:
            if objNewRawdataStep0011Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0012_and_step0013_from_step0011(objNewRawdataStep0011Path)
                objHandledInputPaths.add(objNewRawdataStep0011Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0012 from step0011: {0}. Detail = {1}".format(
                        objNewRawdataStep0011Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0012PrepaidCommutePaths:
        for objNewRawdataStep0012PrepaidCommutePath in objNewRawdataStep0012PrepaidCommutePaths:
            if objNewRawdataStep0012PrepaidCommutePath.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0013_prepayed_commute_from_step0012(
                    objNewRawdataStep0012PrepaidCommutePath
                )
                objHandledInputPaths.add(objNewRawdataStep0012PrepaidCommutePath.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process prepaid commute step0013 from step0012: {0}. Detail = {1}".format(
                        objNewRawdataStep0012PrepaidCommutePath,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0012Paths:
        for objNewRawdataStep0012Path in objNewRawdataStep0012Paths:
            if objNewRawdataStep0012Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0013_from_step0012(objNewRawdataStep0012Path)
                objHandledInputPaths.add(objNewRawdataStep0012Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0013 from step0012: {0}. Detail = {1}".format(
                        objNewRawdataStep0012Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0013Paths:
        for objNewRawdataStep0013Path in objNewRawdataStep0013Paths:
            if "前払通勤交通費按分表" in objNewRawdataStep0013Path.name:
                continue
            if objNewRawdataStep0013Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0014_from_step0013(objNewRawdataStep0013Path)
                objHandledInputPaths.add(objNewRawdataStep0013Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0014 from step0013: {0}. Detail = {1}".format(
                        objNewRawdataStep0013Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0014Paths:
        for objNewRawdataStep0014Path in objNewRawdataStep0014Paths:
            if objNewRawdataStep0014Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0015_from_step0014(objNewRawdataStep0014Path)
                objHandledInputPaths.add(objNewRawdataStep0014Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0015 from step0014: {0}. Detail = {1}".format(
                        objNewRawdataStep0014Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0015Paths:
        for objNewRawdataStep0015Path in objNewRawdataStep0015Paths:
            if objNewRawdataStep0015Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0016_from_step0015(objNewRawdataStep0015Path)
                objHandledInputPaths.add(objNewRawdataStep0015Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0016 from step0015: {0}. Detail = {1}".format(
                        objNewRawdataStep0015Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0016Paths:
        for objNewRawdataStep0016Path in objNewRawdataStep0016Paths:
            if objNewRawdataStep0016Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0017_from_step0016(objNewRawdataStep0016Path)
                objHandledInputPaths.add(objNewRawdataStep0016Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0017 from step0016: {0}. Detail = {1}".format(
                        objNewRawdataStep0016Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0017Paths:
        for objNewRawdataStep0017Path in objNewRawdataStep0017Paths:
            if objNewRawdataStep0017Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0018_from_step0017(objNewRawdataStep0017Path)
                objHandledInputPaths.add(objNewRawdataStep0017Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0018 from step0017: {0}. Detail = {1}".format(
                        objNewRawdataStep0017Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0018Paths:
        for objNewRawdataStep0018Path in objNewRawdataStep0018Paths:
            if objNewRawdataStep0018Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0019_from_step0018(objNewRawdataStep0018Path)
                objHandledInputPaths.add(objNewRawdataStep0018Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0019 from step0018: {0}. Detail = {1}".format(
                        objNewRawdataStep0018Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0019Paths:
        for objNewRawdataStep0019Path in objNewRawdataStep0019Paths:
            if objNewRawdataStep0019Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0020_from_step0019(objNewRawdataStep0019Path)
                objHandledInputPaths.add(objNewRawdataStep0019Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0020 from step0019: {0}. Detail = {1}".format(
                        objNewRawdataStep0019Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0020Paths:
        for objNewRawdataStep0020Path in objNewRawdataStep0020Paths:
            if objNewRawdataStep0020Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0021_from_step0020_with_org_table(objNewRawdataStep0020Path)
                objHandledInputPaths.add(objNewRawdataStep0020Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0021 from step0020: {0}. Detail = {1}".format(
                        objNewRawdataStep0020Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0021Paths:
        for objNewRawdataStep0021Path in objNewRawdataStep0021Paths:
            if objNewRawdataStep0021Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0022_from_step0021(objNewRawdataStep0021Path)
                objHandledInputPaths.add(objNewRawdataStep0021Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0022 from step0021: {0}. Detail = {1}".format(
                        objNewRawdataStep0021Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0022Paths:
        for objNewRawdataStep0022Path in objNewRawdataStep0022Paths:
            if objNewRawdataStep0022Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0023_from_step0022_with_org_table(objNewRawdataStep0022Path)
                objHandledInputPaths.add(objNewRawdataStep0022Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0023 from step0022: {0}. Detail = {1}".format(
                        objNewRawdataStep0022Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0023Paths:
        for objNewRawdataStep0023Path in objNewRawdataStep0023Paths:
            if objNewRawdataStep0023Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0024_from_step0023_with_org_table(objNewRawdataStep0023Path)
                objHandledInputPaths.add(objNewRawdataStep0023Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0024 from step0023: {0}. Detail = {1}".format(
                        objNewRawdataStep0023Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0024Paths:
        for objNewRawdataStep0024Path in objNewRawdataStep0024Paths:
            if objNewRawdataStep0024Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0025_from_step0024_with_org_table(objNewRawdataStep0024Path)
                objHandledInputPaths.add(objNewRawdataStep0024Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0025 from step0024: {0}. Detail = {1}".format(
                        objNewRawdataStep0024Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0025Paths:
        for objNewRawdataStep0025Path in objNewRawdataStep0025Paths:
            if objNewRawdataStep0025Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0026_from_step0025(objNewRawdataStep0025Path)
                objHandledInputPaths.add(objNewRawdataStep0025Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0026 from step0025: {0}. Detail = {1}".format(
                        objNewRawdataStep0025Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0026Paths:
        for objNewRawdataStep0026Path in objNewRawdataStep0026Paths:
            if objNewRawdataStep0026Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0027_from_step0026(objNewRawdataStep0026Path)
                objHandledInputPaths.add(objNewRawdataStep0026Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0027 from step0026: {0}. Detail = {1}".format(
                        objNewRawdataStep0026Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0027Paths:
        for objNewRawdataStep0027Path in objNewRawdataStep0027Paths:
            if objNewRawdataStep0027Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0028_from_step0027(objNewRawdataStep0027Path)
                objHandledInputPaths.add(objNewRawdataStep0027Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0028 from step0027: {0}. Detail = {1}".format(
                        objNewRawdataStep0027Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0028Paths:
        for objNewRawdataStep0028Path in objNewRawdataStep0028Paths:
            if objNewRawdataStep0028Path.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0029_from_step0028(objNewRawdataStep0028Path)
                objHandledInputPaths.add(objNewRawdataStep0028Path.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0029 from step0028: {0}. Detail = {1}".format(
                        objNewRawdataStep0028Path,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0014StatutoryWelfarePaths:
        for objNewRawdataStep0014StatutoryWelfarePath in objNewRawdataStep0014StatutoryWelfarePaths:
            if objNewRawdataStep0014StatutoryWelfarePath.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0015_statutory_welfare_from_step0014_statutory_welfare(
                    objNewRawdataStep0014StatutoryWelfarePath
                )
                objHandledInputPaths.add(objNewRawdataStep0014StatutoryWelfarePath.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0015 from step0014 statutory welfare: {0}. Detail = {1}".format(
                        objNewRawdataStep0014StatutoryWelfarePath,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0015StatutoryWelfarePaths:
        for objNewRawdataStep0015StatutoryWelfarePath in objNewRawdataStep0015StatutoryWelfarePaths:
            if objNewRawdataStep0015StatutoryWelfarePath.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0016_statutory_welfare_from_step0015_statutory_welfare(
                    objNewRawdataStep0015StatutoryWelfarePath
                )
                objHandledInputPaths.add(objNewRawdataStep0015StatutoryWelfarePath.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0016 from step0015 statutory welfare: {0}. Detail = {1}".format(
                        objNewRawdataStep0015StatutoryWelfarePath,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0016StatutoryWelfarePaths:
        for objNewRawdataStep0016StatutoryWelfarePath in objNewRawdataStep0016StatutoryWelfarePaths:
            if objNewRawdataStep0016StatutoryWelfarePath.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0017_statutory_welfare_from_step0016_statutory_welfare(
                    objNewRawdataStep0016StatutoryWelfarePath
                )
                objHandledInputPaths.add(objNewRawdataStep0016StatutoryWelfarePath.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0017 from step0016 statutory welfare: {0}. Detail = {1}".format(
                        objNewRawdataStep0016StatutoryWelfarePath,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0017StatutoryWelfarePaths:
        for objNewRawdataStep0017StatutoryWelfarePath in objNewRawdataStep0017StatutoryWelfarePaths:
            if objNewRawdataStep0017StatutoryWelfarePath.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0018_statutory_welfare_from_step0017_statutory_welfare(
                    objNewRawdataStep0017StatutoryWelfarePath
                )
                objHandledInputPaths.add(objNewRawdataStep0017StatutoryWelfarePath.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0018 from step0017 statutory welfare: {0}. Detail = {1}".format(
                        objNewRawdataStep0017StatutoryWelfarePath,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0018StatutoryWelfarePaths:
        for objNewRawdataStep0018StatutoryWelfarePath in objNewRawdataStep0018StatutoryWelfarePaths:
            if objNewRawdataStep0018StatutoryWelfarePath.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0019_statutory_welfare_from_step0018_statutory_welfare(
                    objNewRawdataStep0018StatutoryWelfarePath
                )
                objHandledInputPaths.add(objNewRawdataStep0018StatutoryWelfarePath.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0019 from step0018 statutory welfare: {0}. Detail = {1}".format(
                        objNewRawdataStep0018StatutoryWelfarePath,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0019StatutoryWelfarePaths:
        for objNewRawdataStep0019StatutoryWelfarePath in objNewRawdataStep0019StatutoryWelfarePaths:
            if objNewRawdataStep0019StatutoryWelfarePath.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0020_statutory_welfare_from_step0019_statutory_welfare(
                    objNewRawdataStep0019StatutoryWelfarePath
                )
                objHandledInputPaths.add(objNewRawdataStep0019StatutoryWelfarePath.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0020 from step0019 statutory welfare: {0}. Detail = {1}".format(
                        objNewRawdataStep0019StatutoryWelfarePath,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0020StatutoryWelfarePaths:
        for objNewRawdataStep0020StatutoryWelfarePath in objNewRawdataStep0020StatutoryWelfarePaths:
            if objNewRawdataStep0020StatutoryWelfarePath.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0021_statutory_welfare_from_step0020_statutory_welfare(
                    objNewRawdataStep0020StatutoryWelfarePath
                )
                objHandledInputPaths.add(objNewRawdataStep0020StatutoryWelfarePath.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0021 from step0020 statutory welfare: {0}. Detail = {1}".format(
                        objNewRawdataStep0020StatutoryWelfarePath,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0021StatutoryWelfarePaths:
        for objNewRawdataStep0021StatutoryWelfarePath in objNewRawdataStep0021StatutoryWelfarePaths:
            if objNewRawdataStep0021StatutoryWelfarePath.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0022_statutory_welfare_from_step0021_statutory_welfare_with_org_table(
                    objNewRawdataStep0021StatutoryWelfarePath
                )
                objHandledInputPaths.add(objNewRawdataStep0021StatutoryWelfarePath.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0022 from step0021 statutory welfare: {0}. Detail = {1}".format(
                        objNewRawdataStep0021StatutoryWelfarePath,
                        objException,
                    )
                )
                iExitCode = 1

    if objNewRawdataStep0022StatutoryWelfarePaths:
        for objNewRawdataStep0022StatutoryWelfarePath in objNewRawdataStep0022StatutoryWelfarePaths:
            if objNewRawdataStep0022StatutoryWelfarePath.resolve() in objHandledInputPaths:
                continue
            try:
                process_new_rawdata_step0023_statutory_welfare_from_step0022_statutory_welfare(
                    objNewRawdataStep0022StatutoryWelfarePath
                )
                objHandledInputPaths.add(objNewRawdataStep0022StatutoryWelfarePath.resolve())
            except Exception as objException:
                print(
                    "Error: failed to process step0023 from step0022 statutory welfare: {0}. Detail = {1}".format(
                        objNewRawdataStep0022StatutoryWelfarePath,
                        objException,
                    )
                )
                iExitCode = 1


    for pszInputXlsxPath in objArgs.pszInputXlsxPaths:
        try:
            objResolvedInputPath: Path = resolve_existing_input_path(pszInputXlsxPath)
            if objResolvedInputPath.resolve() in objHandledInputPaths:
                continue
            if SALARY_PAYMENT_STEP0001_FILE_PATTERN.match(objResolvedInputPath.name) is not None:
                continue
            process_single_input(pszInputXlsxPath)
        except Exception as objException:
            print(
                "Error: failed to process input file: {0}. Detail = {1}".format(
                    pszInputXlsxPath,
                    objException,
                )
            )
            iExitCode = 1
            continue

    return iExitCode


if __name__ == "__main__":
    raise SystemExit(main())
